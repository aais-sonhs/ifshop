"""
Centralized Excel export utility for ifshop.
Usage: from core.excel_export import excel_response
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, date
from decimal import Decimal


# ===== STYLE PRESETS =====
HEADER_FILL = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1a73e8')
SUBTITLE_FONT = Font(name='Arial', italic=True, size=10, color='666666')
DATA_FONT = Font(name='Arial', size=10)
MONEY_FONT = Font(name='Arial', size=10, bold=True, color='28a745')
TOTAL_FILL = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
TOTAL_FONT = Font(name='Arial', bold=True, size=11, color='1b5e20')
THIN_BORDER = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)
ALT_FILL = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')


def _format_value(val):
    """Convert Python values to Excel-friendly format."""
    if val is None:
        return ''
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, (date, datetime)):
        return val.strftime('%d/%m/%Y')
    return val


def excel_response(title, subtitle, columns, rows, filename,
                   money_cols=None, total_row=None):
    """
    Build a styled Excel response.

    Args:
        title: Sheet title (e.g. 'DANH SÁCH PHIẾU THU')
        subtitle: Sub-info (e.g. 'Xuất ngày 14/04/2026')
        columns: list of {'key': str, 'label': str, 'width': int}
        rows: list of dicts matching column keys
        filename: download filename (without .xlsx)
        money_cols: list of column keys that should be formatted as money
        total_row: dict of {col_key: value} for a totals row at the bottom

    Returns:
        HttpResponse with xlsx attachment.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit

    money_cols = money_cols or []
    num_cols = len(columns)

    # ----- TITLE ROW -----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ----- SUBTITLE ROW -----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    # ----- HEADER ROW -----
    header_row = 4
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=col['label'])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col.get('width', 15)
    ws.row_dimensions[header_row].height = 28

    # ----- DATA ROWS -----
    for ri, row_data in enumerate(rows):
        excel_row = header_row + 1 + ri
        for ci, col in enumerate(columns, 1):
            val = _format_value(row_data.get(col['key'], ''))
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Money formatting
            if col['key'] in money_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.font = MONEY_FONT
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # Zebra striping
            if ri % 2 == 1:
                cell.fill = ALT_FILL

    # ----- TOTAL ROW -----
    if total_row:
        total_excel_row = header_row + 1 + len(rows)
        for ci, col in enumerate(columns, 1):
            val = total_row.get(col['key'], '')
            if val:
                val = _format_value(val)
            cell = ws.cell(row=total_excel_row, column=ci, value=val or '')
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            if col['key'] in money_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # ----- FREEZE & FILTER -----
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(num_cols)}{header_row + len(rows)}'
    ws.freeze_panes = f'A{header_row + 1}'

    # ----- RESPONSE -----
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_filename = filename.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}.xlsx"'
    wb.save(response)
    return response
