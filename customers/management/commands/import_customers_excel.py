"""Import khách hàng và các địa chỉ nhận hàng từ file Excel dạng Sapo."""

import re
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from customers.models import Customer, CustomerAddress, CustomerGroup
from system_management.models import Store


EXPECTED_HEADERS = (
    'Tên khách hàng *',
    'Mã khách hàng',
    'Mã nhóm khách hàng',
    'Điện thoại',
    'Giới tính',
    'Địa chỉ',
    'Tỉnh thành',
    'Quận huyện',
    'Phường xã',
    'SL đơn hàng',
    'Tổng SL sản phẩm đã mua',
    'Tổng SL sản phẩm hoàn trả',
    'Ngày mua cuối cùng',
    'Địa chỉ - SĐT',
    'Người liên hệ - SĐT',
)

GENDER_MAP = {
    'khác': 0,
    'nam': 1,
    'nữ': 2,
}

CUSTOMER_KIND_BY_GROUP = {
    'BANLE': Customer.CUSTOMER_KIND_RETAIL,
    'BANBUON': Customer.CUSTOMER_KIND_WHOLESALE,
}


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_text(value):
    return re.sub(r'\s+', ' ', _cell_text(value)).strip().casefold()


def _phone_key(value):
    raw_value = _cell_text(value)
    digits = re.sub(r'\D+', '', raw_value)
    return digits or raw_value.casefold()


def _address_key(address, phone):
    return (_normalized_text(address), _phone_key(phone))


def _location_core(value):
    value = _normalized_text(value)
    return re.sub(
        r'^(?:tỉnh|thành phố|tp\.?|quận|huyện|phường|xã|thị trấn|thị xã)\s+',
        '',
        value,
    )


def _full_delivery_address(address, province='', district='', ward=''):
    """Ghép địa chỉ phụ với địa giới, tránh lặp lại thành phần đã có."""
    parts = []
    searchable = ''
    for value in (address, ward, district, province):
        value = _cell_text(value)
        location_key = _location_core(value)
        if not value or (location_key and location_key in searchable):
            continue
        parts.append(value)
        searchable = _normalized_text(', '.join(parts))
    return ', '.join(parts)


def _decimal_value(value, row_number, field_name):
    if value in (None, ''):
        return Decimal('0')
    try:
        return Decimal(str(value).replace(',', '').strip())
    except (InvalidOperation, ValueError, AttributeError) as exc:
        raise CommandError(
            f'Dòng {row_number}: {field_name} không phải là số hợp lệ: {value!r}'
        ) from exc


def _integer_value(value, row_number, field_name):
    parsed = _decimal_value(value, row_number, field_name)
    if parsed != parsed.to_integral_value():
        raise CommandError(f'Dòng {row_number}: {field_name} phải là số nguyên: {value!r}')
    return int(parsed)


def _datetime_value(value, row_number):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value
    text_value = _cell_text(value)
    for date_format in ('%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try:
            return datetime.strptime(text_value, date_format)
        except ValueError:
            continue
    raise CommandError(f'Dòng {row_number}: Ngày mua cuối cùng không hợp lệ: {value!r}')


class Command(BaseCommand):
    help = (
        'Import khách hàng từ Excel. Mặc định chỉ xem trước; thêm --apply để ghi dữ liệu. '
        'Dữ liệu đang có được giữ lại, địa chỉ từ file được gộp và chống trùng.'
    )

    def add_arguments(self, parser):
        parser.add_argument('excel_file', help='Đường dẫn file Excel khách hàng.')
        parser.add_argument('--store-id', type=int, required=True, help='ID cửa hàng nhận dữ liệu.')
        parser.add_argument('--user-id', type=int, required=True, help='ID người tạo khách hàng mới.')
        parser.add_argument(
            '--apply',
            action='store_true',
            default=False,
            help='Thực sự ghi dữ liệu. Không truyền cờ này sẽ chỉ chạy thử.',
        )

    def _read_records(self, excel_path):
        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError(f'Không thể đọc file Excel: {exc}') from exc

        worksheet = workbook.active
        headers = tuple(_cell_text(cell.value) for cell in worksheet[1])
        if headers[:len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
            raise CommandError(
                'Cấu trúc cột không đúng mẫu khách hàng. '
                f'Cần: {", ".join(EXPECTED_HEADERS)}'
            )

        records = []
        current_record = None
        seen_codes = set()
        continuation_count = 0

        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
            start=2,
        ):
            name = _cell_text(values[0])
            address = _cell_text(values[5])

            if not name:
                if not address:
                    continue
                if current_record is None:
                    raise CommandError(
                        f'Dòng {row_number}: có địa chỉ phụ nhưng chưa có khách hàng đứng trước.'
                    )
                current_record['delivery_addresses'].append({
                    'row_number': row_number,
                    'address': address,
                    'province': _cell_text(values[6]),
                    'district': _cell_text(values[7]),
                    'ward': _cell_text(values[8]),
                    'phone': _cell_text(values[13]),
                })
                continuation_count += 1
                continue

            code = _cell_text(values[1])
            if not code:
                raise CommandError(f'Dòng {row_number}: khách hàng {name!r} chưa có mã.')
            if code in seen_codes:
                raise CommandError(f'Dòng {row_number}: mã khách hàng {code!r} bị trùng trong file.')
            if len(code) > Customer._meta.get_field('code').max_length:
                raise CommandError(f'Dòng {row_number}: mã khách hàng {code!r} quá dài.')
            if len(name) > Customer._meta.get_field('name').max_length:
                raise CommandError(f'Dòng {row_number}: tên khách hàng {name!r} quá dài.')

            gender_text = _normalized_text(values[4])
            if gender_text not in GENDER_MAP:
                raise CommandError(f'Dòng {row_number}: giới tính không hợp lệ: {values[4]!r}')

            current_record = {
                'row_number': row_number,
                'name': name,
                'code': code,
                'group_code': _cell_text(values[2]).upper(),
                'phone': _cell_text(values[3]),
                'gender': GENDER_MAP[gender_text],
                'address': address,
                'province': _cell_text(values[6]),
                'district': _cell_text(values[7]),
                'ward': _cell_text(values[8]),
                'order_count': _integer_value(values[9], row_number, 'SL đơn hàng'),
                'total_product_quantity': _decimal_value(
                    values[10], row_number, 'Tổng SL sản phẩm đã mua'
                ),
                'total_returned_product_quantity': _decimal_value(
                    values[11], row_number, 'Tổng SL sản phẩm hoàn trả'
                ),
                'last_purchase_at': _datetime_value(values[12], row_number),
                'address_phone': _cell_text(values[13]),
                'contact_phone': _cell_text(values[14]),
                'delivery_addresses': [],
            }
            records.append(current_record)
            seen_codes.add(code)

        if not records:
            raise CommandError('File không có khách hàng để import.')
        return records, continuation_count

    @staticmethod
    def _candidate_addresses(record, customer, original_address, existing_addresses):
        candidates = []
        known_keys = {
            _address_key(customer.address, customer.phone),
            *(_address_key(item.address, item.phone) for item in existing_addresses),
        }

        primary_full_address = _full_delivery_address(
            record['address'], record['province'], record['district'], record['ward']
        )
        primary_phone = record['address_phone'] or record['phone']
        primary_file_key = _address_key(primary_full_address, primary_phone)

        main_address_changed = (
            bool(original_address)
            and bool(record['address'])
            and _normalized_text(original_address) != _normalized_text(record['address'])
        )
        address_phone_changed = (
            bool(record['address_phone'])
            and _phone_key(record['address_phone']) != _phone_key(customer.phone)
        )
        if primary_full_address and (main_address_changed or address_phone_changed):
            if primary_file_key not in known_keys:
                candidates.append({
                    'label': 'Địa chỉ chính từ Excel',
                    'address': primary_full_address,
                    'phone': primary_phone[:CustomerAddress._meta.get_field('phone').max_length],
                })
                known_keys.add(primary_file_key)

        # Dùng địa chỉ chính trong file làm mốc chống lặp cho các dòng phụ,
        # kể cả khi địa chỉ đó không cần tạo thêm một CustomerAddress.
        if primary_full_address:
            known_keys.add(primary_file_key)

        for index, item in enumerate(record['delivery_addresses'], start=2):
            full_address = _full_delivery_address(
                item['address'], item['province'], item['district'], item['ward']
            )
            phone = item['phone'][:CustomerAddress._meta.get_field('phone').max_length]
            key = _address_key(full_address, phone)
            if not full_address or key in known_keys:
                continue
            candidates.append({
                'label': f'Địa chỉ phụ từ Excel {index}',
                'address': full_address,
                'phone': phone,
            })
            known_keys.add(key)

        return candidates

    def handle(self, *args, **options):
        excel_path = Path(options['excel_file']).expanduser().resolve()
        if not excel_path.is_file():
            raise CommandError(f'Không tìm thấy file: {excel_path}')

        try:
            store = Store.objects.get(id=options['store_id'])
        except Store.DoesNotExist as exc:
            raise CommandError(f'Không tìm thấy cửa hàng ID {options["store_id"]}.') from exc
        try:
            creator = User.objects.get(id=options['user_id'], is_active=True)
        except User.DoesNotExist as exc:
            raise CommandError(f'Không tìm thấy user hoạt động ID {options["user_id"]}.') from exc

        records, continuation_count = self._read_records(excel_path)
        group_codes = {record['group_code'] for record in records if record['group_code']}
        groups = {
            group.code.upper(): group
            for group in CustomerGroup.objects.filter(code__in=group_codes)
            if group.code
        }
        missing_groups = sorted(group_codes - set(groups))
        if missing_groups:
            raise CommandError(f'Chưa có nhóm khách hàng: {", ".join(missing_groups)}')

        codes = [record['code'] for record in records]
        existing_by_code = {
            customer.code: customer
            for customer in Customer.all_objects.filter(code__in=codes).select_related('store')
        }
        unavailable = [
            customer.code
            for customer in existing_by_code.values()
            if customer.is_deleted or customer.store_id not in (None, store.id)
        ]
        if unavailable:
            raise CommandError(
                'Các mã đã bị xóa hoặc thuộc cửa hàng khác: '
                + ', '.join(sorted(unavailable)[:20])
            )

        addresses_by_customer = defaultdict(list)
        existing_ids = [customer.id for customer in existing_by_code.values()]
        for address in CustomerAddress.objects.filter(customer_id__in=existing_ids).order_by(
            'customer_id', 'sort_order', 'id'
        ):
            addresses_by_customer[address.customer_id].append(address)

        apply_changes = options['apply']
        stats = Counter()
        pending_addresses = []
        phone_limit = Customer._meta.get_field('phone').max_length
        contact_phone_limit = Customer._meta.get_field('contact_phone').max_length

        context = transaction.atomic() if apply_changes else transaction.atomic()
        with context:
            for record in records:
                customer = existing_by_code.get(record['code'])
                is_new = customer is None
                original_address = customer.address if customer else ''
                update_fields = []

                if is_new:
                    customer = Customer(
                        store=store,
                        created_by=creator,
                        code=record['code'],
                        name=record['name'],
                        phone=record['phone'][:phone_limit],
                        gender=record['gender'],
                        address=record['address'],
                        province=record['province'],
                        district=record['district'],
                        ward=record['ward'],
                        contact_phone=record['contact_phone'][:contact_phone_limit],
                        group=groups.get(record['group_code']),
                        customer_kind=CUSTOMER_KIND_BY_GROUP.get(record['group_code'], ''),
                        order_count=record['order_count'],
                        total_product_quantity=record['total_product_quantity'],
                        total_returned_product_quantity=record['total_returned_product_quantity'],
                        last_purchase_at=record['last_purchase_at'],
                        imported_legacy_metrics=True,
                        is_active=True,
                    )
                    stats['customers_created'] += 1
                    if apply_changes:
                        customer.save()
                        existing_by_code[customer.code] = customer
                else:
                    same_main_address = (
                        not original_address
                        or not record['address']
                        or _normalized_text(original_address) == _normalized_text(record['address'])
                    )
                    fill_values = {
                        'phone': record['phone'][:phone_limit],
                        'address': record['address'],
                        'contact_phone': record['contact_phone'][:contact_phone_limit],
                        'last_purchase_at': record['last_purchase_at'],
                    }
                    if same_main_address:
                        fill_values.update({
                            'province': record['province'],
                            'district': record['district'],
                            'ward': record['ward'],
                        })
                    for field_name, file_value in fill_values.items():
                        if getattr(customer, field_name) in (None, '') and file_value not in (None, ''):
                            setattr(customer, field_name, file_value)
                            update_fields.append(field_name)
                    if not customer.group_id and record['group_code']:
                        customer.group = groups[record['group_code']]
                        update_fields.append('group')
                    if not customer.customer_kind and record['group_code'] in CUSTOMER_KIND_BY_GROUP:
                        customer.customer_kind = CUSTOMER_KIND_BY_GROUP[record['group_code']]
                        update_fields.append('customer_kind')

                    if update_fields:
                        stats['customers_updated'] += 1
                        if apply_changes:
                            customer.save(update_fields=[*update_fields, 'updated_at'])
                    else:
                        stats['customers_unchanged'] += 1

                if len(record['phone']) > phone_limit:
                    stats['customer_phones_truncated'] += 1

                current_addresses = addresses_by_customer.get(customer.id, []) if customer.id else []
                address_candidates = self._candidate_addresses(
                    record, customer, original_address, current_addresses
                )
                stats['addresses_created'] += len(address_candidates)
                stats['addresses_skipped'] += (
                    len(record['delivery_addresses'])
                    + int(bool(record['address']) and (
                        _normalized_text(original_address) != _normalized_text(record['address'])
                        or (
                            bool(record['address_phone'])
                            and _phone_key(record['address_phone']) != _phone_key(customer.phone)
                        )
                    ))
                    - len(address_candidates)
                )

                if apply_changes and address_candidates:
                    next_sort_order = max(
                        (address.sort_order for address in current_addresses),
                        default=-1,
                    ) + 1
                    for offset, item in enumerate(address_candidates):
                        pending_addresses.append(CustomerAddress(
                            customer=customer,
                            label=item['label'],
                            address=item['address'],
                            phone=item['phone'],
                            sort_order=next_sort_order + offset,
                        ))

            if apply_changes and pending_addresses:
                CustomerAddress.objects.bulk_create(pending_addresses, batch_size=500)

            if not apply_changes:
                transaction.set_rollback(True)

        mode = 'ĐÃ GHI DỮ LIỆU' if apply_changes else 'CHẠY THỬ - CHƯA GHI DỮ LIỆU'
        self.stdout.write(self.style.SUCCESS(mode) if apply_changes else self.style.WARNING(mode))
        self.stdout.write(f'File: {excel_path.name}')
        self.stdout.write(f'Cửa hàng: {store.code} - {store.name}')
        self.stdout.write(f'Khách hàng trong file: {len(records):,}')
        self.stdout.write(f'Dòng địa chỉ phụ trong file: {continuation_count:,}')
        self.stdout.write(f'Khách hàng mới: {stats["customers_created"]:,}')
        self.stdout.write(f'Khách hàng được bổ sung trường trống: {stats["customers_updated"]:,}')
        self.stdout.write(f'Khách hàng giữ nguyên: {stats["customers_unchanged"]:,}')
        self.stdout.write(f'Địa chỉ nhận hàng thêm mới: {stats["addresses_created"]:,}')
        self.stdout.write(f'Địa chỉ trùng được bỏ qua: {max(stats["addresses_skipped"], 0):,}')
        if stats['customer_phones_truncated']:
            self.stdout.write(self.style.WARNING(
                f'{stats["customer_phones_truncated"]:,} SĐT khách dài hơn {phone_limit} ký tự; '
                'bản đầy đủ vẫn được giữ trong SĐT của địa chỉ nhận hàng khi file có địa chỉ.'
            ))
