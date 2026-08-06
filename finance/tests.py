import json
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.core import mail
from django.db.models import Sum
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from customers.models import Customer
from finance.models import (
    CashBook,
    ExpenseClassification,
    FinanceCategory,
    FinancialPlan,
    FinancialPlanAllocation,
    FinancialPlanItem,
    FinancialPlanRevision,
    Payment,
    PaymentMethodOption,
    Receipt,
    SupplierPaymentSchedule,
)
from finance.financial_alerts import run_due_financial_alerts
from orders.models import Order
from products.models import GoodsReceipt, PurchaseReturn, Supplier, Warehouse
from system_management.models import Brand, ModulePermission, RoleGroup, Store, UserProfile


class FinanceFlowTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.brand = Brand.objects.create(name='Finance Brand')
        cls.store = Store.objects.create(brand=cls.brand, name='Finance Store A', code='FSA')
        cls.other_store = Store.objects.create(brand=cls.brand, name='Finance Store B', code='FSB')

        cls.user = User.objects.create_user(username='finance_a', password='pass123')
        cls.other_user = User.objects.create_user(username='finance_b', password='pass123')
        UserProfile.objects.create(user=cls.user, store=cls.store)
        UserProfile.objects.create(user=cls.other_user, store=cls.other_store)

        cls.customer = Customer.objects.create(
            store=cls.store,
            code='FKH001',
            name='Finance Customer A',
            created_by=cls.user,
        )
        cls.other_customer = Customer.objects.create(
            store=cls.other_store,
            code='FKH002',
            name='Finance Customer B',
            created_by=cls.other_user,
        )

        cls.supplier = Supplier.objects.create(
            code='NCC001',
            name='Supplier A',
            created_by=cls.user,
        )
        cls.warehouse = Warehouse.objects.create(store=cls.store, code='FKHO-A', name='Kho Finance A')
        cls.other_warehouse = Warehouse.objects.create(
            store=cls.other_store,
            code='FKHO-B',
            name='Kho Finance B',
        )

    def setUp(self):
        self.client.force_login(self.user)

    def _create_order(self, code, store=None, customer=None, warehouse=None, created_by=None):
        return Order.objects.create(
            code=code,
            store=store or self.store,
            customer=customer or self.customer,
            warehouse=warehouse or self.warehouse,
            total_amount=100,
            final_amount=100,
            order_date=date.today(),
            created_by=created_by or self.user,
        )

    def _create_goods_receipt(self, code, store=None, supplier=None, warehouse=None, created_by=None):
        return GoodsReceipt.objects.create(
            code=code,
            supplier=supplier or self.supplier,
            warehouse=warehouse or self.warehouse,
            total_amount=100,
            receipt_date=date.today(),
            created_by=created_by or self.user,
        )

    def test_delete_payment_refunds_cashbook_balance(self):
        cash_book = CashBook.objects.create(name='Quỹ A', balance=Decimal('800'))
        payment = Payment.objects.create(
            code='PC-001',
            store=self.store,
            cash_book=cash_book,
            amount=Decimal('200'),
            payment_date=date.today(),
            status=1,
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_delete_payment'),
            data=json.dumps({'id': payment.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())

        cash_book.refresh_from_db()
        deleted_payment = Payment.all_objects.get(id=payment.id)
        self.assertEqual(cash_book.balance, Decimal('1000'))
        self.assertTrue(deleted_payment.is_deleted)

    def test_payment_list_returns_paginated_meta_when_requested(self):
        today = date.today()
        for index in range(11):
            Payment.objects.create(
                code=f'PC-PAGE-{index:02d}',
                store=self.store,
                amount=Decimal('100'),
                payment_date=today - timedelta(days=10 - index),
                status=1,
                created_by=self.user,
            )

        response = self.client.get(
            reverse('api_get_payments'),
            data={'page': 2, 'page_size': 10},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['page'], 2)
        self.assertEqual(payload['meta']['page_size'], 10)
        self.assertEqual(payload['meta']['page_count'], 1)
        self.assertEqual(payload['meta']['total_pages'], 2)
        self.assertEqual(payload['meta']['total_filtered_count'], 11)
        self.assertEqual(payload['meta']['start_index'], 11)
        self.assertEqual(payload['meta']['end_index'], 11)
        self.assertFalse(payload['meta']['has_next'])
        self.assertEqual([item['code'] for item in payload['data']], ['PC-PAGE-00'])
        self.assertEqual(payload['meta']['next_code'], 'PC-001')

    def test_payment_list_without_pagination_keeps_legacy_full_response(self):
        for index in range(11):
            Payment.objects.create(
                code=f'PC-LEGACY-{index:02d}',
                store=self.store,
                amount=Decimal('100'),
                payment_date=date.today() - timedelta(days=index),
                status=1,
                created_by=self.user,
            )

        response = self.client.get(reverse('api_get_payments'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertNotIn('meta', payload)
        self.assertEqual(len(payload['data']), 11)
        self.assertEqual(payload['next_code'], 'PC-001')

    def test_draft_linked_payment_is_displayed_from_current_goods_receipt_total(self):
        goods_receipt = GoodsReceipt.objects.create(
            code='PN-DRAFT-RECALCULATE',
            supplier=self.supplier,
            warehouse=self.warehouse,
            total_amount=Decimal('1000'),
            receipt_date=date.today(),
            status=1,
            created_by=self.user,
        )
        payment = Payment.objects.create(
            code='PC-DRAFT-RECALCULATE',
            store=self.store,
            goods_receipt=goods_receipt,
            amount=Decimal('333'),  # Dữ liệu Nháp cũ đã chỉnh tay.
            promotion_mode='amount',
            promotion_amount=Decimal('100'),
            promotion_percent=Decimal('0'),
            payment_date=date.today(),
            status=0,
            created_by=self.user,
        )

        payload = self.client.get(reverse('api_get_payments')).json()['data']
        row = next(item for item in payload if item['id'] == payment.id)

        self.assertEqual(row['gross_amount'], 1000.0)
        self.assertEqual(row['promotion_amount'], 100.0)
        self.assertEqual(row['promotion_percent'], 10.0)
        self.assertEqual(row['amount'], 900.0)

    def test_payment_list_sorts_by_payment_date_in_requested_direction(self):
        today = date.today()
        for index, days_ago in enumerate((1, 3, 2), start=1):
            Payment.objects.create(
                code=f'PC-SORT-{index}',
                store=self.store,
                amount=Decimal('100'),
                payment_date=today - timedelta(days=days_ago),
                status=0,
                created_by=self.user,
            )

        ascending_response = self.client.get(
            reverse('api_get_payments'),
            data={'payment_date_order': 'asc'},
        )
        descending_response = self.client.get(
            reverse('api_get_payments'),
            data={'payment_date_order': 'desc'},
        )

        self.assertEqual(
            [item['code'] for item in ascending_response.json()['data']],
            ['PC-SORT-2', 'PC-SORT-3', 'PC-SORT-1'],
        )
        self.assertEqual(
            [item['code'] for item in descending_response.json()['data']],
            ['PC-SORT-1', 'PC-SORT-3', 'PC-SORT-2'],
        )

    def test_payment_list_order_stays_stable_after_edit(self):
        payments = [
            Payment.objects.create(
                code=f'PC-STABLE-{index}',
                store=self.store,
                amount=Decimal('100'),
                payment_date=date.today(),
                status=0,
                created_by=self.user,
            )
            for index in range(1, 4)
        ]
        expected_codes = [payment.code for payment in reversed(payments)]

        before_edit = self.client.get(reverse('api_get_payments'))
        self.assertEqual(
            [item['code'] for item in before_edit.json()['data']],
            expected_codes,
        )

        edited_payment = payments[1]
        edit_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'id': edited_payment.id,
                'code': edited_payment.code,
                'amount': 100,
                'payment_date': date.today().isoformat(),
                'status': 0,
                'payment_method': 2,
                'description': 'Sửa nhưng không đổi vị trí',
            }),
            content_type='application/json',
        )
        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())

        after_edit = self.client.get(reverse('api_get_payments'))
        self.assertEqual(
            [item['code'] for item in after_edit.json()['data']],
            expected_codes,
        )

    def test_payment_list_filters_full_queryset_and_export_uses_same_filters(self):
        today = date.today()
        category = FinanceCategory.objects.create(name='Chi phí lọc', type=2)
        other_category = FinanceCategory.objects.create(name='Chi phí khác', type=2)
        cash_book = CashBook.objects.create(name='Quỹ lọc')
        other_cash_book = CashBook.objects.create(name='Quỹ khác')
        bank_method = PaymentMethodOption.objects.create(
            code='PAYMENT_FILTER_BANK',
            name='Ngân hàng lọc',
            legacy_type=2,
        )
        cash_method = PaymentMethodOption.objects.create(
            code='PAYMENT_FILTER_CASH',
            name='Tiền mặt lọc',
            legacy_type=1,
        )
        goods_receipt = self._create_goods_receipt(code='PN-PAYMENT-FILTER')

        Payment.objects.create(
            code='PC-PAYMENT-FILTER-MATCH',
            store=self.store,
            category=category,
            cash_book=cash_book,
            supplier=self.supplier,
            goods_receipt=goods_receipt,
            amount=Decimal('500'),
            payment_date=today,
            status=1,
            payment_method=2,
            payment_method_option=bank_method,
            description='Chi phí cần tìm',
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-PAYMENT-FILTER-DRAFT',
            store=self.store,
            category=other_category,
            cash_book=other_cash_book,
            supplier=self.supplier,
            amount=Decimal('100'),
            payment_date=today,
            status=0,
            payment_method=1,
            payment_method_option=cash_method,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-PAYMENT-FILTER-OUTSIDE',
            store=self.store,
            category=category,
            cash_book=cash_book,
            supplier=self.supplier,
            goods_receipt=goods_receipt,
            amount=Decimal('500'),
            payment_date=today - timedelta(days=5),
            status=1,
            payment_method=2,
            payment_method_option=bank_method,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-PAYMENT-FILTER-FOREIGN',
            store=self.other_store,
            category=category,
            cash_book=cash_book,
            supplier=self.supplier,
            amount=Decimal('500'),
            payment_date=today,
            status=1,
            payment_method=2,
            payment_method_option=bank_method,
            created_by=self.other_user,
        )

        response = self.client.get(reverse('api_get_payments'), data={
            'search': 'FILTER-MATCH',
            'date_from': today.isoformat(),
            'date_to': today.isoformat(),
            'status': '1',
            'category_id': category.id,
            'supplier_id': self.supplier.id,
            'cash_book_id': cash_book.id,
            'payment_method_option_id': bank_method.id,
            'payment_type': '2',
            'goods_receipt_state': 'yes',
            'amount_from': '400',
            'amount_to': '600',
            'store_id': self.store.id,
            'page': 1,
            'page_size': 10,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([item['code'] for item in payload['data']], ['PC-PAYMENT-FILTER-MATCH'])
        self.assertEqual(payload['meta']['total_filtered_count'], 1)
        self.assertEqual(payload['meta']['total_all_count'], 3)

        draft_response = self.client.get(reverse('api_get_payments'), data={
            'status': '0',
            'goods_receipt_state': 'no',
            'page': 1,
            'page_size': 10,
        })
        self.assertEqual(
            [item['code'] for item in draft_response.json()['data']],
            ['PC-PAYMENT-FILTER-DRAFT'],
        )

        excel_response = self.client.get(reverse('export_payments_excel'), data={
            'status': '0',
            'goods_receipt_state': 'no',
        })
        self.assertEqual(excel_response.status_code, 200)
        worksheet = load_workbook(BytesIO(excel_response.content), data_only=True).active
        headers = [cell.value for cell in worksheet[4]]
        self.assertIn('Phiếu nhập', headers)
        self.assertIn('Trạng thái', headers)
        self.assertIn('Người duyệt', headers)
        self.assertIn('Thời gian duyệt', headers)
        code_column = headers.index('Mã phiếu') + 1
        status_column = headers.index('Trạng thái') + 1
        self.assertEqual(worksheet.cell(row=5, column=code_column).value, 'PC-PAYMENT-FILTER-DRAFT')
        self.assertEqual(worksheet.cell(row=5, column=status_column).value, 'Nháp')
        self.assertNotIn(
            'PC-PAYMENT-FILTER-MATCH',
            [worksheet.cell(row=row, column=code_column).value for row in range(5, worksheet.max_row + 1)],
        )

    def test_payment_page_exposes_filter_controls(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        response = self.client.get(reverse('payment_tbl'), {
            'date_from': '2026-07-01',
            'date_to': '2026-07-20',
            'status': '1',
            'store_id': self.store.id,
        })

        self.assertEqual(response.status_code, 200)
        for control_id in [
            'payment_filters',
            'f_search',
            'f_date_from',
            'f_date_to',
            'f_status',
            'f_category',
            'f_expense_classification',
            'f_supplier',
            'f_cashbook',
            'f_method',
            'f_goods_receipt_state',
            'f_payment_type',
            'f_amount_from',
            'f_amount_to',
            'f_store',
            'btn_apply_filters',
            'btn_clear_filters',
        ]:
            self.assertContains(response, f'id="{control_id}"')
        self.assertContains(response, 'buildPaymentExportUrl')
        self.assertContains(response, 'var PAYMENT_URL_FILTERS')
        self.assertContains(response, "params.get('date_from')")
        self.assertContains(response, "params.get('date_to')")
        self.assertContains(response, "params.get('status')")
        self.assertContains(response, "params.get('store_id')")
        self.assertContains(response, 'applyPaymentUrlFilters();')
        self.assertNotContains(response, '<th>Người tạo</th>', html=True)
        self.assertContains(response, '<th>Người duyệt</th>', html=True)
        self.assertNotContains(response, "d.created_by || ''")
        self.assertContains(response, "d.approved_by || ''")
        self.assertContains(response, 'Tiền phiếu chi')
        self.assertContains(response, 'Còn phải chi cho phiếu nhập')
        self.assertContains(response, 'Khuyến mãi (% / tiền)')
        self.assertContains(response, 'id="inp_promotion_mode"')
        self.assertContains(response, 'id="inp_promotion_amount"')
        self.assertContains(response, 'id="inp_promotion_percent"')
        self.assertContains(response, 'id="inp_actual_amount"')
        self.assertContains(response, 'function syncPaymentPromotion()')
        self.assertContains(response, 'function setPaymentGrossInputMode(isLinkedReceipt)')
        self.assertContains(response, ".prop('readonly', !!isLinkedReceipt)")
        self.assertContains(response, 'Tự lấy giá trị phiếu nhập còn lại sau hàng trả')
        self.assertContains(response, 'id="inp_expense_classification_id"')
        self.assertContains(response, "$('#inp_status').val('1')")
        self.assertContains(response, 'btn-approve-payment')
        self.assertContains(response, '/api/payments/approve/')
        self.assertContains(response, 'd-inline-flex align-items-center flex-nowrap text-nowrap')
        self.assertContains(response, 'btn-action-group d-inline-flex flex-nowrap')
        self.assertContains(response, 'style="gap:6px;"')

    def test_quick_approve_payment_uses_priority_defaults_and_only_deducts_once(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        later_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi ưu tiên sau khi duyệt',
            type=2,
            sort_order=20,
        )
        priority_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi ưu tiên khi duyệt',
            type=2,
            sort_order=2,
        )
        later_classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=priority_category,
            name='Phân loại ưu tiên sau khi duyệt',
            sort_order=30,
        )
        priority_classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=priority_category,
            name='Phân loại ưu tiên khi duyệt',
            sort_order=3,
        )
        inactive_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ ngừng sử dụng',
            sort_order=0,
            balance=Decimal('5000'),
            is_active=False,
        )
        later_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ ưu tiên sau khi duyệt',
            sort_order=40,
            balance=Decimal('5000'),
        )
        priority_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ ưu tiên khi duyệt',
            sort_order=4,
            balance=Decimal('5000'),
        )
        payment = Payment.objects.create(
            code='PC-DUYET-UU-TIEN-001',
            store=self.store,
            amount=Decimal('1200'),
            payment_date=date.today(),
            status=0,
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_approve_payment'),
            data=json.dumps({'id': payment.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        payment.refresh_from_db()
        priority_cashbook.refresh_from_db()
        later_cashbook.refresh_from_db()
        inactive_cashbook.refresh_from_db()
        self.assertEqual(payment.status, 1)
        self.assertEqual(payment.category, priority_category)
        self.assertEqual(payment.expense_classification, priority_classification)
        self.assertEqual(payment.cash_book, priority_cashbook)
        self.assertEqual(payment.approved_by, self.user)
        self.assertIsNotNone(payment.approved_at)
        self.assertEqual(priority_cashbook.balance, Decimal('3800'))
        self.assertEqual(later_cashbook.balance, Decimal('5000'))
        self.assertEqual(inactive_cashbook.balance, Decimal('5000'))
        self.assertNotEqual(payment.category, later_category)
        self.assertNotEqual(payment.expense_classification, later_classification)

        second_response = self.client.post(
            reverse('api_approve_payment'),
            data=json.dumps({'id': payment.id}),
            content_type='application/json',
        )
        self.assertEqual(second_response.json()['status'], 'error')
        self.assertIn('Nháp', second_response.json()['message'])
        priority_cashbook.refresh_from_db()
        self.assertEqual(priority_cashbook.balance, Decimal('3800'))

    def test_quick_approve_preserves_choices_and_recalculates_linked_receipt_amount(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        priority_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi mặc định',
            type=2,
            sort_order=1,
        )
        selected_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi đã chọn',
            type=2,
            sort_order=20,
        )
        selected_classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=selected_category,
            name='Phân loại đã chọn',
            sort_order=20,
        )
        priority_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ mặc định',
            sort_order=1,
            balance=Decimal('5000'),
        )
        selected_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ đã chọn',
            sort_order=20,
            balance=Decimal('5000'),
        )
        goods_receipt = self._create_goods_receipt('PN-DUYET-NHANH-001')
        goods_receipt.total_amount = Decimal('1500')
        goods_receipt.save(update_fields=['total_amount'])
        payment = Payment.objects.create(
            code='PC-DUYET-NHANH-LINKED-001',
            store=self.store,
            category=selected_category,
            expense_classification=selected_classification,
            cash_book=selected_cashbook,
            supplier=self.supplier,
            goods_receipt=goods_receipt,
            amount=Decimal('900'),
            promotion_mode='amount',
            promotion_amount=Decimal('100'),
            payment_date=date.today(),
            status=0,
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_approve_payment'),
            data=json.dumps({'id': payment.id}),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        payment.refresh_from_db()
        selected_cashbook.refresh_from_db()
        priority_cashbook.refresh_from_db()
        self.assertEqual(payment.category, selected_category)
        self.assertEqual(payment.expense_classification, selected_classification)
        self.assertEqual(payment.cash_book, selected_cashbook)
        self.assertNotEqual(payment.category, priority_category)
        self.assertEqual(payment.amount, Decimal('1400'))
        self.assertEqual(selected_cashbook.balance, Decimal('3600'))
        self.assertEqual(priority_cashbook.balance, Decimal('5000'))

    def test_finance_category_delete_and_deactivate_flow(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        page_response = self.client.get(reverse('category_tbl'))
        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, 'id="fc_is_active"')
        self.assertContains(page_response, 'deleteFinanceCategory')
        self.assertContains(page_response, 'Ngừng sử dụng')

        unused_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Danh mục chưa phát sinh',
            type=2,
        )
        delete_unused_response = self.client.post(
            reverse('api_delete_finance_category'),
            data=json.dumps({'id': unused_category.id}),
            content_type='application/json',
        )
        self.assertEqual(delete_unused_response.json()['status'], 'ok')
        self.assertFalse(
            FinanceCategory.objects.filter(id=unused_category.id).exists(),
        )
        self.assertTrue(
            FinanceCategory.all_objects.filter(id=unused_category.id).exists(),
        )

        used_expense_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Danh mục chi đã phát sinh',
            type=2,
        )
        payment = Payment.objects.create(
            code='PC-DANH-MUC-DA-DUNG',
            store=self.store,
            category=used_expense_category,
            amount=Decimal('100000'),
            payment_date=date.today(),
            created_by=self.user,
        )
        payment.delete()
        ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=used_expense_category,
            name='Phân loại con đã dùng',
        )
        plan = FinancialPlan.objects.create(
            code='KH-DANH-MUC-DA-DUNG',
            name='Kế hoạch kiểm tra xóa danh mục',
            brand=self.brand,
            store=self.store,
            start_date=date.today(),
            end_date=date.today(),
            created_by=self.user,
        )
        plan_item = FinancialPlanItem.objects.create(
            plan=plan,
            direction=2,
            category=used_expense_category,
            planned_amount=Decimal('100000'),
        )
        plan_item.delete()

        delete_used_expense_response = self.client.post(
            reverse('api_delete_finance_category'),
            data=json.dumps({'id': used_expense_category.id}),
            content_type='application/json',
        )
        self.assertEqual(delete_used_expense_response.json()['status'], 'error')
        self.assertIn('phiếu chi', delete_used_expense_response.json()['message'])
        self.assertIn('phân loại chi', delete_used_expense_response.json()['message'])
        self.assertIn('khoản kế hoạch', delete_used_expense_response.json()['message'])
        self.assertIn('Ngừng sử dụng', delete_used_expense_response.json()['message'])

        deactivate_response = self.client.post(
            reverse('api_save_finance_category'),
            data=json.dumps({
                'id': used_expense_category.id,
                'name': used_expense_category.name,
                'type': used_expense_category.type,
                'description': used_expense_category.description,
                'is_active': False,
            }),
            content_type='application/json',
        )
        self.assertEqual(deactivate_response.json()['status'], 'ok')
        used_expense_category.refresh_from_db()
        self.assertFalse(used_expense_category.is_active)

        used_income_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Danh mục thu đã phát sinh',
            type=1,
        )
        receipt = Receipt.objects.create(
            code='PT-DANH-MUC-DA-DUNG',
            store=self.store,
            category=used_income_category,
            amount=Decimal('100000'),
            receipt_date=date.today(),
            created_by=self.user,
        )
        receipt.delete()
        delete_used_income_response = self.client.post(
            reverse('api_delete_finance_category'),
            data=json.dumps({'id': used_income_category.id}),
            content_type='application/json',
        )
        self.assertEqual(delete_used_income_response.json()['status'], 'error')
        self.assertIn('phiếu thu', delete_used_income_response.json()['message'])

    def test_finance_master_priority_order(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        page_response = self.client.get(reverse('category_tbl'))
        self.assertEqual(page_response.status_code, 200)
        for control_id in ('fc_sort_order', 'cb_sort_order', 'ec_sort_order'):
            self.assertContains(page_response, f'id="{control_id}"')
        self.assertContains(page_response, 'Thứ tự ưu tiên')
        self.assertContains(page_response, 'Số nhỏ hiển thị trước')
        self.assertContains(page_response, 'finance-master-sort-order')
        self.assertContains(page_response, "financeMasterSortOrderInput('finance_category', d)")
        self.assertContains(page_response, "financeMasterSortOrderInput('expense_classification', d)")
        self.assertContains(page_response, "financeMasterSortOrderInput('cashbook', d)")
        self.assertContains(page_response, reverse('api_update_finance_master_sort_order'))

        later_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi ưu tiên sau',
            type=2,
            sort_order=20,
        )
        earlier_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi ưu tiên trước',
            type=2,
            sort_order=5,
        )
        category_response = self.client.get(reverse('api_get_finance_categories'))
        category_rows = [
            row for row in category_response.json()['data']
            if row['id'] in {later_category.id, earlier_category.id}
        ]
        self.assertEqual(
            [row['id'] for row in category_rows],
            [earlier_category.id, later_category.id],
        )
        self.assertEqual([row['sort_order'] for row in category_rows], [5, 20])

        update_category_response = self.client.post(
            reverse('api_save_finance_category'),
            data=json.dumps({
                'id': later_category.id,
                'name': later_category.name,
                'type': later_category.type,
                'description': '',
                'sort_order': '2',
                'is_active': True,
            }),
            content_type='application/json',
        )
        self.assertEqual(update_category_response.json()['status'], 'ok')
        later_category.refresh_from_db()
        self.assertEqual(later_category.sort_order, 2)

        later_classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=earlier_category,
            name='Phân loại ưu tiên sau',
            sort_order=30,
        )
        earlier_classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=earlier_category,
            name='Phân loại ưu tiên trước',
            sort_order=3,
        )
        classification_response = self.client.get(
            reverse('api_get_expense_classifications'),
        )
        classification_rows = [
            row for row in classification_response.json()['data']
            if row['id'] in {later_classification.id, earlier_classification.id}
        ]
        self.assertEqual(
            [row['id'] for row in classification_rows],
            [earlier_classification.id, later_classification.id],
        )
        self.assertEqual(
            [row['sort_order'] for row in classification_rows],
            [3, 30],
        )

        update_classification_response = self.client.post(
            reverse('api_save_expense_classification'),
            data=json.dumps({
                'id': later_classification.id,
                'name': later_classification.name,
                'parent_category_id': earlier_category.id,
                'description': '',
                'sort_order': '1',
                'is_active': True,
            }),
            content_type='application/json',
        )
        self.assertEqual(update_classification_response.json()['status'], 'ok')
        later_classification.refresh_from_db()
        self.assertEqual(later_classification.sort_order, 1)

        later_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ ưu tiên sau',
            sort_order=40,
        )
        earlier_cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ ưu tiên trước',
            sort_order=4,
        )
        cashbook_response = self.client.get(reverse('api_get_cashbooks'))
        cashbook_rows = [
            row for row in cashbook_response.json()['data']
            if row['id'] in {later_cashbook.id, earlier_cashbook.id}
        ]
        self.assertEqual(
            [row['id'] for row in cashbook_rows],
            [earlier_cashbook.id, later_cashbook.id],
        )
        self.assertEqual([row['sort_order'] for row in cashbook_rows], [4, 40])

        update_cashbook_response = self.client.post(
            reverse('api_save_cashbook'),
            data=json.dumps({
                'id': later_cashbook.id,
                'name': later_cashbook.name,
                'description': '',
                'sort_order': '0',
            }),
            content_type='application/json',
        )
        self.assertEqual(update_cashbook_response.json()['status'], 'ok')
        later_cashbook.refresh_from_db()
        self.assertEqual(later_cashbook.sort_order, 0)

        invalid_priority_response = self.client.post(
            reverse('api_save_cashbook'),
            data=json.dumps({
                'id': earlier_cashbook.id,
                'name': earlier_cashbook.name,
                'sort_order': '-1',
            }),
            content_type='application/json',
        )
        self.assertEqual(invalid_priority_response.json()['status'], 'error')
        self.assertIn('từ 0 đến 9999', invalid_priority_response.json()['message'])
        earlier_cashbook.refresh_from_db()
        self.assertEqual(earlier_cashbook.sort_order, 4)

        payment_page_response = self.client.get(reverse('payment_tbl'))
        payment_page_html = payment_page_response.content.decode()
        self.assertLess(
            payment_page_html.index(later_category.name),
            payment_page_html.index(earlier_category.name),
        )
        self.assertLess(
            payment_page_html.index(later_classification.name),
            payment_page_html.index(earlier_classification.name),
        )
        self.assertLess(
            payment_page_html.index(later_cashbook.name),
            payment_page_html.index(earlier_cashbook.name),
        )

    def test_inline_finance_master_priority_updates_only_sort_order(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi giữ nguyên tên',
            type=2,
            description='Giữ nguyên mô tả danh mục',
            sort_order=10,
        )
        classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=category,
            name='Phân loại giữ nguyên tên',
            description='Giữ nguyên mô tả phân loại',
            sort_order=20,
        )
        cashbook = CashBook.objects.create(
            brand=self.brand,
            name='Quỹ giữ nguyên tên',
            description='Giữ nguyên mô tả quỹ',
            sort_order=30,
            balance=Decimal('12345'),
        )

        for master_type, item, new_sort_order in (
            ('finance_category', category, 3),
            ('expense_classification', classification, 2),
            ('cashbook', cashbook, 1),
        ):
            response = self.client.post(
                reverse('api_update_finance_master_sort_order'),
                data=json.dumps({
                    'master_type': master_type,
                    'id': item.id,
                    'sort_order': new_sort_order,
                    'name': 'Tên không được phép ghi đè',
                }),
                content_type='application/json',
            )
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
            item.refresh_from_db()
            self.assertEqual(item.sort_order, new_sort_order)

        self.assertEqual(category.name, 'Chi giữ nguyên tên')
        self.assertEqual(category.description, 'Giữ nguyên mô tả danh mục')
        self.assertEqual(classification.name, 'Phân loại giữ nguyên tên')
        self.assertEqual(classification.description, 'Giữ nguyên mô tả phân loại')
        self.assertEqual(classification.parent_category, category)
        self.assertEqual(cashbook.name, 'Quỹ giữ nguyên tên')
        self.assertEqual(cashbook.description, 'Giữ nguyên mô tả quỹ')
        self.assertEqual(cashbook.balance, Decimal('12345'))

        invalid_response = self.client.post(
            reverse('api_update_finance_master_sort_order'),
            data=json.dumps({
                'master_type': 'cashbook',
                'id': cashbook.id,
                'sort_order': '-1',
            }),
            content_type='application/json',
        )
        self.assertEqual(invalid_response.json()['status'], 'error')
        cashbook.refresh_from_db()
        self.assertEqual(cashbook.sort_order, 1)

    def test_expense_classification_settings_and_payment_flow(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        expense_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi văn phòng',
            type=2,
        )

        page_response = self.client.get(reverse('category_tbl'))
        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, 'Danh mục hệ thống')
        self.assertContains(page_response, 'Phân loại chi')
        self.assertContains(page_response, 'Danh mục cha')
        self.assertContains(page_response, 'Chi thường xuyên')
        self.assertContains(
            page_response,
            'btn-action-group d-inline-flex flex-nowrap align-items-center',
            count=2,
        )
        self.assertContains(page_response, 'style="gap:4px;"', count=2)

        create_response = self.client.post(
            reverse('api_save_expense_classification'),
            data=json.dumps({
                'name': 'Chi thường xuyên',
                'parent_category_id': expense_category.id,
                'description': 'Các khoản chi vận hành định kỳ',
                'is_active': True,
            }),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(create_response.json()['status'], 'ok')
        classification = ExpenseClassification.objects.get(name='Chi thường xuyên')
        self.assertEqual(classification.brand, self.brand)
        self.assertEqual(classification.parent_category, expense_category)

        payment_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-PHAN-LOAI-001',
                'category_id': expense_category.id,
                'expense_classification_id': classification.id,
                'gross_amount': '1000000',
                'promotion_mode': 'amount',
                'promotion_amount': '0',
                'payment_date': '2026-08-03',
                'status': 0,
                'payment_method': 2,
            }),
            content_type='application/json',
        )
        self.assertEqual(payment_response.status_code, 200)
        self.assertEqual(payment_response.json()['status'], 'ok')
        payment = Payment.objects.get(code='PC-PHAN-LOAI-001')
        self.assertEqual(payment.category, expense_category)
        self.assertEqual(payment.expense_classification, classification)

        other_expense_category = FinanceCategory.objects.create(
            brand=self.brand,
            name='Chi nhân sự',
            type=2,
        )
        mismatch_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-PHAN-LOAI-SAI-DANH-MUC',
                'category_id': other_expense_category.id,
                'expense_classification_id': classification.id,
                'gross_amount': '100000',
                'payment_date': '2026-08-03',
                'status': 0,
                'payment_method': 2,
            }),
            content_type='application/json',
        )
        self.assertEqual(mismatch_response.json()['status'], 'error')
        self.assertIn('chỉ thuộc Danh mục cha', mismatch_response.json()['message'])
        self.assertFalse(Payment.objects.filter(code='PC-PHAN-LOAI-SAI-DANH-MUC').exists())

        list_response = self.client.get(reverse('api_get_payments'), {
            'expense_classification_id': classification.id,
            'page': 1,
            'page_size': 25,
        })
        self.assertEqual(list_response.status_code, 200)
        row = next(item for item in list_response.json()['data'] if item['id'] == payment.id)
        self.assertEqual(row['expense_classification'], 'Chi thường xuyên')
        self.assertEqual(row['expense_classification_id'], classification.id)

        export_response = self.client.get(reverse('export_payments_excel'), {
            'expense_classification_id': classification.id,
        })
        worksheet = load_workbook(BytesIO(export_response.content), data_only=True).active
        headers = [cell.value for cell in worksheet[4]]
        self.assertIn('Phân loại chi', headers)
        classification_column = headers.index('Phân loại chi') + 1
        self.assertEqual(worksheet.cell(row=5, column=classification_column).value, 'Chi thường xuyên')

        deactivate_response = self.client.post(
            reverse('api_save_expense_classification'),
            data=json.dumps({
                'id': classification.id,
                'name': classification.name,
                'parent_category_id': expense_category.id,
                'description': classification.description,
                'is_active': False,
            }),
            content_type='application/json',
        )
        self.assertEqual(deactivate_response.json()['status'], 'ok')

        edit_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'id': payment.id,
                'code': payment.code,
                'category_id': expense_category.id,
                'expense_classification_id': classification.id,
                'gross_amount': '1000000',
                'promotion_mode': 'amount',
                'promotion_amount': '0',
                'payment_date': '2026-08-03',
                'status': 0,
                'payment_method': 2,
            }),
            content_type='application/json',
        )
        self.assertEqual(edit_response.json()['status'], 'ok')
        payment.refresh_from_db()
        self.assertEqual(payment.expense_classification, classification)

        delete_response = self.client.post(
            reverse('api_delete_expense_classification'),
            data=json.dumps({'id': classification.id}),
            content_type='application/json',
        )
        self.assertEqual(delete_response.json()['status'], 'error')
        self.assertIn('Ngừng sử dụng', delete_response.json()['message'])

        payment.delete()
        delete_after_payment_soft_delete_response = self.client.post(
            reverse('api_delete_expense_classification'),
            data=json.dumps({'id': classification.id}),
            content_type='application/json',
        )
        self.assertEqual(
            delete_after_payment_soft_delete_response.json()['status'], 'error',
        )
        self.assertIn(
            'Ngừng sử dụng',
            delete_after_payment_soft_delete_response.json()['message'],
        )

        unused_classification = ExpenseClassification.objects.create(
            brand=self.brand,
            parent_category=expense_category,
            name='Chi chưa phát sinh',
        )
        delete_unused_response = self.client.post(
            reverse('api_delete_expense_classification'),
            data=json.dumps({'id': unused_classification.id}),
            content_type='application/json',
        )
        self.assertEqual(delete_unused_response.json()['status'], 'ok')
        self.assertFalse(
            ExpenseClassification.objects.filter(id=unused_classification.id).exists(),
        )
        self.assertTrue(
            ExpenseClassification.all_objects.filter(id=unused_classification.id).exists(),
        )

    def test_payment_rejects_expense_classification_from_another_brand(self):
        other_brand = Brand.objects.create(name='Thương hiệu không thuộc quyền')
        foreign_classification = ExpenseClassification.objects.create(
            brand=other_brand,
            name='Chi ngoài phạm vi',
        )

        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-PHAN-LOAI-SAI',
                'expense_classification_id': foreign_classification.id,
                'gross_amount': '100000',
                'payment_date': '2026-08-03',
                'status': 0,
                'payment_method': 2,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Phân loại chi không thuộc thương hiệu', response.json()['message'])
        self.assertFalse(Payment.objects.filter(code='PC-PHAN-LOAI-SAI').exists())

    def test_payment_form_cashbook_options_include_current_balance(self):
        cash_book = CashBook.objects.create(
            name='Quỹ tiền mặt hiển thị',
            balance=Decimal('100000000'),
        )

        response = self.client.get(reverse('payment_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'value="{cash_book.id}" data-balance="100000000"')
        self.assertContains(response, 'Quỹ tiền mặt hiển thị - 100.000.000đ')

    def test_payment_create_and_edit_only_confirm_close_when_form_changed(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        response = self.client.get(reverse('payment_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'id="modal_form" tabindex="-1" data-confirm-close="dirty" '
            'data-backdrop="static" data-keyboard="false"',
        )
        self.assertContains(response, 'function getPaymentFormCloseSnapshot()')
        self.assertContains(response, 'function startPaymentFormCloseTracking()')
        self.assertContains(response, 'function capturePaymentFormCloseBaseline(trackingToken)')
        self.assertContains(response, 'function paymentFormHasUnsavedChanges()')
        self.assertContains(
            response,
            'var closeTrackingToken = startPaymentFormCloseTracking();',
            count=2,
        )
        self.assertContains(
            response,
            'schedulePaymentFormCloseBaselineCapture(closeTrackingToken);',
            count=2,
        )
        for field_name in (
            'code', 'category_id', 'expense_classification_id', 'supplier_id', 'goods_receipt_id',
            'cash_book_id', 'amount', 'promotion_mode', 'promotion_amount',
            'promotion_percent', 'payment_date',
            'payment_method_option_id', 'status', 'description', 'note',
        ):
            self.assertContains(response, f'{field_name}:')
        self.assertContains(response, 'id="btn_sort_payment_date"')
        self.assertContains(response, "var PAYMENT_DATE_SORT_DIRECTION = 'desc';")
        self.assertContains(
            response,
            "PAYMENT_DATE_SORT_DIRECTION = PAYMENT_DATE_SORT_DIRECTION === 'desc' ? 'asc' : 'desc';",
        )

    def test_receipt_create_and_edit_only_confirm_close_when_form_changed(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        response = self.client.get(reverse('receipt_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'id="modal_form" tabindex="-1" data-confirm-close="dirty" '
            'data-backdrop="static" data-keyboard="false"',
        )
        self.assertContains(response, 'function getReceiptFormCloseSnapshot()')
        self.assertContains(response, 'function startReceiptFormCloseTracking()')
        self.assertContains(response, 'function captureReceiptFormCloseBaseline(trackingToken)')
        self.assertContains(response, 'function receiptFormHasUnsavedChanges()')
        self.assertContains(
            response,
            'var closeTrackingToken = startReceiptFormCloseTracking();',
            count=2,
        )
        self.assertContains(
            response,
            'scheduleReceiptFormCloseBaselineCapture(closeTrackingToken);',
            count=2,
        )
        for field_name in (
            'code', 'category_id', 'customer_id', 'order_id',
            'cash_book_id', 'receipt_date', 'amount',
            'payment_method_option_id', 'status', 'description', 'note',
        ):
            self.assertContains(response, f'{field_name}:')
        self.assertContains(response, 'id="btn_sort_receipt_date"')
        self.assertContains(response, "var receiptDateSortDirection = 'desc';")
        self.assertContains(
            response,
            "receiptDateSortDirection = receiptDateSortDirection === 'desc' ? 'asc' : 'desc';",
        )

    def test_receipt_edit_refreshes_payment_method_select2_display(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        response = self.client.get(reverse('receipt_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            ".val(receipt.payment_method_option_id || '')\n"
            "        .trigger('change.select2');",
        )

    def test_save_payment_auto_generates_code_when_blank(self):
        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': '',
                'amount': 100,
                'payment_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        payment = Payment.objects.get(amount=Decimal('100'))
        self.assertEqual(payment.code, 'PC-001')
        self.assertEqual(payment.store_id, self.store.id)

    def test_payment_approval_uses_logged_in_user_and_clears_when_cancelled(self):
        payment = Payment.objects.create(
            code='PC-APPROVAL-USER',
            store=self.store,
            amount=Decimal('100'),
            payment_date=date.today(),
            status=0,
            created_by=self.other_user,
        )

        approval_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'id': payment.id,
                'code': payment.code,
                'amount': '100',
                'payment_date': payment.payment_date.isoformat(),
                'status': 1,
            }),
            content_type='application/json',
        )

        self.assertEqual(approval_response.json()['status'], 'ok', msg=approval_response.content.decode())
        payment.refresh_from_db()
        self.assertEqual(payment.created_by_id, self.other_user.id)
        self.assertEqual(payment.approved_by_id, self.user.id)
        self.assertIsNotNone(payment.approved_at)

        list_payload = self.client.get(reverse('api_get_payments')).json()['data']
        payment_row = next(row for row in list_payload if row['id'] == payment.id)
        self.assertEqual(payment_row['created_by'], self.other_user.username)
        self.assertEqual(payment_row['approved_by'], self.user.username)
        self.assertTrue(payment_row['approved_at'])

        cancel_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'id': payment.id,
                'code': payment.code,
                'amount': '100',
                'payment_date': payment.payment_date.isoformat(),
                'status': 2,
            }),
            content_type='application/json',
        )
        self.assertEqual(cancel_response.json()['status'], 'ok', msg=cancel_response.content.decode())
        payment.refresh_from_db()
        self.assertIsNone(payment.approved_by_id)
        self.assertIsNone(payment.approved_at)

    def test_manual_completed_payment_uses_logged_in_user_as_creator_and_approver(self):
        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-MANUAL-SELF-APPROVED',
                'amount': '250',
                'payment_date': date.today().isoformat(),
                'status': 1,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        payment = Payment.objects.get(code='PC-MANUAL-SELF-APPROVED')
        self.assertEqual(payment.created_by_id, self.user.id)
        self.assertEqual(payment.approved_by_id, self.user.id)
        self.assertIsNotNone(payment.approved_at)

    def test_payment_amount_promotion_reduces_cash_and_settles_supplier_debt(self):
        receipt = GoodsReceipt.objects.create(
            code='PN-PROMOTION-AMOUNT',
            supplier=self.supplier,
            warehouse=self.warehouse,
            total_amount=Decimal('1000'),
            receipt_date=date.today(),
            status=1,
            created_by=self.user,
        )
        cash_book = CashBook.objects.create(
            name='Quỹ thanh toán có KM',
            balance=Decimal('2000'),
        )
        payment = Payment.objects.create(
            code='PC-PROMOTION-AMOUNT',
            store=self.store,
            supplier=self.supplier,
            goods_receipt=receipt,
            amount=Decimal('1000'),
            payment_date=date.today(),
            status=0,
            created_by=self.other_user,
        )

        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'id': payment.id,
                'code': payment.code,
                'cash_book_id': cash_book.id,
                'supplier_id': self.supplier.id,
                'goods_receipt_id': receipt.id,
                'gross_amount': '777',  # Backend phải bỏ qua và lấy tổng phiếu nhập 1.000.
                'promotion_mode': 'amount',
                'promotion_amount': '150',
                'promotion_percent': '0',
                'amount': '999999',  # Server phải tự tính lại, không tin client.
                'payment_date': date.today().isoformat(),
                'status': 1,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        payment.refresh_from_db()
        receipt.refresh_from_db()
        cash_book.refresh_from_db()
        self.assertEqual(payment.amount, Decimal('850'))
        self.assertEqual(payment.promotion_mode, 'amount')
        self.assertEqual(payment.promotion_amount, Decimal('150'))
        self.assertEqual(payment.promotion_percent, Decimal('15.00'))
        self.assertEqual(payment.approved_by_id, self.user.id)
        self.assertEqual(cash_book.balance, Decimal('1150'))
        self.assertEqual(receipt.total_amount, Decimal('1000'))

        payment_row = next(
            row for row in self.client.get(reverse('api_get_payments')).json()['data']
            if row['id'] == payment.id
        )
        self.assertEqual(payment_row['gross_amount'], 1000.0)
        self.assertEqual(payment_row['promotion_amount'], 150.0)
        self.assertEqual(payment_row['amount'], 850.0)

        debt_payload = self.client.get(
            reverse('api_get_supplier_debts'),
            {'q': receipt.code, 'payment_state': 'all'},
        ).json()
        debt_row = debt_payload['data'][0]
        self.assertEqual(debt_row['cash_paid_amount'], 850.0)
        self.assertEqual(debt_row['promotion_amount'], 150.0)
        self.assertEqual(debt_row['paid_amount'], 1000.0)
        self.assertEqual(debt_row['debt_amount'], 0.0)
        self.assertEqual(debt_row['payment_state'], 'settled')

    def test_payment_percent_promotion_is_calculated_and_rounded_by_server(self):
        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-PROMOTION-PERCENT',
                'gross_amount': '999',
                'promotion_mode': 'percent',
                'promotion_percent': '12.5',
                'promotion_amount': '0',
                'amount': '0',
                'payment_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        payment = Payment.objects.get(code='PC-PROMOTION-PERCENT')
        self.assertEqual(payment.promotion_percent, Decimal('12.50'))
        self.assertEqual(payment.promotion_amount, Decimal('125'))
        self.assertEqual(payment.amount, Decimal('874'))

    def test_payment_rejects_promotion_greater_than_gross_without_changing_cashbook(self):
        cash_book = CashBook.objects.create(name='Quỹ không đổi khi lỗi KM', balance=Decimal('2000'))
        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-PROMOTION-INVALID',
                'cash_book_id': cash_book.id,
                'gross_amount': '1000',
                'promotion_mode': 'amount',
                'promotion_amount': '1001',
                'payment_date': date.today().isoformat(),
                'status': 1,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('không được lớn hơn', response.json()['message'])
        self.assertFalse(Payment.objects.filter(code='PC-PROMOTION-INVALID').exists())
        cash_book.refresh_from_db()
        self.assertEqual(cash_book.balance, Decimal('2000'))

    def test_finance_entries_api_returns_paginated_combined_rows(self):
        today = date.today()
        for index in range(6):
            Receipt.objects.create(
                code=f'PT-LIST-{index:02d}',
                store=self.store,
                customer=self.customer,
                amount=Decimal('100'),
                receipt_date=today - timedelta(days=index),
                status=1,
                created_by=self.user,
            )
        for index in range(5):
            Payment.objects.create(
                code=f'PC-LIST-{index:02d}',
                store=self.store,
                amount=Decimal('50'),
                payment_date=today - timedelta(days=index + 6),
                status=1,
                created_by=self.user,
            )

        response = self.client.get(
            reverse('api_get_finance_entries'),
            data={'page': 2, 'page_size': 10},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['page'], 2)
        self.assertEqual(payload['meta']['page_size'], 10)
        self.assertEqual(payload['meta']['page_count'], 1)
        self.assertEqual(payload['meta']['total_pages'], 2)
        self.assertEqual(payload['meta']['total_filtered_count'], 11)
        self.assertEqual(payload['meta']['start_index'], 11)
        self.assertEqual(payload['meta']['end_index'], 11)
        self.assertFalse(payload['meta']['has_next'])
        self.assertEqual(payload['data'][0]['code'], 'PC-LIST-04')
        self.assertEqual(payload['data'][0]['type'], 'Chi')
        self.assertEqual(payload['data'][0]['status_display'], 'Hoàn thành')

    def test_finance_entries_api_filters_by_type(self):
        Receipt.objects.create(
            code='PT-FILTER-001',
            store=self.store,
            customer=self.customer,
            amount=Decimal('100'),
            receipt_date=date.today(),
            status=1,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-FILTER-001',
            store=self.store,
            amount=Decimal('50'),
            payment_date=date.today(),
            status=1,
            created_by=self.user,
        )

        response = self.client.get(
            reverse('api_get_finance_entries'),
            data={'type': 'thu', 'page': 1, 'page_size': 10},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['total_filtered_count'], 1)
        self.assertEqual([item['code'] for item in payload['data']], ['PT-FILTER-001'])
        self.assertEqual(payload['data'][0]['type'], 'Thu')

    def test_supplier_debt_api_calculates_returns_and_completed_linked_payments(self):
        receipt = GoodsReceipt.objects.create(
            code='PN-DEBT-001',
            supplier=self.supplier,
            warehouse=self.warehouse,
            total_amount=Decimal('1000'),
            receipt_date=date.today(),
            status=1,
            created_by=self.user,
        )
        PurchaseReturn.objects.create(
            code='THN-DEBT-001',
            goods_receipt=receipt,
            supplier=self.supplier,
            warehouse=self.warehouse,
            total_amount=Decimal('100'),
            return_date=date.today(),
            status=1,
            created_by=self.user,
        )
        PurchaseReturn.objects.create(
            code='THN-DEBT-DRAFT',
            goods_receipt=receipt,
            supplier=self.supplier,
            warehouse=self.warehouse,
            total_amount=Decimal('50'),
            return_date=date.today(),
            status=0,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-DEBT-PAID',
            store=self.store,
            supplier=self.supplier,
            goods_receipt=receipt,
            amount=Decimal('300'),
            payment_date=date.today(),
            status=1,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-DEBT-DRAFT',
            store=self.store,
            supplier=self.supplier,
            goods_receipt=receipt,
            amount=Decimal('200'),
            payment_date=date.today(),
            status=0,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-DEBT-UNLINKED',
            store=self.store,
            supplier=self.supplier,
            amount=Decimal('500'),
            payment_date=date.today(),
            status=1,
            created_by=self.user,
        )
        GoodsReceipt.objects.create(
            code='PN-DEBT-FOREIGN',
            supplier=self.supplier,
            warehouse=self.other_warehouse,
            total_amount=Decimal('999'),
            receipt_date=date.today(),
            status=1,
            created_by=self.other_user,
        )

        response = self.client.get(reverse('api_get_supplier_debts'), {
            'q': 'PN-DEBT-001',
            'payment_state': 'outstanding',
            'page': 1,
            'page_size': 10,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['total_filtered_count'], 1)
        self.assertEqual(payload['meta']['total_all_count'], 1)
        row = payload['data'][0]
        self.assertEqual(row['code'], 'PN-DEBT-001')
        self.assertEqual(row['original_amount'], 1000.0)
        self.assertEqual(row['returned_amount'], 100.0)
        self.assertEqual(row['payable_amount'], 900.0)
        self.assertEqual(row['paid_amount'], 300.0)
        self.assertEqual(row['debt_amount'], 600.0)
        self.assertEqual(row['payment_state'], 'partial')
        self.assertEqual(row['payment_codes'], ['PC-DEBT-PAID'])
        self.assertEqual(payload['totals']['debt_amount'], 600.0)
        self.assertEqual(payload['totals']['debt_document_count'], 1)
        self.assertEqual(payload['totals']['overall_debt_amount'], 600.0)

    def test_supplier_debt_page_is_available_from_finance_menu(self):
        response = self.client.get(reverse('supplier_debt_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Công nợ nhà cung cấp')
        self.assertContains(response, 'id="supplier_debt_tbl"')
        self.assertContains(response, 'id="supplier_debt_state"')
        self.assertContains(response, 'id="supplier_debt_cashbook_tbl"')
        self.assertContains(response, 'id="supplier_debt_cashbook_total"')
        self.assertContains(response, 'id="supplier_debt_after_payment_card"')
        self.assertContains(response, 'Còn sau trả nợ')
        self.assertContains(response, 'Thiếu để trả nợ')
        self.assertContains(response, 'function updateSupplierDebtAfterPayment()')
        self.assertContains(response, 'Số dư hiện tại theo quỹ')
        self.assertContains(response, 'function loadSupplierDebtCashbooks()')
        self.assertContains(response, 'Lập phiếu chi')
        self.assertContains(response, reverse('supplier_debt_tbl'))

    def test_supplier_debt_api_filters_payment_states_and_paginates(self):
        receipts = []
        for index in range(11):
            receipts.append(GoodsReceipt.objects.create(
                code=f'PN-DEBT-PAGE-{index:02d}',
                supplier=self.supplier,
                warehouse=self.warehouse,
                total_amount=Decimal('100'),
                receipt_date=date.today() - timedelta(days=index),
                status=1,
                created_by=self.user,
            ))
        Payment.objects.create(
            code='PC-DEBT-PARTIAL',
            store=self.store,
            supplier=self.supplier,
            goods_receipt=receipts[0],
            amount=Decimal('40'),
            payment_date=date.today(),
            status=1,
            created_by=self.user,
        )
        Payment.objects.create(
            code='PC-DEBT-SETTLED',
            store=self.store,
            supplier=self.supplier,
            goods_receipt=receipts[1],
            amount=Decimal('100'),
            payment_date=date.today(),
            status=1,
            created_by=self.user,
        )

        outstanding = self.client.get(reverse('api_get_supplier_debts'))
        partial = self.client.get(reverse('api_get_supplier_debts'), {'payment_state': 'partial'})
        settled = self.client.get(reverse('api_get_supplier_debts'), {'payment_state': 'settled'})
        second_page = self.client.get(reverse('api_get_supplier_debts'), {
            'payment_state': 'all',
            'page': 2,
            'page_size': 10,
        })

        self.assertEqual(outstanding.json()['meta']['total_filtered_count'], 10)
        self.assertEqual(partial.json()['meta']['total_filtered_count'], 1)
        self.assertEqual(partial.json()['data'][0]['payment_state'], 'partial')
        self.assertEqual(settled.json()['meta']['total_filtered_count'], 1)
        self.assertEqual(settled.json()['data'][0]['payment_state'], 'settled')
        page_payload = second_page.json()
        self.assertEqual(page_payload['meta']['page'], 2)
        self.assertEqual(page_payload['meta']['total_pages'], 2)
        self.assertEqual(page_payload['meta']['start_index'], 11)
        self.assertEqual(page_payload['meta']['end_index'], 11)
        self.assertEqual(len(page_payload['data']), 1)

    def test_payment_page_supports_prefill_from_supplier_debt(self):
        response = self.client.get(reverse('payment_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "params.get('create_goods_receipt_id')")
        self.assertContains(response, 'function openPaymentForSupplierDebt()')
        self.assertContains(response, "$('#inp_goods_receipt_id').val(goodsReceiptId).trigger('change');")
        self.assertContains(response, 'openPaymentForSupplierDebt();')

    def test_save_receipt_rejects_foreign_order(self):
        other_order = self._create_order(
            code='DH-FOREIGN-001',
            store=self.other_store,
            customer=self.other_customer,
            warehouse=self.other_warehouse,
            created_by=self.other_user,
        )

        response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'code': 'PT-FOREIGN-001',
                'order_id': other_order.id,
                'amount': 100,
                'receipt_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Không tìm thấy đơn hàng', payload['message'])
        self.assertFalse(Receipt.objects.filter(code='PT-FOREIGN-001').exists())

    def test_save_receipt_rejects_foreign_customer_without_order(self):
        response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'code': 'PT-FOREIGN-CUSTOMER',
                'customer_id': self.other_customer.id,
                'amount': 100,
                'receipt_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Khách hàng', payload['message'])
        self.assertFalse(Receipt.objects.filter(code='PT-FOREIGN-CUSTOMER').exists())

    def test_save_receipt_rejects_customer_that_mismatches_order(self):
        order = self._create_order(code='DH-RECEIPT-MISMATCH')

        response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'code': 'PT-MISMATCH-CUSTOMER',
                'order_id': order.id,
                'customer_id': self.other_customer.id,
                'amount': 100,
                'receipt_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Khách hàng', payload['message'])
        self.assertFalse(Receipt.objects.filter(code='PT-MISMATCH-CUSTOMER').exists())

    def test_save_receipt_accepts_string_ids_when_customer_matches_order(self):
        order = self._create_order(code='DH-RECEIPT-STRING-IDS')

        response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'code': 'PT-STRING-IDS',
                'order_id': str(order.id),
                'customer_id': str(self.customer.id),
                'amount': 100,
                'receipt_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        receipt = Receipt.objects.get(code='PT-STRING-IDS')
        self.assertEqual(receipt.order_id, order.id)
        self.assertEqual(receipt.customer_id, self.customer.id)
        self.assertEqual(receipt.store_id, self.store.id)

    def test_save_receipt_cannot_change_linked_order_when_editing(self):
        original_order = self._create_order(code='DH-RECEIPT-LOCKED-1')
        other_order = self._create_order(code='DH-RECEIPT-LOCKED-2')
        receipt = Receipt.objects.create(
            code='PT-LOCKED-ORDER',
            store=self.store,
            customer=self.customer,
            order=original_order,
            amount=Decimal('50'),
            receipt_date=date.today(),
            status=0,
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'id': receipt.id,
                'code': receipt.code,
                'category_id': None,
                'customer_id': self.customer.id,
                'order_id': other_order.id,
                'amount': 60,
                'receipt_date': date.today().isoformat(),
                'status': 0,
                'description': 'Thu thêm',
                'note': 'Không được đổi đơn',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('đơn hàng', payload['message'].lower())

        receipt.refresh_from_db()
        self.assertEqual(receipt.order_id, original_order.id)
        self.assertEqual(receipt.amount, Decimal('50'))

    def test_delete_receipt_endpoint_rejects_deletion(self):
        receipt = Receipt.objects.create(
            code='PT-NO-DELETE',
            store=self.store,
            customer=self.customer,
            amount=Decimal('50'),
            receipt_date=date.today(),
            status=0,
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_delete_receipt'),
            data=json.dumps({'id': receipt.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('không được xóa', payload['message'].lower())
        self.assertTrue(Receipt.objects.filter(id=receipt.id).exists())

    def test_save_receipt_partial_edit_preserves_fixed_and_payment_fields(self):
        order = self._create_order(code='DH-RECEIPT-PARTIAL-EDIT')
        cash_book = CashBook.objects.create(name='Quỹ giữ nguyên', balance=Decimal('1000'))
        receipt = Receipt.objects.create(
            code='PT-PARTIAL-EDIT',
            store=self.store,
            customer=self.customer,
            order=order,
            cash_book=cash_book,
            amount=Decimal('75'),
            receipt_date=date.today(),
            status=0,
            payment_method=1,
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'id': receipt.id,
                'description': 'Chỉ sửa diễn giải',
                'note': 'Giữ nguyên đơn và hình thức thanh toán',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())

        receipt.refresh_from_db()
        self.assertEqual(receipt.order_id, order.id)
        self.assertEqual(receipt.customer_id, self.customer.id)
        self.assertEqual(receipt.cash_book_id, cash_book.id)
        self.assertEqual(receipt.amount, Decimal('75'))
        self.assertEqual(receipt.payment_method, 1)
        self.assertEqual(receipt.description, 'Chỉ sửa diễn giải')
        self.assertEqual(receipt.note, 'Giữ nguyên đơn và hình thức thanh toán')

    def test_receipt_list_order_stays_stable_after_edit(self):
        receipts = [
            Receipt.objects.create(
                code=f'PT-STABLE-{index}',
                store=self.store,
                customer=self.customer,
                amount=Decimal('100'),
                receipt_date=date.today(),
                status=0,
                created_by=self.user,
            )
            for index in range(1, 4)
        ]
        expected_codes = [receipt.code for receipt in reversed(receipts)]

        before_edit = self.client.get(reverse('api_get_receipts'))
        self.assertEqual(before_edit.status_code, 200)
        self.assertEqual(
            [item['code'] for item in before_edit.json()['data']],
            expected_codes,
        )

        edited_receipt = receipts[1]
        edit_response = self.client.post(
            reverse('api_save_receipt'),
            data=json.dumps({
                'id': edited_receipt.id,
                'description': 'Sửa nhưng không đổi vị trí',
            }),
            content_type='application/json',
        )
        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())

        after_edit = self.client.get(reverse('api_get_receipts'))
        self.assertEqual(after_edit.status_code, 200)
        self.assertEqual(
            [item['code'] for item in after_edit.json()['data']],
            expected_codes,
        )

    def test_export_receipts_excel_includes_note_column(self):
        Receipt.objects.create(
            code='PT-EXPORT-NOTE',
            store=self.store,
            customer=self.customer,
            amount=Decimal('125000'),
            receipt_date=date.today(),
            status=1,
            description='Thu tiền đơn test',
            note='Ghi chú cần xuất Excel',
            created_by=self.user,
        )

        response = self.client.get(reverse('export_receipts_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[4]]
        self.assertIn('Ghi chú', headers)
        note_index = headers.index('Ghi chú') + 1
        self.assertEqual(worksheet.cell(row=5, column=note_index).value, 'Ghi chú cần xuất Excel')

    def test_export_receipts_excel_includes_cashbook_and_method_summaries(self):
        cashbook_a = CashBook.objects.create(name='Tài khoản A')
        cashbook_b = CashBook.objects.create(name='Tài khoản B')
        bank_method = PaymentMethodOption.objects.create(
            code='EXPORT_BANK',
            name='Chuyển khoản ngân hàng',
            legacy_type=2,
        )
        cash_method = PaymentMethodOption.objects.create(
            code='EXPORT_CASH',
            name='Tiền mặt tại quầy',
            legacy_type=1,
        )
        receipt_values = [
            ('PT-EXPORT-A-BANK', cashbook_a, bank_method, '100000', 1),
            ('PT-EXPORT-A-CASH', cashbook_a, cash_method, '50000', 1),
            ('PT-EXPORT-B-BANK', cashbook_b, bank_method, '25000', 1),
            # Tab Tất cả trên màn hình chỉ cộng phiếu hoàn thành vào dashboard.
            ('PT-EXPORT-DRAFT', cashbook_b, cash_method, '999000', 0),
        ]
        for code, cashbook, method, amount, status in receipt_values:
            Receipt.objects.create(
                code=code,
                store=self.store,
                customer=self.customer,
                cash_book=cashbook,
                payment_method_option=method,
                payment_method=method.legacy_type,
                amount=Decimal(amount),
                receipt_date=date.today(),
                status=status,
                created_by=self.user,
            )

        response = self.client.get(reverse('export_receipts_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ['DANH SÁCH PHIẾU THU', 'Tiền về từng tài khoản', 'Theo hình thức nhận'],
        )

        cashbook_sheet = workbook['Tiền về từng tài khoản']
        self.assertEqual(
            [cell.value for cell in cashbook_sheet[4]],
            ['STT', 'Tài khoản', 'Số phiếu', 'Tổng tiền', 'Tỷ trọng (%)'],
        )
        self.assertEqual(
            [cashbook_sheet.cell(row=5, column=column).value for column in range(2, 5)],
            ['Tài khoản A', 2, 150000],
        )
        self.assertEqual(
            [cashbook_sheet.cell(row=6, column=column).value for column in range(2, 5)],
            ['Tài khoản B', 1, 25000],
        )
        self.assertEqual(
            [cashbook_sheet.cell(row=7, column=column).value for column in range(2, 5)],
            ['TỔNG CỘNG', 3, 175000],
        )

        method_sheet = workbook['Theo hình thức nhận']
        self.assertEqual(method_sheet.cell(row=4, column=2).value, 'Hình thức nhận')
        self.assertEqual(
            [method_sheet.cell(row=5, column=column).value for column in range(2, 5)],
            ['Chuyển khoản ngân hàng', 2, 125000],
        )
        self.assertEqual(
            [method_sheet.cell(row=6, column=column).value for column in range(2, 5)],
            ['Tiền mặt tại quầy', 1, 50000],
        )
        self.assertIn('Phiếu hoàn thành', method_sheet.cell(row=2, column=1).value)

    def test_brand_owner_can_create_payment_method_option(self):
        owner = User.objects.create_user(username='finance_owner', password='pass123')
        self.brand.owner = owner
        self.brand.save(update_fields=['owner'])
        cash_book = CashBook.objects.create(name='Tài khoản MoMo', balance=Decimal('0'))
        self.client.force_login(owner)

        response = self.client.post(
            reverse('api_save_payment_method'),
            data=json.dumps({
                'code': 'momo_test',
                'name': 'Ví MoMo test',
                'legacy_type': 3,
                'default_cash_book_id': cash_book.id,
                'sort_order': 10,
                'is_active': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        method = PaymentMethodOption.objects.get(code='MOMO_TEST')
        self.assertEqual(payload['method']['id'], method.id)
        self.assertEqual(method.default_cash_book_id, cash_book.id)

    def test_brand_owner_can_reorder_payment_methods_in_bulk(self):
        owner = User.objects.create_user(username='finance_order_owner', password='pass123')
        self.brand.owner = owner
        self.brand.save(update_fields=['owner'])
        first = PaymentMethodOption.objects.create(code='ORDER_FIRST', name='Phương thức đầu', sort_order=0)
        second = PaymentMethodOption.objects.create(code='ORDER_SECOND', name='Phương thức sau', sort_order=1)
        self.client.force_login(owner)

        response = self.client.post(
            reverse('api_reorder_payment_methods'),
            data=json.dumps({
                'items': [
                    {'id': first.id, 'sort_order': 20},
                    {'id': second.id, 'sort_order': 5},
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.sort_order, 20)
        self.assertEqual(second.sort_order, 5)

    def test_regular_staff_cannot_reorder_payment_methods(self):
        method = PaymentMethodOption.objects.create(
            code='ORDER_STAFF', name='Phương thức nhân viên', sort_order=0,
        )

        response = self.client.post(
            reverse('api_reorder_payment_methods'),
            data=json.dumps({'items': [{'id': method.id, 'sort_order': 10}]}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['status'], 'error')

    def test_reorder_payment_methods_rejects_non_integer_without_partial_update(self):
        owner = User.objects.create_user(username='finance_invalid_order_owner', password='pass123')
        self.brand.owner = owner
        self.brand.save(update_fields=['owner'])
        first = PaymentMethodOption.objects.create(code='ORDER_VALID', name='Phương thức hợp lệ', sort_order=1)
        second = PaymentMethodOption.objects.create(code='ORDER_INVALID', name='Phương thức không hợp lệ', sort_order=2)
        self.client.force_login(owner)

        response = self.client.post(
            reverse('api_reorder_payment_methods'),
            data=json.dumps({
                'items': [
                    {'id': first.id, 'sort_order': 10},
                    {'id': second.id, 'sort_order': 2.5},
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.sort_order, 1)
        self.assertEqual(second.sort_order, 2)

    def test_reorder_payment_methods_accepts_browser_csrf_request(self):
        owner = User.objects.create_user(username='finance_csrf_order_owner', password='pass123')
        self.brand.owner = owner
        self.brand.save(update_fields=['owner'])
        method = PaymentMethodOption.objects.create(code='ORDER_CSRF', name='Phương thức CSRF', sort_order=0)
        browser = Client(enforce_csrf_checks=True)
        browser.force_login(owner)
        page = browser.get(reverse('setting_payment_methods'))
        csrf_cookie = page.cookies.get('csrftoken')
        self.assertIsNotNone(csrf_cookie)

        response = browser.post(
            reverse('api_reorder_payment_methods'),
            data=json.dumps({'items': [{'id': method.id, 'sort_order': 8}]}),
            content_type='application/json',
            HTTP_X_CSRFTOKEN=csrf_cookie.value,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        method.refresh_from_db()
        self.assertEqual(method.sort_order, 8)

    def test_payment_method_settings_exposes_inline_sort_order_editor(self):
        owner = User.objects.create_user(username='finance_inline_owner', password='pass123')
        self.brand.owner = owner
        self.brand.save(update_fields=['owner'])
        self.client.force_login(owner)

        response = self.client.get(reverse('setting_payment_methods'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="btn_save_order"')
        self.assertContains(response, 'pm-sort-order')
        self.assertContains(response, '/api/payment-methods/reorder/')

    def test_save_payment_assigns_store_from_goods_receipt(self):
        goods_receipt = self._create_goods_receipt(code='PN-001')
        cash_book = CashBook.objects.create(name='Quỹ B', balance=Decimal('1000'))

        response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'code': 'PC-STORE-001',
                'goods_receipt_id': goods_receipt.id,
                'cash_book_id': cash_book.id,
                'amount': 100,
                'payment_date': date.today().isoformat(),
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')

        payment = Payment.objects.get(code='PC-STORE-001')
        self.assertEqual(payment.store_id, self.store.id)
        self.assertEqual(payment.goods_receipt_id, goods_receipt.id)

    def test_regular_staff_cannot_save_cashbook(self):
        response = self.client.post(
            reverse('api_save_cashbook'),
            data=json.dumps({
                'name': 'Quỹ staff',
                'description': 'Không được tạo',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['status'], 'error')

    def test_cashbook_page_hides_create_button_for_regular_staff(self):
        response = self.client.get(reverse('cashbook_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="btn_add_cashbook"')
        self.assertNotContains(response, 'id="modal_cashbook"')

    def test_cashbook_page_exposes_code_filter(self):
        response = self.client.get(reverse('cashbook_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="filter_code"')
        self.assertContains(response, 'placeholder="Mã phiếu / phiếu nhập"')
        self.assertContains(response, 'id="filter_entry_type"')
        self.assertContains(response, '-- Tất cả thu/chi --')
        self.assertContains(response, 'Thu − Chi lũy kế')
        self.assertContains(response, 'Thu − Chi theo bộ lọc')
        self.assertContains(response, 'id="cashbook_balance_tbl"')
        self.assertContains(response, 'Số dư hiện tại theo quỹ')
        self.assertContains(response, 'renderCashbookBalances(r3[0].data || [])')
        self.assertContains(response, 'var showRunningBalance = !entryType;')
        self.assertContains(response, "$('#running_balance_header').toggle(showRunningBalance);")
        self.assertContains(response, "$('#net_filter_summary_col').toggle(showRunningBalance);")
        self.assertContains(response, 'id="cashbook_entry_page_size"')
        self.assertContains(response, 'id="cashbook_entry_pagination"')
        self.assertContains(response, 'function renderCashbookEntryPagination(')
        self.assertContains(response, 'CASHBOOK_FILTERED_ENTRIES = entries;')
        self.assertContains(response, "var goodsReceiptCode = String(d.goods_receipt || '').toLowerCase();")
        self.assertContains(response, "paymentCode.indexOf(code) === -1 && goodsReceiptCode.indexOf(code) === -1")
        self.assertContains(response, "if(entryType === 'payment') return;")
        self.assertContains(response, "if(entryType === 'receipt') return;")
        self.assertContains(response, "String(d.code || '').toLowerCase().indexOf(code)")

    def test_cashbook_page_shows_create_button_to_brand_owner(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        response = self.client.get(reverse('cashbook_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="btn_add_cashbook"')
        self.assertContains(response, 'id="modal_cashbook"')

    def _create_current_month_plan(self, name='Kế hoạch kiểm thử'):
        today = date.today()
        response = self.client.post(
            reverse('api_save_financial_plan'),
            data=json.dumps({
                'name': name,
                'period_type': 'month',
                'period_value': today.strftime('%Y-%m'),
                'store_id': self.store.id,
                'status': 1,
            }),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        return FinancialPlan.objects.get(id=response.json()['id'])

    def test_financial_plan_page_exposes_budget_schedule_and_forecast_sections(self):
        response = self.client.get(reverse('financial_plan_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Kế hoạch so với thực tế')
        self.assertContains(response, 'Bước tiếp theo: nhập số tiền kế hoạch')
        self.assertContains(response, 'Nhập khoản Thu / Chi')
        self.assertContains(response, 'id="fpi_form_amount"')
        self.assertContains(response, 'Số tiền kế hoạch (VNĐ)')
        self.assertContains(response, 'function fpDistributeAnnualAmount(totalAmount)')
        self.assertContains(response, 'id="btn_distribute_annual_amount"')
        self.assertContains(response, 'id="fpi_category_empty"')
        self.assertContains(response, 'id="financial_plan_locked"')
        self.assertContains(response, 'function fpParseMoney(value)')
        self.assertContains(response, "fpLoad(r.id,function(){fpOpenPlanItemModal();})")
        self.assertContains(response, 'Lịch thanh toán nhà cung cấp')
        self.assertContains(response, 'Dự báo dòng tiền tương lai')
        self.assertContains(response, 'Cảnh báo tài chính')

    def test_financial_plan_item_requires_positive_planned_amount(self):
        plan = self._create_current_month_plan()
        category = FinanceCategory.objects.create(name='Khoản cần nhập tiền', type=1)

        response = self.client.post(
            reverse('api_save_financial_plan_item'),
            data=json.dumps({
                'plan_id': plan.id,
                'direction': 1,
                'category_id': category.id,
                'planned_amount': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Số tiền kế hoạch phải lớn hơn 0', response.json()['message'])
        self.assertFalse(FinancialPlanItem.objects.filter(plan=plan, category=category).exists())

    def test_financial_plan_rejects_duplicate_store_period(self):
        self._create_current_month_plan()

        response = self.client.post(
            reverse('api_save_financial_plan'),
            data=json.dumps({
                'name': 'Kế hoạch trùng kỳ',
                'period_type': 'month',
                'period_value': date.today().strftime('%Y-%m'),
                'store_id': self.store.id,
                'status': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Đã có kế hoạch tháng', response.json()['message'])

    def test_financial_plan_rejects_invalid_alert_email(self):
        response = self.client.post(
            reverse('api_save_financial_plan'),
            data=json.dumps({
                'period_type': 'month',
                'period_value': date.today().strftime('%Y-%m'),
                'store_id': self.store.id,
                'alert_enabled': True,
                'alert_email_recipients': 'email-khong-hop-le',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Email cảnh báo không hợp lệ', response.json()['message'])

    def test_locked_financial_plan_cannot_be_deleted(self):
        plan = self._create_current_month_plan()
        plan.status = 2
        plan.save(update_fields=['status'])

        response = self.client.post(
            reverse('api_delete_financial_plan'),
            data=json.dumps({'id': plan.id}),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Kế hoạch đã khóa', response.json()['message'])
        self.assertTrue(FinancialPlan.objects.filter(id=plan.id).exists())

    def test_company_plan_is_not_visible_to_another_brand(self):
        other_brand = Brand.objects.create(name='Finance Brand khác')
        other_plan = FinancialPlan.objects.create(
            code='KHTC-BRAND-KHAC',
            name='Kế hoạch công ty khác',
            brand=other_brand,
            store=None,
            period_type='month',
            start_date=date.today().replace(day=1),
            end_date=date.today(),
            status=1,
            created_by=self.other_user,
        )

        payload = self.client.get(reverse('api_get_financial_plans')).json()

        self.assertNotIn(other_plan.id, [row['id'] for row in payload['plans']])

    def test_financial_plan_view_only_role_cannot_create_plan(self):
        group = Group.objects.create(name='Finance chỉ xem')
        role = RoleGroup.objects.create(
            brand=self.brand,
            name='Finance chỉ xem',
            group=group,
        )
        ModulePermission.objects.create(
            role_group=role,
            module='finance',
            action='view',
            is_allowed=True,
        )
        self.user.groups.add(group)

        page_response = self.client.get(reverse('financial_plan_tbl'))
        save_response = self.client.post(
            reverse('api_save_financial_plan'),
            data=json.dumps({
                'period_type': 'month',
                'period_value': date.today().strftime('%Y-%m'),
                'store_id': self.store.id,
            }),
            content_type='application/json',
        )

        self.assertEqual(page_response.status_code, 200)
        self.assertEqual(save_response.status_code, 403)
        self.assertIn('không có quyền lập mới', save_response.json()['message'])

    def test_yearly_financial_plan_and_budget_item_are_saved_with_correct_period(self):
        year = date.today().year + 1
        plan_response = self.client.post(
            reverse('api_save_financial_plan'),
            data=json.dumps({
                'period_type': 'year',
                'period_value': str(year),
                'store_id': self.store.id,
                'status': 1,
            }),
            content_type='application/json',
        )
        self.assertEqual(plan_response.json()['status'], 'ok', msg=plan_response.content.decode())
        plan = FinancialPlan.objects.get(id=plan_response.json()['id'])
        self.assertEqual(plan.start_date, date(year, 1, 1))
        self.assertEqual(plan.end_date, date(year, 12, 31))

        category = FinanceCategory.objects.create(name='Ngân sách năm', type=2)
        item_response = self.client.post(
            reverse('api_save_financial_plan_item'),
            data=json.dumps({
                'plan_id': plan.id,
                'direction': 2,
                'category_id': category.id,
                'planned_amount': 12000000,
                'expected_date': f'{year}-12-20',
                'include_in_forecast': True,
            }),
            content_type='application/json',
        )
        self.assertEqual(item_response.json()['status'], 'ok', msg=item_response.content.decode())
        item = FinancialPlanItem.objects.get(id=item_response.json()['id'])
        self.assertEqual(item.planned_amount, Decimal('12000000'))
        self.assertEqual(item.expected_date, date(year, 12, 20))

    def test_financial_plan_compares_completed_receipts_and_payments_with_budget(self):
        plan = self._create_current_month_plan()
        income_category = FinanceCategory.objects.create(name='Thu kế hoạch', type=1)
        expense_category = FinanceCategory.objects.create(name='Chi kế hoạch', type=2)
        FinancialPlanItem.objects.create(
            plan=plan, direction=1, category=income_category, planned_amount=Decimal('1000'),
        )
        FinancialPlanItem.objects.create(
            plan=plan, direction=2, category=expense_category, planned_amount=Decimal('600'),
        )
        Receipt.objects.create(
            code='PT-PLAN-ACTUAL', store=self.store, category=income_category,
            amount=Decimal('700'), receipt_date=date.today(), status=1, created_by=self.user,
        )
        Payment.objects.create(
            code='PC-PLAN-ACTUAL', store=self.store, category=expense_category,
            amount=Decimal('250'), payment_date=date.today(), status=1, created_by=self.user,
        )
        # Nháp không được tính vào thực tế.
        Payment.objects.create(
            code='PC-PLAN-DRAFT', store=self.store, category=expense_category,
            amount=Decimal('999'), payment_date=date.today(), status=0, created_by=self.user,
        )

        payload = self.client.get(
            reverse('api_get_financial_plans'), {'plan_id': plan.id},
        ).json()['dashboard']

        self.assertEqual(payload['summary']['planned_income'], 1000.0)
        self.assertEqual(payload['summary']['actual_income'], 700.0)
        self.assertEqual(payload['summary']['planned_expense'], 600.0)
        self.assertEqual(payload['summary']['actual_expense'], 250.0)
        expense_row = next(row for row in payload['items'] if row['direction'] == 2)
        self.assertEqual(expense_row['variance'], -350.0)

    def test_financial_plan_actual_amount_respects_selected_cashbook(self):
        plan = self._create_current_month_plan()
        category = FinanceCategory.objects.create(name='Thu theo quỹ', type=1)
        cashbook_a = CashBook.objects.create(name='Quỹ kế hoạch A')
        cashbook_b = CashBook.objects.create(name='Quỹ kế hoạch B')
        FinancialPlanItem.objects.create(
            plan=plan,
            direction=1,
            category=category,
            cash_book=cashbook_a,
            planned_amount=Decimal('1000'),
        )
        Receipt.objects.create(
            code='PT-PLAN-CASH-A', store=self.store, category=category,
            cash_book=cashbook_a, amount=Decimal('300'), receipt_date=date.today(),
            status=1, created_by=self.user,
        )
        Receipt.objects.create(
            code='PT-PLAN-CASH-B', store=self.store, category=category,
            cash_book=cashbook_b, amount=Decimal('700'), receipt_date=date.today(),
            status=1, created_by=self.user,
        )

        payload = self.client.get(
            reverse('api_get_financial_plans'), {'plan_id': plan.id},
        ).json()['dashboard']

        self.assertEqual(payload['summary']['actual_income'], 1000.0)
        self.assertEqual(payload['items'][0]['actual_amount'], 300.0)

        details = self.client.get(
            reverse('api_financial_plan_item_details', args=[payload['items'][0]['id']]),
        ).json()
        self.assertEqual(details['total'], 300.0)
        self.assertEqual([row['code'] for row in details['rows']], ['PT-PLAN-CASH-A'])

    def test_financial_forecast_warns_when_future_budget_makes_cash_negative(self):
        plan = self._create_current_month_plan()
        cashbook = CashBook.objects.create(name='Quỹ dự báo', balance=Decimal('100'))
        category = FinanceCategory.objects.create(name='Chi tương lai', type=2)
        FinancialPlanItem.objects.create(
            plan=plan,
            direction=2,
            category=category,
            cash_book=cashbook,
            planned_amount=Decimal('500'),
            expected_date=plan.end_date,
        )

        dashboard = self.client.get(
            reverse('api_get_financial_plans'), {'plan_id': plan.id},
        ).json()['dashboard']

        self.assertEqual(dashboard['summary']['current_balance'], 100.0)
        self.assertEqual(dashboard['summary']['forecast_balance'], -400.0)
        self.assertTrue(any(alert['type'] == 'cash_shortage' for alert in dashboard['alerts']))

    def test_supplier_schedule_creates_draft_payment_and_approval_marks_schedule_paid(self):
        plan = self._create_current_month_plan()
        category = FinanceCategory.objects.create(name='Nhập hàng', type=2)
        item = FinancialPlanItem.objects.create(
            plan=plan, direction=2, category=category, planned_amount=Decimal('1000'),
        )
        cashbook = CashBook.objects.create(name='Quỹ trả NCC', balance=Decimal('1000'))
        receipt = GoodsReceipt.objects.create(
            code='PN-SCHEDULE-001', supplier=self.supplier, warehouse=self.warehouse,
            status=1, total_amount=Decimal('1000'), receipt_date=date.today(), created_by=self.user,
        )

        create_response = self.client.post(
            reverse('api_save_supplier_payment_schedule'),
            data=json.dumps({
                'plan_id': plan.id,
                'plan_item_id': item.id,
                'goods_receipt_id': receipt.id,
                'due_date': date.today().isoformat(),
                'gross_amount': 1000,
                'promotion_mode': 'percent',
                'promotion_percent': 10,
                'cash_book_id': cashbook.id,
                'priority': 1,
            }),
            content_type='application/json',
        )

        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        schedule = SupplierPaymentSchedule.objects.get(id=create_response.json()['id'])
        payment = schedule.payment
        self.assertEqual(schedule.amount, Decimal('900'))
        self.assertEqual(schedule.promotion_amount, Decimal('100'))
        self.assertEqual(payment.status, 0)
        self.assertEqual(payment.amount, Decimal('900'))
        self.assertEqual(payment.goods_receipt_id, receipt.id)

        approve_response = self.client.post(
            reverse('api_save_payment'),
            data=json.dumps({
                'id': payment.id,
                'code': payment.code,
                'category_id': category.id,
                'cash_book_id': cashbook.id,
                'supplier_id': self.supplier.id,
                'goods_receipt_id': receipt.id,
                'gross_amount': 999999,
                'promotion_mode': 'percent',
                'promotion_percent': 10,
                'payment_date': date.today().isoformat(),
                'payment_method': 2,
                'status': 1,
            }),
            content_type='application/json',
        )

        self.assertEqual(approve_response.json()['status'], 'ok', msg=approve_response.content.decode())
        schedule.refresh_from_db()
        payment.refresh_from_db()
        cashbook.refresh_from_db()
        self.assertEqual(payment.amount, Decimal('900'))
        self.assertEqual(payment.approved_by_id, self.user.id)
        self.assertEqual(schedule.status, 1)
        self.assertEqual(cashbook.balance, Decimal('100'))

    def test_supplier_schedule_rejects_amount_over_remaining_goods_receipt_debt(self):
        plan = self._create_current_month_plan()
        receipt = GoodsReceipt.objects.create(
            code='PN-SCHEDULE-LIMIT', supplier=self.supplier, warehouse=self.warehouse,
            status=1, total_amount=Decimal('500'), receipt_date=date.today(), created_by=self.user,
        )
        common = {
            'plan_id': plan.id,
            'goods_receipt_id': receipt.id,
            'due_date': date.today().isoformat(),
            'promotion_mode': 'amount',
            'promotion_amount': 0,
            'priority': 3,
        }
        first_response = self.client.post(
            reverse('api_save_supplier_payment_schedule'),
            data=json.dumps(dict(common, gross_amount=400)),
            content_type='application/json',
        )
        second_response = self.client.post(
            reverse('api_save_supplier_payment_schedule'),
            data=json.dumps(dict(common, gross_amount=200)),
            content_type='application/json',
        )

        self.assertEqual(first_response.json()['status'], 'ok', msg=first_response.content.decode())
        self.assertEqual(second_response.json()['status'], 'error')
        self.assertIn('100', second_response.json()['message'])
        self.assertEqual(SupplierPaymentSchedule.objects.filter(goods_receipt=receipt).count(), 1)

    def test_yearly_budget_allocations_drive_monthly_comparison_and_revision_history(self):
        year = date.today().year
        plan_response = self.client.post(
            reverse('api_save_financial_plan'),
            data=json.dumps({
                'period_type': 'year',
                'period_value': str(year),
                'store_id': self.store.id,
                'status': 1,
            }),
            content_type='application/json',
        )
        plan = FinancialPlan.objects.get(id=plan_response.json()['id'])
        category = FinanceCategory.objects.create(name='Chi phân bổ tháng', type=2)
        allocations = [{
            'month': f'{year}-{month:02d}',
            'planned_amount': month * 100,
            'expected_date': f'{year}-{month:02d}-15',
        } for month in range(1, 13)]

        response = self.client.post(
            reverse('api_save_financial_plan_item'),
            data=json.dumps({
                'plan_id': plan.id,
                'direction': 2,
                'category_id': category.id,
                'planned_amount': 0,
                'allocations': allocations,
                'include_in_forecast': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        item = FinancialPlanItem.objects.get(id=response.json()['id'])
        self.assertEqual(item.planned_amount, Decimal('7800'))
        self.assertEqual(FinancialPlanAllocation.objects.filter(item=item).count(), 12)
        dashboard = self.client.get(
            reverse('api_get_financial_plans'), {'plan_id': plan.id},
        ).json()['dashboard']
        self.assertEqual(len(dashboard['monthly_summary']), 12)
        self.assertEqual(dashboard['monthly_summary'][0]['planned_expense'], 100.0)
        self.assertGreaterEqual(len(dashboard['revisions']), 2)

        missing_reason_response = self.client.post(
            reverse('api_save_financial_plan_item'),
            data=json.dumps({
                'id': item.id,
                'plan_id': plan.id,
                'direction': 2,
                'category_id': category.id,
                'planned_amount': 7800,
                'allocations': allocations,
            }),
            content_type='application/json',
        )
        self.assertEqual(missing_reason_response.json()['status'], 'error')
        self.assertIn('lý do', missing_reason_response.json()['message'].lower())

    def test_forecast_is_calculated_for_each_cashbook_and_respects_minimum_balance(self):
        plan = self._create_current_month_plan()
        cashbook = CashBook.objects.create(
            name='Quỹ an toàn', balance=Decimal('1000'), minimum_balance=Decimal('500'),
        )
        category = FinanceCategory.objects.create(name='Chi theo quỹ', type=2)
        FinancialPlanItem.objects.create(
            plan=plan,
            direction=2,
            category=category,
            cash_book=cashbook,
            planned_amount=Decimal('700'),
            expected_date=plan.end_date,
        )

        dashboard = self.client.get(
            reverse('api_get_financial_plans'), {'plan_id': plan.id},
        ).json()['dashboard']
        row = next(
            row for row in dashboard['cashbook_forecasts']
            if row['cash_book_id'] == cashbook.id
        )

        self.assertEqual(row['current_balance'], 1000.0)
        self.assertEqual(row['minimum_balance'], 500.0)
        self.assertEqual(row['forecast_balance'], 300.0)
        self.assertTrue(row['shortage_date'])
        self.assertTrue(any(alert['type'] == 'cashbook_shortage' for alert in dashboard['alerts']))

    def test_auto_schedule_splits_supplier_debt_using_priority_and_safe_cash_balances(self):
        plan = self._create_current_month_plan()
        self.supplier.payment_priority = 1
        self.supplier.payment_term_days = 0
        self.supplier.save(update_fields=['payment_priority', 'payment_term_days'])
        cashbook_a = CashBook.objects.create(
            name='Quỹ đề xuất A', balance=Decimal('600'), minimum_balance=Decimal('100'),
        )
        cashbook_b = CashBook.objects.create(
            name='Quỹ đề xuất B', balance=Decimal('400'), minimum_balance=Decimal('100'),
        )
        income_category = FinanceCategory.objects.create(name='Thu dự kiến cho lịch', type=1)
        import_category = FinanceCategory.objects.create(name='Nhập hàng', type=2)
        income_item = FinancialPlanItem.objects.create(
            plan=plan, direction=1, category=income_category, cash_book=cashbook_a,
            planned_amount=Decimal('300'), expected_date=min(date.today() + timedelta(days=1), plan.end_date),
        )
        FinancialPlanItem.objects.create(
            plan=plan, direction=2, category=import_category,
            planned_amount=Decimal('1000'), expected_date=plan.end_date,
        )
        receipt = GoodsReceipt.objects.create(
            code='PN-AUTO-SPLIT', supplier=self.supplier, warehouse=self.warehouse,
            status=1, total_amount=Decimal('1000'), receipt_date=date.today(), created_by=self.user,
        )

        suggestion_response = self.client.get(
            reverse('api_suggest_supplier_payment_schedules'), {'plan_id': plan.id},
        )
        payload = suggestion_response.json()
        self.assertEqual(payload['status'], 'ok', msg=suggestion_response.content.decode())
        funded = [row for row in payload['suggestions'] if not row['insufficient']]
        self.assertGreaterEqual(len(funded), 2)
        self.assertEqual(sum(row['gross_amount'] for row in funded), 1000.0)
        self.assertEqual(payload['summary']['unfunded_amount'], 0)
        self.assertEqual(funded[0]['priority'], 1)

        apply_response = self.client.post(
            reverse('api_apply_supplier_payment_suggestions'),
            data=json.dumps({
                'plan_id': plan.id,
                'suggestion_keys': [row['key'] for row in funded],
                'revision_reason': 'Chấp nhận lịch tự động kiểm thử',
            }),
            content_type='application/json',
        )
        self.assertEqual(apply_response.json()['status'], 'ok', msg=apply_response.content.decode())
        schedules = SupplierPaymentSchedule.objects.filter(goods_receipt=receipt)
        self.assertEqual(schedules.count(), len(funded))
        self.assertEqual(
            schedules.aggregate(total=Sum('gross_amount'))['total'], Decimal('1000'),
        )
        self.assertFalse(schedules.exclude(source='automatic').exists())
        self.assertEqual(Payment.objects.filter(goods_receipt=receipt, status=0).count(), len(funded))
        self.assertEqual(income_item.cash_book_id, cashbook_a.id)

    def test_plan_item_detail_returns_completed_documents_only(self):
        plan = self._create_current_month_plan()
        category = FinanceCategory.objects.create(name='Chi xem chi tiết', type=2)
        item = FinancialPlanItem.objects.create(
            plan=plan, direction=2, category=category, planned_amount=Decimal('500'),
        )
        Payment.objects.create(
            code='PC-DETAIL-DONE', store=self.store, category=category,
            amount=Decimal('200'), payment_date=date.today(), status=1, created_by=self.user,
        )
        Payment.objects.create(
            code='PC-DETAIL-DRAFT', store=self.store, category=category,
            amount=Decimal('300'), payment_date=date.today(), status=0, created_by=self.user,
        )

        response = self.client.get(reverse('api_financial_plan_item_details', args=[item.id]))

        self.assertEqual(response.json()['status'], 'ok')
        self.assertEqual(response.json()['total'], 200.0)
        self.assertEqual([row['code'] for row in response.json()['rows']], ['PC-DETAIL-DONE'])

    def test_financial_alert_scheduler_sends_due_supplier_and_cash_warning_email(self):
        plan = self._create_current_month_plan()
        plan.alert_enabled = True
        plan.alert_lead_days = '3'
        plan.alert_email_recipients = 'ketoan@example.com'
        plan.save(update_fields=['alert_enabled', 'alert_lead_days', 'alert_email_recipients'])
        due_date = min(date.today() + timedelta(days=3), plan.end_date)
        actual_lead = max((due_date - date.today()).days, 0)
        plan.alert_lead_days = str(actual_lead)
        plan.save(update_fields=['alert_lead_days'])
        SupplierPaymentSchedule.objects.create(
            code='LCT-EMAIL-001', plan=plan, store=self.store, supplier=self.supplier,
            due_date=due_date, gross_amount=Decimal('100'), amount=Decimal('100'),
            priority=1, status=0, created_by=self.user,
        )

        result = run_due_financial_alerts(
            now=datetime.combine(date.today(), time(9, 0)),
        )

        self.assertEqual(result['totals']['sent'], 1)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn('Cảnh báo tài chính', mail.outbox[0].subject)
        plan.refresh_from_db()
        self.assertIsNotNone(plan.last_alert_run_at)
        self.assertIsNotNone(plan.last_alert_sent)

    def test_financial_alert_scheduler_can_retry_after_send_error(self):
        plan = self._create_current_month_plan()
        plan.alert_enabled = True
        plan.alert_lead_days = '0'
        plan.alert_email_recipients = 'ketoan@example.com'
        plan.save(update_fields=['alert_enabled', 'alert_lead_days', 'alert_email_recipients'])
        SupplierPaymentSchedule.objects.create(
            code='LCT-EMAIL-RETRY', plan=plan, store=self.store, supplier=self.supplier,
            due_date=date.today(), gross_amount=Decimal('100'), amount=Decimal('100'),
            priority=1, status=0, created_by=self.user,
        )
        run_time = datetime.combine(date.today(), time(9, 0))

        with patch(
            'finance.financial_alerts._send_plan_alert_email',
            side_effect=RuntimeError('SMTP tạm thời lỗi'),
        ):
            failed_result = run_due_financial_alerts(now=run_time)

        plan.refresh_from_db()
        self.assertEqual(failed_result['totals']['error'], 1)
        self.assertIsNone(plan.last_alert_run_at)

        retry_result = run_due_financial_alerts(now=run_time)

        self.assertEqual(retry_result['totals']['sent'], 1)
