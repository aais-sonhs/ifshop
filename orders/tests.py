import json
import calendar
from io import BytesIO
from datetime import date, datetime, timedelta

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from customers.models import Customer, CustomerAddress
from finance.models import CashBook, Payment, PaymentMethodOption, Receipt
from finance.services import update_order_payment_status
from orders.models import (
    Order, OrderEditHistory, OrderItem, OrderReturn, OrderReturnExchangeItem, Packaging,
    OrderReturnItem, Quotation, WarrantyCertificate,
)
from products.models import ComboItem, Product, ProductLocation, ProductStock, ProductVariant, Warehouse
from system_management.models import Brand, BusinessConfig, PrintTemplate, Store, UserProfile


class OrderRiskFlowTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username='order_owner', password='pass123')
        cls.brand = Brand.objects.create(name='Test Brand', owner=cls.owner)
        cls.store = Store.objects.create(brand=cls.brand, name='Store A', code='STA')
        cls.other_store = Store.objects.create(brand=cls.brand, name='Store B', code='STB')

        cls.user = User.objects.create_user(username='seller_a', password='pass123')
        cls.other_user = User.objects.create_user(username='seller_b', password='pass123')
        UserProfile.objects.create(user=cls.user, store=cls.store)
        UserProfile.objects.create(user=cls.other_user, store=cls.other_store)

        cls.customer = Customer.objects.create(
            store=cls.store,
            code='KH001',
            name='Customer A',
            created_by=cls.user,
        )
        cls.other_customer = Customer.objects.create(
            store=cls.other_store,
            code='KH002',
            name='Customer B',
            created_by=cls.other_user,
        )

        cls.warehouse = Warehouse.objects.create(store=cls.store, code='KHO-A', name='Kho A')
        cls.other_warehouse = Warehouse.objects.create(
            store=cls.other_store,
            code='KHO-B',
            name='Kho B',
        )
        cls.product = Product.objects.create(
            store=cls.store,
            code='SP-ORDER-001',
            name='Sản phẩm test đơn hàng',
            created_by=cls.user,
        )
        cls.other_product = Product.objects.create(
            store=cls.other_store,
            code='SP-ORDER-OTHER',
            name='Sản phẩm store khác',
            created_by=cls.other_user,
        )
        cls.cashbook = CashBook.objects.create(
            name='Tiền mặt test đơn hàng',
            balance=1000000,
        )
        cls.payment_method = PaymentMethodOption.objects.create(
            code='TMORDERTEST',
            name='Tiền mặt test',
            legacy_type=1,
            default_cash_book=cls.cashbook,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def _create_quotation(self, code, store=None, customer=None, created_by=None, status=3):
        return Quotation.objects.create(
            code=code,
            store=store or self.store,
            customer=customer or self.customer,
            status=status,
            total_amount=100,
            final_amount=100,
            quotation_date=date.today(),
            created_by=created_by or self.user,
        )

    def _create_order(
        self,
        code,
        quotation=None,
        store=None,
        customer=None,
        warehouse=None,
        created_by=None,
        status=1,
    ):
        store = store or self.store
        return Order.objects.create(
            code=code,
            store=store,
            quotation=quotation,
            customer=customer or self.customer,
            warehouse=warehouse or self.warehouse,
            status=status,
            total_amount=100,
            final_amount=100,
            order_date=date.today(),
            created_by=created_by or self.user,
        )

    def test_order_detail_exposes_related_print_brands(self):
        print_label = Brand.objects.create(
            name='Z Brand Print',
            owner=self.owner,
            brand_type=Brand.TYPE_PRINT_LABEL,
            print_priority=20,
        )
        priority_label = Brand.objects.create(
            name='A Priority Brand Print',
            owner=self.owner,
            brand_type=Brand.TYPE_PRINT_LABEL,
            print_priority=1,
        )
        order = self._create_order(code='DH-BRAND-001')

        response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        brand_ids = [item['id'] for item in payload['available_print_brands']]
        self.assertEqual(brand_ids, [priority_label.id, print_label.id])
        self.assertNotIn(self.brand.id, brand_ids)
        self.assertEqual(payload['order']['store_brand_id'], self.brand.id)

    def test_print_order_can_switch_brand_and_persist_issuing_brand(self):
        print_label = Brand.objects.create(
            name='Z Brand Invoice',
            owner=self.owner,
            brand_type=Brand.TYPE_PRINT_LABEL,
        )
        order = self._create_order(code='DH-BRAND-002')

        response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order', 'brand_id': print_label.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, print_label.name)
        order.refresh_from_db()
        self.assertEqual(order.issuing_brand_id, print_label.id)

    def test_export_print_does_not_prefill_creator_signature_name(self):
        self.user.first_name = 'Nguyễn Văn'
        self.user.last_name = 'Người Lập'
        self.user.save(update_fields=['first_name', 'last_name'])
        order = self._create_order(code='DH-EXPORT-SIGNATURE')

        response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'export', 'source': 'order'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Người lập phiếu')
        self.assertContains(response, '(Ký và ghi rõ họ tên)')
        self.assertNotContains(response, 'Nguyễn Văn Người Lập')

    def test_packing_a5_print_uses_packing_labels_without_prices(self):
        location = ProductLocation.objects.create(name='Kệ A1')
        self.product.unit = 'Thùng'
        self.product.location = location
        self.product.save(update_fields=['unit', 'location'])
        order = self._create_order(code='DH-PACKING-A5')
        Packaging.objects.create(
            code='DG-PACKING-A5',
            order=order,
            status=2,
            packed_by=self.user,
            packed_at=datetime(2026, 7, 21, 10, 3),
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=125000,
            total_price=250000,
        )

        response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'packing', 'source': 'order'},
        )

        self.assertEqual(response.status_code, 200)
        for expected in (
            'Phiếu đóng hàng A5',
            'PHIẾU ĐÓNG HÀNG',
            'size: A5 portrait',
            'Kho lấy hàng',
            'Giờ đóng: ....................',
            'Thông tin giao hàng',
            'Mã hàng',
            'Sản phẩm / quy cách',
            'SL đóng',
            'Vị trí',
            'Kệ A1',
            'Kiểm / ghi chú',
            'Người soạn hàng',
            'Người đóng gói',
            'Người kiểm tra',
            'Số kiện: ....................',
            self.product.code,
            self.product.name,
            'Thùng',
        ):
            self.assertContains(response, expected)
        content = response.content.decode()
        self.assertLess(content.index('SL đóng'), content.index('ĐVT'))
        self.assertLess(content.index('ĐVT'), content.index('Vị trí'))
        self.assertLess(content.index('Vị trí'), content.index('Kiểm / ghi chú'))
        self.assertLess(content.index('Tổng: 1 dòng hàng'), content.index('Số kiện: ....................'))
        self.assertContains(response, '<td class="text-center"><span class="pk-check-box"></span></td>', html=True)
        self.assertContains(response, '2')
        for unexpected in ('Đơn giá', 'Thành tiền', 'PHIẾU XUẤT KHO', 'Thủ kho'):
            self.assertNotContains(response, unexpected)

    def test_order_page_offers_packing_a5_print_option(self):
        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-print-type="packing"')
        self.assertContains(response, 'Phiếu đóng hàng A5')

    def test_order_page_applies_date_filters_from_query_string_before_loading(self):
        response = self.client.get(reverse('order_tbl'), {
            'from_date': '2026-07-21',
            'to_date': '2026-07-21',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'function applyOrderListQueryFilters(params)')
        self.assertContains(response, 'applyOrderListQueryFilters(pageParams);')
        self.assertContains(response, "['from_date', 'to_date']")
        content = response.content.decode()
        self.assertLess(
            content.index('applyOrderListQueryFilters(pageParams);'),
            content.index('loadData();', content.index('applyOrderListQueryFilters(pageParams);')),
        )

    def test_order_page_ignores_loss_warning_for_fully_returned_order(self):
        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'var ORDER_FULLY_RETURNED = false;')
        self.assertContains(response, 'function setCurrentOrderReturnState(order, items)')
        self.assertContains(response, 'if(!ORDER_FULLY_RETURNED && costPrice>0 && effectivePrice<costPrice)')
        self.assertContains(response, 'var amountOnlyComplete = !hasReturnItems && soldAmount > 0')
        self.assertContains(response, 'setCurrentOrderReturnState(o, res.items);')
        self.assertContains(response, 'resetCurrentOrderReturnState();')

    def test_k80_print_uses_four_product_quantity_price_and_total_columns(self):
        self.product.unit = 'Bộ 12c'
        self.product.save(update_fields=['unit'])
        order = self._create_order(code='DH-K80-ITEM-LAYOUT')
        order.total_amount = 250000
        order.final_amount = 250000
        order.save(update_fields=['total_amount', 'final_amount'])
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=125000,
            total_price=250000,
        )

        response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'k80', 'source': 'order'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<th class="item-product-cell">Tên SP/Mã hàng</th>', html=True)
        self.assertContains(response, '<th class="item-quantity-cell">SL</th>', html=True)
        self.assertContains(response, 'Đơn giá')
        self.assertContains(response, 'Thành tiền')
        self.assertContains(response, 'class="item-row"')
        self.assertContains(response, 'class="item-product-name"')
        self.assertContains(response, 'class="item-product-code"')
        self.assertNotContains(response, 'class="item-name-row"')
        self.assertNotContains(response, 'class="item-value-row"')
        self.assertContains(response, 'border-bottom: 1px dashed #aaa;')
        self.assertContains(response, 'Bộ 12c')
        self.assertContains(response, '125.000đ')
        self.assertContains(response, '250.000đ')

    def test_print_quotation_can_switch_brand_and_persist_issuing_brand(self):
        print_label = Brand.objects.create(
            name='Z Brand Quote',
            owner=self.owner,
            brand_type=Brand.TYPE_PRINT_LABEL,
        )
        quotation = self._create_quotation(code='BG-BRAND-001', status=1)

        response = self.client.get(
            reverse('api_print_order'),
            {'id': quotation.id, 'type': 'quotation', 'source': 'quotation', 'brand_id': print_label.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, print_label.name)
        quotation.refresh_from_db()
        self.assertEqual(quotation.issuing_brand_id, print_label.id)

    def test_print_order_keeps_company_template_when_switching_print_label(self):
        print_label = Brand.objects.create(
            name='Z Brand Header',
            owner=self.owner,
            brand_type=Brand.TYPE_PRINT_LABEL,
        )
        PrintTemplate.objects.update_or_create(
            brand=self.brand,
            template_type='a4',
            defaults={'title': 'Mau cong ty goc'},
        )
        PrintTemplate.objects.update_or_create(
            brand=print_label,
            template_type='a4',
            defaults={'title': 'Mau nhan hieu khong duoc dung'},
        )
        order = self._create_order(code='DH-BRAND-004')

        response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order', 'brand_id': print_label.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Mau cong ty goc')
        self.assertNotContains(response, 'Mau nhan hieu khong duoc dung')

    def test_save_order_rejects_foreign_issuing_brand(self):
        other_owner = User.objects.create_user(username='foreign_brand_owner', password='pass123')
        foreign_brand = Brand.objects.create(name='Foreign Invoice Brand', owner=other_owner)
        order = self._create_order(code='DH-BRAND-003')

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({'id': order.id, 'issuing_brand_id': foreign_brand.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertEqual(response.json()['message'], 'Nhãn hàng không thuộc phạm vi được phép sử dụng.')

    def test_save_order_allows_status_change_when_draft_receipt_exists(self):
        order = self._create_order(code='DH-PAID-001', status=0)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=50,
            total_price=100,
        )
        Receipt.objects.create(
            code='PT-PAID-001',
            store=self.store,
            customer=self.customer,
            order=order,
            amount=100,
            receipt_date=date.today(),
            status=1,
            description='Phiếu thu thủ công',
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 2,
                    'unit_price': 50,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['order_status'], 1)
        self.assertEqual(payload['payment_status'], 2)

        order.refresh_from_db()
        self.assertEqual(order.status, 1)
        self.assertEqual(order.final_amount, 100)

    def test_order_other_fee_is_editable_and_included_in_final_amount(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-OTHER-FEE-001',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 5,
                'shipping_fee': 10,
                'other_fee': 25,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        order = Order.objects.get(id=response.json()['order_id'])
        self.assertEqual(float(order.other_fee), 25.0)
        self.assertEqual(float(order.final_amount), 130.0)

        detail = self.client.get(reverse('api_get_order_detail'), {'id': order.id}).json()
        self.assertEqual(detail['order']['other_fee'], 25.0)
        list_row = next(
            row for row in self.client.get(reverse('api_get_orders')).json()['data']
            if row['id'] == order.id
        )
        self.assertEqual(list_row['other_fee'], 25.0)

        edit_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'discount_amount': 5,
                'shipping_fee': 10,
                'other_fee': 30,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())
        order.refresh_from_db()
        self.assertEqual(float(order.other_fee), 30.0)
        self.assertEqual(float(order.final_amount), 135.0)
        history = OrderEditHistory.objects.filter(order=order, action='update').first()
        self.assertIn('Chi phí khác', history.summary)

        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order'},
        )
        self.assertContains(print_response, 'Chi phí khác:')

    def test_quick_create_customer_persists_customer_kind(self):
        response = self.client.post(
            reverse('api_quick_create_customer'),
            data=json.dumps({
                'name': 'Khách tạo nhanh',
                'phone': '0909555666',
                'customer_kind': Customer.CUSTOMER_KIND_WHOLESALE,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        customer = Customer.objects.get(code=payload['customer']['code'])
        self.assertEqual(customer.customer_kind, Customer.CUSTOMER_KIND_WHOLESALE)
        self.assertEqual(customer.store_id, self.store.id)

    def test_products_select_keeps_negative_stock_values(self):
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse,
            quantity=-3,
        )

        response = self.client.get(reverse('api_get_products_for_select'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == self.product.id)
        warehouse_key = str(self.warehouse.id)
        self.assertEqual(row['stocks'][warehouse_key], -3.0)
        self.assertEqual(row['sellable_stocks'][warehouse_key], -3.0)
        self.assertEqual(row['total_stock'], -3.0)
        self.assertEqual(row['total_sellable_stock'], -3.0)

    def test_next_order_code_skips_soft_deleted_order(self):
        order = self._create_order(code='DH-016', status=0)
        order.delete()

        response = self.client.get(reverse('api_next_order_code'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['code'], 'DH017')
        self.assertEqual(Order.all_objects.get(id=order.id).code, 'DH-016')

    def test_order_list_always_labels_legacy_guest_customers(self):
        blank_customer = Customer.objects.create(
            store=self.store,
            code='KH-LEGACY-BLANK',
            name='',
            created_by=self.user,
        )
        without_customer = Order.objects.create(
            code='DH-GUEST-NULL',
            store=self.store,
            customer=None,
            warehouse=self.warehouse,
            status=5,
            order_date=date.today(),
            created_by=self.user,
        )
        blank_customer_order = Order.objects.create(
            code='DH-GUEST-BLANK',
            store=self.store,
            customer=blank_customer,
            warehouse=self.warehouse,
            status=5,
            order_date=date.today(),
            created_by=self.user,
        )

        rows = {
            row['code']: row
            for row in self.client.get(reverse('api_get_orders')).json()['data']
        }
        self.assertEqual(rows[without_customer.code]['customer'], 'Khách lẻ / khách vãng lai')
        self.assertEqual(rows[blank_customer_order.code]['customer'], 'Khách lẻ / khách vãng lai')
        self.assertTrue(rows[without_customer.code]['customer_is_guest'])
        self.assertFalse(rows[blank_customer_order.code]['customer_is_guest'])

    def test_save_order_auto_advances_when_requested_code_belongs_to_deleted_order(self):
        order = self._create_order(code='DH-016', status=0)
        order.delete()

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-016',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 0,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['order_code'], 'DH017')
        self.assertEqual(Order.all_objects.get(id=order.id).code, 'DH-016')

    def test_products_select_exposes_product_retail_price_as_default_price(self):
        self.product.selling_price = 12000000
        self.product.save(update_fields=['selling_price'])
        ProductVariant.objects.create(
            product=self.product,
            size_name='Legacy',
            sku='SP-ORDER-001-LEGACY',
            selling_price=10990000,
        )

        response = self.client.get(reverse('api_get_products_for_select'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == self.product.id)
        self.assertEqual(row['retail_price'], 12000000.0)
        self.assertEqual(row['price'], 12000000.0)
        self.assertEqual(row['variants'][0]['selling_price'], 10990000.0)
        self.assertEqual(row['variants'][0]['retail_price'], 12000000.0)

    def test_products_select_exposes_product_note_for_order_form(self):
        self.product.note = 'Tặng kèm dây nguồn'
        self.product.specification = 'Thùng 24 chai - 330ml'
        self.product.save(update_fields=['note', 'specification'])

        response = self.client.get(reverse('api_get_products_for_select'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('no-cache', response.headers.get('Cache-Control', ''))
        row = next(item for item in response.json()['data'] if item['id'] == self.product.id)
        self.assertEqual(row['note'], 'Tặng kèm dây nguồn')
        self.assertEqual(row['specification'], 'Thùng 24 chai - 330ml')

    def test_order_detail_exposes_product_note_for_view(self):
        self.product.note = 'Tặng kèm dây nguồn'
        self.product.specification = 'Thùng 24 chai - 330ml'
        self.product.save(update_fields=['note', 'specification'])
        order = self._create_order(code='DH-VIEW-NOTE')
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})

        self.assertEqual(response.status_code, 200)
        detail_item = response.json()['items'][0]
        self.assertEqual(detail_item['product_note'], 'Tặng kèm dây nguồn')
        self.assertEqual(detail_item['effective_note'], 'Tặng kèm dây nguồn')
        self.assertEqual(detail_item['specification'], 'Thùng 24 chai - 330ml')

        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order'},
        )
        self.assertEqual(print_response.status_code, 200)
        self.assertNotContains(print_response, 'Thùng 24 chai - 330ml')

    def test_order_item_note_defaults_from_product_then_can_be_customized_per_order(self):
        self.product.note = 'Ghi chú mặc định của sản phẩm'
        self.product.save(update_fields=['note'])

        def create_order(code):
            response = self.client.post(
                reverse('api_save_order'),
                data=json.dumps({
                    'code': code,
                    'customer_id': self.customer.id,
                    'warehouse_id': self.warehouse.id,
                    'order_date': date.today().isoformat(),
                    'discount_amount': 0,
                    'shipping_fee': 0,
                    'status': 1,
                    'note': '',
                    'tags': '',
                    'pay_mode': 'none',
                    'payment_amount': 0,
                    'payment_lines': [],
                    'items': [{
                        'product_id': self.product.id,
                        'variant_id': None,
                        'quantity': 1,
                        'unit_price': 100,
                        'discount_percent': 0,
                    }],
                }),
                content_type='application/json',
            )
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
            return Order.objects.get(id=response.json()['order_id'])

        first_order = create_order('DH-ITEM-NOTE-FIRST')
        second_order = create_order('DH-ITEM-NOTE-SECOND')
        self.assertEqual(first_order.items.get().note, 'Ghi chú mặc định của sản phẩm')
        self.assertEqual(second_order.items.get().note, 'Ghi chú mặc định của sản phẩm')

        edit_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': first_order.id,
                'code': first_order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': first_order.order_date.isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                    'note': 'Ghi chú chỉ dành cho đơn thứ nhất',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())
        first_order.refresh_from_db()
        second_order.refresh_from_db()
        self.product.refresh_from_db()
        self.assertEqual(first_order.items.get().note, 'Ghi chú chỉ dành cho đơn thứ nhất')
        self.assertEqual(second_order.items.get().note, 'Ghi chú mặc định của sản phẩm')
        self.assertEqual(self.product.note, 'Ghi chú mặc định của sản phẩm')

        first_detail = self.client.get(reverse('api_get_order_detail'), {'id': first_order.id}).json()
        second_detail = self.client.get(reverse('api_get_order_detail'), {'id': second_order.id}).json()
        self.assertEqual(first_detail['items'][0]['note'], 'Ghi chú chỉ dành cho đơn thứ nhất')
        self.assertEqual(second_detail['items'][0]['note'], 'Ghi chú mặc định của sản phẩm')
        first_print = self.client.get(
            reverse('api_print_order'),
            {'id': first_order.id, 'type': 'a4', 'source': 'order'},
        )
        second_print = self.client.get(
            reverse('api_print_order'),
            {'id': second_order.id, 'type': 'a4', 'source': 'order'},
        )
        self.assertContains(first_print, '(Ghi chú chỉ dành cho đơn thứ nhất)')
        self.assertNotContains(first_print, '(Ghi chú mặc định của sản phẩm)')
        self.assertContains(second_print, '(Ghi chú mặc định của sản phẩm)')

    def test_order_item_note_can_be_cleared_without_falling_back_to_product_note(self):
        self.product.note = 'Ghi chú gốc không được hiện lại'
        self.product.save(update_fields=['note'])
        order = self._create_order(code='DH-ITEM-NOTE-CLEAR', status=1)
        item = OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
            note='',
        )

        detail_response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})
        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order'},
        )

        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.json()['items'][0]['note'], '')
        self.assertEqual(detail_response.json()['items'][0]['effective_note'], '')
        self.assertNotContains(print_response, 'Ghi chú gốc không được hiện lại')
        item.refresh_from_db()
        self.assertEqual(item.note, '')

    def test_order_detail_and_prints_show_customer_delivery_address(self):
        self.customer.address = '123 Nguyễn Huệ, Quận 1, TP.HCM'
        self.customer.save(update_fields=['address'])
        order = self._create_order(code='DH-SHIP-ADDRESS')

        detail_response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})
        k80_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'k80', 'source': 'order'},
        )
        a4_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order'},
        )

        self.assertEqual(detail_response.json()['order']['customer_address'], self.customer.address)
        self.assertContains(k80_response, 'ĐC giao:')
        self.assertContains(k80_response, self.customer.address)
        self.assertContains(a4_response, 'Địa chỉ giao:')
        self.assertContains(a4_response, self.customer.address)

    def test_order_page_exposes_all_saved_customer_delivery_addresses(self):
        self.customer.address = 'Địa chỉ mặc định'
        self.customer.save(update_fields=['address'])
        CustomerAddress.objects.create(
            customer=self.customer,
            label='Kho Hà Nội',
            address='Số 1 Tràng Tiền, Hà Nội',
            phone='0901000001',
            sort_order=0,
        )
        CustomerAddress.objects.create(
            customer=self.customer,
            label='Chi nhánh 2',
            address='Số 2 Nguyễn Huệ, TP.HCM',
            phone='0902000002',
            sort_order=1,
        )
        self.client.force_login(self.owner)

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        customer = next(item for item in response.context['customers'] if item['id'] == self.customer.id)
        self.assertEqual(
            customer['delivery_addresses'],
            [
                {'label': 'Kho Hà Nội', 'address': 'Số 1 Tràng Tiền, Hà Nội', 'phone': '0901000001'},
                {'label': 'Chi nhánh 2', 'address': 'Số 2 Nguyễn Huệ, TP.HCM', 'phone': '0902000002'},
            ],
        )
        self.assertContains(response, 'Kho Hà Nội')
        self.assertContains(response, 'Chi nhánh 2')
        self.assertContains(response, '<option value="custom">Nhập địa chỉ khác</option>', html=True)
        self.assertContains(response, "$('#inp_customer_address_choice').html(html).prop('disabled', false)")
        self.assertContains(response, 'id="inp_shipping_phone"')
        self.assertContains(response, "$('#inp_shipping_address,#inp_shipping_phone').val('')")

    def test_order_page_exposes_distinct_shipping_addresses_from_previous_orders(self):
        first = self._create_order(code='DH-ADDRESS-HISTORY-1')
        first.shipping_address = 'Kho công trình Quận 7'
        first.shipping_phone = '0907000001'
        first.save(update_fields=['shipping_address', 'shipping_phone'])
        duplicate = self._create_order(code='DH-ADDRESS-HISTORY-2')
        duplicate.shipping_address = 'Kho công trình Quận 7'
        duplicate.shipping_phone = '0907000002'
        duplicate.save(update_fields=['shipping_address', 'shipping_phone'])
        second = self._create_order(code='DH-ADDRESS-HISTORY-3')
        second.shipping_address = 'Chi nhánh Thủ Đức'
        second.shipping_phone = '0909000003'
        second.save(update_fields=['shipping_address', 'shipping_phone'])
        other_customer_order = self._create_order(
            code='DH-ADDRESS-OTHER-CUSTOMER', customer=self.other_customer,
            store=self.other_store, warehouse=self.other_warehouse,
            created_by=self.other_user,
        )
        other_customer_order.shipping_address = 'Địa chỉ không thuộc khách này'
        other_customer_order.save(update_fields=['shipping_address'])

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        customer = next(item for item in response.context['customers'] if item['id'] == self.customer.id)
        self.assertEqual(
            [item['address'] for item in customer['historical_addresses']],
            ['Chi nhánh Thủ Đức', 'Kho công trình Quận 7', 'Kho công trình Quận 7'],
        )
        self.assertEqual(
            [item['phone'] for item in customer['historical_addresses']],
            ['0909000003', '0907000002', '0907000001'],
        )
        self.assertContains(response, 'Chi nhánh Thủ Đức')
        self.assertContains(response, 'Kho công trình Quận 7')
        self.assertNotContains(response, 'Địa chỉ không thuộc khách này')

    def test_order_page_exposes_copy_action_for_every_order_and_quick_view(self):
        self.client.force_login(self.owner)

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Mã đơn, tên khách hàng, SĐT, ghi chú, tag...')
        self.assertContains(response, 'Tạo đơn mới giống đơn này')
        self.assertContains(response, 'id="btn_quick_view_copy"')
        self.assertContains(response, 'class="quick-view-product-code-link"')
        self.assertContains(response, 'target="_blank" rel="noopener"')
        self.assertContains(response, 'renderQuickViewProductSpecification(it.specification)')
        self.assertContains(response, "'<td><div class=\"quick-view-product-name\">' + escapeHtml(it.product_name || it.item_name || '') + '</div>' + productMetaHtml")
        self.assertContains(response, 'item.effective_note')
        self.assertContains(response, '<strong>Note:</strong>')
        self.assertNotContains(response, 'quick-view-order-edit-link')
        self.assertNotContains(response, "pageParams.get('edit_order')")
        self.assertContains(response, "$('#inp_status').val('1')")
        self.assertContains(response, 'clearPaymentLines()')

    def test_order_page_exposes_amount_and_percent_discount_modes(self):
        self.client.force_login(self.owner)

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="inp_discount_mode"')
        self.assertContains(response, '<option value="amount">Số tiền</option>', html=True)
        self.assertContains(response, '<option value="percent">Phần trăm</option>', html=True)
        self.assertContains(response, 'id="lbl_discount_conversion"')
        self.assertContains(response, 'CK (% hoặc tiền)')
        self.assertContains(response, 'class="order-summary-value-column"')
        self.assertContains(response, 'order-discount-input-group')
        self.assertContains(response, 'Chiết khấu (% hoặc tiền)')
        self.assertContains(response, 'class="order-line-discount-group"')
        self.assertContains(response, 'class="custom-select inp_disc_mode"')
        self.assertContains(response, 'flex-direction: column;')
        self.assertNotContains(response, 'input-group input-group-sm order-line-discount-group')
        self.assertContains(response, '<option value="amount"', count=2)
        self.assertContains(response, 'var itemDiscountAmount = Number(it.discount_amount || 0);')
        self.assertContains(response, "$(document).on('input.orderLineDiscount', '#items_tbl .inp_disc'")
        self.assertContains(response, "$(document).on('blur.orderLineDiscount', '#items_tbl .inp_disc'")
        self.assertContains(response, "$(document).on('keydown.orderLineDiscount', '#items_tbl .inp_disc'")
        self.assertContains(response, "raw.replace(/\\B(?=(\\d{3})+(?!\\d))/g, '.')")
        self.assertContains(response, "if(event.key !== 'Enter') return;")

    def test_create_and_edit_order_persist_custom_shipping_address(self):
        self.customer.address = 'Địa chỉ mặc định của khách'
        self.customer.save(update_fields=['address'])
        create_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-CUSTOM-SHIPPING',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'shipping_address': 'Kho nhận hàng số 1',
                'shipping_phone': '0901111111',
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        order = Order.objects.get(code='DH-CUSTOM-SHIPPING')
        self.assertEqual(order.shipping_address, 'Kho nhận hàng số 1')
        self.assertEqual(order.shipping_phone, '0901111111')

        edit_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'shipping_address': 'Nhà người thân, hẻm 12',
                'shipping_phone': '0902222222',
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())
        order.refresh_from_db()
        self.assertEqual(order.shipping_address, 'Nhà người thân, hẻm 12')
        self.assertEqual(order.shipping_phone, '0902222222')

        detail_response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})
        self.assertEqual(detail_response.json()['order']['shipping_address'], 'Nhà người thân, hẻm 12')
        self.assertEqual(detail_response.json()['order']['shipping_phone'], '0902222222')
        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order'},
        )
        self.assertContains(print_response, 'Nhà người thân, hẻm 12')
        self.assertContains(print_response, '0902222222')
        self.assertNotContains(print_response, 'Địa chỉ mặc định của khách')

    def test_edit_order_same_address_new_phone_creates_customer_shipping_point(self):
        self.customer.address = '25 Nguyễn Huệ, Quận 1'
        self.customer.phone = '0901000001'
        self.customer.save(update_fields=['address', 'phone'])
        order = self._create_order(code='DH-SAME-ADDRESS-NEW-PHONE')

        payload = {
            'id': order.id,
            'code': order.code,
            'customer_id': self.customer.id,
            'warehouse_id': self.warehouse.id,
            'order_date': order.order_date.isoformat(),
            'shipping_address': '25 Nguyễn Huệ, Quận 1',
            'shipping_phone': '0902000002',
            'status': 1,
            'items': [{
                'product_id': self.product.id,
                'quantity': 1,
                'unit_price': 100,
                'discount_percent': 0,
            }],
        }
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps(payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        self.assertTrue(response.json()['shipping_point_created'])
        self.assertEqual(response.json()['shipping_point']['phone'], '0902000002')
        self.assertTrue(CustomerAddress.objects.filter(
            customer=self.customer,
            address='25 Nguyễn Huệ, Quận 1',
            phone='0902000002',
        ).exists())
        self.customer.refresh_from_db()
        self.assertEqual(self.customer.phone, '0901000001')

        duplicate_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps(payload),
            content_type='application/json',
        )
        self.assertEqual(duplicate_response.json()['status'], 'ok', msg=duplicate_response.content.decode())
        self.assertFalse(duplicate_response.json()['shipping_point_created'])
        self.assertEqual(CustomerAddress.objects.filter(
            customer=self.customer,
            address='25 Nguyễn Huệ, Quận 1',
            phone='0902000002',
        ).count(), 1)

    def test_create_order_same_address_new_phone_creates_customer_shipping_point(self):
        self.customer.address = '80 Trần Hưng Đạo, Quận 5'
        self.customer.phone = '0911000001'
        self.customer.save(update_fields=['address', 'phone'])

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-CREATE-SAME-ADDRESS-NEW-PHONE',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'shipping_address': '80 Trần Hưng Đạo, Quận 5',
                'shipping_phone': '0911000002',
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        self.assertTrue(response.json()['shipping_point_created'])
        self.assertTrue(CustomerAddress.objects.filter(
            customer=self.customer,
            address='80 Trần Hưng Đạo, Quận 5',
            phone='0911000002',
        ).exists())
        self.customer.refresh_from_db()
        self.assertEqual(self.customer.phone, '0911000001')

    def test_quotation_prints_show_customer_address(self):
        self.customer.address = '456 Lê Lợi, Quận 3, TP.HCM'
        self.customer.save(update_fields=['address'])
        quotation = self._create_quotation(code='BG-SHIP-ADDRESS', status=1)

        for print_type in ('quotation', 'quotation_a4'):
            with self.subTest(print_type=print_type):
                response = self.client.get(
                    reverse('api_print_order'),
                    {'id': quotation.id, 'type': print_type, 'source': 'quotation'},
                )
                self.assertContains(response, 'Địa chỉ giao:')
                self.assertContains(response, self.customer.address)

    def test_quotation_validity_setting_hides_field_list_column_and_print_block(self):
        quotation = self._create_quotation(code='BG-HIDE-VALIDITY', status=1)
        quotation.valid_until = date.today() + timedelta(days=30)
        quotation.save(update_fields=['valid_until'])
        config = BusinessConfig.get_config(brand=self.brand)
        config.opt_quotation_validity = False
        config.save(update_fields=['opt_quotation_validity'])

        self.client.force_login(self.owner)
        list_response = self.client.get(reverse('quotation_tbl'))
        self.assertEqual(list_response.status_code, 200)
        self.assertNotContains(list_response, '<th data-col="valid">Hiệu lực</th>', html=True)
        self.assertNotContains(list_response, '<label>Hiệu lực đến</label>', html=True)
        self.assertContains(list_response, 'type="hidden" id="inp_valid_until"')

        order_response = self.client.get(reverse('order_tbl'))
        self.assertEqual(order_response.status_code, 200)
        self.assertNotContains(order_response, '<label>Hiệu lực đến</label>', html=True)
        self.assertContains(order_response, 'type="hidden" id="inp_valid_until"')

        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': quotation.id, 'type': 'quotation_a4', 'source': 'quotation'},
        )
        self.assertEqual(print_response.status_code, 200)
        self.assertNotContains(print_response, 'Hiệu lực:')

    def test_order_form_exposes_item_sequence_sort_and_independent_scroll_area(self):
        self.client.force_login(self.owner)

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="order_item_sort"')
        self.assertContains(response, '<option value="desc">STT: cao → thấp</option>', html=True)
        self.assertContains(response, 'id="order_items_scroll"')
        self.assertContains(response, 'id="order_items_summary"')
        self.assertContains(response, 'modal-dialog-scrollable order-form-dialog')
        self.assertContains(response, 'max-width: 1600px;')
        self.assertContains(response, 'overscroll-behavior: contain;')
        self.assertContains(response, 'getOrderItemRowsInSequenceOrder().forEach(function(row)')
        self.assertContains(response, 'var itemSequence=ensureOrderItemSequence($row);')
        self.assertContains(response, 'sequence:itemSequence,', count=2)
        self.assertContains(response, 'var persistedSequence = parseInt(item && item.sequence, 10) || 0;')
        self.assertNotContains(response, "$(row).attr('data-item-sequence', index + 1);")
        self.assertNotContains(response, 'addItemRow(item, {prepend: true')

    def test_order_form_exposes_product_specification_line(self):
        self.client.force_login(self.owner)

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'order-item-product-specification')
        self.assertContains(response, 'renderOrderProductSpecification')
        self.assertContains(response, 'data-specification')
        self.assertContains(response, 'specification: product.specification ||')

    def test_order_form_product_code_opens_product_editor_in_new_tab(self):
        self.client.force_login(self.owner)

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('no-cache', response.headers.get('Cache-Control', ''))
        self.assertContains(response, "$.ajax({url: '/api/products-select/', type: 'GET', cache: false})")
        self.assertContains(response, 'class="order-product-edit-link"')
        self.assertContains(response, 'target="_blank" rel="noopener"')
        self.assertContains(response, '/product-tbl/?edit_product_id=')
        self.assertContains(response, 'function buildOrderProductEditUrl(productId)')
        self.assertContains(response, 'updateOrderRowProductEditLink($row)')
        self.assertContains(response, 'class="order-product-meta-line"')
        self.assertContains(response, 'data-specification-value=')
        self.assertContains(response, "escapeHtml(normalizedSpecification || '—')", count=2)
        self.assertContains(response, "setOrderRowProductSpecification($row, product.specification || '')")
        self.assertContains(response, '#items_tbl .order-product-edit-link-wrap')
        self.assertContains(response, '#items_tbl .order-item-product-specification')
        self.assertContains(response, '#items_tbl .order-product-meta-line')
        self.assertContains(response, 'flex-wrap: wrap;')

    def test_order_item_sequence_stays_fixed_when_quantity_is_edited_and_order_reopened(self):
        second_product = Product.objects.create(
            store=self.store,
            code='SP-ORDER-002',
            name='Sản phẩm nhập thứ hai',
            created_by=self.user,
        )
        create_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-FIXED-SEQUENCE',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 1,
                'items': [
                    {
                        'sequence': 1,
                        'product_id': self.product.id,
                        'quantity': 1,
                        'unit_price': 100,
                        'discount_percent': 0,
                    },
                    {
                        'sequence': 2,
                        'product_id': second_product.id,
                        'quantity': 1,
                        'unit_price': 200,
                        'discount_percent': 0,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        order = Order.objects.get(id=create_response.json()['order_id'])
        self.assertEqual(
            list(order.items.values_list('sequence', 'product_id')),
            [(1, self.product.id), (2, second_product.id)],
        )

        first_detail = self.client.get(reverse('api_get_order_detail'), {'id': order.id}).json()
        self.assertEqual(
            [(item['sequence'], item['product_id']) for item in first_detail['items']],
            [(1, self.product.id), (2, second_product.id)],
        )

        # Sửa số lượng khi form đang hiển thị STT cao xuống thấp: nội dung đổi nhưng STT dòng giữ nguyên.
        edit_response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'status': 1,
                'items': [
                    {
                        'sequence': 2,
                        'product_id': second_product.id,
                        'quantity': 3,
                        'unit_price': 200,
                        'discount_percent': 0,
                    },
                    {
                        'sequence': 1,
                        'product_id': self.product.id,
                        'quantity': 5,
                        'unit_price': 100,
                        'discount_percent': 0,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())
        reopened_detail = self.client.get(reverse('api_get_order_detail'), {'id': order.id}).json()
        self.assertEqual(
            [(item['sequence'], item['product_id'], item['quantity']) for item in reopened_detail['items']],
            [(1, self.product.id, 5.0), (2, second_product.id, 3.0)],
        )

    def test_order_item_model_assigns_sequence_in_creation_order(self):
        second_product = Product.objects.create(
            store=self.store,
            code='SP-MODEL-SEQUENCE-002',
            name='Sản phẩm model thứ hai',
            created_by=self.user,
        )
        order = self._create_order(code='DH-MODEL-SEQUENCE')

        first_item = OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )
        second_item = OrderItem.objects.create(
            order=order,
            product=second_product,
            quantity=1,
            unit_price=200,
            total_price=200,
        )
        first_item.quantity = 4
        first_item.total_price = 400
        first_item.save(update_fields=['quantity', 'total_price'])

        first_item.refresh_from_db()
        second_item.refresh_from_db()
        self.assertEqual((first_item.sequence, second_item.sequence), (1, 2))
        self.assertEqual(first_item.quantity, 4)
        self.assertEqual(
            list(order.items.values_list('product_id', flat=True)),
            [self.product.id, second_product.id],
        )

    def test_order_salesperson_is_searchable_beside_shipping_address(self):
        self.user.first_name = 'Lan'
        self.user.last_name = 'Anh'
        self.user.save(update_fields=['first_name', 'last_name'])
        historical_order = self._create_order(code='DH-SALESPERSON-HISTORY', status=1)
        historical_order.salesperson = 'Nhân viên lịch sử'
        historical_order.save(update_fields=['salesperson'])
        foreign_order = self._create_order(
            code='DH-SALESPERSON-FOREIGN',
            store=self.other_store,
            customer=self.other_customer,
            warehouse=self.other_warehouse,
            created_by=self.other_user,
            status=1,
        )
        foreign_order.salesperson = 'Nhân viên cửa hàng khác'
        foreign_order.save(update_fields=['salesperson'])
        config = BusinessConfig.get_config(brand=self.brand)
        config.opt_order_salesperson = True
        config.save(update_fields=['opt_order_salesperson'])

        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="shipping_salesperson_row"')
        self.assertContains(response, 'id="shipping_address_section"')
        self.assertContains(response, 'id="order_salesperson_field"')
        self.assertContains(response, 'id="inp_salesperson"')
        self.assertContains(response, 'Tìm mã hoặc tên NV...')
        self.assertContains(response, 'seller_a - Lan Anh')
        self.assertContains(response, 'Nhân viên lịch sử')
        self.assertNotContains(response, 'Nhân viên cửa hàng khác')
        self.assertNotContains(response, 'id="order_staff_section"')
        self.assertContains(response, 'id="order_item_search_section"')
        self.assertContains(response, 'id="order_item_search"')
        self.assertContains(response, 'Tìm theo mã hoặc tên sản phẩm đã có trong đơn...')
        self.assertContains(response, 'id="quick_search_product"')
        self.assertContains(response, 'function filterOrderItemRows()')
        self.assertContains(response, 'function focusExistingOrderItemSearch()')
        self.assertContains(response, "setOrderFormInitialFocus('edit')")
        self.assertContains(response, "setOrderFormInitialFocus('create')")
        self.assertContains(response, "mode === 'edit' && focusExistingOrderItemSearch()")

    def test_order_status_action_buttons_have_spacing(self):
        response = self.client.get(reverse('order_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'class="d-flex align-items-center" id="status_flow_actions" style="gap:3px;"',
        )
        self.assertContains(response, 'class="btn btn-sm btn-outline-primary" id="btn_status_next"')
        self.assertContains(response, 'class="btn btn-sm btn-outline-success" id="btn_status_save"')

    def test_products_select_exposes_combo_components_and_component_based_stock(self):
        self.product.cost_price = 100
        self.product.selling_price = 150
        self.product.save(update_fields=['cost_price', 'selling_price'])
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=5)
        combo = Product.objects.create(
            store=self.store,
            code='CB-ORDER-001',
            name='Combo bán hàng',
            unit='Bo',
            is_combo=True,
            cost_price=200,
            selling_price=250,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=2)

        response = self.client.get(reverse('api_get_products_for_select'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == combo.id)
        warehouse_key = str(self.warehouse.id)
        self.assertTrue(row['is_combo'])
        self.assertEqual(row['stocks'][warehouse_key], 2.0)
        self.assertEqual(row['total_stock'], 2.0)
        self.assertEqual(row['combo_items'][0]['product_code'], self.product.code)
        self.assertEqual(row['combo_items'][0]['unit'], self.product.unit)
        self.assertEqual(row['combo_items'][0]['line_cost'], 200.0)
        self.assertEqual(row['combo_items'][0]['total_stock'], 5.0)

    def test_get_orders_returns_paginated_meta(self):
        created_orders = []
        for idx in range(12):
            created_orders.append(self._create_order(code=f'DH-PAGE-{idx:02d}', status=1))

        response = self.client.get(reverse('api_get_orders'), {'page': 2, 'page_size': 10})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['page'], 2)
        self.assertEqual(payload['meta']['page_size'], 10)
        self.assertEqual(payload['meta']['total_filtered_count'], 12)
        self.assertEqual(payload['meta']['total_pages'], 2)
        self.assertEqual(payload['meta']['start_index'], 11)
        self.assertEqual(payload['meta']['end_index'], 12)
        self.assertEqual(len(payload['data']), 2)
        expected_ids = [order.id for order in list(reversed(created_orders))[10:12]]
        self.assertEqual([row['id'] for row in payload['data']], expected_ids)

    def test_get_orders_finds_order_code_without_separators(self):
        matching_order = self._create_order(code='DH-CODE-045', status=1)
        self._create_order(code='DH-CODE-046', status=1)

        for search_term in ('dhcode045', '# DH-CODE-045'):
            with self.subTest(search_term=search_term):
                response = self.client.get(reverse('api_get_orders'), {'text': search_term})

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    [row['id'] for row in response.json()['data']],
                    [matching_order.id],
                )

    def test_get_orders_sorts_by_created_at_not_latest_update(self):
        first_created = self._create_order(code='DH-CREATED-FIRST', status=1)
        second_created = self._create_order(code='DH-CREATED-SECOND', status=1)
        now = timezone.now()
        Order.objects.filter(id=first_created.id).update(
            created_at=now - timedelta(hours=2),
            updated_at=now,
        )
        Order.objects.filter(id=second_created.id).update(
            created_at=now - timedelta(hours=1),
            updated_at=now - timedelta(hours=3),
        )

        response = self.client.get(reverse('api_get_orders'))

        self.assertEqual(response.status_code, 200)
        returned_ids = [row['id'] for row in response.json()['data']]
        self.assertEqual(returned_ids[:2], [second_created.id, first_created.id])

    def test_get_orders_filters_by_product_on_server(self):
        other_same_store_product = Product.objects.create(
            store=self.store,
            code='SP-ORDER-FILTER',
            name='Sản phẩm lọc đơn',
            created_by=self.user,
        )
        first_order = self._create_order(code='DH-FILTER-001', status=1)
        second_order = self._create_order(code='DH-FILTER-002', status=1)
        OrderItem.objects.create(order=first_order, product=self.product, quantity=1, unit_price=100, total_price=100)
        OrderItem.objects.create(order=second_order, product=other_same_store_product, quantity=1, unit_price=100, total_price=100)

        response = self.client.get(reverse('api_get_orders'), {'product': 'FILTER'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['total_filtered_count'], 1)
        self.assertEqual([row['id'] for row in payload['data']], [second_order.id])

    def test_get_orders_filters_by_variant_and_multiple_product_terms(self):
        product = Product.objects.create(
            store=self.store,
            code='SP-AO-001',
            barcode='893000000001',
            name='Áo thun cổ tròn',
            specification='Cotton co giãn',
            sapo_id='SAPO-7788',
            created_by=self.user,
        )
        variant = ProductVariant.objects.create(
            product=product,
            size_name='Size XL',
            sku='SKU-AO-XL',
            barcode='VARIANT-893',
        )
        matching_order = self._create_order(code='DH-VARIANT-FILTER', status=1)
        OrderItem.objects.create(
            order=matching_order,
            product=product,
            variant=variant,
            quantity=1,
            unit_price=100,
            total_price=100,
            note='in logo sau lưng',
        )
        self._create_order(code='DH-VARIANT-NOT-MATCH', status=1)

        for search_term in ('SKU-AO-XL', 'VARIANT-893', 'Size XL', 'Cotton giãn', 'SAPO-7788', 'logo lưng'):
            with self.subTest(search_term=search_term):
                response = self.client.get(reverse('api_get_orders'), {'product': search_term})
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    [row['id'] for row in response.json()['data']],
                    [matching_order.id],
                )

    def test_get_orders_combines_customer_product_and_pending_export_filters(self):
        selected_customer = Customer.objects.create(
            store=self.store,
            code='KH-ORDER-FILTER',
            name='Khách cần lọc đơn',
            phone='0909123456',
            created_by=self.user,
        )
        selected_product = Product.objects.create(
            store=self.store,
            code='SP-CUSTOMER-FILTER',
            name='Sản phẩm khách đã mua',
            created_by=self.user,
        )
        matching_order = self._create_order(
            code='DH-COMBINED-MATCH',
            customer=selected_customer,
            status=2,
        )
        wrong_customer_order = self._create_order(code='DH-COMBINED-WRONG-CUSTOMER', status=2)
        wrong_product_order = self._create_order(
            code='DH-COMBINED-WRONG-PRODUCT',
            customer=selected_customer,
            status=3,
        )
        exported_order = self._create_order(
            code='DH-COMBINED-EXPORTED',
            customer=selected_customer,
            status=4,
        )
        for order, product in (
            (matching_order, selected_product),
            (wrong_customer_order, selected_product),
            (wrong_product_order, self.product),
            (exported_order, selected_product),
        ):
            OrderItem.objects.create(
                order=order,
                product=product,
                quantity=1,
                unit_price=100,
                total_price=100,
            )

        response = self.client.get(reverse('api_get_orders'), {
            'customer': selected_customer.id,
            'product': 'CUSTOMER-FILTER',
            'export_status': 'pending',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['total_filtered_count'], 1)
        self.assertEqual([row['id'] for row in payload['data']], [matching_order.id])

    def test_get_orders_pending_export_search_includes_quotations_and_unexported_orders(self):
        pending_orders = [
            self._create_order(code=f'DH-PENDING-SEARCH-{status}', status=status)
            for status in (0, 1, 2, 3)
        ]
        exported_order = self._create_order(code='DH-PENDING-SEARCH-4', status=4)
        completed_order = self._create_order(code='DH-PENDING-SEARCH-5', status=5)
        canceled_order = self._create_order(code='DH-PENDING-SEARCH-6', status=6)

        response = self.client.get(reverse('api_get_orders'), {
            'text': 'PENDING-SEARCH',
            'export_status': 'pending',
        })

        self.assertEqual(response.status_code, 200)
        returned_ids = {row['id'] for row in response.json()['data']}
        self.assertEqual(returned_ids, {order.id for order in pending_orders})
        self.assertNotIn(exported_order.id, returned_ids)
        self.assertNotIn(completed_order.id, returned_ids)
        self.assertNotIn(canceled_order.id, returned_ids)

    def test_get_orders_exported_filter_includes_exported_and_completed_only(self):
        pending_order = self._create_order(code='DH-EXPORT-FILTER-PENDING', status=3)
        exported_order = self._create_order(code='DH-EXPORT-FILTER-DONE', status=4)
        completed_order = self._create_order(code='DH-EXPORT-FILTER-COMPLETE', status=5)
        canceled_order = self._create_order(code='DH-EXPORT-FILTER-CANCELED', status=6)

        response = self.client.get(reverse('api_get_orders'), {'export_status': 'exported'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        returned_ids = {row['id'] for row in payload['data']}
        self.assertEqual(returned_ids, {exported_order.id, completed_order.id})
        self.assertNotIn(pending_order.id, returned_ids)
        self.assertNotIn(canceled_order.id, returned_ids)

    def test_export_orders_excel_uses_customer_product_and_export_filters(self):
        selected_customer = Customer.objects.create(
            store=self.store,
            code='KH-EXPORT-FILTER',
            name='Khách xuất file lọc',
            created_by=self.user,
        )
        selected_product = Product.objects.create(
            store=self.store,
            code='SP-EXPORT-FILTER',
            name='Sản phẩm xuất file lọc',
            created_by=self.user,
        )
        matching_order = self._create_order(
            code='DH-EXCEL-FILTER-MATCH',
            customer=selected_customer,
            status=2,
        )
        exported_order = self._create_order(
            code='DH-EXCEL-FILTER-EXPORTED',
            customer=selected_customer,
            status=4,
        )
        for order in (matching_order, exported_order):
            OrderItem.objects.create(
                order=order,
                product=selected_product,
                quantity=1,
                unit_price=100,
                total_price=100,
            )

        response = self.client.get('/api/orders/export-excel/', {
            'customer': selected_customer.id,
            'product': 'SP-EXPORT-FILTER',
            'export_status': 'pending',
        })

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        worksheet = workbook.active
        exported_codes = {
            worksheet.cell(row=row_index, column=2).value
            for row_index in range(2, worksheet.max_row + 1)
        }
        self.assertIn(matching_order.code, exported_codes)
        self.assertNotIn(exported_order.code, exported_codes)

    def test_export_orders_excel_adds_goods_totals_grouped_by_customer(self):
        customer_a = Customer.objects.create(
            store=self.store,
            code='KH-EXCEL-A',
            name='Khách tổng hợp A',
            phone='0901000001',
            created_by=self.user,
        )
        customer_b = Customer.objects.create(
            store=self.store,
            code='KH-EXCEL-B',
            name='Khách tổng hợp B',
            phone='0901000002',
            created_by=self.user,
        )
        order_specs = [
            ('DH-EXCEL-SUMMARY-A1', customer_a, 5, 120000),
            ('DH-EXCEL-SUMMARY-A2', customer_a, 5, 180000),
            ('DH-EXCEL-SUMMARY-B1', customer_b, 5, 150000),
            ('DH-EXCEL-SUMMARY-EXCLUDED', customer_a, 6, 999000),
        ]
        for code, customer, status, goods_total in order_specs:
            order = self._create_order(code=code, customer=customer, status=status)
            order.total_amount = goods_total
            order.final_amount = goods_total
            order.save(update_fields=['total_amount', 'final_amount'])

        response = self.client.get('/api/orders/export-excel/', {'status': '5'})

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ['DANH SÁCH ĐƠN HÀNG', 'Tổng tiền theo khách'],
        )

        summary_sheet = workbook['Tổng tiền theo khách']
        self.assertEqual(
            [cell.value for cell in summary_sheet[4]],
            ['STT', 'Mã khách hàng', 'Khách hàng', 'Số điện thoại', 'Số đơn hàng', 'Tổng tiền hàng'],
        )
        self.assertEqual(
            [summary_sheet.cell(row=5, column=column).value for column in range(1, 7)],
            [1, customer_a.code, customer_a.name, customer_a.phone, 2, 300000],
        )
        self.assertEqual(
            [summary_sheet.cell(row=6, column=column).value for column in range(1, 7)],
            [2, customer_b.code, customer_b.name, customer_b.phone, 1, 150000],
        )
        self.assertEqual(
            [summary_sheet.cell(row=7, column=column).value for column in (3, 5, 6)],
            ['TỔNG CỘNG', 3, 450000],
        )

    def test_get_orders_status_counts_ignore_active_status_filter(self):
        self._create_order(code='DH-COUNT-001', status=1)
        self._create_order(code='DH-COUNT-002', status=1)
        completed_order = self._create_order(code='DH-COUNT-003', status=5)

        response = self.client.get(reverse('api_get_orders'), {'status': 1})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(all(row['status'] == 1 for row in payload['data']))
        self.assertEqual(payload['meta']['status_counts']['all'], 3)
        self.assertEqual(payload['meta']['status_counts']['1'], 2)
        self.assertEqual(payload['meta']['status_counts']['5'], 1)
        self.assertNotIn(completed_order.id, [row['id'] for row in payload['data']])

    def test_save_order_allows_financial_edit_before_completion_when_receipt_exists(self):
        order = self._create_order(code='DH-PAID-EDIT-001', status=0)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )
        Receipt.objects.create(
            code='PT-PAID-EDIT-001',
            store=self.store,
            customer=self.customer,
            order=order,
            amount=100,
            receipt_date=date.today(),
            status=1,
            description='Phiếu thu thủ công',
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'status': 1,
                'discount_amount': 10,
                'shipping_fee': 0,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order.refresh_from_db()
        self.assertEqual(order.status, 1)
        self.assertEqual(order.discount_amount, 10)
        self.assertEqual(order.final_amount, 90)
        self.assertEqual(order.payment_status, 2)

    def test_save_order_clamps_final_amount_to_zero_when_discount_exceeds_total(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-DISCOUNT-CLAMP',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 150,
                'shipping_fee': 0,
                'status': 0,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(id=payload['order_id'])
        self.assertEqual(float(order.total_amount), 100.0)
        self.assertEqual(float(order.final_amount), 0.0)
        self.assertEqual(order.payment_status, 2)

    def test_save_order_converts_percent_discount_to_amount_and_preserves_mode(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-DISCOUNT-PERCENT',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_mode': 'percent',
                'discount_percent': 5,
                'discount_amount': 999,
                'shipping_fee': 0,
                'other_fee': 0,
                'status': 1,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 2,
                    'unit_price': 500,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(code='DH-DISCOUNT-PERCENT')
        self.assertEqual(order.discount_mode, 'percent')
        self.assertEqual(float(order.discount_percent), 5.0)
        self.assertEqual(float(order.discount_amount), 50.0)
        self.assertEqual(float(order.final_amount), 950.0)

        detail_response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})
        detail = detail_response.json()['order']
        self.assertEqual(detail['discount_mode'], 'percent')
        self.assertEqual(detail['discount_percent'], 5.0)
        self.assertEqual(detail['discount_amount'], 50.0)

    def test_save_order_line_accepts_amount_discount_and_preserves_mode(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-LINE-DISCOUNT-AMOUNT',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_mode': 'amount',
                'discount_amount': 0,
                'shipping_fee': 0,
                'other_fee': 0,
                'status': 1,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 2,
                    'unit_price': 500,
                    'discount_mode': 'amount',
                    'discount_amount': 150,
                    'discount_percent': 99,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(code='DH-LINE-DISCOUNT-AMOUNT')
        item = order.items.get()
        self.assertEqual(float(order.total_amount), 850.0)
        self.assertEqual(float(order.final_amount), 850.0)
        self.assertEqual(item.discount_mode, 'amount')
        self.assertEqual(float(item.discount_amount), 150.0)
        self.assertEqual(float(item.discount_percent), 15.0)
        self.assertEqual(float(item.total_price), 850.0)

        detail_response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})
        detail_item = detail_response.json()['items'][0]
        self.assertEqual(detail_item['discount_mode'], 'amount')
        self.assertEqual(detail_item['discount_amount'], 150.0)
        self.assertEqual(detail_item['discount_percent'], 15.0)

        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'a4', 'source': 'order'},
        )
        self.assertContains(print_response, '150đ')

    def test_save_quotation_line_accepts_amount_discount_and_preserves_mode(self):
        response = self.client.post(
            reverse('api_save_quotation'),
            data=json.dumps({
                'code': 'BG-LINE-DISCOUNT-AMOUNT',
                'customer_id': self.customer.id,
                'quotation_date': date.today().isoformat(),
                'discount_mode': 'amount',
                'discount_amount': 0,
                'shipping_fee': 0,
                'other_fee': 0,
                'status': 0,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 2,
                    'unit_price': 300,
                    'discount_mode': 'amount',
                    'discount_amount': 120,
                    'discount_percent': 88,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        quotation = Quotation.objects.get(code='BG-LINE-DISCOUNT-AMOUNT')
        item = quotation.items.get()
        self.assertEqual(float(quotation.total_amount), 480.0)
        self.assertEqual(float(quotation.final_amount), 480.0)
        self.assertEqual(item.discount_mode, 'amount')
        self.assertEqual(float(item.discount_amount), 120.0)
        self.assertEqual(float(item.discount_percent), 20.0)
        self.assertEqual(float(item.total_price), 480.0)

        self.product.specification = 'Thùng 24 chai - 330ml'
        self.product.save(update_fields=['specification'])

        detail_response = self.client.get(reverse('api_get_quotation_detail'), {'id': quotation.id})
        detail_item = detail_response.json()['items'][0]
        self.assertEqual(detail_item['discount_mode'], 'amount')
        self.assertEqual(detail_item['discount_amount'], 120.0)
        self.assertEqual(detail_item['discount_percent'], 20.0)
        self.assertEqual(detail_item['specification'], 'Thùng 24 chai - 330ml')

        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': quotation.id, 'type': 'quotation_a4', 'source': 'quotation'},
        )
        self.assertContains(print_response, '120đ')

    def test_partial_payment_converts_quote_status_to_order(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-PARTIAL-TO-ORDER',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 0,
                'note': '',
                'tags': '',
                'pay_mode': 'partial',
                'payment_amount': 40,
                'payment_lines': [{'amount': 40, 'payment_method': 2}],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(id=payload['order_id'])
        self.assertEqual(order.status, 1)
        self.assertEqual(order.payment_status, 1)
        self.assertEqual(float(order.paid_amount), 40.0)

    def test_export_with_full_payment_auto_completes_and_deducts_stock_once(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=5)

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-EXPORT-PAID',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 4,
                'note': '',
                'tags': '',
                'pay_mode': 'full',
                'payment_amount': 100,
                'payment_lines': [{
                    'amount': 100,
                    'payment_method': 1,
                    'payment_method_option_id': self.payment_method.id,
                    'cash_book_id': self.cashbook.id,
                }],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(id=payload['order_id'])
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse)
        self.assertEqual(order.status, 5)
        self.assertEqual(order.payment_status, 2)
        self.assertEqual(float(stock.quantity), 4.0)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 1000100.0)

    def test_new_order_cannot_start_completed_even_when_fully_paid(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-COMPLETE-WITHOUT-EXPORT',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 5,
                'note': '',
                'tags': '',
                'pay_mode': 'full',
                'payment_amount': 100,
                'payment_lines': [{'amount': 100, 'payment_method': 2}],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('xuất kho', payload['message'])
        self.assertFalse(Order.objects.filter(code='DH-COMPLETE-WITHOUT-EXPORT').exists())

    def test_exported_order_cannot_complete_until_fully_paid(self):
        order = self._create_order(code='DH-EXPORT-PARTIAL', status=4)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )
        Receipt.objects.create(
            code='PT-EXPORT-PARTIAL',
            store=self.store,
            customer=self.customer,
            order=order,
            amount=40,
            receipt_date=date.today(),
            status=1,
            description='Thu một phần',
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 5,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('thanh toán đủ', payload['message'])

        order.refresh_from_db()
        self.assertEqual(order.status, 4)

    def test_exported_order_rejects_price_change_and_item_removal(self):
        order = self._create_order(code='DH-EXPORT-LOCKED', status=4)
        item = OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        base_payload = {
            'id': order.id,
            'code': order.code,
            'customer_id': self.customer.id,
            'warehouse_id': self.warehouse.id,
            'order_date': order.order_date.isoformat(),
            'discount_amount': 0,
            'shipping_fee': 0,
            'status': 4,
            'note': '',
            'tags': '',
            'pay_mode': 'none',
            'payment_amount': 0,
            'payment_lines': [],
        }

        changed_price_payload = {
            **base_payload,
            'items': [{
                'product_id': self.product.id,
                'variant_id': None,
                'quantity': 1,
                'unit_price': 90,
                'discount_percent': 0,
            }],
        }
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps(changed_price_payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('đã xuất kho', payload['message'])
        self.assertIn('sản phẩm/giá/số lượng', payload['message'])
        item.refresh_from_db()
        self.assertEqual(item.unit_price, 100)

        removed_item_payload = {**base_payload, 'items': []}
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps(removed_item_payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('đã xuất kho', payload['message'])
        self.assertEqual(order.items.count(), 1)

    def test_save_order_allows_below_cost_with_warning(self):
        self.product.cost_price = 150
        self.product.import_price = 150
        self.product.save(update_fields=['cost_price', 'import_price'])

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-BELOW-COST-WARN',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertIn('Cảnh báo', payload['message'])

        order = Order.objects.get(code='DH-BELOW-COST-WARN')
        self.assertTrue(order.below_listed_price_warning)
        self.assertTrue(order.items.first().is_below_listed)

    def test_save_order_allows_zero_price_gift_with_below_cost_warning(self):
        self.product.cost_price = 150
        self.product.import_price = 150
        self.product.save(update_fields=['cost_price', 'import_price'])

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-GIFT-ZERO-PRICE',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 1,
                'note': 'Tặng khách',
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 0,
                    'discount_percent': 0,
                    'note': 'Hàng tặng',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertIn('Đơn vẫn được lưu', payload['message'])

        order = Order.objects.get(code='DH-GIFT-ZERO-PRICE')
        item = order.items.get(product=self.product)
        self.assertEqual(float(order.final_amount), 0.0)
        self.assertEqual(float(item.unit_price), 0.0)
        self.assertEqual(float(item.cost_price), 150.0)
        self.assertTrue(order.below_listed_price_warning)
        self.assertTrue(item.is_below_listed)

    def test_save_order_falls_back_to_import_price_when_cost_price_is_zero(self):
        self.product.cost_price = 0
        self.product.import_price = 75
        self.product.save(update_fields=['cost_price', 'import_price'])

        product_response = self.client.get(reverse('api_get_products_for_select'))
        product_row = next(
            item for item in product_response.json()['data']
            if item['id'] == self.product.id
        )
        self.assertEqual(product_row['cost_price'], 75.0)

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-COST-FALLBACK-001',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'other_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 2,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        item = Order.objects.get(id=response.json()['order_id']).items.get()
        self.assertEqual(float(item.cost_price), 75.0)

    def test_save_order_calculates_combo_cost_from_component_import_prices(self):
        self.product.cost_price = 0
        self.product.import_price = 40
        self.product.save(update_fields=['cost_price', 'import_price'])
        combo = Product.objects.create(
            store=self.store,
            code='CB-COST-FALLBACK-001',
            name='Combo tính vốn từ thành phần',
            unit='Bộ',
            is_combo=True,
            cost_price=0,
            import_price=0,
            selling_price=120,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=2)

        product_response = self.client.get(reverse('api_get_products_for_select'))
        combo_row = next(item for item in product_response.json()['data'] if item['id'] == combo.id)
        self.assertEqual(combo_row['cost_price'], 80.0)
        self.assertEqual(combo_row['combo_items'][0]['line_cost'], 80.0)

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-COMBO-COST-FALLBACK-001',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'other_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': combo.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 120,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        item = Order.objects.get(id=response.json()['order_id']).items.get()
        self.assertEqual(float(item.cost_price), 80.0)

    def test_completed_order_note_can_be_updated(self):
        order = self._create_order(code='DH-COMPLETE-NOTE', status=5)

        response = self.client.post(
            reverse('api_update_order_note'),
            data=json.dumps({'id': order.id, 'note': 'Ghi chú sau hoàn thành'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order.refresh_from_db()
        self.assertEqual(order.note, 'Ghi chú sau hoàn thành')

    def test_canceled_order_note_remains_locked(self):
        order = self._create_order(code='DH-CANCEL-NOTE', status=6)

        response = self.client.post(
            reverse('api_update_order_note'),
            data=json.dumps({'id': order.id, 'note': 'Không được sửa'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Hủy', payload['message'])

        order.refresh_from_db()
        self.assertNotEqual(order.note, 'Không được sửa')

    def test_exported_order_note_remains_locked(self):
        order = self._create_order(code='DH-EXPORT-NOTE', status=4)

        response = self.client.post(
            reverse('api_update_order_note'),
            data=json.dumps({'id': order.id, 'note': 'Không được sửa sau xuất kho'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('xuất kho', payload['message'])

        order.refresh_from_db()
        self.assertNotEqual(order.note, 'Không được sửa sau xuất kho')

    def test_cancel_order_reopens_linked_quotation(self):
        quotation = self._create_quotation(code='BG-CANCEL-001')
        order = self._create_order(code='DH-CANCEL-001', quotation=quotation)

        response = self.client.post(
            reverse('api_cancel_order'),
            data=json.dumps({'id': order.id, 'reason': 'Kiểm tra test'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')

        order.refresh_from_db()
        quotation.refresh_from_db()
        self.assertEqual(order.status, 6)
        self.assertEqual(order.payment_status, 0)
        self.assertEqual(order.paid_amount, 0)
        self.assertEqual(quotation.status, 2)

    def test_delete_order_soft_deletes_and_reopens_linked_quotation(self):
        quotation = self._create_quotation(code='BG-DELETE-001')
        order = self._create_order(code='DH-DELETE-001', quotation=quotation)

        response = self.client.post(
            reverse('api_delete_order'),
            data=json.dumps({'id': order.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')

        deleted_order = Order.all_objects.get(id=order.id)
        quotation.refresh_from_db()
        self.assertTrue(deleted_order.is_deleted)
        self.assertEqual(quotation.status, 2)

    def test_bulk_cancel_reopens_linked_quotation(self):
        quotation = self._create_quotation(code='BG-BULK-001')
        order = self._create_order(code='DH-BULK-001', quotation=quotation)

        response = self.client.post(
            reverse('api_bulk_cancel_orders'),
            data=json.dumps({'ids': [order.id], 'reason': 'Bulk test'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')

        order.refresh_from_db()
        quotation.refresh_from_db()
        self.assertEqual(order.status, 6)
        self.assertEqual(quotation.status, 2)

    def test_bulk_collect_returns_total_collected_for_reconciliation(self):
        first_order = self._create_order(code='DH-BULK-COLLECT-001')
        second_order = self._create_order(code='DH-BULK-COLLECT-002')
        second_order.final_amount = 150
        second_order.save(update_fields=['final_amount'])
        Receipt.objects.create(
            code='PT-BULK-COLLECT-PARTIAL',
            store=self.store,
            customer=self.customer,
            order=second_order,
            amount=50,
            receipt_date=date.today(),
            status=1,
            description='Thu một phần',
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_bulk_collect_orders'),
            data=json.dumps({'ids': [first_order.id, second_order.id], 'payment_method': 2}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['collected_count'], 2)
        self.assertEqual(payload['total_collected'], 200.0)
        self.assertIn('Tổng đã thu', payload['message'])

        first_order.refresh_from_db()
        second_order.refresh_from_db()
        self.assertEqual(float(first_order.paid_amount), 100.0)
        self.assertEqual(float(second_order.paid_amount), 150.0)
        self.assertEqual(first_order.payment_status, 2)
        self.assertEqual(second_order.payment_status, 2)

    def test_bulk_collect_accepts_line_amounts_across_customers(self):
        other_customer_same_store = Customer.objects.create(
            store=self.store,
            code='KH003',
            name='Customer C',
            created_by=self.user,
        )
        first_order = self._create_order(code='DH-BULK-LINE-001', customer=self.customer)
        second_order = self._create_order(
            code='DH-BULK-LINE-002',
            customer=other_customer_same_store,
        )
        second_order.final_amount = 200
        second_order.save(update_fields=['final_amount'])

        response = self.client.post(
            reverse('api_bulk_collect_orders'),
            data=json.dumps({
                'ids': [first_order.id, second_order.id],
                'payment_method_option_id': self.payment_method.id,
                'payments': [
                    {'order_id': first_order.id, 'amount': 40},
                    {'order_id': second_order.id, 'amount': 200},
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['collected_count'], 2)
        self.assertEqual(payload['total_collected'], 240.0)

        first_order.refresh_from_db()
        second_order.refresh_from_db()
        self.assertEqual(float(first_order.paid_amount), 40.0)
        self.assertEqual(first_order.payment_status, 1)
        self.assertEqual(float(second_order.paid_amount), 200.0)
        self.assertEqual(second_order.payment_status, 2)

    def test_bulk_collect_can_use_payment_date_before_order_creation_and_complete_order(self):
        order = self._create_order(code='DH-BULK-BACKDATE-001', status=4)
        actual_payment_date = date.today() - timedelta(days=1)

        response = self.client.post(
            reverse('api_bulk_collect_orders'),
            data=json.dumps({
                'ids': [order.id],
                'payments': [{'order_id': order.id, 'amount': 100}],
                'payment_method_option_id': self.payment_method.id,
                'receipt_date': actual_payment_date.isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        order.refresh_from_db()
        self.assertEqual(order.status, 5)
        receipt = Receipt.objects.get(order=order, status=1)
        self.assertEqual(receipt.receipt_date, actual_payment_date)
        history = OrderEditHistory.objects.get(order=order, action='bulk_collect')
        self.assertIn(actual_payment_date.strftime('%d/%m/%Y'), history.summary)

    def test_collect_order_payment_accepts_multiple_methods_without_saving_order(self):
        order = self._create_order(code='DH-COLLECT-ONE-001', status=1)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.post(
            reverse('api_collect_order_payment'),
            data=json.dumps({
                'order_id': order.id,
                'payment_lines': [
                    {
                        'amount': 40,
                        'payment_method': 1,
                        'payment_method_option_id': self.payment_method.id,
                        'cash_book_id': self.cashbook.id,
                    },
                    {
                        'amount': 60,
                        'payment_method': 1,
                        'payment_method_option_id': self.payment_method.id,
                        'cash_book_id': self.cashbook.id,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order.refresh_from_db()
        self.assertEqual(order.status, 1)
        self.assertEqual(order.payment_status, 2)
        self.assertEqual(float(order.paid_amount), 100.0)
        self.assertEqual(Receipt.objects.filter(order=order, status=1).count(), 2)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 1000100.0)

    def test_collect_order_payment_accepts_date_before_order_creation(self):
        order = self._create_order(code='DH-COLLECT-DATE-BEFORE')
        actual_payment_date = date.today() - timedelta(days=1)

        response = self.client.post(
            reverse('api_collect_order_payment'),
            data=json.dumps({
                'order_id': order.id,
                'receipt_date': actual_payment_date.isoformat(),
                'payment_lines': [{'amount': 100, 'payment_method': 2}],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        receipt = Receipt.objects.get(order=order, status=1)
        self.assertEqual(receipt.receipt_date, actual_payment_date)

    def test_collect_order_payment_rejects_future_date(self):
        order = self._create_order(code='DH-COLLECT-DATE-FUTURE')
        invalid_date = date.today() + timedelta(days=1)

        response = self.client.post(
            reverse('api_collect_order_payment'),
            data=json.dumps({
                'order_id': order.id,
                'receipt_date': invalid_date.isoformat(),
                'payment_lines': [{'amount': 100, 'payment_method': 2}],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('vượt quá ngày hiện tại', payload['message'])
        self.assertFalse(Receipt.objects.filter(order=order).exists())

    def test_collect_order_payment_can_be_called_repeatedly_without_saving_order(self):
        order = self._create_order(code='DH-COLLECT-REPEAT-001', status=1)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        first_response = self.client.post(
            reverse('api_collect_order_payment'),
            data=json.dumps({
                'order_id': order.id,
                'payment_lines': [{
                    'amount': 40,
                    'payment_method': 1,
                    'payment_method_option_id': self.payment_method.id,
                    'cash_book_id': self.cashbook.id,
                }],
            }),
            content_type='application/json',
        )
        second_response = self.client.post(
            reverse('api_collect_order_payment'),
            data=json.dumps({
                'order_id': order.id,
                'payment_lines': [{
                    'amount': 60,
                    'payment_method': 1,
                    'payment_method_option_id': self.payment_method.id,
                    'cash_book_id': self.cashbook.id,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(first_response.status_code, 200)
        first_payload = first_response.json()
        self.assertEqual(first_payload['status'], 'ok', msg=first_response.content.decode())
        self.assertEqual(first_payload['paid_amount'], 40.0)
        self.assertEqual(first_payload['remaining_amount'], 60.0)
        self.assertEqual(first_payload['payment_status'], 1)

        self.assertEqual(second_response.status_code, 200)
        second_payload = second_response.json()
        self.assertEqual(second_payload['status'], 'ok', msg=second_response.content.decode())
        self.assertEqual(second_payload['paid_amount'], 100.0)
        self.assertEqual(second_payload['remaining_amount'], 0.0)
        self.assertEqual(second_payload['payment_status'], 2)

        order.refresh_from_db()
        self.assertEqual(float(order.paid_amount), 100.0)
        self.assertEqual(order.payment_status, 2)
        self.assertEqual(Receipt.objects.filter(order=order, status=1).count(), 2)
        history_summaries = list(
            OrderEditHistory.objects.filter(order=order, action='payment').values_list('summary', flat=True)
        )
        self.assertTrue(any('còn phải thu 60' in summary for summary in history_summaries))
        self.assertTrue(any('đã thanh toán đủ' in summary for summary in history_summaries))

    def test_export_order_stock_is_independent_from_payment(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=5)
        order = self._create_order(code='DH-EXPORT-NO-PAY-001', status=1)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=50,
            total_price=100,
        )

        response = self.client.post(
            reverse('api_export_order_stock'),
            data=json.dumps({'order_id': order.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order.refresh_from_db()
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse)
        self.assertEqual(order.status, 4)
        self.assertEqual(order.payment_status, 0)
        self.assertEqual(float(stock.quantity), 3.0)
        self.assertFalse(Receipt.objects.filter(order=order).exists())

    def test_export_order_stock_persists_warehouse_selected_in_form(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=5)
        order = Order.objects.create(
            code='DH-EXPORT-FORM-WAREHOUSE-001',
            store=self.store,
            customer=self.customer,
            warehouse=None,
            status=1,
            total_amount=100,
            final_amount=100,
            order_date=date.today(),
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.post(
            reverse('api_export_order_stock'),
            data=json.dumps({'order_id': order.id, 'warehouse_id': self.warehouse.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        order.refresh_from_db()
        self.assertEqual(order.warehouse_id, self.warehouse.id)
        self.assertEqual(order.status, 4)

    def test_explicit_order_stock_export_allows_negative_stock(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Negative stock disabled for automatic flows',
            opt_allow_negative_stock=False,
        )
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse,
            quantity=0,
        )
        order = self._create_order(code='DH-EXPLICIT-NEGATIVE-STOCK', status=1)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=50,
            total_price=100,
        )

        response = self.client.post(
            reverse('api_export_order_stock'),
            data=json.dumps({'order_id': order.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        order.refresh_from_db()
        stock.refresh_from_db()
        self.assertEqual(order.status, 4)
        self.assertEqual(float(stock.quantity), -2.0)

    def test_update_order_status_api_moves_steps_without_full_order_save(self):
        order = self._create_order(code='DH-STATUS-FAST-001', status=1)

        first_response = self.client.post(
            reverse('api_update_order_status'),
            data=json.dumps({'order_id': order.id, 'status': 2}),
            content_type='application/json',
        )
        second_response = self.client.post(
            reverse('api_update_order_status'),
            data=json.dumps({'order_id': order.id, 'status': 3}),
            content_type='application/json',
        )

        self.assertEqual(first_response.json()['status'], 'ok', msg=first_response.content.decode())
        self.assertEqual(second_response.json()['status'], 'ok', msg=second_response.content.decode())
        order.refresh_from_db()
        self.assertEqual(order.status, 3)

    def test_save_order_preserves_original_creator_when_edited_by_another_user(self):
        self.user.first_name = 'Ngọc'
        self.user.save(update_fields=['first_name'])
        editor = User.objects.create_user(
            username='tan_hop',
            password='pass123',
            first_name='Tấn',
            last_name='Hợp',
        )
        UserProfile.objects.create(user=editor, store=self.store)
        order = self._create_order(code='DH-CREATOR-FIX-001', status=1, created_by=self.user)
        order.creator_name = self.user.get_full_name()
        order.save(update_fields=['creator_name'])

        self.client.force_login(editor)
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 1,
                'note': 'Tấn Hợp sửa đơn',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        order.refresh_from_db()
        self.assertEqual(order.created_by_id, self.user.id)
        self.assertEqual(order.creator_name, 'Ngọc')

        history = OrderEditHistory.objects.filter(order=order, action='update').first()
        self.assertIsNotNone(history)
        self.assertEqual(history.actor_id, editor.id)

    def test_order_list_creator_display_uses_original_created_by(self):
        self.user.first_name = 'Ngọc'
        self.user.save(update_fields=['first_name'])
        order = self._create_order(code='DH-CREATOR-DISPLAY-001', status=1, created_by=self.user)
        order.creator_name = 'Ngân'
        order.save(update_fields=['creator_name'])

        response = self.client.get(reverse('api_get_orders'), {'text': order.code})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload['data'] if item['id'] == order.id)
        self.assertEqual(row['creator_name'], 'Ngọc')

    def test_save_order_history_describes_item_quantity_change(self):
        order = self._create_order(code='DH-HISTORY-ITEM-001', status=1)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=50,
            total_price=100,
        )

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 3,
                    'unit_price': 50,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        history = OrderEditHistory.objects.filter(order=order, action='update').first()
        self.assertIsNotNone(history)
        self.assertIn('Sửa SP', history.summary)
        self.assertIn('SL 2 → 3', history.summary)
        self.assertIn('Tổng thanh toán: 100đ → 150đ', history.summary)
        self.assertLess(
            history.summary.index('Sửa SP'),
            history.summary.index('Tổng thanh toán'),
        )

    def test_save_order_history_prioritizes_deleted_item_before_amount_change(self):
        deleted_product = Product.objects.create(
            store=self.store,
            code='SP-ORDER-DELETE',
            name='Sản phẩm bị xóa khỏi đơn',
            created_by=self.user,
        )
        order = self._create_order(code='DH-HISTORY-DELETE-001', status=1)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )
        OrderItem.objects.create(
            order=order,
            product=deleted_product,
            quantity=1,
            unit_price=50,
            total_price=50,
        )
        order.total_amount = 150
        order.final_amount = 150
        order.save(update_fields=['total_amount', 'final_amount'])

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'id': order.id,
                'code': order.code,
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': order.order_date.isoformat(),
                'discount_amount': 0,
                'shipping_fee': 0,
                'status': 1,
                'note': '',
                'tags': '',
                'pay_mode': 'none',
                'payment_amount': 0,
                'payment_lines': [],
                'items': [{
                    'product_id': self.product.id,
                    'variant_id': None,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        history = OrderEditHistory.objects.filter(order=order, action='update').first()
        self.assertIsNotNone(history)
        self.assertIn('Xóa SP "SP-ORDER-DELETE - Sản phẩm bị xóa khỏi đơn"', history.summary)
        self.assertIn('thành tiền 50đ', history.summary)
        self.assertIn('Tổng thanh toán: 150đ → 100đ', history.summary)
        self.assertLess(
            history.summary.index('Xóa SP'),
            history.summary.index('Tổng thanh toán'),
        )

    def test_store_scope_blocks_foreign_order_detail(self):
        quotation = self._create_quotation(
            code='BG-OTHER-001',
            store=self.other_store,
            customer=self.other_customer,
            created_by=self.other_user,
        )
        other_order = self._create_order(
            code='DH-OTHER-001',
            quotation=quotation,
            store=self.other_store,
            customer=self.other_customer,
            warehouse=self.other_warehouse,
            created_by=self.other_user,
        )

        response = self.client.get(reverse('api_get_order_detail'), {'id': other_order.id})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertEqual(payload['message'], 'Không tìm thấy')

    def test_save_order_rejects_foreign_warehouse(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-FOREIGN-WH',
                'customer_id': self.customer.id,
                'warehouse_id': self.other_warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 0,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Kho xuất', payload['message'])
        self.assertFalse(Order.objects.filter(code='DH-FOREIGN-WH').exists())

    def test_save_order_rejects_foreign_product(self):
        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-FOREIGN-PRODUCT',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 0,
                'items': [{
                    'product_id': self.other_product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('sản phẩm', payload['message'].lower())
        self.assertFalse(Order.objects.filter(code='DH-FOREIGN-PRODUCT').exists())

    def test_new_exported_order_is_saved_without_export_when_negative_stock_disabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Negative stock disabled',
            opt_allow_negative_stock=False,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=0)

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-NEG-STOCK',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 4,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertTrue(payload['stock_export_deferred'])
        self.assertIn('Tồn kho không đủ', payload['stock_export_warning'])
        order = Order.objects.get(code='DH-NEG-STOCK')
        self.assertEqual(order.status, 3)
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse)
        self.assertEqual(float(stock.quantity), 0.0)

    def test_save_exported_order_allows_negative_stock_when_enabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Negative stock enabled',
            opt_allow_negative_stock=True,
        )
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse,
            quantity=0,
        )

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-ALLOW-NEG-STOCK',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 4,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock.refresh_from_db()
        self.assertEqual(float(stock.quantity), -1.0)

    def test_stock_export_uses_warehouse_brand_negative_stock_config(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Warehouse negative stock enabled',
            opt_allow_negative_stock=True,
        )
        mismatched_brand = Brand.objects.create(name='Legacy Order Brand', owner=self.owner)
        mismatched_store = Store.objects.create(
            brand=mismatched_brand,
            name='Legacy Order Store',
            code='LEGACY-STOCK',
        )
        BusinessConfig.objects.create(
            brand=mismatched_brand,
            business_name='Legacy negative stock disabled',
            opt_allow_negative_stock=False,
        )
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse,
            quantity=0,
        )
        order = self._create_order(
            code='DH-WAREHOUSE-CONFIG',
            store=mismatched_store,
            warehouse=self.warehouse,
            status=1,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        from orders.views import _apply_order_stock_adjustment
        _apply_order_stock_adjustment(order, direction=-1)

        stock.refresh_from_db()
        self.assertEqual(float(stock.quantity), -1.0)

    def test_pos_checkout_rejects_negative_stock_when_disabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='POS negative stock disabled',
            opt_allow_negative_stock=False,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=0)

        response = self.client.post(
            reverse('api_pos_checkout'),
            data=json.dumps({
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
                'discount_amount': 0,
                'paid_amount': 100,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Tồn kho không đủ', payload['message'])
        self.assertFalse(Order.objects.filter(code__startswith='POS-').exists())

    def test_pos_checkout_without_payment_stays_exported(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=1)

        response = self.client.post(
            reverse('api_pos_checkout'),
            data=json.dumps({
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
                'discount_amount': 0,
                'paid_amount': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(id=payload['order_id'])
        self.assertEqual(order.status, 4)
        self.assertEqual(order.payment_status, 0)
        self.assertEqual(float(order.paid_amount), 0.0)
        self.assertEqual(order.receipts.count(), 0)

    def test_quotation_detail_blocks_foreign_quotation(self):
        quotation = self._create_quotation(
            code='BG-FOREIGN-DETAIL',
            store=self.other_store,
            customer=self.other_customer,
            created_by=self.other_user,
        )

        response = self.client.get(reverse('api_get_quotation_detail'), {'id': quotation.id})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertEqual(payload['message'], 'Không tìm thấy')

    def test_brand_owner_cannot_save_order_with_product_from_other_store_than_warehouse(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('api_save_order'),
            data=json.dumps({
                'code': 'DH-CROSS-STORE-PRODUCT',
                'customer_id': self.customer.id,
                'warehouse_id': self.warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 0,
                'items': [{
                    'product_id': self.other_product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'discount_percent': 0,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('không cùng cửa hàng', payload['message'])
        self.assertFalse(Order.objects.filter(code='DH-CROSS-STORE-PRODUCT').exists())

    def test_save_order_return_requires_order_and_syncs_order_fields(self):
        order = self._create_order(code='DH-RETURN-001', status=5)

        missing_order_response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': 'TH-RETURN-001',
                'return_date': date.today().isoformat(),
                'total_refund': 50,
                'status': 2,
            }),
            content_type='application/json',
        )

        self.assertEqual(missing_order_response.status_code, 200)
        missing_payload = missing_order_response.json()
        self.assertEqual(missing_payload['status'], 'error')
        self.assertIn('đơn hàng gốc', missing_payload['message'])

        response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': 'TH-RETURN-002',
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'total_refund': 50,
                'status': 2,
                'reason': 'Khách đổi ý',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order_return = OrderReturn.objects.get(code='TH-RETURN-002')
        self.assertEqual(order_return.order_id, order.id)
        self.assertEqual(order_return.customer_id, order.customer_id)
        self.assertEqual(order_return.warehouse_id, order.warehouse_id)
        self.assertEqual(order_return.status, 2)
        self.assertIn('PC-TH-RETURN-002', payload['message'])

        refund_payment = Payment.objects.get(reference=f'ORDER_RETURN:{order_return.id}')
        self.assertEqual(refund_payment.code, 'PC-TH-RETURN-002')
        self.assertEqual(refund_payment.customer_id, order.customer_id)
        self.assertEqual(refund_payment.payment_method_option_id, self.payment_method.id)
        self.assertEqual(float(refund_payment.amount), 50.0)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 999950.0)

    def test_save_order_return_with_items_restocks_exported_order(self):
        order = self._create_order(code='DH-RETURN-ITEMS-001', status=4)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=50,
            total_price=100,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=3)

        response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': '',
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'status': 2,
                'reason': 'Khách trả một phần',
                'payment_method_option_id': self.payment_method.id,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 45,
                    'reason': 'Hàng lỗi',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertTrue(payload['return_code'].startswith('TH-'))

        order_return = OrderReturn.objects.get(id=payload['return_id'])
        return_item = OrderReturnItem.objects.get(order_return=order_return)
        self.assertEqual(return_item.product_id, self.product.id)
        self.assertEqual(return_item.quantity, 1)
        self.assertEqual(float(order_return.total_refund), 45.0)
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse)
        self.assertEqual(float(stock.quantity), 4.0)
        refund_payment = Payment.objects.get(reference=f'ORDER_RETURN:{order_return.id}')
        self.assertEqual(float(refund_payment.amount), 45.0)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 999955.0)

    def test_save_order_return_accepts_selected_cashbook_when_method_has_no_default(self):
        payment_method = PaymentMethodOption.objects.create(
            code='RETURN_NO_DEFAULT',
            name='Hoàn tiền không có quỹ mặc định',
            legacy_type=3,
        )
        order = self._create_order(code='DH-RETURN-EXPLICIT-CASHBOOK', status=5)

        response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': 'TH-EXPLICIT-CASHBOOK',
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'total_refund': 50,
                'status': 2,
                'payment_method_option_id': payment_method.id,
                'cash_book_id': self.cashbook.id,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        order_return = OrderReturn.objects.get(code='TH-EXPLICIT-CASHBOOK')
        refund_payment = Payment.objects.get(reference=f'ORDER_RETURN:{order_return.id}')
        self.assertEqual(refund_payment.payment_method_option_id, payment_method.id)
        self.assertEqual(refund_payment.cash_book_id, self.cashbook.id)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 999950.0)

    def test_order_return_form_exposes_cashbook_selector(self):
        response = self.client.get(reverse('order_return_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="inp_cash_book_id"')
        self.assertContains(response, self.cashbook.name)

        order_list_response = self.client.get(reverse('order_tbl'))
        self.assertEqual(order_list_response.status_code, 200)
        self.assertContains(order_list_response, 'id="quick_return_cash_book"')
        self.assertContains(order_list_response, self.cashbook.name)
        self.assertContains(order_list_response, 'id="modal_warranty_items"')
        self.assertContains(order_list_response, 'warranty-modal-content')

    def test_save_order_return_with_exchange_items_collects_difference_and_moves_stock(self):
        exchange_product = Product.objects.create(
            store=self.store,
            code='SP-EXCHANGE-001',
            name='Sản phẩm đổi mới',
            created_by=self.user,
        )
        order = self._create_order(code='DH-RETURN-EXCHANGE-001', status=4)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=50,
            total_price=100,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=3)
        ProductStock.objects.create(product=exchange_product, warehouse=self.warehouse, quantity=5)

        response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': 'TH-EXCHANGE-001',
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'status': 2,
                'reason': 'Khách đổi sang mẫu khác',
                'payment_method_option_id': self.payment_method.id,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 45,
                    'reason': 'Hàng lỗi',
                }],
                'exchange_items': [{
                    'product_id': exchange_product.id,
                    'quantity': 1,
                    'unit_price': 70,
                    'discount_percent': 0,
                    'note': 'Đổi mẫu mới',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['total_refund'], 0.0)
        self.assertEqual(payload['amount_due'], 25.0)

        order_return = OrderReturn.objects.get(code='TH-EXCHANGE-001')
        self.assertEqual(float(order_return.return_amount), 45.0)
        self.assertEqual(float(order_return.exchange_amount), 70.0)
        self.assertEqual(float(order_return.amount_due), 25.0)
        self.assertEqual(OrderReturnExchangeItem.objects.filter(order_return=order_return).count(), 1)
        self.assertIsNotNone(order_return.exchange_order_id)
        exchange_order = order_return.exchange_order
        self.assertEqual(exchange_order.customer_id, order.customer_id)
        self.assertEqual(exchange_order.warehouse_id, order.warehouse_id)
        self.assertEqual(float(exchange_order.total_amount), 70.0)
        self.assertEqual(float(exchange_order.discount_amount), 45.0)
        self.assertEqual(float(exchange_order.final_amount), 25.0)
        self.assertEqual(exchange_order.items.count(), 1)
        returned_stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse)
        exchanged_stock = ProductStock.objects.get(product=exchange_product, warehouse=self.warehouse)
        self.assertEqual(float(returned_stock.quantity), 4.0)
        self.assertEqual(float(exchanged_stock.quantity), 4.0)

        due_receipt = Receipt.objects.get(reference=f'ORDER_RETURN_DUE:{order_return.id}')
        self.assertEqual(float(due_receipt.amount), 25.0)
        self.assertEqual(due_receipt.order_id, exchange_order.id)
        self.assertFalse(Payment.objects.filter(reference=f'ORDER_RETURN:{order_return.id}', status=1).exists())
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 1000025.0)
        history = OrderEditHistory.objects.filter(order=order, action='return').first()
        self.assertIn('hàng đổi', history.summary)
        self.assertIn('khách còn thanh toán 25đ', history.summary)
        self.assertIn(exchange_order.code, history.summary)

        orders_payload = self.client.get(reverse('api_get_orders')).json()['data']
        original_row = next(row for row in orders_payload if row['id'] == order.id)
        exchange_row = next(row for row in orders_payload if row['id'] == exchange_order.id)
        self.assertTrue(original_row['has_returns'])
        self.assertTrue(original_row['has_exchange_returns'])
        self.assertEqual(original_row['returns'][0]['code'], order_return.code)
        self.assertEqual(original_row['returns'][0]['exchange_order_id'], exchange_order.id)
        self.assertTrue(exchange_row['is_exchange_order'])
        self.assertEqual(exchange_row['exchange_source_return_code'], order_return.code)
        self.assertEqual(exchange_row['exchange_original_order_code'], order.code)

        detail_payload = self.client.get(reverse('api_get_order_detail'), {'id': order.id}).json()
        self.assertEqual(detail_payload['status'], 'ok')
        return_detail = detail_payload['order']['returns'][0]
        self.assertEqual(return_detail['exchange_order_code'], exchange_order.code)
        self.assertEqual(return_detail['reason'], 'Khách đổi sang mẫu khác')
        self.assertEqual(return_detail['created_by_name'], self.user.username)
        self.assertEqual(return_detail['return_items'][0]['product_code'], self.product.code)
        self.assertEqual(return_detail['return_items'][0]['quantity'], 1.0)
        self.assertEqual(return_detail['exchange_items'][0]['product_code'], exchange_product.code)
        self.assertEqual(return_detail['exchange_items'][0]['quantity'], 1.0)
        self.assertEqual(return_detail['exchange_items'][0]['note'], 'Đổi mẫu mới')

        return_a4 = self.client.get(
            reverse('api_print_order_return'),
            {'id': order_return.id, 'type': 'a4'},
        )
        self.assertEqual(return_a4.status_code, 200)
        self.assertContains(return_a4, 'PHIẾU HOÀN / ĐỔI HÀNG')
        self.assertContains(return_a4, 'DANH MỤC HÀNG KHÁCH TRẢ LẠI')
        self.assertContains(return_a4, self.product.code)
        self.assertContains(return_a4, exchange_product.code)
        self.assertContains(return_a4, 'Khách đổi sang mẫu khác')

        return_k80 = self.client.get(
            reverse('api_print_order_return'),
            {'id': order_return.id, 'type': 'k80'},
        )
        self.assertEqual(return_k80.status_code, 200)
        self.assertContains(return_k80, 'HÀNG KHÁCH TRẢ')
        self.assertContains(return_k80, self.product.code)

    def test_save_order_return_allows_standalone_compensation_refund(self):
        order = self._create_order(code='DH-RETURN-COMPENSATION', status=5)

        response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': 'TH-COMPENSATION-001',
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'status': 2,
                'reason': 'Hàng vỡ cần đền bù',
                'payment_method_option_id': self.payment_method.id,
                'compensation_amount': 30,
                'items': [],
                'exchange_items': [],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['total_refund'], 30.0)
        self.assertEqual(payload['amount_due'], 0.0)

        order_return = OrderReturn.objects.get(code='TH-COMPENSATION-001')
        self.assertEqual(float(order_return.compensation_amount), 30.0)
        self.assertEqual(order_return.items.count(), 0)
        self.assertEqual(order_return.exchange_items.count(), 0)
        refund_payment = Payment.objects.get(reference=f'ORDER_RETURN:{order_return.id}')
        self.assertEqual(float(refund_payment.amount), 30.0)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 999970.0)

    def test_order_return_list_keeps_legacy_orphan_in_scope(self):
        orphan_return = OrderReturn.objects.create(
            code='TH-LEGACY-001',
            customer=self.customer,
            warehouse=self.warehouse,
            status=0,
            total_refund=25,
            return_date=date.today(),
            created_by=self.user,
        )
        OrderReturn.objects.create(
            code='TH-LEGACY-OTHER',
            customer=self.other_customer,
            warehouse=self.other_warehouse,
            status=0,
            total_refund=30,
            return_date=date.today(),
            created_by=self.other_user,
        )

        response = self.client.get(reverse('api_get_order_returns'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        rows = payload['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['id'], orphan_return.id)
        self.assertEqual(rows[0]['order'], '(Thiếu đơn gốc)')

    def test_save_order_return_allows_linking_legacy_orphan(self):
        order = self._create_order(code='DH-RETURN-LEGACY', status=5)
        orphan_return = OrderReturn.objects.create(
            code='TH-LEGACY-EDIT',
            customer=self.customer,
            warehouse=self.warehouse,
            status=0,
            total_refund=40,
            return_date=date.today(),
            created_by=self.user,
        )

        response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'id': orphan_return.id,
                'code': orphan_return.code,
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'total_refund': 40,
                'status': 2,
                'reason': 'Bổ sung đơn gốc',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        orphan_return.refresh_from_db()
        self.assertEqual(orphan_return.order_id, order.id)
        self.assertEqual(orphan_return.customer_id, order.customer_id)
        self.assertEqual(orphan_return.warehouse_id, order.warehouse_id)
        self.assertEqual(orphan_return.status, 2)

    def test_save_order_return_updates_refund_payment_and_restores_cashbook_on_cancel(self):
        order = self._create_order(code='DH-RETURN-UPDATE', status=5)

        create_response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'code': 'TH-RETURN-UPDATE',
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'total_refund': 50,
                'status': 2,
                'payment_method_option_id': self.payment_method.id,
            }),
            content_type='application/json',
        )
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        order_return = OrderReturn.objects.get(code='TH-RETURN-UPDATE')

        update_response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'id': order_return.id,
                'code': order_return.code,
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'total_refund': 80,
                'status': 2,
                'payment_method_option_id': self.payment_method.id,
            }),
            content_type='application/json',
        )
        self.assertEqual(update_response.json()['status'], 'ok', msg=update_response.content.decode())

        refund_payment = Payment.objects.get(reference=f'ORDER_RETURN:{order_return.id}')
        self.assertEqual(float(refund_payment.amount), 80.0)
        self.assertEqual(Payment.objects.filter(reference=f'ORDER_RETURN:{order_return.id}').count(), 1)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 999920.0)

        cancel_response = self.client.post(
            reverse('api_save_order_return'),
            data=json.dumps({
                'id': order_return.id,
                'code': order_return.code,
                'order_id': order.id,
                'return_date': date.today().isoformat(),
                'total_refund': 80,
                'status': 3,
                'payment_method_option_id': self.payment_method.id,
            }),
            content_type='application/json',
        )
        self.assertEqual(cancel_response.json()['status'], 'ok', msg=cancel_response.content.decode())

        refund_payment.refresh_from_db()
        self.assertEqual(refund_payment.status, 2)
        self.cashbook.refresh_from_db()
        self.assertEqual(float(self.cashbook.balance), 1000000.0)

    def test_pos_checkout_clamps_negative_total_to_zero(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=1)

        response = self.client.post(
            reverse('api_pos_checkout'),
            data=json.dumps({
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 100,
                    'total_price': 100,
                    'discount_percent': 0,
                }],
                'total_amount': 100,
                'discount_amount': 150,
                'final_amount': -50,
                'paid_amount': -50,
                'status': 5,
                'payment_status': 2,
                'note': 'POS âm',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        order = Order.objects.get(id=payload['order_id'])
        self.assertEqual(float(order.total_amount), 100.0)
        self.assertEqual(float(order.final_amount), 0.0)
        self.assertEqual(float(order.paid_amount), 0.0)
        self.assertEqual(order.payment_status, 2)
        self.assertEqual(order.status, 5)
        self.assertEqual(order.receipts.count(), 0)

    def test_payment_sync_demotes_completed_order_when_unpaid(self):
        order = self._create_order(code='DH-COMPLETE-UNPAID', status=5)

        update_order_payment_status(order)

        order.refresh_from_db()
        self.assertEqual(order.status, 4)
        self.assertEqual(order.payment_status, 0)

    def test_print_warranty_uses_custom_selected_items(self):
        order = self._create_order(code='DH-WARRANTY-CUSTOM', status=5)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=10,
            unit_price=100,
            total_price=1000,
        )

        response = self.client.get(
            reverse('api_print_order'),
            {
                'id': order.id,
                'type': 'warranty',
                'source': 'order',
                'warranty_items': json.dumps([{
                    'code': 'MAY-BH',
                    'name': 'Máy cần bảo hành',
                    'unit': 'Cái',
                    'quantity': 1,
                    'serial': 'SN001',
                    'warranty_term': '12 tháng',
                    'note': 'Chỉ bảo hành máy',
                }]),
            },
        )

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('Máy cần bảo hành', content)
        self.assertIn('SN001', content)
        self.assertIn('12 tháng', content)
        self.assertNotIn('Sản phẩm test đơn hàng', content)

    def test_saved_warranty_uses_product_defaults_and_keeps_historical_snapshot(self):
        self.product.refresh_from_db()
        self.product.warranty_period_months = 12
        self.product.warranty_policy = 'Chính sách bảo hành tại thời điểm bán'
        self.product.save(update_fields=['warranty_period_months', 'warranty_policy'])
        order = self._create_order(code='DH-WARRANTY-SAVED', status=4)
        order_item = OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=100,
            total_price=200,
        )

        response = self.client.post(
            reverse('api_save_order_warranty'),
            data=json.dumps({
                'order_id': order.id,
                'items': [{
                    'order_item_id': order_item.id,
                    'quantity': 1,
                    'serial': 'SERIAL-001',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        certificate = WarrantyCertificate.objects.get(order=order)
        warranty_item = certificate.items.get()
        self.assertEqual(certificate.code, f'PBH-{order.code}')
        self.assertEqual(warranty_item.order_item_id, order_item.id)
        self.assertEqual(warranty_item.serial, 'SERIAL-001')
        self.assertEqual(warranty_item.warranty_period_months, 12)
        self.assertEqual(warranty_item.warranty_term, '12 tháng')
        self.assertEqual(warranty_item.warranty_policy, 'Chính sách bảo hành tại thời điểm bán')
        self.assertEqual(warranty_item.warranty_start_date, date.today())
        expected_end_day = min(
            date.today().day,
            calendar.monthrange(date.today().year + 1, date.today().month)[1],
        )
        self.assertEqual(
            warranty_item.warranty_end_date,
            date(date.today().year + 1, date.today().month, expected_end_day),
        )
        self.assertTrue(order.history_entries.filter(action='warranty').exists())

        self.product.warranty_period_months = 6
        self.product.warranty_policy = 'Chính sách mới không được đổi phiếu cũ'
        self.product.save(update_fields=['warranty_period_months', 'warranty_policy'])

        detail_response = self.client.get(reverse('api_get_order_detail'), {'id': order.id})
        detail_payload = detail_response.json()
        self.assertTrue(detail_payload['can_save_warranty'])
        self.assertEqual(detail_payload['warranty_certificate']['code'], certificate.code)
        self.assertEqual(
            detail_payload['warranty_certificate']['items'][0]['warranty_policy'],
            'Chính sách bảo hành tại thời điểm bán',
        )

        list_payload = self.client.get(reverse('api_get_orders')).json()
        order_row = next(row for row in list_payload['data'] if row['id'] == order.id)
        self.assertTrue(order_row['has_warranty_certificate'])

        print_response = self.client.get(
            reverse('api_print_order'),
            {'id': order.id, 'type': 'warranty', 'source': 'order'},
        )
        print_content = print_response.content.decode()
        self.assertIn(certificate.code, print_content)
        self.assertIn('SERIAL-001', print_content)
        self.assertIn('Chính sách bảo hành tại thời điểm bán', print_content)
        self.assertNotIn('Chính sách mới không được đổi phiếu cũ', print_content)

    def test_warranty_cannot_be_saved_before_order_stock_export(self):
        order = self._create_order(code='DH-WARRANTY-NOT-EXPORTED', status=3)
        order_item = OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.post(
            reverse('api_save_order_warranty'),
            data=json.dumps({
                'order_id': order.id,
                'items': [{'order_item_id': order_item.id, 'quantity': 1}],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Xuất kho', response.json()['message'])
        self.assertFalse(WarrantyCertificate.objects.filter(order=order).exists())

    def test_print_warranty_empty_custom_items_does_not_fallback_to_all_products(self):
        order = self._create_order(code='DH-WARRANTY-EMPTY', status=5)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.get(
            reverse('api_print_order'),
            {
                'id': order.id,
                'type': 'warranty',
                'source': 'order',
                'warranty_items': json.dumps([]),
            },
        )

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('Chưa chọn sản phẩm bảo hành', content)
        self.assertNotIn('Sản phẩm test đơn hàng', content)

    def test_print_order_can_show_and_hide_combo_components(self):
        combo = Product.objects.create(
            store=self.store,
            code='CB-PRINT-001',
            name='Combo in đơn',
            unit='Bo',
            is_combo=True,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=2)
        order = self._create_order(code='DH-COMBO-PRINT', status=5)
        OrderItem.objects.create(
            order=order,
            product=combo,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.get(reverse('api_print_order'), {'id': order.id, 'type': 'a4', 'source': 'order'})

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('Combo in đơn', content)
        self.assertIn('SP-ORDER-001 - Sản phẩm test đơn hàng', content)

        PrintTemplate.objects.update_or_create(
            brand=self.brand,
            template_type='a4',
            defaults={'title': 'Hóa đơn A4', 'show_combo_components': False},
        )
        hidden_response = self.client.get(reverse('api_print_order'), {'id': order.id, 'type': 'a4', 'source': 'order'})

        self.assertEqual(hidden_response.status_code, 200)
        hidden_content = hidden_response.content.decode()
        self.assertIn('Combo in đơn', hidden_content)
        self.assertNotIn('SP-ORDER-001 - Sản phẩm test đơn hàng', hidden_content)

    def test_print_order_shows_product_note_below_item_name(self):
        self.product.note = 'Tặng kèm dây nguồn'
        self.product.save(update_fields=['note'])
        order = self._create_order(code='DH-PRINT-NOTE', status=5)
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.get(reverse('api_print_order'), {'id': order.id, 'type': 'a4', 'source': 'order'})

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('Sản phẩm test đơn hàng', content)
        self.assertIn('(Tặng kèm dây nguồn)', content)
