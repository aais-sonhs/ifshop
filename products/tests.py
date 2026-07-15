import json
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from products.models import (
    ComboItem,
    GoodsReceipt,
    GoodsReceiptItem,
    PurchaseReturn,
    PurchaseReturnItem,
    Product,
    ProductCategory,
    ProductLocation,
    ProductStock,
    ProductVariant,
    StockCheck,
    StockCheckItem,
    StockTransfer,
    StockTransferItem,
    Supplier,
    Warehouse,
)
from system_management.models import Brand, BusinessConfig, Store, UserProfile


class ProductInventoryFlowTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username='products_owner', password='pass123')
        cls.brand = Brand.objects.create(name='Products Brand', owner=cls.owner)
        cls.store = Store.objects.create(brand=cls.brand, name='Products Store', code='PST')
        cls.other_store = Store.objects.create(brand=cls.brand, name='Other Products Store', code='OPS')
        cls.user = User.objects.create_user(username='products_user', password='pass123')
        UserProfile.objects.create(user=cls.user, store=cls.store)
        cls.other_user = User.objects.create_user(username='other_products_user', password='pass123')
        UserProfile.objects.create(user=cls.other_user, store=cls.other_store)

        cls.warehouse_a = Warehouse.objects.create(store=cls.store, code='KHO-A1', name='Kho A1')
        cls.warehouse_b = Warehouse.objects.create(store=cls.store, code='KHO-B1', name='Kho B1')
        cls.other_warehouse = Warehouse.objects.create(store=cls.other_store, code='KHO-OTHER', name='Kho khác')
        cls.supplier = Supplier.objects.create(code='SUP-001', name='Supplier Products', created_by=cls.user)
        cls.product = Product.objects.create(
            store=cls.store,
            code='SP-001',
            name='San pham test',
            created_by=cls.user,
        )
        cls.other_product = Product.objects.create(
            store=cls.other_store,
            code='SP-OTHER-001',
            name='San pham store khac',
            created_by=cls.other_user,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def _build_product_import_upload(self, rows, headers=None, header_row=1):
        headers = headers or [
            'Mã SP',
            'Tên sản phẩm',
            'Barcode',
            'Danh mục',
            'ĐVT',
            'Quy cách',
            'Loại sản phẩm',
            'Tính chất',
            'Giá nhập',
            'Giá vốn',
            'Giá bán lẻ',
            'Giá sỉ KBH',
            'Giá sỉ BH',
            'Tồn kho',
            'Tồn tối thiểu',
            'Tồn tối đa',
            'Trạng thái',
            'NCC',
            'Vị trí',
            'Mô tả',
        ]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if header_row > 1:
            sheet.cell(row=1, column=1, value='DANH SÁCH SẢN PHẨM')
            sheet.cell(row=2, column=1, value='Mẫu import từ export')
        for col_index, header in enumerate(headers, 1):
            sheet.cell(row=header_row, column=col_index, value=header)
        for row_index, row in enumerate(rows, header_row + 1):
            for col_index, value in enumerate(row, 1):
                sheet.cell(row=row_index, column=col_index, value=value)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return SimpleUploadedFile(
            'products.xlsx',
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _create_completed_goods_receipt(
        self,
        code='P-RETURN-SOURCE',
        quantity=Decimal('10'),
        unit_price=Decimal('100'),
        product=None,
        warehouse=None,
        supplier=None,
        created_by=None,
    ):
        product = product or self.product
        warehouse = warehouse or self.warehouse_a
        supplier = supplier or self.supplier
        created_by = created_by or self.user
        receipt = GoodsReceipt.objects.create(
            code=code,
            supplier=supplier,
            warehouse=warehouse,
            receipt_date=date.today(),
            status=1,
            total_amount=quantity * unit_price,
            created_by=created_by,
        )
        item = GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=product,
            quantity=quantity,
            unit_price=unit_price,
            total_price=quantity * unit_price,
        )
        return receipt, item

    def _post_purchase_return(
        self,
        receipt,
        receipt_item,
        quantity,
        status=1,
        return_id=None,
        code='',
    ):
        return self.client.post(
            reverse('api_save_purchase_return'),
            data=json.dumps({
                'id': return_id,
                'code': code,
                'goods_receipt_id': receipt.id,
                'return_date': date.today().isoformat(),
                'status': status,
                'reason': 'Nha cung cap giao nham hang',
                'note': 'Theo doi tra hang nhap',
                'items': [{
                    'goods_receipt_item_id': receipt_item.id,
                    'quantity': str(quantity),
                    'unit_price': str(receipt_item.unit_price),
                }],
            }),
            content_type='application/json',
        )

    def test_save_product_auto_generates_code_when_blank(self):
        response = self.client.post(
            reverse('api_save_product'),
            data={
                'name': 'San pham tao nhanh',
                'unit': 'Cai',
                'cost_price': '1000',
                'import_price': '1000',
                'selling_price': '1500',
                'variants': '[]',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        product = Product.objects.get(name='San pham tao nhanh')
        self.assertRegex(product.code, r'^SP\d{3}$')
        self.assertEqual(payload['product']['id'], product.id)
        self.assertEqual(payload['product']['code'], product.code)
        self.assertEqual(product.store_id, self.store.id)

    def test_product_list_exposes_category_and_product_type_levels(self):
        category = ProductCategory.objects.create(name='May moc')
        product_type = ProductCategory.objects.create(name='May xay', parent=category)
        self.product.category = product_type
        self.product.save(update_fields=['category'])

        response = self.client.get(reverse('api_get_products'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == self.product.id)
        self.assertEqual(row['category'], 'May moc')
        self.assertEqual(row['category_id'], category.id)
        self.assertEqual(row['product_type'], 'May xay')
        self.assertEqual(row['product_type_id'], product_type.id)
        self.assertEqual(row['category_record_id'], product_type.id)

    def test_product_table_groups_related_fields_into_compact_columns(self):
        response = self.client.get(reverse('product_tbl'))

        self.assertEqual(response.status_code, 200)
        for column_key in ('spec_unit', 'category_type', 'import_cost', 'wholesale_prices'):
            self.assertContains(response, f'data-col="{column_key}"')
        for old_column_key in (
            'spec', 'unit', 'category', 'product_type',
            'import_price', 'cost_price', 'ws_no_warranty', 'ws_warranty',
        ):
            self.assertNotContains(response, f'data-col="{old_column_key}"')
        self.assertContains(response, 'function renderProductStackedCell(rows, extraClass)')
        self.assertContains(response, '.product-combined-header small{margin-top:2px;color:#fff !important;')
        self.assertContains(response, 'colspan="15"')

    def test_product_form_uses_near_full_width_dialog(self):
        response = self.client.get(reverse('product_tbl'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'class="modal-dialog modal-xl modal-dialog-scrollable product-form-dialog"',
        )
        self.assertContains(
            response,
            '#modal_form .product-form-dialog{width:calc(100vw - 24px);max-width:1600px;margin:12px auto;}',
        )
        for section_title in (
            'Thông tin cơ bản', 'Phân loại &amp; kho', 'Giá sản phẩm',
            'Hình ảnh', 'Tồn theo kho', 'Mô tả &amp; ghi chú',
        ):
            self.assertContains(response, section_title)
        self.assertContains(response, 'class="product-option-grid"')
        self.assertContains(response, 'class="product-money-input is-primary"')
        for field_id in (
            'inp_code', 'inp_name', 'inp_barcode', 'inp_unit', 'inp_specification',
            'inp_category_id', 'inp_product_type_id', 'inp_supplier_id', 'inp_location_id',
            'inp_import_price', 'inp_cost_price', 'inp_selling_price',
            'inp_wholesale_no_warranty', 'inp_wholesale_warranty',
        ):
            self.assertContains(response, f'id="{field_id}"', count=1)

    def test_product_list_exposes_and_updates_inline_note(self):
        self.product.note = 'In kem phu kien'
        self.product.save(update_fields=['note'])

        response = self.client.get(reverse('api_get_products'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == self.product.id)
        self.assertEqual(row['note'], 'In kem phu kien')

        update_response = self.client.post(
            reverse('api_update_product_note'),
            data=json.dumps({
                'id': self.product.id,
                'note': 'Hien note duoi ten san pham',
            }),
            content_type='application/json',
        )

        self.assertEqual(update_response.status_code, 200)
        payload = update_response.json()
        self.assertEqual(payload['status'], 'ok', msg=update_response.content.decode())

        self.product.refresh_from_db()
        self.assertEqual(self.product.note, 'Hien note duoi ten san pham')

    def test_product_list_keeps_zero_and_negative_stock_by_warehouse(self):
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('0'),
        )
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_b,
            quantity=Decimal('-2'),
        )

        response = self.client.get(reverse('api_get_products'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == self.product.id)
        stock_by_warehouse = {
            item['warehouse_id']: item['quantity']
            for item in row['stock_by_warehouse']
        }
        self.assertEqual(stock_by_warehouse[self.warehouse_a.id], 0.0)
        self.assertEqual(stock_by_warehouse[self.warehouse_b.id], -2.0)
        self.assertEqual(row['total_stock'], -2.0)

    def test_product_list_sorts_total_stock_before_pagination(self):
        for index in range(12):
            product = Product.objects.create(
                store=self.store,
                code=f'STOCK-SORT-{index:02d}',
                name=f'San pham sap xep ton {index:02d}',
                created_by=self.user,
            )
            ProductStock.objects.create(
                product=product,
                warehouse=self.warehouse_a,
                quantity=Decimal(str(index)),
            )

        descending_page_one = self.client.get(
            reverse('api_get_products'),
            data={'text': 'STOCK-SORT', 'stock_sort': 'desc', 'page_size': 10, 'page': 1},
        ).json()
        descending_page_two = self.client.get(
            reverse('api_get_products'),
            data={'text': 'STOCK-SORT', 'stock_sort': 'desc', 'page_size': 10, 'page': 2},
        ).json()
        ascending_page_one = self.client.get(
            reverse('api_get_products'),
            data={'text': 'STOCK-SORT', 'stock_sort': 'asc', 'page_size': 10, 'page': 1},
        ).json()

        self.assertEqual(
            [row['total_stock'] for row in descending_page_one['data']],
            [float(value) for value in range(11, 1, -1)],
        )
        self.assertEqual(
            [row['total_stock'] for row in descending_page_two['data']],
            [1.0, 0.0],
        )
        self.assertEqual(
            [row['total_stock'] for row in ascending_page_one['data']],
            [float(value) for value in range(10)],
        )

    def test_product_list_stock_sort_uses_computed_combo_quantity(self):
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('14'),
        )
        standalone = Product.objects.create(
            store=self.store,
            code='SP-STOCK-SORT-STANDALONE',
            name='San pham ton nam',
            created_by=self.user,
        )
        ProductStock.objects.create(
            product=standalone,
            warehouse=self.warehouse_a,
            quantity=Decimal('5'),
        )
        combo = Product.objects.create(
            store=self.store,
            code='SP-STOCK-SORT-COMBO',
            name='Combo ton bay',
            is_combo=True,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=Decimal('2'))

        response = self.client.get(
            reverse('api_get_products'),
            data={'stock_sort': 'desc'},
        )

        self.assertEqual(response.status_code, 200)
        rows = response.json()['data']
        self.assertEqual([row['id'] for row in rows], [self.product.id, combo.id, standalone.id])
        self.assertEqual([row['total_stock'] for row in rows], [14.0, 7.0, 5.0])

    def test_edit_product_allows_negative_stock_when_enabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Allow negative product edit',
            opt_allow_negative_stock=True,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('2'))

        response = self.client.post(
            reverse('api_save_product'),
            data={
                'id': self.product.id,
                'code': self.product.code,
                'name': self.product.name,
                'unit': self.product.unit,
                'skip_variants': '1',
                'combo_items': '[]',
                'stocks': json.dumps([
                    {'warehouse_id': self.warehouse_a.id, 'quantity': '-5'},
                    {'warehouse_id': self.warehouse_b.id, 'quantity': '1.5'},
                ]),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stocks = {
            stock.warehouse_id: stock.quantity
            for stock in ProductStock.objects.filter(product=self.product)
        }
        self.assertEqual(stocks[self.warehouse_a.id], Decimal('-5'))
        self.assertEqual(stocks[self.warehouse_b.id], Decimal('1.5'))

    def test_edit_product_rejects_negative_stock_when_disabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Reject negative product edit',
            opt_allow_negative_stock=False,
        )
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('2'),
        )

        response = self.client.post(
            reverse('api_save_product'),
            data={
                'id': self.product.id,
                'code': self.product.code,
                'name': self.product.name,
                'unit': self.product.unit,
                'skip_variants': '1',
                'combo_items': '[]',
                'stocks': json.dumps([
                    {'warehouse_id': self.warehouse_a.id, 'quantity': '-1'},
                ]),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('chưa bật cấu hình cho phép tồn âm', response.json()['message'])
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('2'))

    def test_edit_product_rejects_stock_from_another_store(self):
        response = self.client.post(
            reverse('api_save_product'),
            data={
                'id': self.product.id,
                'code': self.product.code,
                'name': self.product.name,
                'unit': self.product.unit,
                'skip_variants': '1',
                'combo_items': '[]',
                'stocks': json.dumps([
                    {'warehouse_id': self.other_warehouse.id, 'quantity': '10'},
                ]),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('không thuộc cửa hàng của sản phẩm', response.json()['message'])
        self.assertFalse(ProductStock.objects.filter(
            product=self.product,
            warehouse=self.other_warehouse,
        ).exists())

    def test_brand_owner_can_create_product_type_under_category(self):
        category = ProductCategory.objects.create(name='Linh kien may moc')
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('api_save_category'),
            data=json.dumps({
                'name': 'May ep',
                'parent_id': category.id,
                'description': 'Loai san pham',
                'is_active': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        product_type = ProductCategory.objects.get(name='May ep')
        self.assertEqual(product_type.parent_id, category.id)
        self.assertEqual(payload['category']['parent_id'], category.id)

    def test_brand_owner_can_quick_create_supplier_with_auto_code(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('api_save_supplier'),
            data=json.dumps({
                'name': 'NCC Tao nhanh',
                'is_active': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        supplier = Supplier.objects.get(name='NCC Tao nhanh')
        self.assertRegex(supplier.code, r'^NCC\d{3}$')
        self.assertEqual(payload['supplier']['id'], supplier.id)
        self.assertEqual(payload['supplier']['code'], supplier.code)

    def test_inventory_user_can_quick_create_supplier_from_goods_receipt(self):
        response = self.client.post(
            reverse('api_quick_create_supplier'),
            data=json.dumps({
                'name': 'NCC tao tai phieu nhap',
                'phone': '0909000111',
                'contact_person': 'Anh Minh',
                'address': '12 Nguyen Trai',
                'note': 'Tao nhanh khi nhap hang',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        supplier = Supplier.objects.get(id=payload['supplier']['id'])
        self.assertRegex(supplier.code, r'^NCC\d{3}$')
        self.assertEqual(supplier.created_by_id, self.user.id)
        self.assertEqual(supplier.phone, '0909000111')
        self.assertEqual(supplier.contact_person, 'Anh Minh')
        self.assertEqual(supplier.address, '12 Nguyen Trai')

    def test_brand_owner_can_quick_create_product_location(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('api_save_location'),
            data=json.dumps({
                'name': 'Ke A1',
                'is_active': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        location = ProductLocation.objects.get(name='Ke A1')
        self.assertEqual(payload['id'], location.id)
        self.assertEqual(payload['name'], location.name)

    def test_save_combo_product_creates_items_and_exposes_combo_metadata(self):
        self.product.cost_price = Decimal('100')
        self.product.selling_price = Decimal('150')
        self.product.save(update_fields=['cost_price', 'selling_price'])
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('5'))
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_b, quantity=Decimal('1'))

        response = self.client.post(
            reverse('api_save_product'),
            data={
                'name': 'Combo test',
                'unit': 'Bo',
                'selling_price': '250.000',
                'wholesale_price_no_warranty': '240.000',
                'wholesale_price_warranty': '260.000',
                'is_combo': '1',
                'combo_items': json.dumps([
                    {'product_id': self.product.id, 'quantity': '2'},
                ]),
                'skip_variants': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())

        combo = Product.objects.get(id=payload['product']['id'])
        self.assertTrue(combo.is_combo)
        self.assertEqual(combo.selling_price, Decimal('250000'))
        self.assertEqual(combo.cost_price, Decimal('200'))
        combo_item = ComboItem.objects.get(combo=combo)
        self.assertEqual(combo_item.product_id, self.product.id)
        self.assertEqual(combo_item.quantity, Decimal('2.00'))

        products_payload = self.client.get(reverse('api_get_products')).json()['data']
        combo_row = next(item for item in products_payload if item['id'] == combo.id)
        component_row = next(item for item in products_payload if item['id'] == self.product.id)
        self.assertTrue(combo_row['is_combo'])
        self.assertEqual(combo_row['combo_items'][0]['product_id'], self.product.id)
        self.assertEqual(combo_row['combo_items'][0]['selling_price'], 150.0)
        self.assertEqual(combo_row['combo_items'][0]['product_code'], self.product.code)
        self.assertEqual(combo_row['combo_items'][0]['unit'], self.product.unit)
        self.assertEqual(combo_row['combo_items'][0]['line_cost'], 200.0)
        self.assertEqual(combo_row['combo_items'][0]['line_total'], 300.0)
        self.assertEqual(combo_row['total_stock'], 2.0)
        self.assertEqual(combo_row['stock_by_warehouse'][0]['warehouse_id'], self.warehouse_a.id)
        self.assertEqual(combo_row['stock_by_warehouse'][0]['quantity'], 2.0)
        self.assertEqual(component_row['combo_parent_count'], 1)
        self.assertEqual(component_row['combo_parents'][0]['id'], combo.id)

    def test_product_list_stock_filter_uses_computed_combo_stock(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('4'))
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_b, quantity=Decimal('2'))
        combo = Product.objects.create(
            store=self.store,
            code='SP-COMBO-STOCK',
            name='Combo ton kho',
            is_combo=True,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=Decimal('2'))

        response = self.client.get(reverse('api_get_products'), data={'stock': 'instock'})

        self.assertEqual(response.status_code, 200)
        ids = [item['id'] for item in response.json()['data']]
        self.assertIn(combo.id, ids)

    def test_product_list_total_stock_price_filter_uses_computed_combo_stock(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('4'))
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_b, quantity=Decimal('2'))
        combo = Product.objects.create(
            store=self.store,
            code='SP-COMBO-TOTAL',
            name='Combo tong ton',
            is_combo=True,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=Decimal('2'))

        response = self.client.get(
            reverse('api_get_products'),
            data={'price_basis': 'total_stock', 'price_from': '3', 'price_to': '3'},
        )

        self.assertEqual(response.status_code, 200)
        ids = [item['id'] for item in response.json()['data']]
        self.assertIn(combo.id, ids)

    def test_product_list_component_filter_ignores_deleted_combo_parents(self):
        combo = Product.objects.create(
            store=self.store,
            code='SP-COMBO-DELETED',
            name='Combo da xoa',
            is_combo=True,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=Decimal('1'))
        combo.delete()

        component_response = self.client.get(reverse('api_get_products'), data={'combo_usage': 'component'})
        standalone_response = self.client.get(reverse('api_get_products'), data={'combo_usage': 'standalone'})

        self.assertEqual(component_response.status_code, 200)
        self.assertEqual(standalone_response.status_code, 200)
        component_ids = [item['id'] for item in component_response.json()['data']]
        standalone_ids = [item['id'] for item in standalone_response.json()['data']]
        self.assertNotIn(self.product.id, component_ids)
        self.assertIn(self.product.id, standalone_ids)

    def test_product_list_text_search_matches_combo_relations(self):
        combo = Product.objects.create(
            store=self.store,
            code='SP-COMBO-SEARCH',
            name='Combo tim kiem',
            is_combo=True,
            created_by=self.user,
        )
        ComboItem.objects.create(combo=combo, product=self.product, quantity=Decimal('1'))

        component_search = self.client.get(reverse('api_get_products'), data={'text': self.product.code})
        combo_search = self.client.get(reverse('api_get_products'), data={'text': combo.code})

        self.assertEqual(component_search.status_code, 200)
        self.assertEqual(combo_search.status_code, 200)
        component_search_ids = [item['id'] for item in component_search.json()['data']]
        combo_search_ids = [item['id'] for item in combo_search.json()['data']]
        self.assertIn(combo.id, component_search_ids)
        self.assertIn(self.product.id, combo_search_ids)

    def test_import_products_excel_creates_and_updates_products_from_export_template(self):
        headers = [
            'STT',
            'Mã SP',
            'Tên sản phẩm',
            'Barcode',
            'Danh mục',
            'ĐVT',
            'Quy cách',
            'Loại sản phẩm',
            'Tính chất',
            'Giá nhập',
            'Giá vốn',
            'Giá bán lẻ',
            'Giá sỉ KBH',
            'Giá sỉ BH',
            'Tồn kho',
            'Tồn tối thiểu',
            'Tồn tối đa',
            'Trạng thái',
            'NCC',
            'Vị trí',
            'Mô tả',
        ]
        upload = self._build_product_import_upload(
            rows=[
                [
                    1,
                    self.product.code,
                    'San pham da sua Excel',
                    'BC-UPDATE',
                    'Thiet bi',
                    'Cai',
                    'Hop 10 cai',
                    'May ep',
                    'Sản phẩm',
                    11000,
                    9000,
                    15000,
                    14000,
                    16000,
                    999,
                    2,
                    20,
                    'Ngừng hoạt động',
                    self.supplier.name,
                    'Ke A1',
                    'Mo ta cap nhat',
                ],
                [
                    2,
                    '',
                    'San pham moi import',
                    'BC-NEW',
                    'Do uong',
                    'Lon',
                    '330ml',
                    'Nuoc ngot',
                    'Cân/đong',
                    '1.200',
                    '1.000',
                    '2.000',
                    '1.800',
                    '2.200',
                    '12,5',
                    1,
                    100,
                    'Đang hoạt động',
                    'NCC Moi Excel',
                    'Ke B2',
                    'Tao tu Excel',
                ],
            ],
            headers=headers,
            header_row=4,
        )

        response = self.client.post(
            reverse('import_products_excel'),
            data={'file': upload, 'import_stock': '1'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['summary']['created'], 1)
        self.assertEqual(payload['summary']['updated'], 1)
        self.assertEqual(payload['summary']['stock_initialized'], 2)

        self.product.refresh_from_db()
        self.assertEqual(self.product.name, 'San pham da sua Excel')
        self.assertEqual(self.product.barcode, 'BC-UPDATE')
        self.assertEqual(self.product.import_price, Decimal('11000'))
        self.assertEqual(self.product.cost_price, Decimal('9000'))
        self.assertEqual(self.product.selling_price, Decimal('15000'))
        self.assertFalse(self.product.is_active)
        self.assertTrue(ProductStock.objects.filter(product=self.product, quantity=Decimal('999')).exists())

        new_product = Product.objects.get(name='San pham moi import')
        self.assertRegex(new_product.code, r'^SP\d{3}$')
        self.assertEqual(new_product.store_id, self.store.id)
        self.assertEqual(new_product.barcode, 'BC-NEW')
        self.assertEqual(new_product.unit, 'Lon')
        self.assertEqual(new_product.specification, '330ml')
        self.assertEqual(new_product.import_price, Decimal('1200'))
        self.assertEqual(new_product.cost_price, Decimal('1000'))
        self.assertEqual(new_product.selling_price, Decimal('2000'))
        self.assertEqual(new_product.wholesale_price_no_warranty, Decimal('1800'))
        self.assertEqual(new_product.wholesale_price_warranty, Decimal('2200'))
        self.assertEqual(new_product.min_stock, 1)
        self.assertEqual(new_product.max_stock, 100)
        self.assertTrue(new_product.is_weight_based)
        self.assertFalse(new_product.is_service)
        self.assertEqual(new_product.description, 'Tao tu Excel')
        self.assertEqual(new_product.supplier.name, 'NCC Moi Excel')
        self.assertEqual(new_product.location.name, 'Ke B2')

        category = ProductCategory.objects.get(name='Do uong', parent__isnull=True)
        product_type = ProductCategory.objects.get(name='Nuoc ngot', parent=category)
        self.assertEqual(new_product.category_id, product_type.id)

        stock = ProductStock.objects.get(product=new_product, warehouse=self.warehouse_a)
        self.assertEqual(stock.quantity, Decimal('12.50'))

    def test_import_products_excel_rejects_product_code_outside_user_store(self):
        upload = self._build_product_import_upload(
            rows=[[self.other_product.code, 'Khong duoc import']],
            headers=['Mã SP', 'Tên sản phẩm'],
        )

        response = self.client.post(
            reverse('import_products_excel'),
            data={'file': upload},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertEqual(payload['summary']['created'], 0)
        self.assertEqual(payload['summary']['updated'], 0)
        self.assertEqual(payload['summary']['errors'], 1)
        self.assertIn('ngoài phạm vi', payload['errors'][0]['message'])

        self.other_product.refresh_from_db()
        self.assertEqual(self.other_product.name, 'San pham store khac')
        self.assertFalse(Product.objects.filter(store=self.store, code=self.other_product.code).exists())

    def test_import_products_excel_updates_existing_product_stock_in_default_warehouse(self):
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('5'),
        )
        upload = self._build_product_import_upload(
            rows=[[
                self.product.code,
                'San pham test sua ton',
                0,
            ]],
            headers=['Mã SP', 'Tên sản phẩm', 'Tồn kho'],
        )

        response = self.client.post(
            reverse('import_products_excel'),
            data={'file': upload, 'import_stock': '1'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['summary']['updated'], 1)
        self.assertEqual(payload['summary']['stock_initialized'], 1)

        self.product.refresh_from_db()
        self.assertEqual(self.product.name, 'San pham test sua ton')
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        self.assertEqual(stock.quantity, Decimal('0'))

    def test_import_products_excel_rejects_existing_stock_split_across_other_warehouses(self):
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('5'),
        )
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_b,
            quantity=Decimal('2'),
        )
        upload = self._build_product_import_upload(
            rows=[[
                self.product.code,
                'San pham test khong duoc sua ton',
                9,
            ]],
            headers=['Mã SP', 'Tên sản phẩm', 'Tồn kho'],
        )

        response = self.client.post(
            reverse('import_products_excel'),
            data={'file': upload, 'import_stock': '1'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertEqual(payload['summary']['updated'], 0)
        self.assertEqual(payload['summary']['errors'], 1)
        self.assertIn('không thể cập nhật tồn từ file Excel', payload['errors'][0]['message'])

        stocks = {
            stock.warehouse_id: stock.quantity
            for stock in ProductStock.objects.filter(product=self.product)
        }
        self.assertEqual(stocks[self.warehouse_a.id], Decimal('5.00'))
        self.assertEqual(stocks[self.warehouse_b.id], Decimal('2.00'))

    def test_import_products_excel_syncs_wholesale_prices_to_existing_variants(self):
        variant_small = ProductVariant.objects.create(
            product=self.product,
            size_name='S',
            sku='SP-001-S',
            import_price=Decimal('1000'),
            wholesale_price_no_warranty=Decimal('1100'),
            wholesale_price_warranty=Decimal('1200'),
        )
        variant_large = ProductVariant.objects.create(
            product=self.product,
            size_name='L',
            sku='SP-001-L',
            import_price=Decimal('2000'),
            wholesale_price_no_warranty=Decimal('2100'),
            wholesale_price_warranty=Decimal('2200'),
        )
        upload = self._build_product_import_upload(
            rows=[[
                self.product.code,
                self.product.name,
                3300,
                4400,
                5500,
            ]],
            headers=['Mã SP', 'Tên sản phẩm', 'Giá nhập', 'Giá sỉ KBH', 'Giá sỉ BH'],
        )

        response = self.client.post(
            reverse('import_products_excel'),
            data={'file': upload},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['summary']['updated'], 1)
        self.assertEqual(payload['summary']['variants_synced'], 2)

        self.product.refresh_from_db()
        variant_small.refresh_from_db()
        variant_large.refresh_from_db()
        self.assertEqual(self.product.import_price, Decimal('3300'))
        self.assertEqual(self.product.wholesale_price_no_warranty, Decimal('4400'))
        self.assertEqual(self.product.wholesale_price_warranty, Decimal('5500'))
        self.assertEqual(variant_small.import_price, Decimal('3300'))
        self.assertEqual(variant_small.wholesale_price_no_warranty, Decimal('4400'))
        self.assertEqual(variant_small.wholesale_price_warranty, Decimal('5500'))
        self.assertEqual(variant_large.import_price, Decimal('3300'))
        self.assertEqual(variant_large.wholesale_price_no_warranty, Decimal('4400'))
        self.assertEqual(variant_large.wholesale_price_warranty, Decimal('5500'))

    def test_product_purchase_history_api_returns_selected_product_receipts_only(self):
        other_same_store_product = Product.objects.create(
            store=self.store,
            code='SP-002',
            name='San pham khac cung store',
            created_by=self.user,
        )
        receipt = GoodsReceipt.objects.create(
            code='P-HISTORY-001',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=date.today(),
            status=1,
            total_amount=Decimal('170'),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('5'),
            unit_price=Decimal('20'),
            total_price=Decimal('100'),
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=other_same_store_product,
            quantity=Decimal('7'),
            unit_price=Decimal('10'),
            total_price=Decimal('70'),
        )

        response = self.client.get(
            reverse('api_product_purchase_history'),
            data={'product_id': self.product.id},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['summary']['total_entries'], 1)
        self.assertEqual(payload['summary']['total_receipts'], 1)
        self.assertEqual(payload['summary']['total_quantity'], 5.0)
        self.assertEqual(payload['summary']['total_amount'], 100.0)
        self.assertEqual(payload['data'][0]['receipt_code'], 'P-HISTORY-001')
        self.assertEqual(payload['receipts'][0]['receipt_code'], 'P-HISTORY-001')
        self.assertEqual(payload['receipts'][0]['quantity'], 5.0)
        self.assertEqual(payload['receipts'][0]['total_price'], 100.0)
        self.assertEqual(len(payload['receipts'][0]['items']), 1)

    def test_product_purchase_history_api_groups_items_by_receipt(self):
        receipt = GoodsReceipt.objects.create(
            code='P-HISTORY-GROUP-001',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=date.today(),
            status=1,
            total_amount=Decimal('250'),
            note='Nhap nhieu dong',
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('5'),
            unit_price=Decimal('20'),
            total_price=Decimal('100'),
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('3'),
            unit_price=Decimal('50'),
            total_price=Decimal('150'),
        )

        response = self.client.get(
            reverse('api_product_purchase_history'),
            data={'product_id': self.product.id},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok', msg=response.content.decode())
        self.assertEqual(payload['summary']['total_entries'], 2)
        self.assertEqual(payload['summary']['total_receipts'], 1)
        self.assertEqual(payload['summary']['total_quantity'], 8.0)
        self.assertEqual(payload['summary']['total_amount'], 250.0)

        receipt_row = payload['receipts'][0]
        self.assertEqual(receipt_row['receipt_code'], 'P-HISTORY-GROUP-001')
        self.assertEqual(receipt_row['item_count'], 2)
        self.assertEqual(receipt_row['quantity'], 8.0)
        self.assertEqual(receipt_row['total_price'], 250.0)
        self.assertEqual(receipt_row['min_unit_price'], 20.0)
        self.assertEqual(receipt_row['max_unit_price'], 50.0)
        self.assertEqual(len(receipt_row['items']), 2)

    def test_delete_goods_receipt_reverts_stock(self):
        receipt = GoodsReceipt.objects.create(
            code='P00001',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=date.today(),
            status=1,
            total_amount=Decimal('50'),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('5'),
            unit_price=Decimal('10'),
            total_price=Decimal('50'),
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('5'))

        response = self.client.post(
            reverse('api_delete_goods_receipt'),
            data=json.dumps({'id': receipt.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())

        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        deleted_receipt = GoodsReceipt.all_objects.get(id=receipt.id)
        self.assertEqual(stock.quantity, Decimal('0'))
        self.assertTrue(deleted_receipt.is_deleted)

    def test_save_completed_goods_receipt_moves_stock_to_new_warehouse(self):
        receipt = GoodsReceipt.objects.create(
            code='P00002',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=date.today(),
            status=1,
            total_amount=Decimal('50'),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('5'),
            unit_price=Decimal('10'),
            total_price=Decimal('50'),
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('5'))

        response = self.client.post(
            reverse('api_save_goods_receipt'),
            data=json.dumps({
                'id': receipt.id,
                'code': receipt.code,
                'supplier_id': self.supplier.id,
                'warehouse_id': self.warehouse_b.id,
                'receipt_date': date.today().isoformat(),
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 5,
                    'unit_price': 10,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())

        stock_a = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        stock_b = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_b)
        self.assertEqual(stock_a.quantity, Decimal('0'))
        self.assertEqual(stock_b.quantity, Decimal('5'))

    def test_completed_purchase_return_deducts_inventory(self):
        receipt, receipt_item = self._create_completed_goods_receipt(code='P-RETURN-COMPLETE')
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )

        response = self._post_purchase_return(receipt, receipt_item, Decimal('3'), status=1)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        purchase_return = PurchaseReturn.objects.get(id=response.json()['id'])
        self.assertEqual(stock.quantity, Decimal('7'))
        self.assertTrue(purchase_return.stock_applied)
        self.assertEqual(purchase_return.total_amount, Decimal('300'))

    def test_draft_purchase_return_does_not_change_inventory(self):
        receipt, receipt_item = self._create_completed_goods_receipt(code='P-RETURN-DRAFT')
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )

        response = self._post_purchase_return(receipt, receipt_item, Decimal('3'), status=0)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        purchase_return = PurchaseReturn.objects.get(id=response.json()['id'])
        self.assertEqual(stock.quantity, Decimal('10'))
        self.assertFalse(purchase_return.stock_applied)

    def test_purchase_return_cannot_exceed_remaining_received_quantity(self):
        receipt, receipt_item = self._create_completed_goods_receipt(code='P-RETURN-LIMIT')
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )
        first_response = self._post_purchase_return(receipt, receipt_item, Decimal('4'), status=1)
        self.assertEqual(first_response.json()['status'], 'ok', msg=first_response.content.decode())

        response = self._post_purchase_return(receipt, receipt_item, Decimal('7'), status=1)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('vượt số còn có thể trả', response.json()['message'])
        self.assertEqual(PurchaseReturn.objects.count(), 1)
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        self.assertEqual(stock.quantity, Decimal('6'))

    def test_edit_completed_purchase_return_reverses_and_reapplies_inventory(self):
        receipt, receipt_item = self._create_completed_goods_receipt(code='P-RETURN-EDIT')
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )
        create_response = self._post_purchase_return(receipt, receipt_item, Decimal('3'), status=1)
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        purchase_return = PurchaseReturn.objects.get(id=create_response.json()['id'])

        response = self._post_purchase_return(
            receipt,
            receipt_item,
            Decimal('5'),
            status=1,
            return_id=purchase_return.id,
            code=purchase_return.code,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        purchase_return.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('5'))
        self.assertEqual(purchase_return.items.get().quantity, Decimal('5'))
        self.assertTrue(purchase_return.stock_applied)

    def test_delete_completed_purchase_return_restores_inventory(self):
        receipt, receipt_item = self._create_completed_goods_receipt(code='P-RETURN-DELETE')
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )
        create_response = self._post_purchase_return(receipt, receipt_item, Decimal('3'), status=1)
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        return_id = create_response.json()['id']

        response = self.client.post(
            reverse('api_delete_purchase_return'),
            data=json.dumps({'id': return_id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        deleted_return = PurchaseReturn.all_objects.get(id=return_id)
        self.assertEqual(stock.quantity, Decimal('10'))
        self.assertTrue(deleted_return.is_deleted)
        self.assertFalse(deleted_return.stock_applied)

    def test_goods_receipt_with_deleted_return_history_cannot_be_edited_or_deleted(self):
        receipt, receipt_item = self._create_completed_goods_receipt(code='P-RETURN-AUDIT')
        ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )
        create_response = self._post_purchase_return(receipt, receipt_item, Decimal('2'), status=0)
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        return_id = create_response.json()['id']
        delete_return_response = self.client.post(
            reverse('api_delete_purchase_return'),
            data=json.dumps({'id': return_id}),
            content_type='application/json',
        )
        self.assertEqual(delete_return_response.json()['status'], 'ok', msg=delete_return_response.content.decode())

        edit_response = self.client.post(
            reverse('api_save_goods_receipt'),
            data=json.dumps({
                'id': receipt.id,
                'code': receipt.code,
                'supplier_id': self.supplier.id,
                'warehouse_id': self.warehouse_a.id,
                'receipt_date': date.today().isoformat(),
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': '9',
                    'unit_price': '100',
                }],
            }),
            content_type='application/json',
        )
        delete_response = self.client.post(
            reverse('api_delete_goods_receipt'),
            data=json.dumps({'id': receipt.id}),
            content_type='application/json',
        )

        self.assertEqual(edit_response.json()['status'], 'error')
        self.assertIn('đã phát sinh trả hàng', edit_response.json()['message'])
        self.assertEqual(delete_response.json()['status'], 'error')
        self.assertIn('đã phát sinh trả hàng', delete_response.json()['message'])
        self.assertTrue(GoodsReceipt.objects.filter(id=receipt.id).exists())
        stock = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        self.assertEqual(stock.quantity, Decimal('10'))

    def test_purchase_return_rejects_goods_receipt_from_another_store(self):
        receipt, receipt_item = self._create_completed_goods_receipt(
            code='P-RETURN-FOREIGN',
            product=self.other_product,
            warehouse=self.other_warehouse,
            created_by=self.other_user,
        )

        response = self._post_purchase_return(receipt, receipt_item, Decimal('1'), status=1)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('Chỉ được trả từ phiếu nhập đã hoàn thành', response.json()['message'])
        self.assertFalse(PurchaseReturn.objects.exists())

    def test_goods_receipt_list_uses_total_item_quantity(self):
        receipt = GoodsReceipt.objects.create(
            code='P00003',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=date.today(),
            status=1,
            total_amount=Decimal('75'),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('5'),
            unit_price=Decimal('10'),
            total_price=Decimal('50'),
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('2.5'),
            unit_price=Decimal('10'),
            total_price=Decimal('25'),
        )

        response = self.client.get(reverse('api_get_goods_receipts'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['id'] == receipt.id)
        self.assertEqual(row['items_count'], 2)
        self.assertEqual(row['total_quantity'], 7.5)
        self.assertEqual(sum(item['quantity'] for item in row['items']), 7.5)

    def test_goods_receipt_list_shows_newest_receipt_first_with_same_receipt_date(self):
        today = date.today()
        now = timezone.now()
        older = GoodsReceipt.objects.create(
            code='P-ORDER-OLDER',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=today,
            status=1,
            total_amount=Decimal('10'),
            created_by=self.user,
        )
        middle = GoodsReceipt.objects.create(
            code='P-ORDER-MIDDLE',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=today,
            status=1,
            total_amount=Decimal('20'),
            created_by=self.user,
        )
        newest = GoodsReceipt.objects.create(
            code='P-ORDER-NEWEST',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=today,
            status=1,
            total_amount=Decimal('30'),
            created_by=self.user,
        )
        GoodsReceipt.objects.filter(id=older.id).update(created_at=now - timedelta(minutes=2))
        GoodsReceipt.objects.filter(id=middle.id).update(created_at=now - timedelta(minutes=1))
        GoodsReceipt.objects.filter(id=newest.id).update(created_at=now)

        response = self.client.get(reverse('api_get_goods_receipts'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [item['code'] for item in response.json()['data'][:3]],
            ['P-ORDER-NEWEST', 'P-ORDER-MIDDLE', 'P-ORDER-OLDER'],
        )

    def test_goods_receipt_list_returns_paginated_meta(self):
        today = date.today()
        now = timezone.now()
        for index in range(11):
            receipt = GoodsReceipt.objects.create(
                code=f'P-PAGE-{index:02d}',
                supplier=self.supplier,
                warehouse=self.warehouse_a,
                receipt_date=today,
                status=1,
                created_by=self.user,
            )
            GoodsReceipt.objects.filter(id=receipt.id).update(created_at=now - timedelta(minutes=10 - index))

        response = self.client.get(
            reverse('api_get_goods_receipts'),
            data={'page': 2, 'page_size': 10},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['page'], 2)
        self.assertEqual(payload['meta']['page_size'], 10)
        self.assertEqual(payload['meta']['page_count'], 1)
        self.assertEqual(payload['meta']['total_pages'], 2)
        self.assertEqual(payload['meta']['total_filtered_count'], 11)
        self.assertEqual(payload['meta']['total_all_count'], 11)
        self.assertEqual(payload['meta']['start_index'], 11)
        self.assertEqual(payload['meta']['end_index'], 11)
        self.assertFalse(payload['meta']['has_next'])
        self.assertEqual([item['code'] for item in payload['data']], ['P-PAGE-00'])

    def test_delete_completed_stock_transfer_reverts_stock(self):
        transfer = StockTransfer.objects.create(
            code='CK-001',
            from_warehouse=self.warehouse_a,
            to_warehouse=self.warehouse_b,
            transfer_date=date.today(),
            status=2,
            created_by=self.user,
        )
        StockTransferItem.objects.create(
            transfer=transfer,
            product=self.product,
            quantity=Decimal('3'),
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('7'))
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_b, quantity=Decimal('3'))

        response = self.client.post(
            reverse('api_delete_stock_transfer'),
            data=json.dumps({'id': transfer.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())

        stock_a = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        stock_b = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_b)
        deleted_transfer = StockTransfer.all_objects.get(id=transfer.id)
        self.assertEqual(stock_a.quantity, Decimal('10'))
        self.assertEqual(stock_b.quantity, Decimal('0'))
        self.assertTrue(deleted_transfer.is_deleted)

    def test_save_completed_stock_transfer_rejects_negative_source_stock_when_disabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Transfer negative disabled',
            opt_allow_negative_stock=False,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('0'))

        response = self.client.post(
            reverse('api_save_stock_transfer'),
            data=json.dumps({
                'code': 'CK-NEG-STOCK',
                'from_warehouse_id': self.warehouse_a.id,
                'to_warehouse_id': self.warehouse_b.id,
                'transfer_date': date.today().isoformat(),
                'status': 2,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Tồn kho không đủ', payload['message'])
        self.assertFalse(StockTransfer.objects.filter(code='CK-NEG-STOCK').exists())

    def test_save_stock_transfer_invalid_item_does_not_revert_existing_stock(self):
        transfer = StockTransfer.objects.create(
            code='CK-ROLLBACK',
            from_warehouse=self.warehouse_a,
            to_warehouse=self.warehouse_b,
            transfer_date=date.today(),
            status=2,
            created_by=self.user,
        )
        StockTransferItem.objects.create(
            transfer=transfer,
            product=self.product,
            quantity=Decimal('3'),
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('7'))
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_b, quantity=Decimal('3'))

        response = self.client.post(
            reverse('api_save_stock_transfer'),
            data=json.dumps({
                'id': transfer.id,
                'code': transfer.code,
                'from_warehouse_id': self.warehouse_a.id,
                'to_warehouse_id': self.warehouse_b.id,
                'transfer_date': date.today().isoformat(),
                'status': 2,
                'items': [{
                    'product_id': self.other_product.id,
                    'quantity': 1,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('sản phẩm', payload['message'].lower())

        stock_a = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a)
        stock_b = ProductStock.objects.get(product=self.product, warehouse=self.warehouse_b)
        self.assertEqual(stock_a.quantity, Decimal('7'))
        self.assertEqual(stock_b.quantity, Decimal('3'))
        self.assertEqual(list(transfer.items.values_list('product_id', 'quantity')), [(self.product.id, Decimal('3.00'))])

    def test_save_stock_check_uses_decimal_quantities(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('5'))

        response = self.client.post(
            reverse('api_save_stock_check'),
            data=json.dumps({
                'code': 'KK-DECIMAL',
                'warehouse_id': self.warehouse_a.id,
                'check_date': date.today().isoformat(),
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'actual_quantity': '3.5',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock_check = StockCheck.objects.get(code='KK-DECIMAL')
        item = StockCheckItem.objects.get(stock_check=stock_check)
        self.assertEqual(item.system_quantity, Decimal('5.00'))
        self.assertEqual(item.actual_quantity, Decimal('3.50'))
        self.assertEqual(item.difference, Decimal('-1.50'))
        self.assertTrue(stock_check.stock_applied)
        self.assertEqual(
            ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a).quantity,
            Decimal('3.50'),
        )

    def test_save_stock_check_auto_generates_code_and_date(self):
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('5.5'))

        response = self.client.post(
            reverse('api_save_stock_check'),
            data=json.dumps({
                'warehouse_id': self.warehouse_a.id,
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'actual_quantity': '7.25',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock_check = StockCheck.objects.get()
        self.assertRegex(stock_check.code, r'^KK\d{5}$')
        self.assertEqual(stock_check.check_date, date.today())
        item = StockCheckItem.objects.get(stock_check=stock_check)
        self.assertEqual(item.system_quantity, Decimal('5.50'))
        self.assertEqual(item.difference, Decimal('1.75'))
        self.assertEqual(
            ProductStock.objects.get(product=self.product, warehouse=self.warehouse_a).quantity,
            Decimal('7.25'),
        )

    def test_draft_stock_check_does_not_change_inventory(self):
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('5'),
        )

        response = self.client.post(
            reverse('api_save_stock_check'),
            data=json.dumps({
                'warehouse_id': self.warehouse_a.id,
                'status': 0,
                'items': [{
                    'product_id': self.product.id,
                    'actual_quantity': '2',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok', msg=response.content.decode())
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('5'))
        self.assertFalse(StockCheck.objects.get().stock_applied)

    def test_edit_completed_stock_check_reapplies_inventory(self):
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )
        create_response = self.client.post(
            reverse('api_save_stock_check'),
            data=json.dumps({
                'code': 'KK-REAPPLY',
                'warehouse_id': self.warehouse_a.id,
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'actual_quantity': '7',
                }],
            }),
            content_type='application/json',
        )
        self.assertEqual(create_response.json()['status'], 'ok', msg=create_response.content.decode())
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('7'))

        stock.quantity = Decimal('5')  # Phát sinh bán thêm 2 sau lần kiểm đầu.
        stock.save(update_fields=['quantity'])
        stock_check = StockCheck.objects.get(code='KK-REAPPLY')
        edit_response = self.client.post(
            reverse('api_save_stock_check'),
            data=json.dumps({
                'id': stock_check.id,
                'code': stock_check.code,
                'warehouse_id': self.warehouse_a.id,
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'actual_quantity': '8',
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(edit_response.json()['status'], 'ok', msg=edit_response.content.decode())
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('8'))
        item = StockCheckItem.objects.get(stock_check=stock_check)
        self.assertEqual(item.system_quantity, Decimal('8'))  # 5 hiện tại - chênh lệch cũ (-3).
        self.assertEqual(item.difference, Decimal('0'))

    def test_delete_completed_stock_check_restores_its_adjustment(self):
        stock = ProductStock.objects.create(
            product=self.product,
            warehouse=self.warehouse_a,
            quantity=Decimal('10'),
        )
        self.client.post(
            reverse('api_save_stock_check'),
            data=json.dumps({
                'code': 'KK-DELETE-APPLIED',
                'warehouse_id': self.warehouse_a.id,
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'actual_quantity': '6',
                }],
            }),
            content_type='application/json',
        )
        stock_check = StockCheck.objects.get(code='KK-DELETE-APPLIED')
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('6'))

        delete_response = self.client.post(
            reverse('api_delete_stock_check'),
            data=json.dumps({'id': stock_check.id}),
            content_type='application/json',
        )

        self.assertEqual(delete_response.json()['status'], 'ok', msg=delete_response.content.decode())
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, Decimal('10'))
        deleted_check = StockCheck.all_objects.get(id=stock_check.id)
        self.assertFalse(deleted_check.stock_applied)

    def test_get_stock_checks_returns_newest_created_first_for_same_date(self):
        StockCheck.objects.create(
            code='KK-OLD',
            warehouse=self.warehouse_a,
            check_date=date.today(),
            status=0,
            created_by=self.user,
        )
        StockCheck.objects.create(
            code='KK-NEW',
            warehouse=self.warehouse_a,
            check_date=date.today(),
            status=0,
            created_by=self.user,
        )

        response = self.client.get(reverse('api_get_stock_checks'))

        self.assertEqual(response.status_code, 200)
        codes = [item['code'] for item in response.json()['data']]
        self.assertEqual(codes[:2], ['KK-NEW', 'KK-OLD'])

    def test_delete_goods_receipt_rejects_negative_stock_when_disabled(self):
        BusinessConfig.objects.create(
            brand=self.brand,
            business_name='Receipt delete negative disabled',
            opt_allow_negative_stock=False,
        )
        receipt = GoodsReceipt.objects.create(
            code='P00004',
            supplier=self.supplier,
            warehouse=self.warehouse_a,
            receipt_date=date.today(),
            status=1,
            total_amount=Decimal('50'),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=Decimal('5'),
            unit_price=Decimal('10'),
            total_price=Decimal('50'),
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse_a, quantity=Decimal('0'))

        response = self.client.post(
            reverse('api_delete_goods_receipt'),
            data=json.dumps({'id': receipt.id}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Tồn kho không đủ', payload['message'])

        receipt.refresh_from_db()
        self.assertFalse(receipt.is_deleted)

    def test_save_goods_receipt_rejects_foreign_warehouse(self):
        response = self.client.post(
            reverse('api_save_goods_receipt'),
            data=json.dumps({
                'code': 'P-FOREIGN-WH',
                'supplier_id': self.supplier.id,
                'warehouse_id': self.other_warehouse.id,
                'receipt_date': date.today().isoformat(),
                'status': 1,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 10,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Kho nhập', payload['message'])
        self.assertFalse(GoodsReceipt.objects.filter(code='P-FOREIGN-WH').exists())

    def test_save_goods_receipt_rejects_foreign_product(self):
        response = self.client.post(
            reverse('api_save_goods_receipt'),
            data=json.dumps({
                'code': 'P-FOREIGN-PRODUCT',
                'supplier_id': self.supplier.id,
                'warehouse_id': self.warehouse_a.id,
                'receipt_date': date.today().isoformat(),
                'status': 1,
                'items': [{
                    'product_id': self.other_product.id,
                    'quantity': 1,
                    'unit_price': 10,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('sản phẩm', payload['message'].lower())
        self.assertFalse(GoodsReceipt.objects.filter(code='P-FOREIGN-PRODUCT').exists())

    def test_save_purchase_order_rejects_foreign_warehouse(self):
        response = self.client.post(
            reverse('api_save_purchase_order'),
            data=json.dumps({
                'code': 'PO-FOREIGN-WH',
                'supplier_id': self.supplier.id,
                'warehouse_id': self.other_warehouse.id,
                'order_date': date.today().isoformat(),
                'status': 0,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                    'unit_price': 10,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('Kho nhập', payload['message'])

    def test_regular_staff_cannot_save_supplier(self):
        response = self.client.post(
            reverse('api_save_supplier'),
            data=json.dumps({
                'code': 'SUP-STAFF',
                'name': 'Supplier Staff',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['status'], 'error')

    def test_brand_owner_cannot_save_receipt_with_product_from_other_store_than_warehouse(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('api_save_goods_receipt'),
            data=json.dumps({
                'code': 'P-CROSS-STORE-PRODUCT',
                'supplier_id': self.supplier.id,
                'warehouse_id': self.warehouse_a.id,
                'receipt_date': date.today().isoformat(),
                'status': 1,
                'items': [{
                    'product_id': self.other_product.id,
                    'quantity': 1,
                    'unit_price': 10,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('không cùng cửa hàng', payload['message'])
        self.assertFalse(GoodsReceipt.objects.filter(code='P-CROSS-STORE-PRODUCT').exists())

    def test_brand_owner_cannot_transfer_between_different_store_warehouses(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('api_save_stock_transfer'),
            data=json.dumps({
                'code': 'CK-CROSS-STORE-WH',
                'from_warehouse_id': self.warehouse_a.id,
                'to_warehouse_id': self.other_warehouse.id,
                'transfer_date': date.today().isoformat(),
                'status': 2,
                'items': [{
                    'product_id': self.product.id,
                    'quantity': 1,
                }],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'error')
        self.assertIn('cùng cửa hàng', payload['message'])
        self.assertFalse(StockTransfer.objects.filter(code='CK-CROSS-STORE-WH').exists())
