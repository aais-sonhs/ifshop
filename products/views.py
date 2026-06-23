import json
import logging
import re
import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_FLOOR
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import transaction
from django.db.models import (
    Case,
    Count,
    DecimalField,
    Exists,
    F,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce
from django.utils import timezone
from .models import (
    Product, ProductCategory, ProductVariant, ProductStock, Supplier, Warehouse,
    GoodsReceipt, GoodsReceiptItem, PurchaseOrder, PurchaseOrderItem,
    StockCheck, StockCheckItem, StockTransfer, StockTransferItem, CostAdjustment,
    ComboItem, ProductLocation
)

from core.store_utils import (
    can_manage_users,
    filter_by_store,
    get_user_store,
    get_managed_store_ids,
    brand_owner_required,
)

logger = logging.getLogger(__name__)


def _forbid_json(message='Bạn không có quyền thực hiện thao tác này'):
    return JsonResponse({'status': 'error', 'message': message}, status=403)


def _generate_next_goods_receipt_code():
    last_receipt = (
        GoodsReceipt.objects
        .select_for_update()
        .filter(code__startswith='P')
        .order_by('-id')
        .first()
    )
    if last_receipt and last_receipt.code and last_receipt.code.startswith('P'):
        try:
            next_number = int(last_receipt.code[1:]) + 1
        except ValueError:
            next_number = GoodsReceipt.objects.filter(code__startswith='P').count() + 1
    else:
        next_number = 1

    candidate = f'P{next_number:05d}'
    while GoodsReceipt.objects.filter(code=candidate).exists():
        next_number += 1
        candidate = f'P{next_number:05d}'
    return candidate


def _generate_next_stock_check_code():
    prefix = 'KK'
    max_number = 0
    for code in (
        StockCheck.all_objects
        .select_for_update()
        .filter(code__startswith=prefix)
        .values_list('code', flat=True)
    ):
        match = re.match(r'^KK-?(\d+)$', code or '', re.IGNORECASE)
        if match:
            max_number = max(max_number, int(match.group(1)))

    next_number = max_number + 1
    while True:
        candidate = f'{prefix}{next_number:05d}'
        if not StockCheck.all_objects.filter(code=candidate).exists():
            return candidate
        next_number += 1


def _generate_next_product_code():
    prefix = 'SP'
    max_number = 0
    for code in Product.all_objects.filter(code__startswith=prefix).values_list('code', flat=True):
        match = re.match(r'^SP-?(\d+)$', code or '', re.IGNORECASE)
        if match:
            max_number = max(max_number, int(match.group(1)))

    next_number = max_number + 1
    while True:
        candidate = f'{prefix}{next_number:03d}'
        if not Product.all_objects.filter(code=candidate).exists():
            return candidate
        next_number += 1


def _generate_next_supplier_code():
    prefix = 'NCC'
    max_number = 0
    for code in Supplier.all_objects.filter(code__startswith=prefix).values_list('code', flat=True):
        match = re.match(r'^NCC-?(\d+)$', code or '', re.IGNORECASE)
        if match:
            max_number = max(max_number, int(match.group(1)))

    next_number = max_number + 1
    while True:
        candidate = f'{prefix}{next_number:03d}'
        if not Supplier.all_objects.filter(code=candidate).exists():
            return candidate
        next_number += 1


def _get_default_store_for_request(request):
    """Return the store new business records should belong to."""
    from system_management.models import Store

    store = get_user_store(request)
    if store:
        return store

    store_ids = get_managed_store_ids(request.user)
    if not store_ids:
        return None
    return Store.objects.filter(id__in=store_ids).order_by('id').first()


def _product_queryset_for_request(request):
    return filter_by_store(Product.objects.all(), request)


def _parse_positive_decimal(value, default='0'):
    try:
        parsed = Decimal(str(value if value not in (None, '') else default))
    except (InvalidOperation, TypeError, ValueError):
        parsed = Decimal(default)
    return parsed


def _to_decimal(value, default='0'):
    try:
        return Decimal(str(value if value not in (None, '') else default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(str(default))


def _to_money_decimal(value, default='0'):
    raw = str(value if value not in (None, '') else default).strip()
    raw = raw.replace('.', '').replace(',', '')
    try:
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(str(default))


def _to_positive_int(value, default, minimum=1, maximum=None):
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError, AttributeError):
        parsed = default
    if parsed < minimum:
        parsed = minimum
    if maximum is not None and parsed > maximum:
        parsed = maximum
    return parsed


def _adjust_stock_quantity(stock, delta):
    """Cộng/trừ tồn kho bằng Decimal để tránh lệch số sau nhiều lần cập nhật."""
    stock.quantity = _to_decimal(stock.quantity) + _to_decimal(delta)
    stock.save(update_fields=['quantity'])


def _get_locked_stock(product_id, warehouse_id):
    stock, _ = ProductStock.objects.select_for_update().get_or_create(
        product_id=product_id,
        warehouse_id=warehouse_id,
        defaults={'quantity': 0},
    )
    return stock


def _allow_negative_stock_for_warehouse_id(warehouse_id):
    if not warehouse_id:
        return False
    try:
        from system_management.models import BusinessConfig

        warehouse = Warehouse.objects.select_related('store__brand').filter(id=warehouse_id).first()
        if not warehouse or not warehouse.store:
            return False
        brand = warehouse.store.brand if warehouse.store.brand_id else None
        return bool(BusinessConfig.get_config(brand=brand).opt_allow_negative_stock)
    except Exception:
        return False


def _adjust_locked_stock(stock, delta, allow_negative=True):
    delta = _to_decimal(delta)
    new_quantity = _to_decimal(stock.quantity) + delta
    if not allow_negative and new_quantity < 0:
        product_name = stock.product.name if getattr(stock, 'product_id', None) and stock.product else 'sản phẩm'
        warehouse_name = stock.warehouse.name if getattr(stock, 'warehouse_id', None) and stock.warehouse else 'kho'
        raise ValueError(f'Tồn kho không đủ cho {product_name} tại {warehouse_name}')
    stock.quantity = new_quantity
    stock.save(update_fields=['quantity'])


def _get_goods_receipt_for_user(request, receipt_id, queryset=None):
    """Lấy phiếu nhập trong phạm vi store mà user được phép thao tác."""
    if not receipt_id:
        return None
    base_queryset = queryset if queryset is not None else GoodsReceipt.objects.all()
    return filter_by_store(base_queryset, request, field_name='warehouse__store').filter(id=receipt_id).first()


def _get_warehouse_for_user(request, warehouse_id):
    if not warehouse_id:
        return None
    return filter_by_store(Warehouse.objects.filter(id=warehouse_id), request).first()


def _get_product_for_user(request, product_id):
    if not product_id:
        return None
    return _product_queryset_for_request(request).filter(id=product_id).first()


def _get_variant_for_product(product, variant_id):
    if not variant_id:
        return None
    return ProductVariant.objects.filter(id=variant_id, product=product).first()


def _ensure_product_matches_store(product, store_id, context='chứng từ'):
    if product and store_id and product.store_id and product.store_id != store_id:
        raise ValueError(f'Sản phẩm "{product.name}" không cùng cửa hàng với {context}.')


def _ensure_warehouses_same_store(from_warehouse, to_warehouse):
    if (
        from_warehouse and to_warehouse and
        from_warehouse.store_id and to_warehouse.store_id and
        from_warehouse.store_id != to_warehouse.store_id
    ):
        raise ValueError('Kho xuất và kho nhập phải thuộc cùng cửa hàng.')


def _get_stock_transfer_for_user(request, transfer_id, queryset=None):
    """Lấy phiếu chuyển kho trong phạm vi store mà user được phép thao tác."""
    if not transfer_id:
        return None
    base_queryset = queryset if queryset is not None else StockTransfer.objects.all()
    return filter_by_store(base_queryset, request, field_name='from_warehouse__store').filter(id=transfer_id).first()


def _get_stock_check_for_user(request, check_id, queryset=None):
    if not check_id:
        return None
    base_queryset = queryset if queryset is not None else StockCheck.objects.all()
    return filter_by_store(base_queryset, request, field_name='warehouse__store').filter(id=check_id).first()


def _get_purchase_order_for_user(request, purchase_order_id, queryset=None):
    if not purchase_order_id:
        return None
    base_queryset = queryset if queryset is not None else PurchaseOrder.objects.all()
    return filter_by_store(base_queryset, request, field_name='warehouse__store').filter(id=purchase_order_id).first()


def _normalize_goods_receipt_items(items_data):
    """Chuẩn hóa danh sách item phiếu nhập về Decimal để tái sử dụng an toàn."""
    normalized_items = []
    total_amount = Decimal('0')

    for item in items_data:
        quantity = _to_decimal(item.get('quantity', 0))
        unit_price = _to_decimal(item.get('unit_price', 0))
        line_total = quantity * unit_price
        normalized_items.append({
            'product_id': item.get('product_id'),
            'variant_id': item.get('variant_id') or None,
            'quantity': quantity,
            'unit_price': unit_price,
            'total_price': line_total,
        })
        total_amount += line_total

    return normalized_items, total_amount


def _calculate_goods_receipt_totals(items):
    """Tính tổng số lượng và tổng tiền trực tiếp từ dòng chi tiết phiếu nhập."""
    total_quantity = Decimal('0')
    total_amount = Decimal('0')
    for item in items:
        quantity = _to_decimal(item.quantity)
        unit_price = _to_decimal(item.unit_price)
        total_quantity += quantity
        total_amount += quantity * unit_price
    return total_quantity, total_amount


def _apply_goods_receipt_stock_adjustment(receipt, warehouse_id, multiplier):
    """Áp biến động tồn kho cho phiếu nhập theo chiều cộng hoặc hoàn tác.

    - `multiplier = 1`: cộng tồn theo item phiếu nhập
    - `multiplier = -1`: hoàn tác phần tồn đã cộng trước đó
    """
    if not warehouse_id:
        return

    allow_negative = True
    if _to_decimal(multiplier) < 0:
        allow_negative = _allow_negative_stock_for_warehouse_id(warehouse_id)

    for item in receipt.items.all():
        stock = _get_locked_stock(item.product_id, warehouse_id)
        _adjust_locked_stock(
            stock,
            _to_decimal(item.quantity) * _to_decimal(multiplier),
            allow_negative=allow_negative,
        )


def _sync_product_cost_after_goods_receipt(product_id, received_quantity, received_price):
    """Cập nhật giá vốn tham chiếu và giá nhập mới nhất sau khi nhập hàng hoàn thành."""
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return

    total_old_stock = Decimal(str(sum(
        float(stock.quantity) for stock in ProductStock.objects.filter(product_id=product_id)
    ))) - received_quantity
    if total_old_stock < 0:
        total_old_stock = Decimal('0')

    total_new_stock = total_old_stock + received_quantity
    if total_new_stock > 0:
        old_cost = product.cost_price or Decimal('0')
        weighted_avg = (
            (total_old_stock * old_cost) + (received_quantity * received_price)
        ) / total_new_stock
        product.cost_price = round(weighted_avg)

    product.import_price = received_price
    product.save(update_fields=['cost_price', 'import_price'])


def _recreate_goods_receipt_items(receipt, normalized_items):
    """Xóa item cũ và tạo lại item mới cho phiếu nhập."""
    receipt.items.all().delete()
    for item in normalized_items:
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product_id=item['product_id'],
            variant_id=item['variant_id'],
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            total_price=item['total_price'],
        )


def _normalize_stock_transfer_items(items_data):
    """Chuẩn hóa danh sách item phiếu chuyển kho về Decimal để tái sử dụng an toàn."""
    normalized_items = []
    for item in items_data:
        normalized_items.append({
            'product_id': item.get('product_id'),
            'variant_id': item.get('variant_id') or None,
            'quantity': _to_decimal(item.get('quantity', 0)),
        })
    return normalized_items


def _apply_stock_transfer_adjustment(transfer, from_warehouse_id, to_warehouse_id, reverse=False):
    """Áp hoặc hoàn tác tồn kho cho phiếu chuyển kho.

    - `reverse=False`: trừ kho xuất, cộng kho nhập
    - `reverse=True`: hoàn tác phiếu đã hoàn thành trước đó
    """
    if not from_warehouse_id or not to_warehouse_id:
        return

    from_multiplier = Decimal('1') if reverse else Decimal('-1')
    to_multiplier = Decimal('-1') if reverse else Decimal('1')

    for item in transfer.items.all():
        quantity = _to_decimal(item.quantity)
        from_stock = _get_locked_stock(item.product_id, from_warehouse_id)
        from_delta = quantity * from_multiplier
        _adjust_locked_stock(
            from_stock,
            from_delta,
            allow_negative=from_delta >= 0 or _allow_negative_stock_for_warehouse_id(from_warehouse_id),
        )

        to_stock = _get_locked_stock(item.product_id, to_warehouse_id)
        to_delta = quantity * to_multiplier
        _adjust_locked_stock(
            to_stock,
            to_delta,
            allow_negative=to_delta >= 0 or _allow_negative_stock_for_warehouse_id(to_warehouse_id),
        )


def _recreate_stock_transfer_items(transfer, normalized_items):
    """Xóa item cũ và tạo lại item mới cho phiếu chuyển kho."""
    transfer.items.all().delete()
    for item in normalized_items:
        StockTransferItem.objects.create(
            transfer=transfer,
            product_id=item['product_id'],
            variant_id=item['variant_id'],
            quantity=item['quantity'],
        )


def _combo_signature(combo_product):
    return {
        (item.product_id, Decimal(str(item.quantity)).normalize())
        for item in combo_product.combo_items.all()
    }


def _validate_combo_items(product, combo_data, request):
    if not isinstance(combo_data, list):
        raise ValueError('Danh sách thành phần combo không hợp lệ.')
    if len(combo_data) > 20:
        raise ValueError('Một combo chỉ nên có tối đa 20 sản phẩm thành phần.')

    seen = set()
    normalized = []
    for item in combo_data:
        product_id = item.get('product_id') if isinstance(item, dict) else None
        if not product_id:
            continue
        product_id = int(product_id)
        if product.id and product_id == product.id:
            raise ValueError('Combo không được chứa chính nó trong thành phần.')
        if product_id in seen:
            raise ValueError('Không được chọn trùng sản phẩm trong cùng một combo.')
        seen.add(product_id)

        quantity = _parse_positive_decimal(item.get('quantity'), '1')
        if quantity <= 0:
            raise ValueError('Số lượng thành phần combo phải lớn hơn 0.')

        component = _product_queryset_for_request(request).filter(
            id=product_id,
            is_active=True,
        ).first()
        if not component:
            raise ValueError('Có sản phẩm thành phần không tồn tại hoặc không thuộc cửa hàng hiện tại.')
        if component.is_combo:
            raise ValueError(f'Không thể dùng combo "{component.name}" làm thành phần của combo khác.')
        normalized.append((component, quantity))

    if not normalized:
        raise ValueError('Vui lòng chọn ít nhất 1 sản phẩm thành phần cho combo.')
    return normalized


def _calculate_combo_cost(normalized_combo_items):
    return sum(
        Decimal(str(component.cost_price or 0)) * quantity
        for component, quantity in normalized_combo_items
    )


def _combo_stock_by_warehouse(product):
    """
    Combo không có tồn kho riêng; tồn khả dụng = min(tồn thành phần / định lượng)
    theo từng kho. Thành phần dịch vụ không giới hạn tồn combo.
    """
    capacities = None
    warehouse_names = {}

    combo_items = product.combo_items.select_related('product').prefetch_related(
        'product__stocks__warehouse'
    )
    for item in combo_items:
        component = item.product
        if not component or component.is_service:
            continue
        required_qty = _parse_positive_decimal(item.quantity, '0')
        if required_qty <= 0:
            continue

        component_capacity = {}
        for stock in component.stocks.all():
            if not stock.warehouse_id:
                continue
            warehouse_names[stock.warehouse_id] = stock.warehouse.name if stock.warehouse else ''
            stock_qty = _parse_positive_decimal(stock.quantity, '0')
            capacity = (stock_qty / required_qty).to_integral_value(rounding=ROUND_FLOOR)
            component_capacity[stock.warehouse_id] = int(capacity)

        if capacities is None:
            capacities = component_capacity
        else:
            warehouse_ids = set(capacities.keys()) | set(component_capacity.keys())
            capacities = {
                warehouse_id: min(capacities.get(warehouse_id, 0), component_capacity.get(warehouse_id, 0))
                for warehouse_id in warehouse_ids
            }

    if capacities is None:
        return [], 0

    rows = []
    total_stock = 0
    for warehouse_id, quantity in sorted(capacities.items(), key=lambda row: warehouse_names.get(row[0], '')):
        total_stock += quantity
        if quantity != 0:
            rows.append({
                'warehouse': warehouse_names.get(warehouse_id, ''),
                'warehouse_id': warehouse_id,
                'quantity': float(quantity),
            })
    return rows, float(total_stock)


def _build_receipt_history_map(products):
    product_ids = [p.id for p in products]
    if not product_ids:
        return {}

    all_receipt_items = (
        GoodsReceiptItem.objects
        .filter(
            product_id__in=product_ids,
            goods_receipt__status=1,
        )
        .select_related(
            'goods_receipt',
            'goods_receipt__supplier',
            'goods_receipt__warehouse',
            'goods_receipt__purchase_order',
            'goods_receipt__created_by',
            'variant',
        )
        .order_by('-goods_receipt__receipt_date', '-goods_receipt__id', '-id')
    )

    receipt_map = {}
    for receipt_item in all_receipt_items:
        receipt_map.setdefault(receipt_item.product_id, []).append(receipt_item)
    return receipt_map


def _annotate_product_list_queryset(queryset):
    latest_receipt_items = (
        GoodsReceiptItem.objects
        .filter(product_id=OuterRef('pk'), goods_receipt__status=1)
        .order_by('-goods_receipt__receipt_date', '-goods_receipt__id', '-id')
    )
    total_stock_subquery = (
        ProductStock.objects
        .filter(product_id=OuterRef('pk'))
        .values('product_id')
        .annotate(total=Coalesce(Sum('quantity'), Value(Decimal('0'))))
        .values('total')[:1]
    )
    return queryset.annotate(
        latest_purchase_date=Subquery(latest_receipt_items.values('goods_receipt__receipt_date')[:1]),
        purchase_price_count=Count(
            'receipt_items__unit_price',
            filter=Q(receipt_items__goods_receipt__status=1),
            distinct=True,
        ),
        total_stock_simple=Coalesce(
            Subquery(total_stock_subquery, output_field=DecimalField(max_digits=15, decimal_places=2)),
            Value(Decimal('0'), output_field=DecimalField(max_digits=15, decimal_places=2)),
        ),
        is_combo_component=Exists(
            ComboItem.objects.filter(
                product_id=OuterRef('pk'),
                combo__is_active=True,
                combo__is_deleted=False,
            )
        ),
        low_stock_threshold=Case(
            When(min_stock__gt=0, then=F('min_stock')),
            default=Value(Decimal('5'), output_field=DecimalField(max_digits=15, decimal_places=2)),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    )


def _apply_product_list_filters(queryset, request, apply_computed_stock_filters=True):
    params = request.GET
    text = (params.get('text') or '').strip()
    category = (params.get('category') or '').strip()
    product_type = (params.get('product_type') or '').strip()
    stock = (params.get('stock') or '').strip()
    supplier = (params.get('supplier') or '').strip()
    location = (params.get('location') or '').strip()
    product_kind = (params.get('type') or '').strip()
    combo_usage = (params.get('combo_usage') or '').strip()
    status = (params.get('status') or '').strip()
    creator = (params.get('creator') or '').strip()
    created_from = (params.get('created_from') or '').strip()
    created_to = (params.get('created_to') or '').strip()
    import_history = (params.get('import_history') or '').strip()
    import_from = (params.get('import_from') or '').strip()
    import_to = (params.get('import_to') or '').strip()
    price_basis = (params.get('price_basis') or 'import_price').strip()
    price_from = (params.get('price_from') or '').strip()
    price_to = (params.get('price_to') or '').strip()

    queryset = _annotate_product_list_queryset(queryset)

    if text:
        combo_match_ids = ComboItem.objects.filter(
            Q(product__code__icontains=text) |
            Q(product__name__icontains=text)
        ).values('combo_id')
        combo_parent_match_ids = ComboItem.objects.filter(
            combo__is_active=True,
            combo__is_deleted=False,
        ).filter(
            Q(combo__code__icontains=text) |
            Q(combo__name__icontains=text)
        ).values('product_id')
        queryset = queryset.filter(
            Q(code__icontains=text) |
            Q(name__icontains=text) |
            Q(barcode__icontains=text) |
            Q(specification__icontains=text) |
            Q(category__name__icontains=text) |
            Q(category__parent__name__icontains=text) |
            Q(supplier__name__icontains=text) |
            Q(location__name__icontains=text) |
            Q(pk__in=combo_match_ids) |
            Q(pk__in=combo_parent_match_ids)
        )

    if category:
        queryset = queryset.filter(
            Q(category_id=category) |
            Q(category__parent_id=category)
        )
    if product_type:
        queryset = queryset.filter(category_id=product_type)
    if supplier:
        queryset = queryset.filter(supplier_id=supplier)
    if location:
        queryset = queryset.filter(location_id=location)

    if product_kind == 'normal':
        queryset = queryset.filter(is_combo=False, is_service=False, is_weight_based=False)
    elif product_kind == 'combo':
        queryset = queryset.filter(is_combo=True)
    elif product_kind == 'service':
        queryset = queryset.filter(is_service=True)
    elif product_kind == 'weight':
        queryset = queryset.filter(is_weight_based=True)

    if combo_usage == 'is_combo':
        queryset = queryset.filter(is_combo=True)
    elif combo_usage == 'component':
        queryset = queryset.filter(is_combo_component=True)
    elif combo_usage == 'standalone':
        queryset = queryset.filter(is_combo=False, is_combo_component=False)

    if status == 'active':
        queryset = queryset.filter(is_active=True)
    elif status == 'inactive':
        queryset = queryset.filter(is_active=False)

    if creator:
        queryset = queryset.filter(created_by_id=creator)
    if created_from:
        queryset = queryset.filter(created_at__date__gte=created_from)
    if created_to:
        queryset = queryset.filter(created_at__date__lte=created_to)

    if apply_computed_stock_filters:
        if stock == 'out':
            queryset = queryset.filter(total_stock_simple__lte=0)
        elif stock == 'instock':
            queryset = queryset.filter(total_stock_simple__gt=0)
        elif stock == 'low':
            queryset = queryset.filter(total_stock_simple__gt=0, total_stock_simple__lte=F('low_stock_threshold'))

    if import_history == 'has_import':
        queryset = queryset.filter(latest_purchase_date__isnull=False)
    elif import_history == 'no_import':
        queryset = queryset.filter(latest_purchase_date__isnull=True)
    elif import_history == 'changed_price':
        queryset = queryset.filter(purchase_price_count__gt=1)

    if import_from:
        queryset = queryset.filter(latest_purchase_date__gte=import_from)
    if import_to:
        queryset = queryset.filter(latest_purchase_date__lte=import_to)

    price_field_map = {
        'import_price': 'import_price',
        'cost_price': 'cost_price',
        'selling_price': 'selling_price',
        'wholesale_price_no_warranty': 'wholesale_price_no_warranty',
        'wholesale_price_warranty': 'wholesale_price_warranty',
        'total_stock': 'total_stock_simple',
    }
    price_field = price_field_map.get(price_basis)
    if not apply_computed_stock_filters and price_basis == 'total_stock':
        price_field = None
    if price_field:
        if price_from not in ('', None):
            queryset = queryset.filter(**{f'{price_field}__gte': _to_decimal(price_from)})
        if price_to not in ('', None):
            queryset = queryset.filter(**{f'{price_field}__lte': _to_decimal(price_to)})

    return queryset


def _prefetch_product_list_queryset(queryset):
    return queryset.select_related(
        'category',
        'category__parent',
        'supplier',
        'location',
        'created_by',
    ).prefetch_related(
        'variants',
        'stocks__warehouse',
        'combo_items__product__stocks__warehouse',
        'in_combos__combo',
        'receipt_items__goods_receipt',
    )


def _serialize_product_list(products):
    receipt_map = _build_receipt_history_map(products)

    data = []
    for p in products:
        category = p.category
        root_category = category.parent if category and category.parent_id else category
        product_type = category if category and category.parent_id else None
        receipt_items = receipt_map.get(p.id, [])
        variants = [{
            'id': v.id,
            'size_name': v.size_name,
            'sku': v.sku,
            'barcode': v.barcode or '',
            'import_price': float(v.import_price),
            'cost_price': float(v.cost_price),
            'selling_price': float(v.selling_price),
            'wholesale_price_no_warranty': float(v.wholesale_price_no_warranty),
            'wholesale_price_warranty': float(v.wholesale_price_warranty),
            'is_active': v.is_active,
        } for v in p.variants.all()]

        combo_items = []
        if p.is_combo:
            combo_items = [{
                'product_id': ci.product_id,
                'product_code': ci.product.code if ci.product else '',
                'product_name': ci.product.name if ci.product else '',
                'product_image': ci.product.image.url if ci.product and ci.product.image else '',
                'unit': ci.product.unit if ci.product else '',
                'is_service': ci.product.is_service if ci.product else False,
                'quantity': float(ci.quantity),
                'total_stock': float(sum(stock.quantity for stock in ci.product.stocks.all())) if ci.product else 0,
                'cost_price': float(ci.product.cost_price) if ci.product else 0,
                'selling_price': float(ci.product.selling_price) if ci.product else 0,
                'line_cost': float(ci.quantity * ci.product.cost_price) if ci.product else 0,
                'line_total': float(ci.quantity * ci.product.selling_price) if ci.product else 0,
            } for ci in p.combo_items.select_related('product').all()]

        combo_parents = []
        for combo_link in p.in_combos.all():
            combo = combo_link.combo
            if not combo or not combo.is_active or getattr(combo, 'is_deleted', False):
                continue
            combo_parents.append({
                'id': combo.id,
                'code': combo.code,
                'name': combo.name,
                'quantity': float(combo_link.quantity),
            })

        stock_by_warehouse = [{
            'warehouse': s.warehouse.name if s.warehouse else '',
            'warehouse_id': s.warehouse_id,
            'quantity': float(s.quantity),
        } for s in p.stocks.select_related('warehouse').all()]

        if p.is_combo:
            stock_by_warehouse, total_stock = _combo_stock_by_warehouse(p)
        else:
            total_stock = float(sum(s.quantity for s in p.stocks.all()))

        current_cost = _calc_on_hand_cost_price(p.id, total_stock, receipt_map)
        effective_cost = current_cost if current_cost > 0 else float(p.cost_price)

        latest_purchase = None
        recent_import_prices = []
        recent_purchase_receipts = []
        purchase_receipt_count = 0
        purchase_total_quantity = 0
        purchase_total_amount = 0
        purchase_price_changed = False
        if receipt_items:
            receipt_summaries = _summarize_purchase_receipts(receipt_items)
            recent_purchase_receipts = receipt_summaries[:3]
            latest_item = receipt_items[0]
            latest_purchase = {
                'receipt_id': latest_item.goods_receipt_id,
                'receipt_code': latest_item.goods_receipt.code if latest_item.goods_receipt else '',
                'receipt_date': latest_item.goods_receipt.receipt_date.strftime('%d/%m/%Y') if latest_item.goods_receipt and latest_item.goods_receipt.receipt_date else '',
                'receipt_date_raw': latest_item.goods_receipt.receipt_date.strftime('%Y-%m-%d') if latest_item.goods_receipt and latest_item.goods_receipt.receipt_date else '',
                'supplier': latest_item.goods_receipt.supplier.name if latest_item.goods_receipt and latest_item.goods_receipt.supplier else '',
                'warehouse': latest_item.goods_receipt.warehouse.name if latest_item.goods_receipt and latest_item.goods_receipt.warehouse else '',
                'variant': latest_item.variant.size_name if latest_item.variant else '',
                'quantity': float(latest_item.quantity),
                'unit_price': float(latest_item.unit_price),
                'total_price': float(latest_item.total_price),
            }
            for item in receipt_items:
                purchase_total_quantity += float(item.quantity or 0)
                purchase_total_amount += float(item.total_price or 0)
            purchase_receipt_count = len(receipt_summaries)
            purchase_price_changed = len({
                float(item.unit_price or 0)
                for item in receipt_items
            }) > 1
            recent_import_prices = [{
                'date': item.goods_receipt.receipt_date.strftime('%d/%m/%Y') if item.goods_receipt and item.goods_receipt.receipt_date else '',
                'price': float(item.unit_price),
                'quantity': float(item.quantity),
            } for item in receipt_items[:3]]

        data.append({
            'id': p.id, 'code': p.code, 'name': p.name, 'barcode': p.barcode or '',
            'category': root_category.name if root_category else '',
            'category_id': root_category.id if root_category else None,
            'category_record_id': p.category_id,
            'product_type': product_type.name if product_type else '',
            'product_type_id': product_type.id if product_type else None,
            'unit': p.unit,
            'import_price': float(p.import_price),
            'cost_price': effective_cost,
            'cost_price_stored': float(p.cost_price),
            'selling_price': float(p.selling_price),
            'retail_price': float(p.selling_price),
            'wholesale_price_no_warranty': float(p.wholesale_price_no_warranty),
            'wholesale_price_warranty': float(p.wholesale_price_warranty),
            'min_stock': p.min_stock, 'max_stock': p.max_stock,
            'supplier': p.supplier.name if p.supplier else '',
            'supplier_id': p.supplier_id,
            'total_stock': total_stock,
            'stock_by_warehouse': stock_by_warehouse,
            'image': p.image.url if p.image else '',
            'description': p.description or '',
            'is_active': p.is_active,
            'is_weight_based': p.is_weight_based,
            'is_service': p.is_service,
            'is_combo': p.is_combo,
            'combo_items': combo_items,
            'combo_parents': combo_parents,
            'combo_parent_count': len(combo_parents),
            'variants': variants,
            'location_id': p.location_id,
            'location': p.location.name if p.location else '',
            'specification': p.specification or '',
            'created_at': p.created_at.strftime('%Y-%m-%d') if p.created_at else '',
            'created_at_display': p.created_at.strftime('%d/%m/%Y %H:%M') if p.created_at else '',
            'creator_id': p.created_by_id,
            'creator_name': (p.created_by.get_full_name() or p.created_by.username) if p.created_by else '',
            'latest_purchase': latest_purchase,
            'recent_purchase_receipts': recent_purchase_receipts,
            'recent_import_prices': recent_import_prices,
            'purchase_receipt_count': purchase_receipt_count,
            'purchase_total_quantity': purchase_total_quantity,
            'purchase_total_amount': purchase_total_amount,
            'purchase_price_changed': purchase_price_changed,
        })
    return data


def _needs_python_product_post_filter(request):
    params = request.GET
    stock = (params.get('stock') or '').strip()
    price_basis = (params.get('price_basis') or 'import_price').strip()
    price_from = (params.get('price_from') or '').strip()
    price_to = (params.get('price_to') or '').strip()
    return bool(stock) or (
        price_basis == 'total_stock' and (price_from not in ('', None) or price_to not in ('', None))
    )


def _apply_python_product_post_filters(items, request):
    params = request.GET
    stock = (params.get('stock') or '').strip()
    price_basis = (params.get('price_basis') or 'import_price').strip()
    price_from = (params.get('price_from') or '').strip()
    price_to = (params.get('price_to') or '').strip()

    price_from_value = float(_to_decimal(price_from)) if price_from not in ('', None) else None
    price_to_value = float(_to_decimal(price_to)) if price_to not in ('', None) else None

    filtered = []
    for item in items:
        total_stock = float(item.get('total_stock') or 0)
        low_stock_threshold = item.get('min_stock') or 0
        low_stock_threshold = float(low_stock_threshold) if low_stock_threshold and low_stock_threshold > 0 else 5.0

        if stock == 'out' and total_stock > 0:
            continue
        if stock == 'instock' and total_stock <= 0:
            continue
        if stock == 'low' and (total_stock <= 0 or total_stock > low_stock_threshold):
            continue

        if price_basis == 'total_stock':
            if price_from_value is not None and total_stock < price_from_value:
                continue
            if price_to_value is not None and total_stock > price_to_value:
                continue

        filtered.append(item)
    return filtered


def _summarize_purchase_receipts(receipt_items, limit=None):
    """Gom lịch sử nhập của sản phẩm theo từng phiếu nhập."""
    receipts = []
    receipt_index = {}

    for item in receipt_items:
        receipt = item.goods_receipt
        if not receipt:
            continue

        receipt_id = receipt.id
        if receipt_id not in receipt_index:
            receipt_index[receipt_id] = {
                'receipt_id': receipt_id,
                'receipt_code': receipt.code,
                'receipt_date': receipt.receipt_date.strftime('%d/%m/%Y') if receipt.receipt_date else '',
                'receipt_date_raw': receipt.receipt_date.strftime('%Y-%m-%d') if receipt.receipt_date else '',
                'supplier': receipt.supplier.name if receipt.supplier else '',
                'warehouse': receipt.warehouse.name if receipt.warehouse else '',
                'purchase_order': receipt.purchase_order.code if getattr(receipt, 'purchase_order', None) else '',
                'created_by': (
                    receipt.created_by.get_full_name() or receipt.created_by.username
                ) if receipt.created_by else '',
                'note': receipt.note or '',
                'receipt_total_amount': float(receipt.total_amount or 0),
                'item_count': 0,
                'quantity': 0.0,
                'total_price': 0.0,
                'min_unit_price': None,
                'max_unit_price': None,
                'avg_unit_price': 0.0,
                'items': [],
            }
            receipts.append(receipt_index[receipt_id])

        row = receipt_index[receipt_id]
        quantity = float(item.quantity or 0)
        unit_price = float(item.unit_price or 0)
        total_price = float(item.total_price or 0)
        row['item_count'] += 1
        row['quantity'] += quantity
        row['total_price'] += total_price
        row['min_unit_price'] = unit_price if row['min_unit_price'] is None else min(row['min_unit_price'], unit_price)
        row['max_unit_price'] = unit_price if row['max_unit_price'] is None else max(row['max_unit_price'], unit_price)
        row['items'].append({
            'variant': item.variant.size_name if item.variant else '',
            'quantity': quantity,
            'unit_price': unit_price,
            'total_price': total_price,
        })

    for row in receipts:
        row['avg_unit_price'] = round(row['total_price'] / row['quantity']) if row['quantity'] > 0 else 0
        row['min_unit_price'] = row['min_unit_price'] or 0
        row['max_unit_price'] = row['max_unit_price'] or 0

    if limit:
        return receipts[:limit]
    return receipts


def _calc_on_hand_cost_price(product_id, total_stock, receipt_map):
    """
    Giá vốn tồn hiện tại = tổng giá trị các lô nhập còn lại / tổng SL tồn hiện tại.
    Duyệt phiếu nhập mới nhất -> cũ nhất để xác định những lô còn nằm trong số tồn.
    """
    total_stock = _to_decimal(total_stock)
    if total_stock <= 0:
        return 0

    receipt_items = receipt_map.get(product_id, [])
    if not receipt_items:
        return 0

    remaining = total_stock
    weighted_sum = Decimal('0')
    for receipt_item in receipt_items:
        quantity = _to_decimal(receipt_item.quantity)
        unit_price = _to_decimal(receipt_item.unit_price)
        if quantity <= 0:
            continue

        allocated = min(remaining, quantity)
        weighted_sum += allocated * unit_price
        remaining -= allocated
        if remaining <= 0:
            break

    return round(weighted_sum / total_stock) if total_stock > 0 else 0


# ============ PAGE VIEWS ============

@login_required(login_url="/login/")
@brand_owner_required
def product_tbl(request):
    store_ids = get_managed_store_ids(request.user)
    categories = list(
        ProductCategory.objects.filter(is_active=True).values('id', 'name', 'parent_id')
    )
    creator_rows = (
        filter_by_store(Product.objects.filter(created_by__isnull=False), request)
        .values('created_by_id', 'created_by__first_name', 'created_by__last_name', 'created_by__username')
        .distinct()
        .order_by('created_by__first_name', 'created_by__last_name', 'created_by__username')
    )
    creators = []
    for row in creator_rows:
        full_name = ' '.join(part for part in [row['created_by__first_name'], row['created_by__last_name']] if part).strip()
        creators.append({
            'id': row['created_by_id'],
            'name': full_name or row['created_by__username'] or 'Không xác định',
        })
    context = {
        'active_tab': 'product_tbl',
        'categories': categories,
        'root_categories': [category for category in categories if not category['parent_id']],
        'product_types': [category for category in categories if category['parent_id']],
        'suppliers': list(Supplier.objects.filter(is_active=True).values('id', 'name')),
        'locations': list(ProductLocation.objects.filter(is_active=True).values('id', 'name')),
        'warehouses': list(Warehouse.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'name')),
        'creators': creators,
    }
    return render(request, "products/product_list.html", context)


@login_required(login_url="/login/")
@brand_owner_required
def warehouse_tbl(request):
    context = {'active_tab': 'warehouse_tbl'}
    return render(request, "products/warehouse_list.html", context)


@login_required(login_url="/login/")
@brand_owner_required
def purchase_order_tbl(request):
    store_ids = get_managed_store_ids(request.user)
    context = {
        'active_tab': 'purchase_order_tbl',
        'suppliers': list(Supplier.objects.filter(is_active=True).values('id', 'name')),
        'warehouses': list(Warehouse.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'name')),
        'products': list(Product.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'code', 'name', 'cost_price', 'unit')),
    }
    return render(request, "products/purchase_order_list.html", context)


@login_required(login_url="/login/")
@brand_owner_required
def goods_receipt_tbl(request):
    store_ids = get_managed_store_ids(request.user)
    context = {
        'active_tab': 'goods_receipt_tbl',
        'suppliers': list(Supplier.objects.filter(is_active=True).values('id', 'name')),
        'warehouses': list(Warehouse.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'name')),
        'products': list(Product.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'code', 'name', 'cost_price', 'unit')),
    }
    return render(request, "products/goods_receipt_list.html", context)


@login_required(login_url="/login/")
@brand_owner_required
def stock_check_tbl(request):
    store_ids = get_managed_store_ids(request.user)
    context = {
        'active_tab': 'stock_check_tbl',
        'warehouses': list(Warehouse.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'name')),
        'products': list(Product.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'code', 'name', 'unit')),
    }
    return render(request, "products/stock_check_list.html", context)


@login_required(login_url="/login/")
@brand_owner_required
def stock_transfer_tbl(request):
    store_ids = get_managed_store_ids(request.user)
    context = {
        'active_tab': 'stock_transfer_tbl',
        'warehouses': list(Warehouse.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'name')),
        'products': list(Product.objects.filter(is_active=True, store_id__in=store_ids).values('id', 'code', 'name', 'unit')),
    }
    return render(request, "products/stock_transfer_list.html", context)


@login_required(login_url="/login/")
@brand_owner_required
def supplier_tbl(request):
    context = {'active_tab': 'supplier_tbl'}
    return render(request, "products/supplier_list.html", context)


@login_required(login_url="/login/")
def cost_adjustment_tbl(request):
    context = {'active_tab': 'cost_adjustment_tbl'}
    return render(request, "products/cost_adjustment_list.html", context)


# ============ API: PRODUCT ============

@login_required(login_url="/login/")
def api_get_products(request):
    page = _to_positive_int(request.GET.get('page'), default=1, minimum=1)
    page_size = _to_positive_int(request.GET.get('page_size'), default=50, minimum=10, maximum=200)

    base_queryset = filter_by_store(Product.objects.all(), request)
    total_all_count = base_queryset.count()
    needs_python_post_filter = _needs_python_product_post_filter(request)
    filtered_queryset = _apply_product_list_filters(
        base_queryset,
        request,
        apply_computed_stock_filters=not needs_python_post_filter,
    ).order_by('name', 'id')
    if needs_python_post_filter:
        # Combo ton kho duoc tinh tu thanh phan, nen mot so bo loc can chay sau khi serialize.
        products = list(_prefetch_product_list_queryset(filtered_queryset))
        filtered_data = _apply_python_product_post_filters(_serialize_product_list(products), request)
        paginator = Paginator(filtered_data, page_size)
        page_obj = paginator.get_page(page)
        data = list(page_obj.object_list)
    else:
        paginator = Paginator(filtered_queryset, page_size)
        page_obj = paginator.get_page(page)
        products = list(_prefetch_product_list_queryset(page_obj.object_list))
        data = _serialize_product_list(products)

    return JsonResponse({
        'data': data,
        'meta': {
            'page': page_obj.number,
            'page_size': page_size,
            'page_count': len(data),
            'total_pages': paginator.num_pages,
            'total_filtered_count': paginator.count,
            'total_all_count': total_all_count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'start_index': page_obj.start_index() if paginator.count else 0,
            'end_index': page_obj.end_index() if paginator.count else 0,
        }
    })


@login_required(login_url="/login/")
def api_get_combo_source_products(request):
    products = list(
        filter_by_store(
            Product.objects.filter(is_combo=False).prefetch_related('stocks__warehouse').order_by('name', 'id'),
            request,
        )
    )
    data = []
    for product in products:
        stock_by_warehouse = [{
            'warehouse_id': stock.warehouse_id,
            'quantity': float(stock.quantity),
        } for stock in product.stocks.all()]
        data.append({
            'id': product.id,
            'code': product.code,
            'name': product.name,
            'selling_price': float(product.selling_price),
            'cost_price': float(product.cost_price),
            'total_stock': float(sum(stock.quantity for stock in product.stocks.all())),
            'stock_by_warehouse': stock_by_warehouse,
            'is_service': product.is_service,
            'is_active': product.is_active,
        })
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_check_product_code(request):
    term = (request.GET.get('term') or '').strip()
    if not term:
        return JsonResponse({'status': 'ok', 'exists': False})

    products = filter_by_store(Product.objects.filter(is_active=True), request)
    product = products.filter(code__iexact=term).first()
    if not product:
        product = products.filter(barcode__iexact=term).first()
    if not product:
        return JsonResponse({'status': 'ok', 'exists': False})

    return JsonResponse({
        'status': 'ok',
        'exists': True,
        'product': {
            'id': product.id,
            'code': product.code,
            'barcode': product.barcode or '',
            'name': product.name,
            'unit': product.unit,
            'selling_price': float(product.selling_price),
            'import_price': float(product.import_price),
            'cost_price': float(product.cost_price),
            'image_url': product.image.url if product.image else '',
            'is_service': product.is_service,
        }
    })


@login_required(login_url="/login/")
def api_save_product(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        with transaction.atomic():
            product_id = request.POST.get('id')
            if product_id:
                product = _product_queryset_for_request(request).get(id=product_id)
                was_combo = product.is_combo
                existing_combo_signature = _combo_signature(product) if was_combo else set()
            else:
                product = Product(created_by=request.user, store=_get_default_store_for_request(request))
                was_combo = False
                existing_combo_signature = set()

            product.code = (request.POST.get('code', '') or '').strip() or (product.code or _generate_next_product_code())
            product.name = (request.POST.get('name', '') or '').strip()
            if not product.name:
                return JsonResponse({'status': 'error', 'message': 'Vui lòng nhập tên SP'})

            product.barcode = request.POST.get('barcode', '')
            product.unit = request.POST.get('unit', 'Cái')
            product.import_price = _to_money_decimal(request.POST.get('import_price', 0))
            product.selling_price = _to_money_decimal(request.POST.get('selling_price', 0))
            product.wholesale_price_no_warranty = _to_money_decimal(request.POST.get('wholesale_price_no_warranty', 0))
            product.wholesale_price_warranty = _to_money_decimal(request.POST.get('wholesale_price_warranty', 0))
            product.min_stock = request.POST.get('min_stock', 0) or 0
            product.max_stock = request.POST.get('max_stock', 0) or 0
            product.description = request.POST.get('description', '')
            product.is_weight_based = request.POST.get('is_weight_based', '0') == '1'
            product.is_service = request.POST.get('is_service', '0') == '1'
            product.is_combo = request.POST.get('is_combo', '0') == '1'
            product.specification = request.POST.get('specification', '') or None
            if product.is_combo:
                product.is_service = False

            cat_id = request.POST.get('category_id')
            product.category_id = cat_id if cat_id else None
            sup_id = request.POST.get('supplier_id')
            product.supplier_id = sup_id if sup_id else None
            loc_id = request.POST.get('location_id')
            product.location_id = loc_id if loc_id else None

            skip_variants = request.POST.get('skip_variants', '0') == '1'
            variants_data = json.loads(request.POST.get('variants', '[]') or '[]')
            combo_data = json.loads(request.POST.get('combo_items', '[]') or '[]')
            normalized_combo_items = _validate_combo_items(product, combo_data, request) if product.is_combo else []
            if was_combo and existing_combo_signature and not product.is_combo:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Combo đã lưu không nên chuyển về sản phẩm thường. Hãy ẩn combo cũ và tạo sản phẩm mới nếu cần.',
                })
            if product.is_combo and was_combo and existing_combo_signature:
                new_signature = {
                    (component.id, quantity.normalize())
                    for component, quantity in normalized_combo_items
                }
                if new_signature != existing_combo_signature:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Combo đã lưu không nên đổi thành phần. Hãy tạo combo mới nếu cần cấu hình khác.',
                    })

            if 'image' in request.FILES:
                product.image = request.FILES['image']

            product.save()

            # Luồng form mới không còn chỉnh biến thể/kích thước.
            # Nếu client không yêu cầu sửa biến thể, giữ nguyên dữ liệu cũ.
            if not skip_variants:
                product.variants.all().delete()
                for v in variants_data:
                    ProductVariant.objects.create(
                        product=product,
                        size_name=v.get('size_name', ''),
                        sku=v.get('sku', ''),
                        barcode=v.get('barcode', ''),
                        import_price=v.get('import_price', 0) or 0,
                        cost_price=v.get('cost_price', 0) or 0,
                        selling_price=v.get('selling_price', 0) or 0,
                        wholesale_price_no_warranty=v.get('wholesale_price_no_warranty', 0) or 0,
                        wholesale_price_warranty=v.get('wholesale_price_warranty', 0) or 0,
                    )

            if product.is_combo:
                if not existing_combo_signature:
                    product.combo_items.all().delete()
                    for component, quantity in normalized_combo_items:
                        ComboItem.objects.create(
                            combo=product,
                            product=component,
                            quantity=quantity,
                        )

                product.cost_price = _calculate_combo_cost(normalized_combo_items)
                product.save(update_fields=['cost_price'])
            else:
                product.combo_items.all().delete()

        return JsonResponse({
            'status': 'ok',
            'message': 'Lưu thành công',
            'product': {
                'id': product.id,
                'code': product.code,
                'name': product.name,
            }
        })
    except Product.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy sản phẩm'})
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_product(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        product = _product_queryset_for_request(request).get(id=data.get('id'))

        # Kiểm tra SP đã được dùng trong đơn hàng chưa
        from orders.models import OrderItem
        order_count = OrderItem.objects.filter(product=product).count()
        if order_count > 0:
            return JsonResponse({
                'status': 'error',
                'message': f'🔒 Không thể xóa "{product.name}" vì đã có {order_count} '
                f'đơn hàng sử dụng sản phẩm này. '
                f'Bạn có thể ẩn sản phẩm bằng cách bỏ tích "Đang hoạt động".'
            })

        product.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Product.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy sản phẩm'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: WAREHOUSE ============

@login_required(login_url="/login/")
def api_get_warehouses(request):
    warehouses = list(
        filter_by_store(Warehouse.objects.all(), request).values(
            'id',
            'code',
            'name',
            'address',
            'is_active',
            manager_username=F('manager__username'),
        )
    )
    data = [{
        'id': row['id'],
        'code': row['code'],
        'name': row['name'],
        'address': row['address'] or '',
        'manager': row['manager_username'] or '',
        'is_active': row['is_active'],
    } for row in warehouses]
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_warehouse(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        wh_id = data.get('id')
        if wh_id:
            wh = _get_warehouse_for_user(request, wh_id)
            if not wh:
                return JsonResponse({'status': 'error', 'message': 'Không tìm thấy kho'})
        else:
            wh = Warehouse()
            wh.store = _get_default_store_for_request(request)
            if not wh.store:
                return JsonResponse({'status': 'error', 'message': 'Tài khoản chưa có phạm vi cửa hàng hợp lệ'})
        wh.code = data.get('code', '')
        wh.name = data.get('name', '')
        wh.address = data.get('address', '')
        wh.is_active = data.get('is_active', True)
        wh.save()
        return JsonResponse({'status': 'ok', 'message': 'Lưu thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_warehouse(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        warehouse = _get_warehouse_for_user(request, data.get('id'))
        if not warehouse:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy kho'})
        warehouse.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: SUPPLIER ============

@login_required(login_url="/login/")
def api_get_suppliers(request):
    suppliers = list(Supplier.objects.values(
        'id',
        'code',
        'name',
        'phone',
        'email',
        'address',
        'tax_code',
        'contact_person',
        'note',
        'is_active',
    ))
    data = [{
        'id': row['id'],
        'code': row['code'],
        'name': row['name'],
        'phone': row['phone'] or '',
        'email': row['email'] or '',
        'address': row['address'] or '',
        'tax_code': row['tax_code'] or '',
        'contact_person': row['contact_person'] or '',
        'note': row['note'] or '',
        'is_active': row['is_active'],
    } for row in suppliers]
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_supplier(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    if not can_manage_users(request.user):
        return _forbid_json('Bạn không có quyền cấu hình nhà cung cấp')
    try:
        data = json.loads(request.body)
        sup_id = data.get('id')
        if sup_id:
            sup = Supplier.objects.get(id=sup_id)
        else:
            sup = Supplier()
            sup.created_by = request.user
        code = (data.get('code') or '').strip() or (sup.code or _generate_next_supplier_code())
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'status': 'error', 'message': 'Vui lòng nhập tên NCC'})
        dup = Supplier.all_objects.filter(code__iexact=code)
        if sup_id:
            dup = dup.exclude(id=sup_id)
        if dup.exists():
            return JsonResponse({'status': 'error', 'message': f'Mã NCC "{code}" đã tồn tại'})

        sup.code = code
        sup.name = name
        sup.phone = data.get('phone', '')
        sup.email = data.get('email', '')
        sup.address = data.get('address', '')
        sup.tax_code = data.get('tax_code', '')
        sup.contact_person = data.get('contact_person', '')
        sup.note = data.get('note', '')
        sup.is_active = data.get('is_active', True)
        sup.save()
        return JsonResponse({
            'status': 'ok',
            'message': 'Lưu thành công',
            'supplier': {
                'id': sup.id,
                'code': sup.code,
                'name': sup.name,
                'phone': sup.phone or '',
                'email': sup.email or '',
                'address': sup.address or '',
                'tax_code': sup.tax_code or '',
                'contact_person': sup.contact_person or '',
                'note': sup.note or '',
                'is_active': sup.is_active,
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_supplier(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    if not can_manage_users(request.user):
        return _forbid_json('Bạn không có quyền cấu hình nhà cung cấp')
    try:
        data = json.loads(request.body)
        Supplier.objects.filter(id=data.get('id')).delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: PRODUCT CATEGORY ============

@login_required(login_url="/login/")
def api_get_categories(request):
    cats = list(ProductCategory.objects.values(
        'id',
        'name',
        'description',
        'parent_id',
        'is_active',
        parent_name=F('parent__name'),
    ))
    data = [{
        'id': row['id'],
        'name': row['name'],
        'description': row['description'] or '',
        'parent_id': row['parent_id'],
        'parent': row['parent_name'] or '',
        'is_active': row['is_active'],
    } for row in cats]
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_category(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    if not can_manage_users(request.user):
        return _forbid_json('Bạn không có quyền cấu hình danh mục sản phẩm')
    try:
        data = json.loads(request.body)
        cat_id = data.get('id')
        if cat_id:
            cat = ProductCategory.objects.get(id=cat_id)
        else:
            cat = ProductCategory()
        parent_id = data.get('parent_id') or None
        if cat_id and parent_id and str(cat_id) == str(parent_id):
            return JsonResponse({'status': 'error', 'message': 'Loại sản phẩm không được chọn chính nó làm danh mục cha'})
        cat.name = (data.get('name') or '').strip()
        cat.description = data.get('description', '')
        cat.parent_id = parent_id
        cat.is_active = bool(data.get('is_active', True))
        if not cat.name:
            return JsonResponse({'status': 'error', 'message': 'Vui lòng nhập tên danh mục/loại sản phẩm'})
        cat.save()
        return JsonResponse({
            'status': 'ok',
            'message': 'Lưu thành công',
            'category': {
                'id': cat.id,
                'name': cat.name,
                'description': cat.description or '',
                'parent_id': cat.parent_id,
                'parent': cat.parent.name if cat.parent else '',
                'is_active': cat.is_active,
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: GOODS RECEIPT ============

def _serialize_goods_receipt_list(receipts):
    data = []
    for r in receipts:
        receipt_items = getattr(r, 'prefetched_items', None)
        if receipt_items is None:
            receipt_items = r.items.select_related('product', 'variant').all()
        receipt_items = list(receipt_items)
        total_quantity, total_amount = _calculate_goods_receipt_totals(receipt_items)
        items = []
        for item in receipt_items:
            quantity = _to_decimal(item.quantity)
            unit_price = _to_decimal(item.unit_price)
            line_total = quantity * unit_price
            items.append({
                'product_id': item.product_id,
                'variant_id': item.variant_id,
                'product_code': item.product.code if item.product else '',
                'product_name': item.product.name if item.product else '',
                'variant_name': item.variant.size_name if item.variant else '',
                'quantity': float(quantity),
                'unit_price': float(unit_price),
                'total_price': float(line_total),
            })
        data.append({
            'id': r.id, 'code': r.code,
            'purchase_order': r.purchase_order.code if r.purchase_order else '',
            'supplier': r.supplier.name if r.supplier else '',
            'supplier_id': r.supplier_id,
            'warehouse': r.warehouse.name if r.warehouse else '',
            'warehouse_id': r.warehouse_id,
            'receipt_date': r.receipt_date.strftime('%Y-%m-%d') if r.receipt_date else '',
            'created_at': r.created_at.strftime('%d/%m/%Y %H:%M:%S') if r.created_at else '',
            'items_count': len(items),
            'total_quantity': float(total_quantity),
            'total_amount': float(total_amount),
            'status': r.status,
            'status_display': r.get_status_display(),
            'note': r.note or '',
            'items': items,
        })
    return data


@login_required(login_url="/login/")
def api_get_goods_receipts(request):
    """Trả về danh sách phiếu nhập trong phạm vi store mà user được phép xem."""
    page = _to_positive_int(request.GET.get('page'), default=1, minimum=1)
    page_size = _to_positive_int(request.GET.get('page_size'), default=50, minimum=10, maximum=200)

    receipts = (
        GoodsReceipt.objects
        .select_related('supplier', 'warehouse', 'purchase_order')
        .order_by('-receipt_date', '-created_at', '-id')
    )
    receipts = filter_by_store(receipts, request, field_name='warehouse__store')
    paginator = Paginator(receipts, page_size)
    page_obj = paginator.get_page(page)
    page_receipts = (
        page_obj.object_list
        .prefetch_related(Prefetch(
            'items',
            queryset=GoodsReceiptItem.objects.select_related('product', 'variant'),
            to_attr='prefetched_items',
        ))
    )
    data = _serialize_goods_receipt_list(page_receipts)

    return JsonResponse({
        'data': data,
        'meta': {
            'page': page_obj.number,
            'page_size': page_size,
            'page_count': len(data),
            'total_pages': paginator.num_pages,
            'total_filtered_count': paginator.count,
            'total_all_count': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'start_index': page_obj.start_index() if paginator.count else 0,
            'end_index': page_obj.end_index() if paginator.count else 0,
        }
    })


@login_required(login_url="/login/")
def api_save_goods_receipt(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            # 1. Nạp phiếu cũ nếu là cập nhật để biết cần hoàn tồn ở kho nào.
            gr_id = data.get('id')
            old_status = None
            old_warehouse_id = None
            if gr_id:
                gr = _get_goods_receipt_for_user(
                    request,
                    gr_id,
                    queryset=GoodsReceipt.objects.select_for_update().prefetch_related('items'),
                )
                if not gr:
                    return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiếu nhập'})
                old_status = gr.status
                old_warehouse_id = gr.warehouse_id
            else:
                gr = GoodsReceipt()
                gr.created_by = request.user

            # Mã phiếu được giữ nguyên khi sửa; chỉ tự tăng khi tạo mới hoặc chưa có mã.
            code = (data.get('code', '') or '').strip()
            if not code:
                code = gr.code or _generate_next_goods_receipt_code()
            gr.code = code

            gr.supplier_id = data.get('supplier_id') or None
            requested_warehouse_id = data.get('warehouse_id') or None
            warehouse = _get_warehouse_for_user(request, requested_warehouse_id) if requested_warehouse_id else None
            if requested_warehouse_id and not warehouse:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Kho nhập không tồn tại hoặc không thuộc phạm vi cửa hàng của bạn.'
                })
            gr.warehouse = warehouse

            # Nếu UI không gửi ngày, giữ nguyên ngày cũ khi sửa; tạo mới thì lấy ngày local hiện tại.
            receipt_date = (data.get('receipt_date', '') or '').strip()
            if not receipt_date:
                if gr.receipt_date:
                    receipt_date = gr.receipt_date.strftime('%Y-%m-%d')
                else:
                    receipt_date = timezone.now().date().strftime('%Y-%m-%d')
            gr.receipt_date = receipt_date

            new_status = int(data.get('status', 1))
            gr.status = new_status
            gr.note = data.get('note', '')

            # 2. Chuẩn hóa item để toàn bộ các bước sau dùng cùng một kiểu dữ liệu.
            normalized_items, total_amount = _normalize_goods_receipt_items(data.get('items', []))
            for item in normalized_items:
                product = _get_product_for_user(request, item['product_id'])
                if not product:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Có sản phẩm không tồn tại hoặc không thuộc cửa hàng hiện tại.'
                    })
                _ensure_product_matches_store(product, warehouse.store_id if warehouse else None, 'kho nhập')
                if item['variant_id'] and not _get_variant_for_product(product, item['variant_id']):
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Biến thể không thuộc sản phẩm "{product.name}".'
                    })
            gr.total_amount = total_amount
            gr.save()

            # 3. Nếu phiếu cũ đã hoàn thành thì hoàn tác tồn theo kho cũ trước khi ghi item mới.
            if old_status == 1 and old_warehouse_id:
                _apply_goods_receipt_stock_adjustment(gr, old_warehouse_id, multiplier=-1)

            # 4. Ghi lại toàn bộ item mới sau khi đã hoàn tác item cũ.
            _recreate_goods_receipt_items(gr, normalized_items)

            # 5. Với phiếu hoàn thành, cộng tồn và đồng bộ giá vốn tham chiếu của sản phẩm.
            if new_status == 1 and gr.warehouse_id:
                _apply_goods_receipt_stock_adjustment(gr, gr.warehouse_id, multiplier=1)
                for item in normalized_items:
                    _sync_product_cost_after_goods_receipt(
                        item['product_id'],
                        item['quantity'],
                        item['unit_price'],
                    )

        return JsonResponse({'status': 'ok', 'message': 'Lưu thành công', 'code': gr.code})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_goods_receipt(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            receipt = _get_goods_receipt_for_user(
                request,
                data.get('id'),
                queryset=GoodsReceipt.objects.select_for_update().prefetch_related('items'),
            )
            if not receipt:
                return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiếu nhập'})

            # Phiếu nhập đã hoàn thành thì xóa phải hoàn tác phần tồn đã cộng trước đó.
            if receipt.status == 1 and receipt.warehouse_id:
                _apply_goods_receipt_stock_adjustment(receipt, receipt.warehouse_id, multiplier=-1)

            receipt.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: STOCK TRANSFER ============

@login_required(login_url="/login/")
def api_get_stock_transfers(request):
    """Trả về danh sách phiếu chuyển kho trong phạm vi store mà user được phép xem."""
    transfers = StockTransfer.objects.select_related('from_warehouse', 'to_warehouse').prefetch_related('items__product').all()
    transfers = filter_by_store(transfers, request, field_name='from_warehouse__store')
    data = []
    for t in transfers:
        items = [{
            'product_id': item.product_id,
            'variant_id': item.variant_id,
            'product_code': item.product.code if item.product else '',
            'product_name': item.product.name if item.product else '',
            'variant_name': item.variant.size_name if item.variant else '',
            'quantity': float(item.quantity),
        } for item in t.items.select_related('product', 'variant').all()]
        data.append({
            'id': t.id, 'code': t.code,
            'from_warehouse': t.from_warehouse.name if t.from_warehouse else '',
            'from_warehouse_id': t.from_warehouse_id,
            'to_warehouse': t.to_warehouse.name if t.to_warehouse else '',
            'to_warehouse_id': t.to_warehouse_id,
            'transfer_date': t.transfer_date.strftime('%Y-%m-%d') if t.transfer_date else '',
            'created_at': t.created_at.strftime('%d/%m/%Y %H:%M:%S') if t.created_at else '',
            'status': t.status, 'status_display': t.get_status_display(),
            'note': t.note or '',
            'items': items,
        })
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_stock_transfer(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            # 1. Nạp phiếu cũ nếu là cập nhật để biết cần hoàn tác tồn theo cặp kho cũ.
            st_id = data.get('id')
            old_status = None
            old_from_warehouse_id = None
            old_to_warehouse_id = None
            if st_id:
                st = _get_stock_transfer_for_user(
                    request,
                    st_id,
                    queryset=StockTransfer.objects.select_for_update().prefetch_related('items'),
                )
                if not st:
                    return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiếu chuyển'})
                old_status = st.status
                old_from_warehouse_id = st.from_warehouse_id
                old_to_warehouse_id = st.to_warehouse_id
            else:
                st = StockTransfer()
                st.created_by = request.user

            # 2. Gán dữ liệu cơ bản từ payload.
            st.code = data.get('code', '')
            from_warehouse_id = data.get('from_warehouse_id') or None
            to_warehouse_id = data.get('to_warehouse_id') or None
            from_warehouse = _get_warehouse_for_user(request, from_warehouse_id) if from_warehouse_id else None
            to_warehouse = _get_warehouse_for_user(request, to_warehouse_id) if to_warehouse_id else None
            if from_warehouse_id and not from_warehouse:
                return JsonResponse({'status': 'error', 'message': 'Kho xuất không thuộc phạm vi cửa hàng của bạn.'})
            if to_warehouse_id and not to_warehouse:
                return JsonResponse({'status': 'error', 'message': 'Kho nhập không thuộc phạm vi cửa hàng của bạn.'})
            _ensure_warehouses_same_store(from_warehouse, to_warehouse)
            st.from_warehouse = from_warehouse
            st.to_warehouse = to_warehouse
            st.transfer_date = data.get('transfer_date')
            new_status = int(data.get('status', 0))
            st.status = new_status
            st.note = data.get('note', '')

            # 3. Validate toàn bộ item trước mọi ghi/hoàn tồn để lỗi không commit nửa chừng.
            normalized_items = _normalize_stock_transfer_items(data.get('items', []))
            for item in normalized_items:
                product = _get_product_for_user(request, item['product_id'])
                if not product:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Có sản phẩm không tồn tại hoặc không thuộc cửa hàng hiện tại.'
                    })
                _ensure_product_matches_store(
                    product,
                    from_warehouse.store_id if from_warehouse else (to_warehouse.store_id if to_warehouse else None),
                    'phiếu chuyển kho',
                )
                if item['variant_id'] and not _get_variant_for_product(product, item['variant_id']):
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Biến thể không thuộc sản phẩm "{product.name}".'
                    })

            st.save()

            # 4. Nếu phiếu cũ đã hoàn thành thì hoàn tác tồn theo cặp kho cũ trước khi ghi item mới.
            if old_status == 2 and old_from_warehouse_id and old_to_warehouse_id:
                _apply_stock_transfer_adjustment(
                    st,
                    old_from_warehouse_id,
                    old_to_warehouse_id,
                    reverse=True,
                )

            # 5. Ghi lại item mới theo payload hiện tại.
            _recreate_stock_transfer_items(st, normalized_items)

            # 6. Chỉ khi phiếu ở trạng thái hoàn thành mới áp biến động kho thực tế.
            if new_status == 2 and st.from_warehouse_id and st.to_warehouse_id:
                _apply_stock_transfer_adjustment(
                    st,
                    st.from_warehouse_id,
                    st.to_warehouse_id,
                    reverse=False,
                )

        return JsonResponse({'status': 'ok', 'message': 'Lưu thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_stock_transfer(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            transfer = _get_stock_transfer_for_user(
                request,
                data.get('id'),
                queryset=StockTransfer.objects.select_for_update().prefetch_related('items'),
            )
            if not transfer:
                return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiếu chuyển'})

            # Phiếu chuyển hoàn thành thì xóa phải hoàn tác tồn ở cả kho xuất và kho nhập.
            if transfer.status == 2 and transfer.from_warehouse_id and transfer.to_warehouse_id:
                _apply_stock_transfer_adjustment(
                    transfer,
                    transfer.from_warehouse_id,
                    transfer.to_warehouse_id,
                    reverse=True,
                )

            transfer.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: STOCK CHECK ============

@login_required(login_url="/login/")
def api_get_stock_checks(request):
    checks = (
        StockCheck.objects
        .select_related('warehouse')
        .prefetch_related('items__product')
        .order_by('-check_date', '-created_at', '-id')
    )
    checks = filter_by_store(checks, request, field_name='warehouse__store')
    data = []
    for c in checks:
        items = [{
            'product_id': item.product_id,
            'variant_id': item.variant_id,
            'product_code': item.product.code if item.product else '',
            'product_name': item.product.name if item.product else '',
            'variant_name': item.variant.size_name if item.variant else '',
            'system_quantity': float(item.system_quantity),
            'actual_quantity': float(item.actual_quantity),
            'difference': float(item.difference),
            'note': item.note or '',
        } for item in c.items.select_related('product', 'variant').all()]
        data.append({
            'id': c.id, 'code': c.code,
            'warehouse': c.warehouse.name if c.warehouse else '',
            'warehouse_id': c.warehouse_id,
            'check_date': c.check_date.strftime('%Y-%m-%d') if c.check_date else '',
            'created_at': c.created_at.strftime('%d/%m/%Y %H:%M:%S') if c.created_at else '',
            'status': c.status, 'status_display': c.get_status_display(),
            'note': c.note or '',
            'items': items,
        })
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_stock_check(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            sc_id = data.get('id')
            if sc_id:
                sc = _get_stock_check_for_user(request, sc_id)
                if not sc:
                    return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiếu kiểm'})
            else:
                sc = StockCheck()
                sc.created_by = request.user

            code = (data.get('code', '') or '').strip()
            if not code:
                code = sc.code or _generate_next_stock_check_code()
            sc.code = code

            warehouse_id = data.get('warehouse_id') or None
            warehouse = _get_warehouse_for_user(request, warehouse_id) if warehouse_id else None
            if warehouse_id and not warehouse:
                return JsonResponse({'status': 'error', 'message': 'Kho kiểm không thuộc phạm vi cửa hàng của bạn.'})
            sc.warehouse = warehouse
            items_data = data.get('items', [])
            normalized_items = []
            for item in items_data:
                product_id = item.get('product_id')
                product = _get_product_for_user(request, product_id)
                if not product:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Có sản phẩm không tồn tại hoặc không thuộc cửa hàng hiện tại.'
                    })
                _ensure_product_matches_store(product, warehouse.store_id if warehouse else None, 'kho kiểm')
                variant_id = item.get('variant_id') or None
                if variant_id and not _get_variant_for_product(product, variant_id):
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Biến thể không thuộc sản phẩm "{product.name}".'
                    })
                actual_qty = _to_decimal(item.get('actual_quantity', 0))
                sys_qty = Decimal('0')
                if warehouse and product_id:
                    try:
                        ps = ProductStock.objects.get(product_id=product_id, warehouse=warehouse)
                        sys_qty = _to_decimal(ps.quantity)
                    except ProductStock.DoesNotExist:
                        sys_qty = Decimal('0')
                normalized_items.append((item, product, variant_id, actual_qty, sys_qty))

            check_date = (data.get('check_date', '') or '').strip()
            if not check_date:
                if sc.check_date:
                    check_date = sc.check_date.strftime('%Y-%m-%d')
                else:
                    check_date = date.today().strftime('%Y-%m-%d')
            sc.check_date = check_date
            sc.status = data.get('status', 0)
            sc.note = data.get('note', '')
            sc.save()

            # Save items
            sc.items.all().delete()
            for item, product, variant_id, actual_qty, sys_qty in normalized_items:
                StockCheckItem.objects.create(
                    stock_check=sc,
                    product=product,
                    variant_id=variant_id,
                    system_quantity=sys_qty,
                    actual_quantity=actual_qty,
                    difference=actual_qty - sys_qty,
                    note=item.get('note', ''),
                )

        return JsonResponse({'status': 'ok', 'message': 'Lưu thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_stock_check(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        stock_check = _get_stock_check_for_user(request, data.get('id'))
        if not stock_check:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiếu kiểm'})
        stock_check.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ============ API: PURCHASE ORDER ============


@login_required(login_url="/login/")
def api_get_purchase_orders(request):
    orders = (
        PurchaseOrder.objects
        .select_related('supplier', 'warehouse')
        .prefetch_related('items__product')
        .order_by('-order_date', '-id')
    )
    orders = filter_by_store(orders, request, field_name='warehouse__store')
    data = []
    for o in orders:
        items = [{
            'product_id': item.product_id,
            'variant_id': item.variant_id,
            'product_code': item.product.code if item.product else '',
            'product_name': item.product.name if item.product else '',
            'variant_name': item.variant.size_name if item.variant else '',
            'quantity': float(item.quantity),
            'received_quantity': float(item.received_quantity),
            'unit_price': float(item.unit_price),
            'total_price': float(item.total_price),
        } for item in o.items.select_related('product', 'variant').all()]
        data.append({
            'id': o.id, 'code': o.code,
            'supplier': o.supplier.name if o.supplier else '',
            'supplier_id': o.supplier_id,
            'warehouse': o.warehouse.name if o.warehouse else '',
            'warehouse_id': o.warehouse_id,
            'order_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else '',
            'expected_date': o.expected_date.strftime('%Y-%m-%d') if o.expected_date else '',
            'created_at': o.created_at.strftime('%d/%m/%Y %H:%M:%S') if o.created_at else '',
            'total_amount': float(o.total_amount),
            'status': o.status, 'status_display': o.get_status_display(),
            'note': o.note or '',
            'items': items,
        })
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_purchase_order(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        po_id = data.get('id')
        if po_id:
            po = _get_purchase_order_for_user(request, po_id)
            if not po:
                return JsonResponse({'status': 'error', 'message': 'Không tìm thấy đơn đặt hàng'})
        else:
            po = PurchaseOrder()
            po.created_by = request.user
        po.code = data.get('code', '')
        po.supplier_id = data.get('supplier_id') or None
        warehouse_id = data.get('warehouse_id') or None
        warehouse = _get_warehouse_for_user(request, warehouse_id) if warehouse_id else None
        if warehouse_id and not warehouse:
            return JsonResponse({'status': 'error', 'message': 'Kho nhập không thuộc phạm vi cửa hàng của bạn.'})
        po.warehouse = warehouse
        po.order_date = data.get('order_date')
        po.expected_date = data.get('expected_date') or None
        po.status = data.get('status', 0)
        po.note = data.get('note', '')

        items_data = data.get('items', [])
        normalized_items = []
        total = 0
        for item in items_data:
            product = _get_product_for_user(request, item.get('product_id'))
            if not product:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Có sản phẩm không tồn tại hoặc không thuộc cửa hàng hiện tại.'
                })
            _ensure_product_matches_store(product, warehouse.store_id if warehouse else None, 'kho nhập')
            variant_id = item.get('variant_id') or None
            if variant_id and not _get_variant_for_product(product, variant_id):
                return JsonResponse({
                    'status': 'error',
                    'message': f'Biến thể không thuộc sản phẩm "{product.name}".'
                })
            qty = float(item.get('quantity', 0))
            price = float(item.get('unit_price', 0))
            total += qty * price
            normalized_items.append((product, variant_id, qty, price))

        po.total_amount = total
        po.save()

        # Delete old items and create new ones
        po.items.all().delete()
        for product, variant_id, qty, price in normalized_items:
            PurchaseOrderItem.objects.create(
                purchase_order=po,
                product=product,
                variant_id=variant_id,
                quantity=qty,
                unit_price=price,
                total_price=qty * price,
            )

        return JsonResponse({'status': 'ok', 'message': 'Lưu thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: DELETE PURCHASE ORDER ============

@login_required(login_url="/login/")
def api_delete_purchase_order(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    try:
        data = json.loads(request.body)
        purchase_order = _get_purchase_order_for_user(request, data.get('id'))
        if not purchase_order:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy đơn đặt hàng'})
        purchase_order.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ API: COST ADJUSTMENT ============

@login_required(login_url="/login/")
def api_get_cost_adjustments(request):
    items = list(
        filter_by_store(
            CostAdjustment.objects.all(),
            request,
            field_name='product__store',
        ).values(
            'id',
            'old_cost',
            'new_cost',
            'reason',
            'adjusted_at',
            product_name=F('product__name'),
            product_code=F('product__code'),
            adjusted_by_username=F('adjusted_by__username'),
        )
    )
    data = [{
        'id': row['id'],
        'product': row['product_name'] or '',
        'product_code': row['product_code'] or '',
        'old_cost': float(row['old_cost']),
        'new_cost': float(row['new_cost']),
        'reason': row['reason'] or '',
        'adjusted_by': row['adjusted_by_username'] or '',
        'adjusted_at': row['adjusted_at'].strftime('%d/%m/%Y %H:%M') if row['adjusted_at'] else '',
    } for row in items]
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_product_purchase_history(request):
    """API: Lịch sử nhập hàng của 1 sản phẩm"""
    product_id = request.GET.get('product_id')
    if not product_id:
        return JsonResponse({'status': 'error', 'message': 'Missing product_id'})

    product = _product_queryset_for_request(request).filter(id=product_id).first()
    if not product:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy sản phẩm'})

    store_ids = get_managed_store_ids(request.user)
    items = GoodsReceiptItem.objects.filter(
        product=product,
        goods_receipt__status=1,
        goods_receipt__warehouse__store_id__in=store_ids,
    ).select_related(
        'goods_receipt',
        'goods_receipt__supplier',
        'goods_receipt__warehouse',
        'goods_receipt__purchase_order',
        'goods_receipt__created_by',
        'variant',
    )

    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    supplier_id = request.GET.get('supplier_id') or ''
    warehouse_id = request.GET.get('warehouse_id') or ''
    receipt_code = (request.GET.get('receipt_code') or '').strip()
    if date_from:
        items = items.filter(goods_receipt__receipt_date__gte=date_from)
    if date_to:
        items = items.filter(goods_receipt__receipt_date__lte=date_to)
    if supplier_id:
        items = items.filter(goods_receipt__supplier_id=supplier_id)
    if warehouse_id:
        items = items.filter(goods_receipt__warehouse_id=warehouse_id)
    if receipt_code:
        items = items.filter(goods_receipt__code__icontains=receipt_code)

    ordered_items = list(items.order_by('-goods_receipt__receipt_date', '-goods_receipt__id', '-id'))
    receipts = _summarize_purchase_receipts(ordered_items)

    data = [{
        'receipt_id': it.goods_receipt_id,
        'receipt_code': it.goods_receipt.code,
        'receipt_date': it.goods_receipt.receipt_date.strftime('%d/%m/%Y') if it.goods_receipt.receipt_date else '',
        'receipt_date_raw': it.goods_receipt.receipt_date.strftime('%Y-%m-%d') if it.goods_receipt.receipt_date else '',
        'supplier': it.goods_receipt.supplier.name if it.goods_receipt.supplier else '',
        'warehouse': it.goods_receipt.warehouse.name if it.goods_receipt.warehouse else '',
        'purchase_order': it.goods_receipt.purchase_order.code if it.goods_receipt.purchase_order else '',
        'created_by': (
            it.goods_receipt.created_by.get_full_name() or it.goods_receipt.created_by.username
        ) if it.goods_receipt.created_by else '',
        'receipt_note': it.goods_receipt.note or '',
        'receipt_total_amount': float(it.goods_receipt.total_amount or 0),
        'variant': it.variant.size_name if it.variant else '',
        'quantity': float(it.quantity),
        'unit_price': float(it.unit_price),
        'total_price': float(it.total_price),
    } for it in ordered_items]

    price_timeline = [{
        'date': it.goods_receipt.receipt_date.strftime('%d/%m/%Y') if it.goods_receipt.receipt_date else '',
        'price': float(it.unit_price),
        'quantity': float(it.quantity),
    } for it in sorted(ordered_items, key=lambda it: (
        it.goods_receipt.receipt_date or date.min,
        it.goods_receipt_id or 0,
        it.id or 0,
    ))]

    # Tổng
    total_qty = sum(d['quantity'] for d in data)
    total_amount = sum(d['total_price'] for d in data)
    avg_price = total_amount / total_qty if total_qty > 0 else 0

    return JsonResponse({
        'status': 'ok',
        'data': data,
        'summary': {
            'total_entries': len(data),
            'total_receipts': len(receipts),
            'total_quantity': total_qty,
            'total_amount': total_amount,
            'avg_unit_price': round(avg_price),
        },
        'receipts': receipts,
        'price_timeline': price_timeline,
    })


@login_required(login_url="/login/")
def api_product_sales_history(request):
    """API: Lịch sử bán hàng (đơn hàng) của 1 sản phẩm"""
    product_id = request.GET.get('product_id')
    if not product_id:
        return JsonResponse({'status': 'error', 'message': 'Missing product_id'})

    product = _product_queryset_for_request(request).filter(id=product_id).first()
    if not product:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy sản phẩm'})

    from orders.models import OrderItem
    store_ids = get_managed_store_ids(request.user)
    items = OrderItem.objects.filter(
        product=product,
        order__store_id__in=store_ids,
        order__status__in=[3, 4, 5],  # Đã duyệt / Đã giao / Hoàn thành
    ).select_related('order', 'order__customer', 'variant')

    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    if date_from:
        items = items.filter(order__order_date__gte=date_from)
    if date_to:
        items = items.filter(order__order_date__lte=date_to)

    items = items.order_by('-order__order_date', '-order__id')

    data = [{
        'order_code': it.order.code,
        'order_date': it.order.order_date.strftime('%d/%m/%Y') if it.order.order_date else '',
        'order_date_raw': it.order.order_date.strftime('%Y-%m-%d') if it.order.order_date else '',
        'customer': it.order.customer.name if it.order.customer else '',
        'variant': it.variant.size_name if it.variant else '',
        'quantity': float(it.quantity),
        'unit_price': float(it.unit_price),
        'discount_percent': float(it.discount_percent),
        'total_price': float(it.total_price),
    } for it in items]

    # Tổng
    total_qty = sum(d['quantity'] for d in data)
    total_amount = sum(d['total_price'] for d in data)
    avg_price = total_amount / total_qty if total_qty > 0 else 0

    return JsonResponse({
        'status': 'ok',
        'data': data,
        'summary': {
            'total_orders': len(set(d['order_code'] for d in data)),
            'total_entries': len(data),
            'total_quantity': total_qty,
            'total_amount': total_amount,
            'avg_unit_price': round(avg_price),
        },
    })

# ============ API: PRODUCT LOCATION ============


@login_required(login_url="/login/")
def api_get_locations(request):
    data = list(ProductLocation.objects.values('id', 'name', 'is_active'))
    return JsonResponse({'data': data})


@login_required(login_url="/login/")
def api_save_location(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    if not can_manage_users(request.user):
        return _forbid_json('Bạn không có quyền cấu hình vị trí sản phẩm')
    try:
        data = json.loads(request.body)
        loc_id = data.get('id')
        name = (data.get('name', '') or '').strip()
        if not name:
            return JsonResponse({'status': 'error', 'message': 'Vui lòng nhập tên vị trí'})
        if loc_id:
            loc = ProductLocation.objects.get(id=loc_id)
        else:
            loc = ProductLocation()
        # Check duplicate name
        dup = ProductLocation.objects.filter(name=name)
        if loc_id:
            dup = dup.exclude(id=loc_id)
        if dup.exists():
            return JsonResponse({'status': 'error', 'message': f'Vị trí "{name}" đã tồn tại'})
        loc.name = name
        loc.is_active = data.get('is_active', True)
        loc.save()
        return JsonResponse({'status': 'ok', 'message': 'Lưu thành công', 'id': loc.id, 'name': loc.name})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="/login/")
def api_delete_location(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})
    if not can_manage_users(request.user):
        return _forbid_json('Bạn không có quyền cấu hình vị trí sản phẩm')
    try:
        data = json.loads(request.body)
        loc = ProductLocation.objects.get(id=data.get('id'))
        # Check if any product is using this location
        product_count = Product.objects.filter(location=loc).count()
        if product_count > 0:
            return JsonResponse({
                'status': 'error',
                'message': f'Không thể xóa vị trí "{loc.name}" vì đang được {product_count} sản phẩm sử dụng.'
            })
        loc.delete()
        return JsonResponse({'status': 'ok', 'message': 'Xóa thành công'})
    except ProductLocation.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy vị trí'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ EXCEL IMPORT ============

PRODUCT_IMPORT_HEADER_ALIASES = {
    'ma sp': 'code',
    'ma san pham': 'code',
    'code': 'code',
    'sku': 'code',
    'ten sp': 'name',
    'ten san pham': 'name',
    'name': 'name',
    'barcode': 'barcode',
    'ma vach': 'barcode',
    'danh muc': 'category',
    'category': 'category',
    'dvt': 'unit',
    'don vi': 'unit',
    'don vi tinh': 'unit',
    'unit': 'unit',
    'quy cach': 'specification',
    'quy cach san pham': 'specification',
    'spec': 'specification',
    'loai san pham': 'product_type',
    'product type': 'product_type',
    'tinh chat': 'product_nature',
    'loai': 'product_nature',
    'product nature': 'product_nature',
    'gia nhap': 'import_price',
    'gia von': 'cost_price',
    'gia ban le': 'selling_price',
    'gia ban': 'selling_price',
    'gia si kbh': 'wholesale_price_no_warranty',
    'gia si khong bao hanh': 'wholesale_price_no_warranty',
    'gia si bh': 'wholesale_price_warranty',
    'gia si co bao hanh': 'wholesale_price_warranty',
    'ton kho': 'stock',
    'ton': 'stock',
    'ton toi thieu': 'min_stock',
    'min stock': 'min_stock',
    'ton toi da': 'max_stock',
    'max stock': 'max_stock',
    'trang thai': 'is_active',
    'status': 'is_active',
    'ncc': 'supplier',
    'nha cung cap': 'supplier',
    'supplier': 'supplier',
    'vi tri': 'location',
    'location': 'location',
    'mo ta': 'description',
    'ghi chu': 'description',
    'description': 'description',
}


def _normalize_excel_text(value):
    text = str(value if value is not None else '').strip().lower()
    text = text.replace('đ', 'd')
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r'[^a-z0-9]+', ' ', text).strip()


def _excel_cell_text(value):
    if value is None:
        return ''
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(value.quantize(Decimal('1')))
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_import_decimal(value, default=Decimal('0'), integer=False):
    if value in (None, ''):
        return default
    if isinstance(value, Decimal):
        parsed = value
    elif isinstance(value, (int, float)):
        parsed = Decimal(str(value))
    else:
        raw = _excel_cell_text(value)
        raw = raw.replace('\xa0', '').replace(' ', '')
        raw = raw.replace('₫', '').replace('đ', '').replace('Đ', '')
        raw = re.sub(r'[^0-9,\.\-]', '', raw)
        if raw in ('', '-'):
            return default
        if raw.count(',') > 1 or raw.count('.') > 1:
            raw = raw.replace(',', '').replace('.', '')
        elif ',' in raw and '.' in raw:
            raw = raw.replace(',', '').replace('.', '')
        elif ',' in raw:
            left, right = raw.rsplit(',', 1)
            raw = left + right if len(right) == 3 else left + '.' + right
        elif '.' in raw:
            left, right = raw.rsplit('.', 1)
            raw = left + right if len(right) == 3 else raw
        try:
            parsed = Decimal(raw)
        except (InvalidOperation, TypeError, ValueError):
            return default
    if integer:
        return parsed.quantize(Decimal('1'), rounding=ROUND_FLOOR)
    return parsed


def _parse_import_int(value, default=0):
    parsed = _parse_import_decimal(value, default=Decimal(str(default)), integer=True)
    try:
        return max(0, int(parsed))
    except (TypeError, ValueError):
        return default


def _parse_import_active(value, current=True):
    raw = _normalize_excel_text(value)
    if not raw:
        return current
    if raw in {'0', 'false', 'no', 'khong', 'tat', 'inactive'}:
        return False
    if any(token in raw for token in ['ngung', 'nghi ban', 'khong hoat dong']):
        return False
    if raw in {'1', 'true', 'yes', 'co', 'active'}:
        return True
    if any(token in raw for token in ['dang hoat dong', 'dang ban']):
        return True
    return current


def _parse_import_product_nature(value):
    raw = _normalize_excel_text(value)
    if not raw:
        return None
    if 'combo' in raw:
        return 'combo'
    if 'dich vu' in raw or 'service' in raw:
        return 'service'
    if 'can dong' in raw or 'khoi luong' in raw or 'weight' in raw:
        return 'weight'
    if 'san pham' in raw or 'vat ly' in raw or 'normal' in raw:
        return 'normal'
    return None


def _find_product_import_header(sheet, max_rows=20):
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True), 1):
        headers = {}
        for col_index, value in enumerate(row, 1):
            key = PRODUCT_IMPORT_HEADER_ALIASES.get(_normalize_excel_text(value))
            if key and key not in headers:
                headers[key] = col_index
        if 'name' in headers and (len(headers) >= 3 or 'code' in headers):
            return row_index, headers
    raise ValueError('Không tìm thấy dòng tiêu đề. File cần có cột "Tên sản phẩm" và nên dùng mẫu xuất Excel từ danh sách sản phẩm.')


def _row_import_value(row, headers, key):
    col_index = headers.get(key)
    if not col_index or col_index > len(row):
        return None
    return row[col_index - 1]


def _row_has_product_import_data(row, headers):
    return any(_excel_cell_text(_row_import_value(row, headers, key)) for key in headers)


def _get_or_create_import_category(name, parent=None):
    name = _excel_cell_text(name)
    if not name:
        return None, 0
    queryset = ProductCategory.objects.filter(name__iexact=name)
    queryset = queryset.filter(parent=parent) if parent else queryset.filter(parent__isnull=True)
    category = queryset.order_by('id').first()
    if category:
        if not category.is_active:
            category.is_active = True
            category.save(update_fields=['is_active'])
        return category, 0
    return ProductCategory.objects.create(name=name, parent=parent, is_active=True), 1


def _resolve_product_import_category(row, headers):
    category_name = _row_import_value(row, headers, 'category')
    product_type_name = _row_import_value(row, headers, 'product_type')
    category, created_count = _get_or_create_import_category(category_name)
    product_type_name = _excel_cell_text(product_type_name)
    if product_type_name:
        product_type, type_created = _get_or_create_import_category(product_type_name, parent=category)
        return product_type, created_count + type_created
    return category, created_count


def _get_or_create_import_supplier(name, user):
    name = _excel_cell_text(name)
    if not name:
        return None, 0
    supplier = Supplier.objects.filter(name__iexact=name).order_by('id').first()
    if supplier:
        if not supplier.is_active:
            supplier.is_active = True
            supplier.save(update_fields=['is_active'])
        return supplier, 0
    supplier = Supplier.objects.create(
        code=_generate_next_supplier_code(),
        name=name,
        created_by=user,
        is_active=True,
    )
    return supplier, 1


def _get_or_create_import_location(name):
    name = _excel_cell_text(name)
    if not name:
        return None, 0
    location = ProductLocation.objects.filter(name__iexact=name).order_by('id').first()
    if location:
        if not location.is_active:
            location.is_active = True
            location.save(update_fields=['is_active'])
        return location, 0
    return ProductLocation.objects.create(name=name, is_active=True), 1


def _get_default_import_warehouse(store):
    if not store:
        return None
    return Warehouse.objects.filter(store=store, is_active=True).order_by('id').first()


def _sync_product_import_stock(product, default_warehouse, stock_quantity, created):
    if not default_warehouse:
        raise ValueError('Không có kho mặc định để nhập tồn từ file Excel')

    stock_quantity = _parse_import_decimal(stock_quantity, default=Decimal('0'))
    existing_stocks = list(
        ProductStock.objects
        .select_for_update()
        .filter(product=product)
        .select_related('warehouse')
    )

    if not created:
        non_default_stocks = [
            stock for stock in existing_stocks
            if stock.warehouse_id != default_warehouse.id and _to_decimal(stock.quantity) != 0
        ]
        if non_default_stocks:
            warehouse_names = ', '.join(
                stock.warehouse.name if stock.warehouse else f'ID {stock.warehouse_id}'
                for stock in non_default_stocks[:3]
            )
            if len(non_default_stocks) > 3:
                warehouse_names += ', ...'
            raise ValueError(
                f'Sản phẩm "{product.code}" đang có tồn ở kho khác ({warehouse_names}), '
                'không thể cập nhật tồn từ file Excel. Hãy dùng kiểm kho hoặc chứng từ kho.'
            )

    default_stock = next(
        (stock for stock in existing_stocks if stock.warehouse_id == default_warehouse.id),
        None,
    )
    if default_stock:
        default_stock.quantity = stock_quantity
        default_stock.save(update_fields=['quantity'])
        return 1

    if stock_quantity == 0 and created:
        return 0

    ProductStock.objects.create(
        product=product,
        warehouse=default_warehouse,
        quantity=stock_quantity,
    )
    return 1


def _sync_variant_prices_from_product_import(product, headers):
    if not product.id:
        return 0

    variant_updates = {}
    if 'import_price' in headers:
        variant_updates['import_price'] = product.import_price
    if 'wholesale_price_no_warranty' in headers:
        variant_updates['wholesale_price_no_warranty'] = product.wholesale_price_no_warranty
    if 'wholesale_price_warranty' in headers:
        variant_updates['wholesale_price_warranty'] = product.wholesale_price_warranty

    if not variant_updates:
        return 0

    return product.variants.update(**variant_updates)


def _upsert_product_import_row(row, headers, request, default_store, default_warehouse, import_stock, seen_codes):
    code = _excel_cell_text(_row_import_value(row, headers, 'code'))
    name = _excel_cell_text(_row_import_value(row, headers, 'name'))
    if not name:
        raise ValueError('Thiếu Tên sản phẩm')

    created = False
    if code:
        code_key = code.lower()
        if code_key in seen_codes:
            raise ValueError(f'Mã SP "{code}" bị trùng trong file')
        seen_codes.add(code_key)

        product = _product_queryset_for_request(request).filter(code__iexact=code).first()
        if not product:
            existing = Product.all_objects.filter(code__iexact=code).first()
            if existing:
                raise ValueError(f'Mã SP "{code}" đã tồn tại ở sản phẩm đã xóa hoặc ngoài phạm vi cửa hàng của bạn')
            product = Product(code=code, created_by=request.user, store=default_store)
            created = True
    else:
        product = Product(
            code=_generate_next_product_code(),
            created_by=request.user,
            store=default_store,
        )
        created = True

    product.name = name
    if 'code' in headers and code:
        product.code = code
    if 'barcode' in headers:
        product.barcode = _excel_cell_text(_row_import_value(row, headers, 'barcode')) or None
    if 'unit' in headers:
        product.unit = _excel_cell_text(_row_import_value(row, headers, 'unit')) or 'Cái'
    elif created and not product.unit:
        product.unit = 'Cái'
    if 'specification' in headers:
        product.specification = _excel_cell_text(_row_import_value(row, headers, 'specification')) or None
    if 'description' in headers:
        product.description = _excel_cell_text(_row_import_value(row, headers, 'description'))
    if 'min_stock' in headers:
        product.min_stock = _parse_import_int(_row_import_value(row, headers, 'min_stock'))
    if 'max_stock' in headers:
        product.max_stock = _parse_import_int(_row_import_value(row, headers, 'max_stock'))
    if 'is_active' in headers:
        product.is_active = _parse_import_active(_row_import_value(row, headers, 'is_active'), current=product.is_active)
    elif created:
        product.is_active = True

    created_categories = 0
    if 'category' in headers or 'product_type' in headers:
        product.category, created_categories = _resolve_product_import_category(row, headers)

    created_suppliers = 0
    if 'supplier' in headers:
        product.supplier, created_suppliers = _get_or_create_import_supplier(
            _row_import_value(row, headers, 'supplier'),
            request.user,
        )

    created_locations = 0
    if 'location' in headers:
        product.location, created_locations = _get_or_create_import_location(_row_import_value(row, headers, 'location'))

    nature = _parse_import_product_nature(_row_import_value(row, headers, 'product_nature'))
    if nature == 'combo':
        if created:
            raise ValueError('Import Excel chưa hỗ trợ tạo combo mới. Hãy tạo combo trên màn hình để khai báo thành phần.')
        if not product.is_combo:
            raise ValueError('Không thể đổi sản phẩm thường thành combo bằng Excel. Hãy tạo combo trên màn hình.')
    elif product.is_combo and nature and nature != 'combo':
        raise ValueError('Không thể đổi combo đã lưu thành sản phẩm thường bằng Excel.')
    elif not product.is_combo:
        if nature == 'service':
            product.is_service = True
            product.is_weight_based = False
        elif nature == 'weight':
            product.is_service = False
            product.is_weight_based = True
        elif nature == 'normal':
            product.is_service = False
            product.is_weight_based = False
        elif created:
            product.is_service = False
            product.is_weight_based = False

    if 'import_price' in headers:
        product.import_price = _parse_import_decimal(_row_import_value(row, headers, 'import_price'), integer=True)
    if 'selling_price' in headers:
        product.selling_price = _parse_import_decimal(_row_import_value(row, headers, 'selling_price'), integer=True)
    if 'wholesale_price_no_warranty' in headers:
        product.wholesale_price_no_warranty = _parse_import_decimal(
            _row_import_value(row, headers, 'wholesale_price_no_warranty'),
            integer=True,
        )
    if 'wholesale_price_warranty' in headers:
        product.wholesale_price_warranty = _parse_import_decimal(
            _row_import_value(row, headers, 'wholesale_price_warranty'),
            integer=True,
        )
    if 'cost_price' in headers and not product.is_combo:
        product.cost_price = _parse_import_decimal(_row_import_value(row, headers, 'cost_price'), integer=True)

    product.save()
    variants_synced = _sync_variant_prices_from_product_import(product, headers)

    stock_initialized = 0
    if import_stock and 'stock' in headers and not product.is_combo and not product.is_service:
        stock_initialized = _sync_product_import_stock(
            product=product,
            default_warehouse=default_warehouse,
            stock_quantity=_row_import_value(row, headers, 'stock'),
            created=created,
        )

    return {
        'created': created,
        'created_categories': created_categories,
        'created_suppliers': created_suppliers,
        'created_locations': created_locations,
        'stock_initialized': stock_initialized,
        'variants_synced': variants_synced,
    }


@login_required(login_url="/login/")
def import_products_excel(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'})

    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'status': 'error', 'message': 'Vui lòng chọn file Excel'})
    if not upload.name.lower().endswith(('.xlsx', '.xlsm')):
        return JsonResponse({'status': 'error', 'message': 'Chỉ hỗ trợ file .xlsx hoặc .xlsm'})

    default_store = _get_default_store_for_request(request)
    if not default_store:
        return JsonResponse({'status': 'error', 'message': 'Tài khoản chưa có phạm vi cửa hàng hợp lệ'})

    import_stock = request.POST.get('import_stock') == '1'
    default_warehouse = _get_default_import_warehouse(default_store) if import_stock else None

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        header_row, headers = _find_product_import_header(sheet)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Không đọc được file Excel: {exc}'})

    summary = {
        'total_rows': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
        'created_categories': 0,
        'created_suppliers': 0,
        'created_locations': 0,
        'stock_initialized': 0,
        'variants_synced': 0,
    }
    errors = []
    seen_codes = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if not _row_has_product_import_data(row, headers):
            summary['skipped'] += 1
            continue
        summary['total_rows'] += 1
        try:
            with transaction.atomic():
                result = _upsert_product_import_row(
                    row,
                    headers,
                    request,
                    default_store,
                    default_warehouse,
                    import_stock,
                    seen_codes,
                )
            if result['created']:
                summary['created'] += 1
            else:
                summary['updated'] += 1
            summary['created_categories'] += result['created_categories']
            summary['created_suppliers'] += result['created_suppliers']
            summary['created_locations'] += result['created_locations']
            summary['stock_initialized'] += result['stock_initialized']
            summary['variants_synced'] += result['variants_synced']
        except Exception as exc:
            summary['errors'] += 1
            if len(errors) < 50:
                errors.append({'row': row_number, 'message': str(exc)})

    success_count = summary['created'] + summary['updated']
    if success_count and summary['errors']:
        status = 'partial'
    elif summary['errors']:
        status = 'error'
    else:
        status = 'ok'

    message = (
        f'Đã import {success_count} sản phẩm '
        f'({summary["created"]} tạo mới, {summary["updated"]} cập nhật)'
    )
    if summary['errors']:
        message += f', {summary["errors"]} dòng lỗi'
    if summary['stock_initialized']:
        message += f', đồng bộ tồn kho cho {summary["stock_initialized"]} sản phẩm'
    if summary['variants_synced']:
        message += f', đồng bộ giá cho {summary["variants_synced"]} biến thể'

    return JsonResponse({
        'status': status,
        'message': message,
        'summary': summary,
        'errors': errors,
    })


# ============ EXCEL EXPORT ============

@login_required(login_url="/login/")
def export_products_excel(request):
    """Xuất danh sách sản phẩm ra Excel"""
    from core.excel_export import excel_response
    from datetime import datetime

    products = Product.objects.select_related('category', 'category__parent', 'supplier', 'location').prefetch_related(
        'stocks__warehouse',
        'combo_items__product__stocks__warehouse',
    ).all()
    products = filter_by_store(products, request)
    receipt_map = _build_receipt_history_map(products)

    columns = [
        {'key': 'stt', 'label': 'STT', 'width': 6},
        {'key': 'code', 'label': 'Mã SP', 'width': 14},
        {'key': 'name', 'label': 'Tên sản phẩm', 'width': 30},
        {'key': 'barcode', 'label': 'Barcode', 'width': 16},
        {'key': 'category', 'label': 'Danh mục', 'width': 16},
        {'key': 'unit', 'label': 'ĐVT', 'width': 8},
        {'key': 'spec', 'label': 'Quy cách', 'width': 14},
        {'key': 'product_type', 'label': 'Loại sản phẩm', 'width': 18},
        {'key': 'product_nature', 'label': 'Tính chất', 'width': 12},
        {'key': 'import_price', 'label': 'Giá nhập', 'width': 14},
        {'key': 'cost_price', 'label': 'Giá vốn', 'width': 14},
        {'key': 'selling_price', 'label': 'Giá bán lẻ', 'width': 14},
        {'key': 'wholesale_no_w', 'label': 'Giá sỉ KBH', 'width': 14},
        {'key': 'wholesale_w', 'label': 'Giá sỉ BH', 'width': 14},
        {'key': 'stock', 'label': 'Tồn kho', 'width': 10},
        {'key': 'min_stock', 'label': 'Tồn tối thiểu', 'width': 12},
        {'key': 'max_stock', 'label': 'Tồn tối đa', 'width': 12},
        {'key': 'status', 'label': 'Trạng thái', 'width': 14},
        {'key': 'supplier', 'label': 'NCC', 'width': 18},
        {'key': 'location', 'label': 'Vị trí', 'width': 14},
        {'key': 'description', 'label': 'Mô tả', 'width': 28},
    ]

    rows = []
    for i, p in enumerate(products, 1):
        category = p.category
        root_category = category.parent if category and category.parent_id else category
        product_type = category if category and category.parent_id else None
        if p.is_combo:
            _, total_stock = _combo_stock_by_warehouse(p)
        else:
            total_stock = float(sum(s.quantity for s in p.stocks.all()))
        current_cost = _calc_on_hand_cost_price(p.id, total_stock, receipt_map)
        effective_cost = current_cost if current_cost > 0 else float(p.cost_price)
        rows.append({
            'stt': i,
            'code': p.code,
            'name': p.name,
            'barcode': p.barcode or '',
            'category': root_category.name if root_category else '',
            'unit': p.unit or '',
            'spec': p.specification or '',
            'product_type': product_type.name if product_type else '',
            'product_nature': 'Combo' if p.is_combo else ('Dịch vụ' if p.is_service else ('Cân/đong' if p.is_weight_based else 'Sản phẩm')),
            'import_price': float(p.import_price),
            'cost_price': effective_cost,
            'selling_price': float(p.selling_price),
            'wholesale_no_w': float(p.wholesale_price_no_warranty),
            'wholesale_w': float(p.wholesale_price_warranty),
            'stock': total_stock,
            'min_stock': p.min_stock,
            'max_stock': p.max_stock,
            'status': 'Đang hoạt động' if p.is_active else 'Ngừng hoạt động',
            'supplier': p.supplier.name if p.supplier else '',
            'location': p.location.name if p.location else '',
            'description': p.description or '',
        })

    return excel_response(
        title='DANH SÁCH SẢN PHẨM',
        subtitle=f'Xuất ngày {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} sản phẩm',
        columns=columns,
        rows=rows,
        filename=f'San_pham_{datetime.now().strftime("%Y%m%d")}',
        money_cols=['import_price', 'cost_price', 'selling_price', 'wholesale_no_w', 'wholesale_w'],
    )


@login_required(login_url="/login/")
def export_goods_receipts_excel(request):
    """Xuất danh sách phiếu nhập kho ra Excel"""
    from core.excel_export import excel_response
    from datetime import datetime

    receipts = GoodsReceipt.objects.select_related('supplier', 'warehouse', 'created_by').prefetch_related('items__product').all()
    receipts = filter_by_store(receipts, request, field_name='warehouse__store')

    columns = [
        {'key': 'stt', 'label': 'STT', 'width': 6},
        {'key': 'code', 'label': 'Mã phiếu', 'width': 14},
        {'key': 'receipt_date', 'label': 'Ngày nhập', 'width': 12},
        {'key': 'supplier', 'label': 'Nhà cung cấp', 'width': 22},
        {'key': 'warehouse', 'label': 'Kho nhập', 'width': 16},
        {'key': 'total_quantity', 'label': 'Số lượng', 'width': 10},
        {'key': 'total_amount', 'label': 'Tổng tiền', 'width': 16},
        {'key': 'status', 'label': 'Trạng thái', 'width': 12},
        {'key': 'created_by', 'label': 'Người tạo', 'width': 14},
        {'key': 'note', 'label': 'Ghi chú', 'width': 24},
    ]

    rows = []
    total = 0
    total_quantity_all = 0
    for i, r in enumerate(receipts, 1):
        total_quantity, receipt_total = _calculate_goods_receipt_totals(r.items.all())
        total += float(receipt_total)
        total_quantity_all += float(total_quantity)
        rows.append({
            'stt': i,
            'code': r.code,
            'receipt_date': r.receipt_date,
            'supplier': r.supplier.name if r.supplier else '',
            'warehouse': r.warehouse.name if r.warehouse else '',
            'total_quantity': float(total_quantity),
            'total_amount': float(receipt_total),
            'status': r.get_status_display(),
            'created_by': r.created_by.get_full_name() or r.created_by.username if r.created_by else '',
            'note': r.note or '',
        })

    return excel_response(
        title='DANH SÁCH PHIẾU NHẬP KHO',
        subtitle=f'Xuất ngày {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} phiếu',
        columns=columns, rows=rows,
        filename=f'Nhap_kho_{datetime.now().strftime("%Y%m%d")}',
        money_cols=['total_amount'],
        total_row={'stt': '', 'code': 'TỔNG CỘNG', 'total_quantity': total_quantity_all, 'total_amount': total},
    )


@login_required(login_url="/login/")
def export_stock_transfers_excel(request):
    """Xuất danh sách chuyển kho ra Excel"""
    from core.excel_export import excel_response
    from datetime import datetime

    transfers = StockTransfer.objects.select_related('from_warehouse', 'to_warehouse', 'created_by').prefetch_related('items__product').all()
    transfers = filter_by_store(transfers, request, field_name='from_warehouse__store')

    columns = [
        {'key': 'stt', 'label': 'STT', 'width': 6},
        {'key': 'code', 'label': 'Mã phiếu', 'width': 14},
        {'key': 'transfer_date', 'label': 'Ngày chuyển', 'width': 12},
        {'key': 'from_warehouse', 'label': 'Kho xuất', 'width': 18},
        {'key': 'to_warehouse', 'label': 'Kho nhập', 'width': 18},
        {'key': 'items_count', 'label': 'Số SP', 'width': 8},
        {'key': 'total_qty', 'label': 'Tổng SL', 'width': 10},
        {'key': 'status', 'label': 'Trạng thái', 'width': 12},
        {'key': 'created_by', 'label': 'Người tạo', 'width': 14},
        {'key': 'note', 'label': 'Ghi chú', 'width': 24},
    ]

    rows = []
    for i, t in enumerate(transfers, 1):
        total_qty = sum(float(item.quantity) for item in t.items.all())
        rows.append({
            'stt': i,
            'code': t.code,
            'transfer_date': t.transfer_date,
            'from_warehouse': t.from_warehouse.name if t.from_warehouse else '',
            'to_warehouse': t.to_warehouse.name if t.to_warehouse else '',
            'items_count': t.items.count(),
            'total_qty': total_qty,
            'status': t.get_status_display(),
            'created_by': t.created_by.get_full_name() or t.created_by.username if t.created_by else '',
            'note': t.note or '',
        })

    return excel_response(
        title='DANH SÁCH PHIẾU CHUYỂN KHO',
        subtitle=f'Xuất ngày {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} phiếu',
        columns=columns, rows=rows,
        filename=f'Chuyen_kho_{datetime.now().strftime("%Y%m%d")}',
    )


@login_required(login_url="/login/")
def export_stock_checks_excel(request):
    """Xuất danh sách kiểm kho ra Excel"""
    from core.excel_export import excel_response
    from datetime import datetime

    checks = StockCheck.objects.select_related('warehouse', 'created_by').prefetch_related('items__product').all()
    checks = filter_by_store(checks, request, field_name='warehouse__store')

    columns = [
        {'key': 'stt', 'label': 'STT', 'width': 6},
        {'key': 'code', 'label': 'Mã phiếu', 'width': 14},
        {'key': 'check_date', 'label': 'Ngày kiểm', 'width': 12},
        {'key': 'warehouse', 'label': 'Kho', 'width': 18},
        {'key': 'items_count', 'label': 'Số SP', 'width': 8},
        {'key': 'total_diff', 'label': 'Chênh lệch SL', 'width': 14},
        {'key': 'status', 'label': 'Trạng thái', 'width': 12},
        {'key': 'created_by', 'label': 'Người tạo', 'width': 14},
        {'key': 'note', 'label': 'Ghi chú', 'width': 24},
    ]

    rows = []
    for i, c in enumerate(checks, 1):
        total_diff = sum(float(item.difference) for item in c.items.all())
        rows.append({
            'stt': i,
            'code': c.code,
            'check_date': c.check_date,
            'warehouse': c.warehouse.name if c.warehouse else '',
            'items_count': c.items.count(),
            'total_diff': total_diff,
            'status': c.get_status_display(),
            'created_by': c.created_by.get_full_name() or c.created_by.username if c.created_by else '',
            'note': c.note or '',
        })

    return excel_response(
        title='DANH SÁCH PHIẾU KIỂM KHO',
        subtitle=f'Xuất ngày {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} phiếu',
        columns=columns, rows=rows,
        filename=f'Kiem_kho_{datetime.now().strftime("%Y%m%d")}',
    )


@login_required(login_url="/login/")
def export_purchase_orders_excel(request):
    """Xuất danh sách đơn đặt hàng nhập ra Excel"""
    from core.excel_export import excel_response
    from datetime import datetime

    orders = PurchaseOrder.objects.select_related('supplier', 'warehouse', 'created_by').prefetch_related('items__product').all()
    orders = filter_by_store(orders, request, field_name='warehouse__store')

    columns = [
        {'key': 'stt', 'label': 'STT', 'width': 6},
        {'key': 'code', 'label': 'Mã đơn', 'width': 14},
        {'key': 'order_date', 'label': 'Ngày đặt', 'width': 12},
        {'key': 'expected_date', 'label': 'Ngày dự kiến', 'width': 12},
        {'key': 'supplier', 'label': 'Nhà cung cấp', 'width': 22},
        {'key': 'warehouse', 'label': 'Kho nhập', 'width': 16},
        {'key': 'items_count', 'label': 'Số SP', 'width': 8},
        {'key': 'total_amount', 'label': 'Tổng tiền', 'width': 16},
        {'key': 'status', 'label': 'Trạng thái', 'width': 12},
        {'key': 'created_by', 'label': 'Người tạo', 'width': 14},
        {'key': 'note', 'label': 'Ghi chú', 'width': 24},
    ]

    rows = []
    total = 0
    for i, o in enumerate(orders, 1):
        total += float(o.total_amount)
        rows.append({
            'stt': i,
            'code': o.code,
            'order_date': o.order_date,
            'expected_date': o.expected_date,
            'supplier': o.supplier.name if o.supplier else '',
            'warehouse': o.warehouse.name if o.warehouse else '',
            'items_count': o.items.count(),
            'total_amount': float(o.total_amount),
            'status': o.get_status_display(),
            'created_by': o.created_by.get_full_name() or o.created_by.username if o.created_by else '',
            'note': o.note or '',
        })

    return excel_response(
        title='DANH SÁCH ĐƠN ĐẶT HÀNG NHẬP',
        subtitle=f'Xuất ngày {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} đơn',
        columns=columns, rows=rows,
        filename=f'Dat_hang_nhap_{datetime.now().strftime("%Y%m%d")}',
        money_cols=['total_amount'],
        total_row={'stt': '', 'code': 'TỔNG CỘNG', 'total_amount': total},
    )


@login_required(login_url="/login/")
def export_suppliers_excel(request):
    """Xuất danh sách nhà cung cấp ra Excel"""
    from core.excel_export import excel_response
    from datetime import datetime

    suppliers = Supplier.objects.all()

    columns = [
        {'key': 'stt', 'label': 'STT', 'width': 6},
        {'key': 'code', 'label': 'Mã NCC', 'width': 12},
        {'key': 'name', 'label': 'Tên nhà cung cấp', 'width': 28},
        {'key': 'phone', 'label': 'SĐT', 'width': 14},
        {'key': 'email', 'label': 'Email', 'width': 22},
        {'key': 'tax_code', 'label': 'MST', 'width': 14},
        {'key': 'contact_person', 'label': 'Người liên hệ', 'width': 18},
        {'key': 'address', 'label': 'Địa chỉ', 'width': 30},
        {'key': 'note', 'label': 'Ghi chú', 'width': 24},
    ]

    rows = []
    for i, s in enumerate(suppliers, 1):
        rows.append({
            'stt': i,
            'code': s.code,
            'name': s.name,
            'phone': s.phone or '',
            'email': s.email or '',
            'tax_code': s.tax_code or '',
            'contact_person': s.contact_person or '',
            'address': s.address or '',
            'note': s.note or '',
        })

    return excel_response(
        title='DANH SÁCH NHÀ CUNG CẤP',
        subtitle=f'Xuất ngày {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} NCC',
        columns=columns, rows=rows,
        filename=f'NCC_{datetime.now().strftime("%Y%m%d")}',
    )
