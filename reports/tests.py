from datetime import date, datetime, time, timedelta
from io import BytesIO

from django.contrib.auth.models import User
from django.core import mail
from django.core.management import call_command
from django.db import connection
from django.test import TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from customers.models import Customer, CustomerGroup
from finance.models import CashBook, Payment, PaymentMethodOption, Receipt
from orders.models import (
    Order,
    OrderItem,
    OrderReturn,
    OrderReturnItem,
    Quotation,
    QuotationItem,
)
from products.models import (
    GoodsReceipt,
    GoodsReceiptItem,
    Product,
    ProductCategory,
    ProductStock,
    ProductVariant,
    PurchaseReturn,
    PurchaseReturnItem,
    StockCheck,
    StockCheckItem,
    StockTransfer,
    StockTransferItem,
    Supplier,
    Warehouse,
)
from system_management.models import Brand, Store, UserProfile
from reports.models import (
    DailyEmailReport,
    StockAlert,
    StockAlertEmailRecipient,
)
from reports.daily_email_reports import collect_daily_email_report_metrics
from reports.email_scheduler import process_daily_email_report, process_stock_alert


class SalesReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.brand = Brand.objects.create(name='Brand Report')
        cls.store = Store.objects.create(brand=cls.brand, name='Store Report', code='SRP')
        cls.user = User.objects.create_user(username='acct_report', password='pass123')
        UserProfile.objects.create(user=cls.user, store=cls.store, position='Kế toán')

        cls.customer = Customer.objects.create(
            store=cls.store,
            code='KH-RP-001',
            name='Khách báo cáo',
            created_by=cls.user,
        )
        cls.warehouse = Warehouse.objects.create(store=cls.store, code='KHO-RP', name='Kho báo cáo')
        cls.product = Product.objects.create(
            store=cls.store,
            code='SP-RP-001',
            name='Sản phẩm báo cáo',
            created_by=cls.user,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def test_api_report_sales_rejects_regular_staff(self):
        staff = User.objects.create_user(username='regular_report_staff', password='pass123')
        UserProfile.objects.create(user=staff, store=self.store, position='Quản lý cửa hàng')
        self.client.force_login(staff)

        response = self.client.get(reverse('api_report_sales'))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['status'], 'error')

    def test_api_report_sales_allows_brand_owner(self):
        owner = User.objects.create_user(username='owner_sales_report', password='pass123')
        Brand.objects.create(name='Owner Sales Report Role', owner=owner)
        self.client.force_login(owner)

        response = self.client.get(reverse('api_report_sales'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok')

    def test_api_report_sales_allows_director_position(self):
        director = User.objects.create_user(username='director_report', password='pass123')
        UserProfile.objects.create(user=director, store=self.store, position='Giám đốc')
        self.client.force_login(director)

        response = self.client.get(reverse('api_report_sales'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok')

    def test_purchase_report_groups_completed_receipts_by_supplier_and_filters_supplier(self):
        today = date.today()
        supplier_a = Supplier.objects.create(code='NCC-RP-A', name='NCC báo cáo A')
        supplier_b = Supplier.objects.create(code='NCC-RP-B', name='NCC báo cáo B')
        for code, supplier, amount, status in [
            ('PN-RP-A1', supplier_a, 100, 1),
            ('PN-RP-A2', supplier_a, 200, 1),
            ('PN-RP-A-DRAFT', supplier_a, 900, 0),
            ('PN-RP-B1', supplier_b, 400, 1),
        ]:
            GoodsReceipt.objects.create(
                code=code,
                supplier=supplier,
                warehouse=self.warehouse,
                receipt_date=today,
                total_amount=amount,
                status=status,
                created_by=self.user,
            )

        response = self.client.get(reverse('api_report_purchases'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        summary_by_supplier = {row['supplier']: row for row in payload['supplier_summary']}
        self.assertEqual(payload['summary']['total_amount'], 700.0)
        self.assertEqual(payload['summary']['total_count'], 3)
        self.assertEqual(payload['summary']['total_suppliers'], 2)
        self.assertEqual(summary_by_supplier[supplier_a.name]['receipt_count'], 2)
        self.assertEqual(summary_by_supplier[supplier_a.name]['total_amount'], 300.0)
        self.assertEqual(summary_by_supplier[supplier_b.name]['total_amount'], 400.0)

        filtered = self.client.get(reverse('api_report_purchases'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'supplier_id': supplier_a.id,
        }).json()
        self.assertEqual(filtered['summary']['total_amount'], 300.0)
        self.assertEqual(filtered['summary']['total_count'], 2)
        self.assertEqual(len(filtered['supplier_summary']), 1)
        self.assertTrue(all(row['supplier'] == supplier_a.name for row in filtered['data']))

    def test_export_purchase_report_includes_supplier_summary_sheet(self):
        today = date.today()
        supplier = Supplier.objects.create(code='NCC-RP-EX', name='NCC xuất báo cáo')
        GoodsReceipt.objects.create(
            code='PN-RP-EX',
            supplier=supplier,
            warehouse=self.warehouse,
            receipt_date=today,
            total_amount=750,
            status=1,
            created_by=self.user,
        )

        response = self.client.get(reverse('export_purchases_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'supplier_id': supplier.id,
        })

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertIn('Tổng hợp NCC', workbook.sheetnames)
        summary_sheet = workbook['Tổng hợp NCC']
        self.assertEqual(summary_sheet['B5'].value, supplier.name)
        self.assertEqual(summary_sheet['C5'].value, 1)
        self.assertEqual(summary_sheet['D5'].value, 750)

    def test_finance_report_adds_completed_payments_and_goods_receipts_in_scope(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        second_store = Store.objects.create(
            brand=self.brand,
            name='Store Report 2',
            code='SRP2',
        )
        second_warehouse = Warehouse.objects.create(
            store=second_store,
            name='Kho báo cáo 2',
            code='KHO-RP-2',
        )
        foreign_brand = Brand.objects.create(name='Foreign Report Brand')
        foreign_store = Store.objects.create(
            brand=foreign_brand,
            name='Foreign Report Store',
            code='FRS',
        )
        foreign_warehouse = Warehouse.objects.create(
            store=foreign_store,
            name='Kho báo cáo ngoài phạm vi',
            code='KHO-RP-FOREIGN',
        )
        to_day = date.today()
        from_day = to_day - timedelta(days=1)
        outside_day = from_day - timedelta(days=1)

        goods_receipt_specs = [
            ('PN-RP-DONE-FROM', self.warehouse, from_day, 4000, 1),
            ('PN-RP-DONE-TO', self.warehouse, to_day, 6000, 1),
            ('PN-RP-SECOND-STORE', second_warehouse, to_day, 20000, 1),
            ('PN-RP-DRAFT', self.warehouse, to_day, 30000, 0),
            ('PN-RP-CANCELED', self.warehouse, to_day, 40000, 2),
            ('PN-RP-OUTSIDE-DATE', self.warehouse, outside_day, 50000, 1),
            ('PN-RP-FOREIGN-STORE', foreign_warehouse, to_day, 60000, 1),
        ]
        for code, warehouse, receipt_date, amount, status in goods_receipt_specs:
            GoodsReceipt.objects.create(
                code=code,
                warehouse=warehouse,
                receipt_date=receipt_date,
                total_amount=amount,
                status=status,
                created_by=self.user,
            )

        receipt_specs = [
            ('PT-RP-DONE-FROM', self.store, from_day, 100, 1),
            ('PT-RP-DONE-TO', self.store, to_day, 250, 1),
            ('PT-RP-SECOND-STORE', second_store, to_day, 600, 1),
            ('PT-RP-DRAFT', self.store, to_day, 900, 0),
            ('PT-RP-CANCELED', self.store, to_day, 800, 2),
            ('PT-RP-OUTSIDE-DATE', self.store, outside_day, 700, 1),
            ('PT-RP-FOREIGN-STORE', foreign_store, to_day, 5000, 1),
        ]
        for code, store, receipt_date, amount, status in receipt_specs:
            Receipt.objects.create(
                code=code,
                store=store,
                amount=amount,
                receipt_date=receipt_date,
                status=status,
                created_by=self.user,
            )

        payment_specs = [
            ('PC-RP-DONE-FROM', self.store, from_day, 40, 1),
            ('PC-RP-DONE-TO', self.store, to_day, 60, 1),
            ('PC-RP-SECOND-STORE', second_store, to_day, 90, 1),
            ('PC-RP-DRAFT', self.store, to_day, 300, 0),
            ('PC-RP-CANCELED', self.store, to_day, 200, 2),
            ('PC-RP-OUTSIDE-DATE', self.store, outside_day, 100, 1),
            ('PC-RP-FOREIGN-STORE', foreign_store, to_day, 500, 1),
        ]
        for code, store, payment_date, amount, status in payment_specs:
            Payment.objects.create(
                code=code,
                store=store,
                amount=amount,
                payment_date=payment_date,
                status=status,
                created_by=self.user,
            )

        response = self.client.get(reverse('api_report_finance'), {
            'from_date': from_day.isoformat(),
            'to_date': to_day.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_income'], 950.0)
        self.assertEqual(payload['summary']['payment_expense'], 190.0)
        self.assertEqual(payload['summary']['goods_receipt_expense'], 30000.0)
        self.assertEqual(payload['summary']['total_expense'], 30190.0)
        self.assertEqual(payload['summary']['net_profit'], -29240.0)
        category_rows = {row['name']: row for row in payload['categories']}
        self.assertEqual(category_rows['Hàng nhập (phiếu nhập)']['expense'], 30000.0)
        self.assertEqual(sum(row['expense'] for row in payload['categories']), 30190.0)

        store_rows = {row['store_id']: row for row in payload['store_breakdown']}
        self.assertEqual(store_rows[self.store.id]['income'], 350.0)
        self.assertEqual(store_rows[self.store.id]['payment_expense'], 100.0)
        self.assertEqual(store_rows[self.store.id]['goods_receipt_expense'], 10000.0)
        self.assertEqual(store_rows[self.store.id]['expense'], 10100.0)
        self.assertEqual(store_rows[self.store.id]['net'], -9750.0)
        self.assertEqual(store_rows[second_store.id]['income'], 600.0)
        self.assertEqual(store_rows[second_store.id]['payment_expense'], 90.0)
        self.assertEqual(store_rows[second_store.id]['goods_receipt_expense'], 20000.0)
        self.assertEqual(store_rows[second_store.id]['expense'], 20090.0)
        self.assertEqual(store_rows[second_store.id]['net'], -19490.0)

        selected_store_payload = self.client.get(reverse('api_report_finance'), {
            'from_date': from_day.isoformat(),
            'to_date': to_day.isoformat(),
            'store_id': self.store.id,
        }).json()
        self.assertEqual(selected_store_payload['summary']['total_income'], 350.0)
        self.assertEqual(selected_store_payload['summary']['payment_expense'], 100.0)
        self.assertEqual(selected_store_payload['summary']['goods_receipt_expense'], 10000.0)
        self.assertEqual(selected_store_payload['summary']['total_expense'], 10100.0)
        self.assertEqual(selected_store_payload['summary']['net_profit'], -9750.0)

        excel_response = self.client.get(reverse('export_finance_excel'), {
            'from_date': from_day.isoformat(),
            'to_date': to_day.isoformat(),
        })
        self.assertEqual(excel_response.status_code, 200)
        worksheet = load_workbook(
            BytesIO(excel_response.content),
            data_only=True,
        )['Thu chi']
        self.assertIn('Tổng phiếu chi: 190đ', worksheet['A3'].value)
        self.assertIn('Tổng hàng nhập: 30,000đ', worksheet['A3'].value)
        self.assertIn('Tổng chi = Tổng phiếu chi + Tổng hàng nhập: 30,190đ', worksheet['A4'].value)

        excel_rows = list(worksheet.iter_rows(values_only=True))
        imported_rows = {
            row[2]: row
            for row in excel_rows
            if row[1] == 'NHẬP HÀNG'
        }
        self.assertEqual(set(imported_rows), {
            'PN-RP-DONE-FROM',
            'PN-RP-DONE-TO',
            'PN-RP-SECOND-STORE',
        })
        self.assertEqual(imported_rows['PN-RP-SECOND-STORE'][6], 20000)
        total_rows = {
            row[1]: row[6]
            for row in excel_rows
            if row[1] in {
                'TỔNG THU',
                'TỔNG PHIẾU CHI',
                'TỔNG HÀNG NHẬP',
                'TỔNG CHI',
                'LÃI/LỖ',
            }
        }
        self.assertEqual(total_rows['TỔNG THU'], 950)
        self.assertEqual(total_rows['TỔNG PHIẾU CHI'], 190)
        self.assertEqual(total_rows['TỔNG HÀNG NHẬP'], 30000)
        self.assertEqual(total_rows['TỔNG CHI'], 30190)
        self.assertEqual(total_rows['LÃI/LỖ'], -29240)

    def test_finance_report_page_shows_expense_formula_cards(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])

        response = self.client.get(reverse('report_finance'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="payment_expense"')
        self.assertContains(response, 'Tổng phiếu chi')
        self.assertContains(response, 'id="goods_receipt_expense"')
        self.assertContains(response, 'Tổng hàng nhập')
        self.assertContains(response, 'id="total_expense"')
        self.assertContains(response, 'Tổng chi (phiếu chi + hàng nhập)')
        self.assertContains(response, 'id="order_debt_link"')
        self.assertContains(response, reverse('report_finance_order_debt'))
        self.assertContains(response, 'target="_blank"')
        self.assertContains(response, 'Xem bảng chi tiết')
        self.assertContains(response, "params.set('sort', 'debt_desc')")
        self.assertContains(response, 'id="payment_expense_link"')
        self.assertContains(response, reverse('payment_tbl'))
        self.assertNotContains(response, 'Xem bảng phiếu chi')
        self.assertContains(response, "params.set('status', '1')")
        self.assertContains(response, 'updatePaymentExpenseLink()')

    def test_customer_report_page_has_pagination(self):
        response = self.client.get(reverse('report_customers'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="customer_report_page_size"')
        self.assertContains(response, '<option value="25">25 dòng</option>', html=True)
        self.assertContains(response, '<option value="200">200 dòng</option>', html=True)
        self.assertContains(response, 'id="customer_report_pagination_summary"')
        self.assertContains(response, 'id="customer_report_pagination"')
        self.assertContains(response, 'ifshop_report_customers_page_size')
        self.assertContains(response, 'function buildCustomerReportPaginationButtons(currentPage, totalPages)')
        self.assertContains(response, 'var pageRows = filtered.slice(start, end);')
        self.assertContains(response, 'customer-report-page-btn')

    def test_customer_report_total_purchase_column_has_sort_toggle(self):
        response = self.client.get(reverse('report_customers'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="customer_total_purchase_heading"')
        self.assertContains(response, 'style="cursor:pointer;" aria-sort="none"')
        self.assertContains(
            response,
            'class="app-sort-toggle-btn" id="customer_total_purchase_sort"',
        )
        self.assertContains(response, 'fas fa-sort-amount-down')
        self.assertContains(response, "customerReportTotalPurchaseSortDirection = 'desc'")
        self.assertContains(response, 'function sortCustomerReportRows(rows)')
        self.assertContains(response, 'function syncCustomerReportTotalPurchaseSort()')
        self.assertContains(response, 'filtered = sortCustomerReportRows(filtered);')
        self.assertContains(response, "customerReportTotalPurchaseSortDirection === 'desc' ? 'asc' : 'desc'")

    def test_customer_report_total_debt_column_has_sort_toggle(self):
        response = self.client.get(reverse('report_customers'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="customer_total_debt_heading"')
        self.assertContains(response, 'style="cursor:pointer;" aria-sort="descending"')
        self.assertContains(
            response,
            'class="app-sort-toggle-btn active" id="customer_total_debt_sort"',
        )
        self.assertContains(response, "customerReportSortField = 'total_debt'")
        self.assertContains(response, "customerReportTotalDebtSortDirection = 'desc'")
        self.assertContains(response, "customerReportSortField === 'total_debt'")
        self.assertContains(response, "customerReportTotalDebtSortDirection === 'desc' ? 'asc' : 'desc'")
        self.assertContains(response, 'function syncCustomerReportTotalDebtSort()')
        self.assertContains(response, 'Number(left.row[sortField] || 0)')

    def test_customer_report_api_combines_legacy_metrics_and_live_orders(self):
        today = date.today()
        legacy_customer = Customer.objects.create(
            store=self.store,
            code='KH-RP-LEGACY-METRICS',
            name='Khách có dữ liệu lịch sử',
            total_purchased=1000,
            total_debt=100,
            order_count=2,
            imported_legacy_metrics=True,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-RP-CUSTOMER-REALIZED',
            store=self.store,
            customer=legacy_customer,
            warehouse=self.warehouse,
            status=4,
            total_amount=500,
            final_amount=500,
            paid_amount=200,
            order_date=today - timedelta(days=2),
            exported_at=datetime.combine(today, time(9, 30)),
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-RP-CUSTOMER-PENDING',
            store=self.store,
            customer=legacy_customer,
            warehouse=self.warehouse,
            status=1,
            total_amount=300,
            final_amount=300,
            paid_amount=0,
            order_date=today,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-RP-CUSTOMER-CANCELED',
            store=self.store,
            customer=legacy_customer,
            warehouse=self.warehouse,
            status=6,
            total_amount=900,
            final_amount=900,
            paid_amount=0,
            order_date=today,
            created_by=self.user,
        )

        response = self.client.get(reverse('api_report_customers'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload['data'] if item['code'] == legacy_customer.code)
        self.assertEqual(row['total_purchased'], 1500.0)
        self.assertEqual(row['total_debt'], 700.0)
        self.assertEqual(row['order_count'], 4)
        self.assertEqual(row['last_order_date'], today.strftime('%d/%m/%Y'))

    def test_customer_report_api_query_count_does_not_grow_per_customer(self):
        today = date.today()
        customers = [
            Customer.objects.create(
                store=self.store,
                code=f'KH-RP-BATCH-{index:02d}',
                name=f'Khách tổng hợp {index:02d}',
                created_by=self.user,
            )
            for index in range(12)
        ]
        Order.objects.bulk_create([
            Order(
                code=f'DH-RP-BATCH-{index:02d}',
                store=self.store,
                customer=customer,
                warehouse=self.warehouse,
                status=5,
                total_amount=100,
                final_amount=100,
                paid_amount=100,
                order_date=today,
                exported_at=datetime.combine(today, time(10, 0)),
                created_by=self.user,
            )
            for index, customer in enumerate(customers)
        ])

        with CaptureQueriesContext(connection) as captured:
            response = self.client.get(reverse('api_report_customers'))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(captured), 15)
        self.assertEqual(len(response.json()['data']), len(customers) + 1)

    def test_customer_report_page_shows_loading_and_api_errors(self):
        response = self.client.get(reverse('report_customers'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Đang tải dữ liệu khách hàng')
        self.assertContains(response, 'function showCustomerReportError(message)')
        self.assertContains(response, '.fail(function(xhr)')

    def test_purchase_supplier_summary_has_pagination(self):
        response = self.client.get(reverse('report_purchases'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="purchase_supplier_page_size"')
        self.assertContains(response, '<option value="25">25 dòng</option>', html=True)
        self.assertContains(response, '<option value="200">200 dòng</option>', html=True)
        self.assertContains(response, 'id="purchase_supplier_pagination_summary"')
        self.assertContains(response, 'id="purchase_supplier_pagination"')
        self.assertContains(response, 'ifshop_report_purchases_supplier_page_size')
        self.assertContains(response, 'function renderPurchaseSupplierSummary()')
        self.assertContains(response, 'purchaseSupplierSummaryRows.slice(start, end)')
        self.assertContains(response, 'purchase-supplier-page-btn')

    def test_purchase_details_have_independent_pagination(self):
        response = self.client.get(reverse('report_purchases'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="purchase_detail_page_size"')
        self.assertContains(response, '<option value="25">25 dòng</option>', html=True)
        self.assertContains(response, '<option value="200">200 dòng</option>', html=True)
        self.assertContains(response, 'id="purchase_detail_pagination_summary"')
        self.assertContains(response, 'id="purchase_detail_pagination"')
        self.assertContains(response, 'ifshop_report_purchases_detail_page_size')
        self.assertContains(response, 'function renderPurchaseDetails()')
        self.assertContains(response, 'purchaseDetailRows.slice(start, end)')
        self.assertContains(response, 'purchase-detail-page-btn')

    def test_finance_order_debt_page_only_lists_positive_debt_and_matches_card(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        today = date.today()
        outside_day = today - timedelta(days=40)
        debt_order = Order.objects.create(
            code='DH-DEBT-DETAIL-001',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=1,
            total_amount=1000,
            final_amount=1000,
            paid_amount=300,
            order_date=today,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-DEBT-PAID',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=500,
            final_amount=500,
            paid_amount=500,
            order_date=today,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-DEBT-OVERPAID',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=500,
            final_amount=500,
            paid_amount=600,
            order_date=today,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-DEBT-CANCELED',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=6,
            payment_status=0,
            total_amount=900,
            final_amount=900,
            paid_amount=0,
            order_date=today,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-DEBT-OUTSIDE',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=0,
            total_amount=800,
            final_amount=800,
            paid_amount=0,
            order_date=outside_day,
            created_by=self.user,
        )
        params = {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        }

        page_response = self.client.get(reverse('report_finance_order_debt'), params)
        api_response = self.client.get(reverse('api_report_finance'), params)

        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, 'class="text-right app-sortable-heading"')
        self.assertContains(page_response, debt_order.code)
        self.assertContains(page_response, f'/order-tbl/?open_order={debt_order.id}')
        self.assertContains(page_response, '700đ')
        for excluded_code in (
            'DH-DEBT-PAID',
            'DH-DEBT-OVERPAID',
            'DH-DEBT-CANCELED',
            'DH-DEBT-OUTSIDE',
        ):
            self.assertNotContains(page_response, excluded_code)
        self.assertEqual(page_response.context['totals']['order_count'], 1)
        self.assertEqual(page_response.context['totals']['debt_amount'], 700)
        self.assertEqual(api_response.status_code, 200)
        self.assertEqual(api_response.json()['summary']['order_debt'], 700.0)

    def test_finance_order_debt_page_is_paginated(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        today = date.today()
        Order.objects.bulk_create([
            Order(
                code=f'DH-DEBT-PAGE-{index:02d}',
                store=self.store,
                customer=self.customer,
                warehouse=self.warehouse,
                status=5,
                payment_status=0,
                total_amount=100,
                final_amount=100,
                paid_amount=0,
                order_date=today,
                created_by=self.user,
            )
            for index in range(31)
        ])

        response = self.client.get(reverse('report_finance_order_debt'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'page': 2,
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['page_obj'].paginator.count, 31)
        self.assertEqual(response.context['page_obj'].paginator.per_page, 25)
        self.assertEqual(len(response.context['page_obj']), 6)
        self.assertContains(response, 'Trang 2/2')
        self.assertContains(response, 'sort=debt_desc')

    def test_finance_order_debt_page_orders_highest_debt_first(self):
        self.brand.owner = self.user
        self.brand.save(update_fields=['owner'])
        today = date.today()
        yesterday = today - timedelta(days=1)
        Order.objects.create(
            code='DH-DEBT-LOW',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=1,
            total_amount=200,
            final_amount=200,
            paid_amount=100,
            order_date=today,
            created_by=self.user,
        )
        Order.objects.create(
            code='DH-DEBT-HIGH',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=1,
            total_amount=1200,
            final_amount=1200,
            paid_amount=200,
            order_date=yesterday,
            created_by=self.user,
        )

        response = self.client.get(reverse('report_finance_order_debt'), {
            'from_date': yesterday.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['filters']['sort'], 'debt_desc')
        self.assertEqual(
            [order.code for order in response.context['page_obj']],
            ['DH-DEBT-HIGH', 'DH-DEBT-LOW'],
        )
        self.assertContains(response, 'Sắp xếp công nợ từ cao đến thấp')

    def test_inventory_report_alert_card_controls_are_available(self):
        owner = User.objects.create_user(username='owner_inventory_report', password='pass123')
        self.brand.owner = owner
        self.brand.save(update_fields=['owner'])
        self.client.force_login(owner)

        response = self.client.get(reverse('report_inventory'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="alert_box"')
        self.assertContains(response, '<option value="all">Tất cả cảnh báo</option>', html=True)
        self.assertContains(response, 'Cần nhập tối thiểu')
        self.assertContains(response, 'id="inventory_alert_filter_notice"')
        self.assertContains(response, 'activateInventoryAlertCard')
        self.assertContains(response, 'class="inventory-product-edit-link" target="_blank"')
        self.assertContains(response, '/product-tbl/?edit_product_id=')
        self.assertContains(response, '_inventoryProductEditorOpened')
        self.assertContains(response, '<th>Tên sản phẩm</th>', count=1)
        self.assertNotContains(response, '<th>Mã SP</th>')
        self.assertContains(response, '<th title="Nhà cung cấp">NCC</th>', html=True)
        self.assertContains(response, 'Giá tính tồn')
        self.assertContains(response, 'Dùng giá nhập')
        self.assertContains(response, 'colspan="13"')
        self.assertContains(response, 'var productIdentityHtml =')
        self.assertContains(response, 'id="inventory_report_tabs"')
        self.assertContains(response, 'id="inventory_product_tab"')
        self.assertContains(response, 'BC kho theo sản phẩm')
        self.assertContains(response, 'id="inventory_category_tab"')
        self.assertContains(response, 'BC kho theo danh mục')
        self.assertContains(response, 'id="inventory_movement_tab"')
        self.assertContains(response, 'BC nhập xuất tồn')
        self.assertContains(response, 'id="inventory_movement_from_date"')
        self.assertContains(response, 'id="inventory_movement_to_date"')
        self.assertContains(response, 'SL nhập trong kỳ')
        self.assertContains(response, 'Giá trị nhập trong kỳ')
        self.assertContains(response, 'SL xuất trong kỳ')
        self.assertContains(response, 'Giá trị xuất (giá vốn)')
        self.assertContains(response, 'id="inventory_movement_tbl"')
        self.assertContains(response, 'function loadInventoryMovementReport()')
        self.assertContains(response, reverse('api_report_inventory_movement'))
        self.assertContains(response, reverse('export_inventory_movement_excel'))
        self.assertContains(response, 'id="inventory_category_tbl"')
        self.assertContains(response, 'function buildInventoryCategoryRows(data)')
        self.assertContains(response, 'function renderInventoryCategoryTable(data)')
        self.assertContains(response, 'id="inventory_category_stock_value_sort"')
        self.assertContains(response, 'id="inventory_category_quantity_sort"')
        self.assertContains(response, 'Bấm để sắp xếp tổng tồn giảm dần')
        self.assertContains(response, 'Đang sắp xếp giá trị tồn giảm dần. Bấm để chuyển sang tăng dần')
        self.assertContains(response, "var _inventoryCategorySortField = 'stock_value';")
        self.assertContains(response, "var _inventoryCategoryStockValueSortDirection = 'desc';")
        self.assertContains(response, "var _inventoryCategoryQuantitySortDirection = 'desc';")
        self.assertContains(response, 'function sortInventoryCategoryRows(rows)')
        self.assertContains(response, 'function syncInventoryCategorySortButtons()')
        self.assertContains(
            response,
            "_inventoryCategoryStockValueSortDirection === 'desc' ? 'asc' : 'desc';",
        )
        self.assertContains(
            response,
            "_inventoryCategoryQuantitySortDirection === 'desc' ? 'asc' : 'desc';",
        )
        self.assertContains(response, "row.products[String(item.product_id)] = true;")
        self.assertContains(response, 'id="inventory_product_page_size"')
        self.assertContains(response, '<option value="25" selected>25 dòng</option>', html=True)
        self.assertContains(response, '<option value="200">200 dòng</option>', html=True)
        self.assertContains(response, 'id="inventory_product_pagination_summary"')
        self.assertContains(response, 'id="inventory_product_pagination"')
        self.assertContains(response, 'function renderInventoryProductPagination(meta)')
        self.assertContains(response, 'function renderInventoryProductTable()')
        self.assertContains(response, '_inventoryProductRows.slice(startOffset, endOffset)')
        self.assertContains(response, 'inventory-product-page-btn')
        self.assertContains(response, 'id="inventory_valuation_sort"')
        self.assertContains(response, 'class="app-sort-toggle-btn inventory-product-sort-btn active"')
        self.assertContains(response, 'data-direction="desc"')
        self.assertContains(response, 'fas fa-sort-amount-down')
        self.assertContains(response, 'id="inventory_stock_sort"')
        self.assertContains(response, 'class="app-sortable-heading"', count=2)
        self.assertContains(response, '.app-sortable-heading {')
        self.assertContains(response, 'white-space: nowrap;')
        self.assertContains(response, 'Bấm để sắp xếp tồn kho giảm dần')
        self.assertContains(response, 'var _inventoryProductSortField = \'valuation_price\';')
        self.assertContains(response, "var _inventoryValuationSortDirection = 'desc';")
        self.assertContains(response, "var _inventoryStockSortDirection = 'desc';")
        self.assertContains(response, 'function sortInventoryProductRows(rows)')
        self.assertContains(response, 'function syncInventoryProductSortButtons()')
        self.assertContains(
            response,
            "var difference = Number(a[_inventoryProductSortField] || 0) - Number(b[_inventoryProductSortField] || 0);",
        )

    def test_inventory_movement_report_reconstructs_period_and_uses_transaction_costs(self):
        self.product.cost_price = 120
        self.product.import_price = 100
        self.product.save(update_fields=['cost_price', 'import_price'])
        from_date = date(2026, 7, 1)
        to_date = date(2026, 7, 31)

        receipt = GoodsReceipt.objects.create(
            code='PN-XNT-IN-PERIOD',
            supplier=None,
            warehouse=self.warehouse,
            status=1,
            receipt_date=date(2026, 7, 5),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=receipt,
            product=self.product,
            quantity=5,
            unit_price=100,
            total_price=500,
        )
        future_receipt = GoodsReceipt.objects.create(
            code='PN-XNT-AFTER-PERIOD',
            supplier=None,
            warehouse=self.warehouse,
            status=1,
            receipt_date=date(2026, 8, 2),
            created_by=self.user,
        )
        GoodsReceiptItem.objects.create(
            goods_receipt=future_receipt,
            product=self.product,
            quantity=4,
            unit_price=110,
            total_price=440,
        )
        order = Order.objects.create(
            code='DH-XNT-OUT-PERIOD',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=4,
            order_date=date(2026, 7, 10),
            exported_at=timezone.make_aware(datetime(2026, 7, 10, 12, 0)),
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=3,
            cost_price=120,
            unit_price=200,
            total_price=600,
        )
        stock_check = StockCheck.objects.create(
            code='KK-XNT-INCREASE',
            warehouse=self.warehouse,
            status=1,
            stock_applied=True,
            check_date=date(2026, 7, 15),
            created_by=self.user,
        )
        StockCheckItem.objects.create(
            stock_check=stock_check,
            product=self.product,
            system_quantity=12,
            actual_quantity=14,
            difference=2,
        )
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=18)

        response = self.client.get(reverse('api_report_inventory_movement'), {
            'from_date': from_date.isoformat(),
            'to_date': to_date.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        row = next(item for item in payload['data'] if item['product_id'] == self.product.id)
        self.assertEqual(row['opening_quantity'], 10.0)
        self.assertEqual(row['import_quantity'], 7.0)
        self.assertEqual(row['import_value'], 740.0)
        self.assertEqual(row['export_quantity'], 3.0)
        self.assertEqual(row['export_value'], 360.0)
        self.assertEqual(row['closing_quantity'], 14.0)
        self.assertEqual(row['opening_value'], 1200.0)
        self.assertEqual(row['closing_value'], 1680.0)

    def test_inventory_movement_report_includes_returns_checks_and_transfers(self):
        self.product.cost_price = 100
        self.product.save(update_fields=['cost_price'])
        other_warehouse = Warehouse.objects.create(
            store=self.store,
            code='KHO-RP-XNT-2',
            name='Kho báo cáo XNT 2',
        )
        period_date = date(2026, 7, 12)

        original_receipt = GoodsReceipt.objects.create(
            code='PN-XNT-ORIGINAL',
            warehouse=self.warehouse,
            status=1,
            receipt_date=date(2026, 6, 1),
            created_by=self.user,
        )
        original_receipt_item = GoodsReceiptItem.objects.create(
            goods_receipt=original_receipt,
            product=self.product,
            quantity=5,
            unit_price=70,
            total_price=350,
        )
        purchase_return = PurchaseReturn.objects.create(
            code='THN-XNT-001',
            goods_receipt=original_receipt,
            warehouse=self.warehouse,
            status=1,
            stock_applied=True,
            return_date=period_date,
            created_by=self.user,
        )
        PurchaseReturnItem.objects.create(
            purchase_return=purchase_return,
            goods_receipt_item=original_receipt_item,
            product=self.product,
            quantity=1,
            unit_price=70,
            total_price=70,
        )

        original_order = Order.objects.create(
            code='DH-XNT-RETURN-ORIGINAL',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=4,
            order_date=date(2026, 6, 10),
            exported_at=timezone.make_aware(datetime(2026, 6, 10, 9, 0)),
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=original_order,
            product=self.product,
            quantity=2,
            cost_price=80,
            unit_price=150,
            total_price=300,
        )
        order_return = OrderReturn.objects.create(
            code='TH-XNT-001',
            order=original_order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            return_date=period_date,
            created_by=self.user,
        )
        OrderReturnItem.objects.create(
            order_return=order_return,
            product=self.product,
            quantity=1,
            unit_price=150,
            total_price=150,
        )

        stock_check = StockCheck.objects.create(
            code='KK-XNT-DECREASE',
            warehouse=self.warehouse,
            status=1,
            stock_applied=True,
            check_date=period_date,
            created_by=self.user,
        )
        StockCheckItem.objects.create(
            stock_check=stock_check,
            product=self.product,
            system_quantity=10,
            actual_quantity=8,
            difference=-2,
        )
        transfer = StockTransfer.objects.create(
            code='CK-XNT-001',
            from_warehouse=self.warehouse,
            to_warehouse=other_warehouse,
            status=2,
            transfer_date=period_date,
            created_by=self.user,
        )
        StockTransferItem.objects.create(transfer=transfer, product=self.product, quantity=3)
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=5)
        ProductStock.objects.create(product=self.product, warehouse=other_warehouse, quantity=3)

        response = self.client.get(reverse('api_report_inventory_movement'), {
            'from_date': '2026-07-01',
            'to_date': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        summary = response.json()['summary']
        self.assertEqual(summary['opening_quantity'], 10.0)
        self.assertEqual(summary['import_quantity'], 4.0)
        self.assertEqual(summary['import_value'], 380.0)
        self.assertEqual(summary['export_quantity'], 6.0)
        self.assertEqual(summary['export_value'], 570.0)
        self.assertEqual(summary['closing_quantity'], 8.0)

    def test_inventory_movement_excel_uses_selected_period(self):
        self.product.cost_price = 100
        self.product.save(update_fields=['cost_price'])
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=2)

        response = self.client.get(reverse('export_inventory_movement_excel'), {
            'from_date': '2026-07-01',
            'to_date': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertIn('BC_Nhap_xuat_ton_20260701_20260731.xlsx', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, 'BÁO CÁO NHẬP XUẤT TỒN')
        self.assertEqual(sheet['A2'].value, 'Kỳ báo cáo: 01/07/2026 - 31/07/2026')
        self.assertEqual(sheet['G4'].value, 'Tồn đầu kỳ')
        self.assertEqual(sheet['I4'].value, 'Nhập trong kỳ')
        self.assertEqual(sheet['K4'].value, 'Xuất trong kỳ')
        self.assertEqual(sheet['M4'].value, 'Tồn cuối kỳ')

    def test_api_inventory_report_exposes_product_supplier(self):
        from products.models import Supplier

        supplier = Supplier.objects.create(code='NCC-RP-STOCK', name='Nhà cung cấp báo cáo tồn')
        self.product.supplier = supplier
        self.product.save(update_fields=['supplier'])
        ProductStock.objects.create(product=self.product, warehouse=self.warehouse, quantity=4)

        response = self.client.get(reverse('api_report_inventory'))

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()['data'] if item['product_id'] == self.product.id)
        self.assertEqual(row['supplier'], supplier.name)

    def test_api_inventory_report_identifies_low_stock_and_restock_quantity(self):
        low_product = Product.objects.create(
            store=self.store,
            code='SP-RP-LOW',
            name='Sản phẩm thiếu tồn',
            min_stock=10,
            max_stock=30,
            created_by=self.user,
        )
        negative_product = Product.objects.create(
            store=self.store,
            code='SP-RP-NEGATIVE',
            name='Sản phẩm tồn âm',
            min_stock=0,
            created_by=self.user,
        )
        high_product = Product.objects.create(
            store=self.store,
            code='SP-RP-HIGH',
            name='Sản phẩm vượt tồn',
            min_stock=2,
            max_stock=20,
            created_by=self.user,
        )
        ProductStock.objects.create(product=low_product, warehouse=self.warehouse, quantity=4)
        ProductStock.objects.create(product=negative_product, warehouse=self.warehouse, quantity=-2)
        ProductStock.objects.create(product=high_product, warehouse=self.warehouse, quantity=25)

        response = self.client.get(reverse('api_report_inventory'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        rows = {row['product_code']: row for row in payload['data']}
        self.assertEqual(rows[low_product.code]['alert_type'], 'danger')
        self.assertEqual(rows[low_product.code]['restock_needed'], 6.0)
        self.assertEqual(rows[negative_product.code]['alert_type'], 'danger')
        self.assertEqual(rows[negative_product.code]['restock_needed'], 2.0)
        self.assertEqual(rows[high_product.code]['alert_type'], 'warning')
        self.assertEqual(rows[high_product.code]['restock_needed'], 0)
        self.assertEqual(payload['summary']['alert_count'], 3)
        self.assertEqual(payload['summary']['low_stock_count'], 2)
        self.assertEqual(payload['summary']['high_stock_count'], 1)

    def test_inventory_value_uses_cost_then_import_without_negative_offset(self):
        positive_product = Product.objects.create(
            store=self.store,
            code='SP-RP-VALUE-POS',
            name='Sản phẩm còn tồn',
            cost_price=120000,
            created_by=self.user,
        )
        negative_product = Product.objects.create(
            store=self.store,
            code='SP-RP-VALUE-NEG',
            name='Sản phẩm âm kho',
            cost_price=50000,
            created_by=self.user,
        )
        import_fallback_product = Product.objects.create(
            store=self.store,
            code='SP-RP-VALUE-IMPORT',
            name='Sản phẩm dùng giá nhập',
            cost_price=0,
            import_price=80000,
            created_by=self.user,
        )
        deleted_product = Product.objects.create(
            store=self.store,
            code='SP-RP-VALUE-DELETED',
            name='Sản phẩm đã xóa',
            cost_price=999999,
            created_by=self.user,
        )
        ProductStock.objects.create(product=positive_product, warehouse=self.warehouse, quantity=3)
        ProductStock.objects.create(product=negative_product, warehouse=self.warehouse, quantity=-2)
        ProductStock.objects.create(product=import_fallback_product, warehouse=self.warehouse, quantity=2)
        ProductStock.objects.create(product=deleted_product, warehouse=self.warehouse, quantity=10)
        deleted_product.delete()

        payload = self.client.get(reverse('api_report_inventory')).json()
        rows = {row['product_code']: row for row in payload['data']}

        self.assertEqual(rows[positive_product.code]['stock_value'], 360000.0)
        self.assertEqual(rows[negative_product.code]['stock_value'], 0.0)
        self.assertEqual(rows[import_fallback_product.code]['cost_price'], 0.0)
        self.assertEqual(rows[import_fallback_product.code]['import_price'], 80000.0)
        self.assertEqual(rows[import_fallback_product.code]['valuation_price'], 80000.0)
        self.assertEqual(rows[import_fallback_product.code]['valuation_source'], 'import_price')
        self.assertEqual(rows[import_fallback_product.code]['stock_value'], 160000.0)
        self.assertNotIn(deleted_product.code, rows)
        self.assertEqual(payload['summary']['total_value'], 520000.0)

    def test_export_inventory_uses_import_price_when_cost_is_zero(self):
        product = Product.objects.create(
            store=self.store,
            code='SP-RP-EXPORT-IMPORT',
            name='Sản phẩm xuất tồn theo giá nhập',
            cost_price=0,
            import_price=90000,
            created_by=self.user,
        )
        ProductStock.objects.create(product=product, warehouse=self.warehouse, quantity=3)

        response = self.client.get(reverse('export_inventory_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook['Tồn kho']
        self.assertEqual(sheet['J4'].value, 'Giá tính tồn')
        product_row = next(
            row for row in range(5, sheet.max_row + 1)
            if sheet.cell(row=row, column=2).value == product.code
        )
        self.assertEqual(sheet.cell(row=product_row, column=10).value, 90000)
        self.assertEqual(sheet.cell(row=product_row, column=11).value, 270000)

    def test_inventory_filters_separate_root_categories_and_product_types(self):
        root = ProductCategory.objects.create(name='Danh mục thiết bị')
        product_type = ProductCategory.objects.create(name='Loại máy xay', parent=root)
        product = Product.objects.create(
            store=self.store,
            code='SP-RP-CATEGORY-TYPE',
            name='Máy xay báo cáo',
            category=product_type,
            created_by=self.user,
        )
        ProductStock.objects.create(product=product, warehouse=self.warehouse, quantity=2)

        payload = self.client.get(reverse('api_report_inventory')).json()
        self.assertIn(root.id, [item['id'] for item in payload['categories']])
        self.assertNotIn(product_type.id, [item['id'] for item in payload['categories']])
        self.assertIn(product_type.id, [item['id'] for item in payload['product_types']])
        inventory_row = next(
            item for item in payload['data']
            if item['product_code'] == product.code
        )
        self.assertEqual(inventory_row['category_id'], root.id)
        self.assertEqual(inventory_row['category_name'], root.name)

        by_category = self.client.get(reverse('api_report_inventory'), {
            'category_id': root.id,
        }).json()
        by_type = self.client.get(reverse('api_report_inventory'), {
            'product_type_id': product_type.id,
        }).json()
        self.assertIn(product.code, [item['product_code'] for item in by_category['data']])
        self.assertEqual([item['product_code'] for item in by_type['data']], [product.code])

    def test_api_report_sales_defaults_to_realized_orders(self):
        today = date.today()
        created_orders = []
        for status, suffix in ((3, 'PACK'), (4, 'EXPORTED'), (5, 'DONE'), (6, 'CANCELLED')):
            order = Order.objects.create(
                code=f'DH-RP-SCOPE-{suffix}',
                store=self.store,
                customer=self.customer,
                warehouse=self.warehouse,
                status=status,
                total_amount=100,
                final_amount=100,
                order_date=today,
                created_by=self.user,
            )
            OrderItem.objects.create(
                order=order,
                product=self.product,
                quantity=1,
                unit_price=100,
                cost_price=60,
                total_price=100,
            )
            created_orders.append(order)

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 2)
        self.assertEqual(payload['timeline'][0]['period_key'], today.isoformat())
        self.assertEqual(
            {row['code'] for row in payload['order_details']},
            {created_orders[1].code, created_orders[2].code},
        )
        self.assertEqual(payload['filters_applied']['order_scope'], 'realized')

    def test_api_report_sales_filters_and_groups_revenue_by_exported_at(self):
        today = date.today()
        yesterday = today - timedelta(days=1)
        exported_today = Order.objects.create(
            code='DH-RP-EXPORTED-TODAY',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=4,
            total_amount=100,
            final_amount=100,
            order_date=yesterday,
            exported_at=datetime.combine(today, time(15, 30)),
            created_by=self.user,
        )
        order_date_today_but_exported_yesterday = Order.objects.create(
            code='DH-RP-ORDER-TODAY-EXPORT-YESTERDAY',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            total_amount=200,
            final_amount=200,
            order_date=today,
            exported_at=datetime.combine(yesterday, time(16, 0)),
            created_by=self.user,
        )
        for order, amount in (
            (exported_today, 100),
            (order_date_today_but_exported_yesterday, 200),
        ):
            OrderItem.objects.create(
                order=order,
                product=self.product,
                quantity=1,
                unit_price=amount,
                cost_price=60,
                total_price=amount,
            )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 1)
        self.assertEqual(payload['order_details'][0]['code'], exported_today.code)
        self.assertEqual(payload['order_details'][0]['date_raw'], today.isoformat())
        self.assertEqual(payload['timeline'][0]['period_key'], today.isoformat())

        staff_response = self.client.get(reverse('api_report_staff_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        self.assertEqual(staff_response.status_code, 200)
        staff_payload = staff_response.json()
        self.assertEqual(staff_payload['summary']['grand_revenue'], 100.0)
        self.assertEqual(staff_payload['summary']['grand_orders'], 1)
        self.assertEqual(staff_payload['staff_data'][0]['orders'][0]['code'], exported_today.code)
        self.assertEqual(
            staff_payload['staff_data'][0]['orders'][0]['date'],
            today.strftime('%d/%m/%Y'),
        )

    def test_sales_report_slow_moving_inventory_uses_latest_realized_sale_across_history(self):
        today = date.today()
        slow_product = Product.objects.create(
            store=self.store,
            code='SP-RP-SLOW-65',
            name='Sản phẩm chậm 65 ngày',
            cost_price=1000,
            created_by=self.user,
        )
        recent_product = Product.objects.create(
            store=self.store,
            code='SP-RP-RECENT-10',
            name='Sản phẩm mới bán',
            cost_price=1500,
            created_by=self.user,
        )
        never_sold_product = Product.objects.create(
            store=self.store,
            code='SP-RP-NEVER-SOLD',
            name='Sản phẩm chưa từng bán',
            import_price=2000,
            created_by=self.user,
        )
        new_never_sold_product = Product.objects.create(
            store=self.store,
            code='SP-RP-NEW-NEVER-SOLD',
            name='Sản phẩm mới tạo chưa bán',
            cost_price=500,
            created_by=self.user,
        )
        zero_stock_product = Product.objects.create(
            store=self.store,
            code='SP-RP-SLOW-ZERO',
            name='Sản phẩm chậm nhưng hết tồn',
            cost_price=3000,
            created_by=self.user,
        )
        ProductStock.objects.create(product=slow_product, warehouse=self.warehouse, quantity=5)
        ProductStock.objects.create(product=recent_product, warehouse=self.warehouse, quantity=2)
        ProductStock.objects.create(product=never_sold_product, warehouse=self.warehouse, quantity=3)
        ProductStock.objects.create(product=new_never_sold_product, warehouse=self.warehouse, quantity=1)
        ProductStock.objects.create(product=zero_stock_product, warehouse=self.warehouse, quantity=0)

        for code, product, days_ago, status in (
            ('DH-RP-SLOW-REALIZED', slow_product, 65, 5),
            ('DH-RP-SLOW-CANCELLED', slow_product, 1, 6),
            ('DH-RP-RECENT-REALIZED', recent_product, 10, 4),
            ('DH-RP-ZERO-REALIZED', zero_stock_product, 100, 5),
        ):
            order = Order.objects.create(
                code=code,
                store=self.store,
                customer=self.customer,
                warehouse=self.warehouse,
                status=status,
                total_amount=100,
                final_amount=100,
                order_date=today - timedelta(days=days_ago),
                created_by=self.user,
            )
            OrderItem.objects.create(
                order=order,
                product=product,
                quantity=1,
                unit_price=100,
                total_price=100,
            )

        response = self.client.get(reverse('api_report_sales'), {
            # Khoảng ngày báo cáo không chứa các đơn cũ nhưng cảnh báo vẫn phải
            # dùng toàn bộ lịch sử bán hàng.
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        rows = {row['product_code']: row for row in payload['slow_moving_products']}
        self.assertEqual(set(rows), {
            slow_product.code,
            recent_product.code,
            never_sold_product.code,
            new_never_sold_product.code,
        })
        self.assertEqual(rows[slow_product.code]['last_sale_date'], (today - timedelta(days=65)).isoformat())
        self.assertEqual(rows[slow_product.code]['days_without_sale'], 65)
        self.assertEqual(rows[slow_product.code]['warning_level'], 'slow')
        self.assertEqual(rows[recent_product.code]['days_without_sale'], 10)
        self.assertEqual(rows[recent_product.code]['warning_level'], 'normal')
        self.assertTrue(rows[never_sold_product.code]['never_sold'])
        self.assertIsNone(rows[never_sold_product.code]['days_without_sale'])
        self.assertEqual(rows[never_sold_product.code]['valuation_source'], 'import_price')
        self.assertEqual(rows[never_sold_product.code]['stock_value'], 6000.0)
        self.assertTrue(rows[new_never_sold_product.code]['never_sold'])
        self.assertIsNone(rows[new_never_sold_product.code]['days_without_sale'])
        self.assertEqual(rows[new_never_sold_product.code]['warning_level'], 'critical')
        self.assertNotIn(zero_stock_product.code, rows)
        self.assertEqual(payload['slow_moving_summary']['total_products'], 4)
        self.assertEqual(payload['slow_moving_summary']['never_sold_count'], 2)
        self.assertEqual(payload['slow_moving_summary']['over_30_count'], 1)
        self.assertEqual(payload['slow_moving_summary']['over_60_count'], 1)
        self.assertEqual(payload['slow_moving_summary']['over_90_count'], 0)
        self.assertEqual(payload['slow_moving_summary']['over_90_stock_value'], 0)

        export_response = self.client.get(reverse('export_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        self.assertIn('Cảnh báo hàng chậm', workbook.sheetnames)
        slow_sheet = workbook['Cảnh báo hàng chậm']
        exported_codes = {
            slow_sheet.cell(row=row, column=2).value
            for row in range(5, slow_sheet.max_row + 1)
        }
        self.assertEqual(exported_codes, {
            slow_product.code,
            never_sold_product.code,
            new_never_sold_product.code,
        })

    def test_sales_report_daily_date_opens_filtered_order_list_in_new_tab(self):
        response = self.client.get(reverse('report_sales'))

        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertContains(response, f'{reverse("product_guide")}#sales-report-guide')
        self.assertContains(response, '<i class="fas fa-book-open"></i> Hướng dẫn', html=True)
        self.assertContains(response, 'id="supplier_sales_section"')
        self.assertContains(response, 'Báo cáo bán hàng theo nhà cung cấp')
        self.assertContains(response, 'id="supplierSalesCollapse"')
        self.assertContains(response, 'data-target="#supplierSalesCollapse"')
        self.assertContains(response, 'supplier-sales-collapse-toggle collapsed')
        self.assertContains(response, 'aria-expanded="false"')
        self.assertContains(response, 'order-detail-collapse-toggle')
        self.assertContains(response, 'order-detail-collapse-icon')
        self.assertContains(response, 'data-target="#orderDetailCollapse"')
        self.assertContains(response, 'id="supplier_sales_tbl"')
        self.assertContains(response, 'id="supplier_sales_tbl" style="width:100%;min-width:1320px;"')
        self.assertContains(response, 'id="chart_supplier_consumption"')
        self.assertContains(response, 'id="chart_supplier_revenue"')
        self.assertContains(response, 'id="daily_chart_column"')
        self.assertNotContains(response, 'id="chart_products"')
        self.assertNotContains(response, 'function renderProductsPieChart(tp)')
        self.assertContains(response, 'Tỷ suất lợi nhuận')
        self.assertContains(response, 'id="ft_profit_margin"')
        self.assertContains(response, 'id="supplier_sales_filter"')
        self.assertContains(response, 'Phần giảm giá chung của đơn tính vào sản phẩm này')
        self.assertContains(response, 'Đây là lần gửi lại/đổi hàng')
        self.assertContains(response, 'Không nên kết luận nhân viên bán dưới giá vốn')
        self.assertContains(
            response,
            'id="profit_margin_badge" class="ml-1 font-weight-bold"',
        )
        self.assertContains(response, 'id="gp_margin" class="ml-1 font-weight-bold"')
        self.assertNotContains(response, 'id="profit_margin_badge" class="badge')
        self.assertNotContains(response, 'id="gp_margin" class="badge')
        self.assertContains(response, 'class="loss-box-action-row"')
        self.assertContains(response, 'row mb-3 overview-info-row')
        self.assertContains(response, 'id="loss_box_col"')
        self.assertContains(
            response,
            '.loss-box-action-row{display:flex;align-items:baseline;gap:.45rem;white-space:nowrap;}',
        )
        self.assertContains(response, "$('#loss_box_col').removeClass('d-none')")
        self.assertContains(response, "$('#loss_box_col').addClass('d-none')")
        self.assertContains(response, 'row mb-3 sales-metric-row', count=2)
        self.assertContains(
            response,
            '.sales-metric-row .small-box{display:flex;width:100%;height:100%;min-height:90px;',
        )
        self.assertContains(
            response,
            'justify-content:center;width:100%;padding:9px 10px 4px;',
        )
        self.assertContains(response, 'id="btn_clear_supplier_filter"')
        self.assertContains(response, 'id="supplier_col_config_container"')
        self.assertContains(response, '/static/js/column_config.js')
        self.assertContains(response, f'ifshop_supplier_sales_columns_v1_user_{self.user.id}')
        self.assertContains(response, 'data-col="net_quantity"')
        self.assertContains(response, 'data-col="net_revenue"')
        self.assertContains(response, 'data-col="top_products"')
        self.assertContains(response, 'window.supplierSalesColConfig.apply()')
        self.assertContains(response, 'function captureSupplierSalesFullWidth()')
        self.assertContains(response, 'minWidth:_supplierSalesStableWidth')
        self.assertContains(response, "width:_supplierSalesStableWidth+'px'")
        self.assertContains(response, 'Tiêu thụ ròng = Số lượng đã bán − Số lượng khách trả')
        self.assertContains(response, 'Lợi nhuận = Doanh thu thuần − Giá vốn thuần')
        self.assertNotContains(response, 'Cho biết hàng của nhà cung cấp nào được bán ra thực tế nhiều nhất')
        self.assertNotContains(response, 'Bấm vào thanh hoặc tên nhà cung cấp để lọc bảng chi tiết')
        self.assertContains(response, 'Công thức: Đã bán − Khách trả')
        self.assertContains(response, 'Biên lợi nhuận: ')
        self.assertContains(response, 'function renderSupplierBreakdown(rows,summary)')
        self.assertContains(response, 'function applySupplierFilter(key,scrollToTable)')
        self.assertContains(response, 'function bindSupplierChartInteraction(canvas,chart,rows)')
        self.assertLess(html.index('id="report_tbl"'), html.index('id="supplier_sales_section"'))
        self.assertLess(html.index('id="report_tbl"'), html.index('id="chart_supplier_consumption"'))
        self.assertLess(html.index('id="chart_supplier_consumption"'), html.index('id="supplier_sales_section"'))
        self.assertLess(html.index('id="supplier_sales_section"'), html.index('id="store_breakdown_section"'))
        self.assertContains(response, 'function getDailyOrdersUrl(dateKey)')
        self.assertContains(response, "'/order-tbl/?from_date='")
        self.assertContains(response, 'renderDailyOrderDate(d)')
        self.assertContains(response, 'sales-daily-order-link')
        self.assertContains(response, 'target="_blank" rel="noopener"')
        self.assertContains(response, 'id="top_products_limit"')
        self.assertContains(response, '<option value="10" selected>10 sản phẩm</option>', html=True)
        self.assertNotContains(response, '>9 sản phẩm</option>')
        self.assertContains(response, '200 sản phẩm')
        self.assertContains(response, 'table-responsive overview-ranking-scroll', count=2)
        self.assertContains(response, 'style="max-height:480px;overflow-y:auto;"')
        self.assertContains(response, 'style="max-height:440px;overflow-y:auto;"')
        self.assertContains(response, '.overview-ranking-scroll th,.overview-ranking-scroll td{white-space:nowrap;}')
        self.assertContains(response, 'id="top_customers_limit"')
        self.assertContains(response, '<option value="10" selected>10 khách hàng</option>', html=True)
        self.assertNotContains(response, '>9 khách hàng</option>')
        self.assertContains(response, '200 khách hàng')
        for paginated_table_id in (
            'report_tbl',
            'top_products_tbl',
            'top_customers_tbl',
            'supplier_sales_tbl',
            'store_breakdown_tbl',
            'order_detail_tbl',
            'daily_finance_tbl',
            'product_detail_tbl',
            'category_detail_tbl',
            'sku_detail_tbl',
            'customer_detail_tbl',
            'customer_kind_tbl',
            'customer_group_tbl',
            'staff_detail_tbl',
            'order_report_tbl',
            'return_order_tbl',
            'return_product_tbl',
            'profit_category_tbl',
            'slow_moving_tbl',
        ):
            self.assertContains(response, f'id="{paginated_table_id}"')
        self.assertContains(response, 'Hàng bị trả lại')
        self.assertContains(response, 'id="df_ft_returns"')
        self.assertContains(response, 'tret+=returns')
        self.assertContains(response, 'Tỉ suất LN gộp')
        self.assertContains(response, 'id="pd_ft_gross_profit"')
        self.assertContains(response, 'id="pd_ft_gross_margin"')
        self.assertContains(response, 'p.gross_profit')
        self.assertContains(response, 'p.gross_margin')
        self.assertContains(
            response,
            'class="app-sort-toggle-btn product-detail-sort-btn"',
            count=2,
        )
        self.assertContains(response, 'data-sort-field="gross_profit"')
        self.assertContains(response, 'data-sort-field="gross_margin"')
        self.assertContains(response, 'function sortProductDetailRows(rows)')
        self.assertContains(response, 'function syncProductDetailSortButtons()')
        self.assertContains(
            response,
            "_productDetailSortDirections={gross_profit:'desc',gross_margin:'desc'}",
        )
        self.assertContains(
            response,
            "_productDetailSortDirections[field]=_productDetailSortDirections[field]==='desc'?'asc':'desc'",
        )
        self.assertContains(response, 'id="cat_ft_gross_profit"')
        self.assertContains(response, 'id="cat_ft_gross_margin"')
        self.assertContains(
            response,
            'class="app-sort-toggle-btn category-detail-sort-btn"',
            count=2,
        )
        self.assertContains(response, 'function sortCategoryDetailRows(rows)')
        self.assertContains(response, 'function syncCategoryDetailSortButtons()')
        self.assertContains(
            response,
            "_categoryDetailSortDirections={gross_profit:'desc',gross_margin:'desc'}",
        )
        self.assertContains(
            response,
            "_categoryDetailSortDirections[field]=_categoryDetailSortDirections[field]==='desc'?'asc':'desc'",
        )
        self.assertContains(
            response,
            'class="app-sort-toggle-btn customer-detail-sort-btn"',
            count=2,
        )
        self.assertContains(response, 'data-sort-field="revenue"')
        self.assertContains(response, 'data-sort-field="profit"')
        self.assertContains(response, 'function sortCustomerDetailRows(rows)')
        self.assertContains(response, 'function syncCustomerDetailSortButtons()')
        self.assertContains(
            response,
            "_customerDetailSortDirections={revenue:'desc',profit:'desc'}",
        )
        self.assertContains(
            response,
            "_customerDetailSortDirections[field]=_customerDetailSortDirections[field]==='desc'?'asc':'desc'",
        )
        self.assertContains(
            response,
            'class="app-sort-toggle-btn order-report-sort-btn"',
            count=2,
        )
        self.assertContains(response, 'function sortOrderReportRows(rows)')
        self.assertContains(response, 'function syncOrderReportSortButtons()')
        self.assertContains(
            response,
            "_orderReportSortDirections={revenue:'desc',profit:'desc'}",
        )
        self.assertContains(
            response,
            "_orderReportSortDirections[field]=_orderReportSortDirections[field]==='desc'?'asc':'desc'",
        )
        self.assertContains(response, 'var SALES_TABLE_DEFAULT_PAGE_SIZE=25;')
        self.assertContains(response, 'var SALES_TABLE_PAGE_SIZE_OPTIONS=[25,50,100,200];')
        self.assertContains(response, 'function initSalesTablePagination()')
        self.assertContains(
            response,
            "$('#reportTabs').siblings('.tab-content').first().find('table[id]')",
        )
        self.assertContains(response, 'new MutationObserver(function(mutations)')
        self.assertContains(response, 'class="sales-table-pagination"')
        self.assertContains(response, 'sales-table-pagination-summary')
        self.assertContains(response, 'sales-table-page-size')
        self.assertContains(response, 'sales-table-page-btn')
        self.assertContains(response, 'initSalesTablePagination();')
        self.assertContains(response, 'class="sales-table-column-config"')
        self.assertContains(response, 'sales-table-column-header')
        self.assertContains(
            response,
            "$(table).closest('.card').children('.card-header').first()",
        )
        self.assertContains(response, 'function getSalesTableColumns(tableId)')
        self.assertContains(response, 'function assignSalesTableColumnKeys(tableId,columns)')
        self.assertContains(response, 'function syncSalesTableColumnSpans(tableId,columns,state)')
        self.assertContains(response, 'function initSalesTableColumnConfigs()')
        self.assertContains(response, "alwaysOn:index===0")
        self.assertContains(
            response,
            "storageKey:'ifshop_sales_columns_'+tableId+'_user_",
        )
        self.assertContains(response, "tableId==='supplier_sales_tbl'")
        self.assertContains(response, 'initSalesTableColumnConfigs();')
        self.assertContains(response, "var productLimit=parseInt($('#top_products_limit').val()||'10',10)||10;")
        self.assertContains(response, "var customerLimit=parseInt($('#top_customers_limit').val()||'10',10)||10;")
        self.assertContains(response, 'renderOverviewRankings();')
        self.assertContains(response, '(_overviewProductRows||[]).slice(0,productLimit)')
        self.assertContains(response, '(_overviewCustomerRows||[]).slice(0,customerLimit)')
        self.assertContains(response, 'id="order_detail_from_date"')
        self.assertContains(response, 'id="order_detail_to_date"')
        self.assertContains(response, 'function setOrderDetailCurrentMonth()')
        self.assertContains(response, 'new Date(today.getFullYear(),today.getMonth()+1,0)')
        self.assertContains(response, 'function getDateFilteredOrderDetails()')
        self.assertContains(response, "orderDate<from")
        self.assertContains(response, "orderDate>to")
        self.assertContains(response, "$('#orderDetailCollapse').on('shown.bs.collapse', refreshOrderDetails)")
        self.assertContains(response, "$('#order_detail_tbl tbody').empty()")
        self.assertContains(response, 'href="#tab_slow_moving"')
        self.assertContains(response, 'Hàng bán chậm')
        self.assertContains(response, 'id="btn_sales_tab_config"')
        self.assertContains(response, '<i class="fas fa-eye mr-1"></i>Ẩn/Hiện', html=True)
        self.assertNotContains(response, 'id="modal_sales_tab_config"')
        self.assertContains(
            response,
            'class="sales-report-tabs-layout d-flex flex-wrap align-items-stretch mb-3"',
        )
        self.assertContains(
            response,
            'class="dropdown-menu dropdown-menu-right cc-menu sales-tab-config-menu"',
        )
        self.assertContains(response, '<span>Hiển thị tab</span>', html=True)
        self.assertContains(
            response,
            '.sales-report-tabs-layout>#reportTabs{order:1;flex:1 1 0;min-width:0;}',
        )
        self.assertContains(
            response,
            'top:calc(100% + 4px)!important;left:auto!important;right:0!important;transform:none!important;',
        )
        self.assertContains(response, 'DT theo ngày', count=2)
        self.assertContains(response, 'DT theo hàng', count=2)
        self.assertContains(response, 'class="sales-tab-visibility-toggle"', count=10)
        self.assertContains(response, 'id="btn_reset_sales_tabs"')
        self.assertContains(response, "$('.sales-tab-config-menu').on('click',function(event)")
        self.assertContains(response, 'event.stopPropagation();')
        self.assertContains(response, 'var SALES_REPORT_TAB_IDS=[')
        self.assertContains(response, "var SALES_REPORT_TAB_STORAGE_KEY='ifshop_sales_visible_tabs_v1_user_")
        self.assertContains(response, 'function loadSalesReportTabVisibility()')
        self.assertContains(response, 'function applySalesReportTabVisibility()')
        self.assertContains(response, 'function showSalesReportTab(tabId)')
        self.assertContains(response, 'visibleCount===1')
        self.assertContains(response, "showSalesReportTab('tab_orders')")
        self.assertContains(response, 'id="slow_moving_threshold"')
        self.assertContains(response, '<option value="30" selected>Chậm từ 30 ngày</option>', html=True)
        self.assertContains(response, 'id="slow_moving_search"')
        self.assertContains(response, 'id="slow_moving_tbl"')
        self.assertContains(response, 'data-sort="days_without_sale"')
        self.assertContains(response, 'data-sort="stock"')
        self.assertContains(response, 'data-sort="stock_value"')
        self.assertContains(response, 'class="text-right app-sortable-heading"', count=11)
        self.assertContains(response, 'class="app-sort-toggle-btn slow-moving-sort"', count=3)
        self.assertContains(response, 'fa-sort-amount-up')
        self.assertContains(response, 'fa-sort-amount-down')
        self.assertContains(response, "_slowMovingSort.direction==='asc'?'desc':'asc'")
        self.assertNotContains(response, '<th>Mức cảnh báo</th>')
        self.assertContains(response, 'Ngày chưa bán')
        self.assertContains(response, 'Hàng chưa từng bán được hiển thị riêng và không quy đổi theo ngày tạo')
        self.assertContains(response, 'không bị giới hạn bởi khoảng ngày báo cáo')
        self.assertContains(response, 'renderSlowMovingTab(res.slow_moving_products||[], res.slow_moving_summary||{})')
        self.assertContains(response, 'function renderSlowMovingRows()')
        self.assertLess(html.index('href="#tab_profit"'), html.index('href="#tab_slow_moving"'))


    def test_api_report_sales_all_active_scope_includes_non_cancelled_orders(self):
        today = date.today()
        expected_codes = set()
        for status, suffix in ((1, 'ORDER'), (3, 'PACK'), (4, 'EXPORTED'), (5, 'DONE'), (6, 'CANCELLED')):
            order = Order.objects.create(
                code=f'DH-RP-ALL-{suffix}',
                store=self.store,
                customer=self.customer,
                warehouse=self.warehouse,
                status=status,
                total_amount=100,
                final_amount=100,
                order_date=today,
                created_by=self.user,
            )
            OrderItem.objects.create(
                order=order,
                product=self.product,
                quantity=1,
                unit_price=100,
                cost_price=60,
                total_price=100,
            )
            if status != 6:
                expected_codes.add(order.code)

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'order_scope': 'all_active',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 4)
        self.assertEqual({row['code'] for row in payload['order_details']}, expected_codes)
        self.assertEqual(payload['filters_applied']['order_scope'], 'all_active')

    def test_api_report_sales_counts_linked_returns(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-001',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            salesperson='Nhân viên A',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=60,
            total_price=100,
        )
        OrderReturn.objects.create(
            code='TH-RP-001',
            order=order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            total_refund=25,
            return_date=today,
            created_by=self.user,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(payload['summary']['total_returns'], 25.0)
        self.assertEqual(payload['summary']['returns_count'], 1)
        self.assertEqual(len(payload['return_orders']), 1)
        self.assertEqual(payload['return_orders'][0]['order_code'], order.code)

    def test_api_report_sales_does_not_warn_for_fully_returned_loss_order(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-FULL-RETURN-LOSS',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            total_amount=100,
            final_amount=100,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=130,
            total_price=100,
        )
        order_return = OrderReturn.objects.create(
            code='TH-RP-FULL-RETURN-LOSS',
            order=order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            total_refund=100,
            return_date=today,
            created_by=self.user,
        )
        OrderReturnItem.objects.create(
            order_return=order_return,
            product=self.product,
            quantity=1,
            unit_price=100,
            total_price=100,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload['order_details'] if item['code'] == order.code)
        self.assertFalse(row['is_loss'])
        self.assertEqual(row['loss_products'], [])
        self.assertEqual(payload['summary']['loss_count'], 0)

        loss_response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'profit_filter': 'loss',
        })
        self.assertEqual(loss_response.status_code, 200)
        loss_payload = loss_response.json()
        self.assertEqual(loss_payload['summary']['loss_count'], 0)
        self.assertNotIn(order.code, [item['code'] for item in loss_payload['order_details']])

    def test_api_report_sales_uses_value_only_return_for_legacy_full_return(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-LEGACY-FULL-RETURN',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            total_amount=100,
            final_amount=100,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=130,
            total_price=100,
        )
        OrderReturn.objects.create(
            code='TH-RP-LEGACY-FULL-RETURN',
            order=order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            return_amount=100,
            total_refund=100,
            return_date=today,
            created_by=self.user,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload['order_details'] if item['code'] == order.code)
        self.assertFalse(row['is_loss'])
        self.assertEqual(row['loss_products'], [])

    def test_api_report_sales_keeps_orphan_return_with_scope_fallback(self):
        today = date.today()
        OrderReturn.objects.create(
            code='TH-RP-ORPHAN',
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            total_refund=30,
            return_date=today,
            created_by=self.user,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(payload['summary']['total_returns'], 30.0)
        self.assertEqual(payload['summary']['returns_count'], 1)
        self.assertEqual(len(payload['return_orders']), 1)
        self.assertEqual(payload['return_orders'][0]['order_code'], '(Thiếu đơn gốc)')
        self.assertEqual(payload['return_orders'][0]['store_name'], self.store.name)

    def test_api_report_sales_allocates_order_level_discount_to_item_scope(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-DISCOUNT',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            discount_amount=20,
            final_amount=80,
            paid_amount=80,
            order_date=today,
            salesperson='Nhân viên A',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=60,
            total_price=100,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'product_id': self.product.id,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(payload['summary']['total_revenue'], 80.0)
        self.assertEqual(payload['summary']['total_profit'], 20.0)
        self.assertEqual(payload['order_details'][0]['revenue'], 80.0)
        self.assertEqual(payload['top_products'][0]['amount'], 80.0)

    def test_api_report_sales_includes_order_other_fee_in_revenue_and_profit(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-OTHER-FEE',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            discount_amount=10,
            shipping_fee=5,
            other_fee=20,
            final_amount=115,
            paid_amount=115,
            order_date=today,
            salesperson='Nhân viên A',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=60,
            total_price=100,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload['order_details'] if item['id'] == order.id)
        self.assertEqual(row['goods_amount'], 100.0)
        self.assertEqual(row['discount_amount'], 10.0)
        self.assertEqual(row['shipping_fee'], 5.0)
        self.assertEqual(row['other_fee'], 20.0)
        self.assertEqual(row['revenue'], 115.0)
        self.assertEqual(row['profit'], 55.0)

    def test_api_report_sales_falls_back_when_legacy_order_item_cost_is_zero(self):
        today = date.today()
        self.product.cost_price = 60
        self.product.import_price = 65
        self.product.save(update_fields=['cost_price', 'import_price'])
        order = Order.objects.create(
            code='DH-RP-LEGACY-ZERO-COST',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            salesperson='Nhân viên A',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=0,
            total_price=100,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload['order_details'] if item['id'] == order.id)
        self.assertEqual(row['cost'], 60.0)
        self.assertEqual(row['profit'], 40.0)

    def test_api_report_sales_includes_sapo_style_sku_details(self):
        today = date.today()
        variant = ProductVariant.objects.create(
            product=self.product,
            size_name='Size A',
            sku='SKU-RP-001-A',
        )
        seller = User.objects.create_user(
            username='sku_report_seller',
            password='pass123',
            first_name='Minh',
            last_name='Ban Hang',
        )
        order = Order.objects.create(
            code='DH-RP-SKU',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=200,
            discount_amount=20,
            final_amount=180,
            paid_amount=180,
            order_date=today,
            salesperson='Nhân viên SKU',
            created_by=seller,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            variant=variant,
            quantity=2,
            unit_price=100,
            cost_price=70,
            total_price=200,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(len(payload['sku_details']), 1)
        row = payload['sku_details'][0]
        self.assertEqual(row['date'], today.strftime('%d/%m/%Y'))
        self.assertEqual(row['customer'], self.customer.name)
        self.assertEqual(row['product_name'], self.product.name)
        self.assertEqual(row['sku'], 'SKU-RP-001-A')
        self.assertEqual(row['order_code'], order.code)
        self.assertEqual(row['salesperson'], 'Nhân viên SKU')
        self.assertEqual(row['revenue'], 180.0)
        self.assertEqual(row['cost'], 140.0)
        self.assertEqual(row['profit'], 40.0)

    def test_api_report_sales_includes_daily_finance_summary(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-DAILY',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=200,
            discount_amount=20,
            final_amount=180,
            paid_amount=180,
            order_date=today,
            salesperson='Nhân viên ngày',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=100,
            cost_price=70,
            total_price=200,
        )
        order_return = OrderReturn.objects.create(
            code='TH-RP-DAILY',
            order=order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            total_refund=30,
            return_date=today,
            created_by=self.user,
        )
        OrderReturnItem.objects.create(
            order_return=order_return,
            product=self.product,
            quantity=1,
            unit_price=30,
            total_price=30,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(payload['summary']['total_goods_amount'], 200.0)
        self.assertEqual(payload['summary']['total_net_revenue'], 150.0)
        self.assertEqual(payload['summary']['total_sales_cost'], 140.0)
        self.assertEqual(payload['summary']['total_return_cost'], 70.0)
        self.assertEqual(payload['summary']['total_net_cost'], 70.0)
        self.assertEqual(payload['summary']['total_gross_profit'], 80.0)
        self.assertEqual(payload['summary']['gross_margin'], 53.3)
        self.assertEqual(len(payload['daily_finance']), 1)
        row = payload['daily_finance'][0]
        self.assertEqual(row['date'], today.strftime('%d/%m/%Y'))
        self.assertEqual(row['goods_amount'], 200.0)
        self.assertEqual(row['revenue'], 180.0)
        self.assertEqual(row['returns'], 30.0)
        self.assertEqual(row['net_revenue'], 150.0)
        self.assertEqual(row['gross_cost'], 140.0)
        self.assertEqual(row['return_cost'], 70.0)
        self.assertEqual(row['cost'], 70.0)
        self.assertEqual(row['gross_profit'], 80.0)
        self.assertEqual(row['gross_margin'], 53.3)
        self.assertEqual(row['net_profit'], 80.0)
        self.assertEqual(payload['daily'][0]['profit_margin'], 53.3)
        product_row = payload['product_breakdown'][0]
        self.assertEqual(product_row['line_profit'], 40.0)
        self.assertEqual(product_row['returns_amount'], 30.0)
        self.assertEqual(product_row['return_cost'], 70.0)
        self.assertEqual(product_row['net_revenue'], 150.0)
        self.assertEqual(product_row['net_cost'], 70.0)
        self.assertEqual(product_row['gross_profit'], 80.0)
        self.assertEqual(product_row['gross_margin'], 53.3)
        category_row = payload['category_breakdown'][0]
        self.assertEqual(category_row['returns_amount'], 30.0)
        self.assertEqual(category_row['return_cost'], 70.0)
        self.assertEqual(category_row['net_revenue'], 150.0)
        self.assertEqual(category_row['net_cost'], 70.0)
        self.assertEqual(category_row['gross_profit'], 80.0)
        self.assertEqual(category_row['gross_margin'], 53.3)

        export_response = self.client.get(reverse('export_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        daily_sheet = workbook['Tổng hợp ngày']
        self.assertEqual(daily_sheet.cell(row=1, column=4).value, 'Hàng bị trả lại')
        self.assertEqual(daily_sheet.cell(row=2, column=4).value, 30)
        self.assertEqual(daily_sheet.cell(row=3, column=4).value, 30)
        product_sheet = workbook['Mặt hàng']
        self.assertEqual(product_sheet.cell(row=1, column=10).value, 'Lợi nhuận gộp')
        self.assertEqual(product_sheet.cell(row=1, column=11).value, 'Tỉ suất LN gộp')
        self.assertEqual(product_sheet.cell(row=2, column=10).value, 80)
        self.assertAlmostEqual(product_sheet.cell(row=2, column=11).value, 0.533)
        category_sheet = workbook['Nhóm mặt hàng']
        self.assertEqual(category_sheet.cell(row=1, column=8).value, 'Lợi nhuận gộp')
        self.assertEqual(category_sheet.cell(row=1, column=9).value, 'Tỉ suất LN gộp')
        self.assertEqual(category_sheet.cell(row=2, column=8).value, 80)
        self.assertAlmostEqual(category_sheet.cell(row=2, column=9).value, 0.533)

    def test_api_report_sales_groups_multiple_products_and_returns_by_supplier(self):
        today = date.today()
        supplier_a = Supplier.objects.create(code='NCC-RP-SALES-A', name='NCC bán hàng A')
        supplier_b = Supplier.objects.create(code='NCC-RP-SALES-B', name='NCC bán hàng B')
        product_a1 = Product.objects.create(
            store=self.store,
            supplier=supplier_a,
            code='SP-RP-NCC-A1',
            name='Sản phẩm NCC A1',
            created_by=self.user,
        )
        product_a2 = Product.objects.create(
            store=self.store,
            supplier=supplier_a,
            code='SP-RP-NCC-A2',
            name='Sản phẩm NCC A2',
            created_by=self.user,
        )
        product_b = Product.objects.create(
            store=self.store,
            supplier=supplier_b,
            code='SP-RP-NCC-B',
            name='Sản phẩm NCC B',
            created_by=self.user,
        )
        first_order = Order.objects.create(
            code='DH-RP-NCC-1',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            total_amount=550,
            final_amount=550,
            order_date=today,
            created_by=self.user,
        )
        for product, quantity, total_price, cost_price in (
            (product_a1, 2, 200, 60),
            (product_a2, 3, 150, 20),
            (product_b, 1, 200, 100),
        ):
            OrderItem.objects.create(
                order=first_order,
                product=product,
                quantity=quantity,
                unit_price=total_price / quantity,
                cost_price=cost_price,
                total_price=total_price,
            )
        second_order = Order.objects.create(
            code='DH-RP-NCC-2',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            total_amount=100,
            final_amount=100,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=second_order,
            product=product_a1,
            quantity=1,
            unit_price=100,
            cost_price=60,
            total_price=100,
        )
        order_return = OrderReturn.objects.create(
            code='TH-RP-NCC-1',
            order=first_order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            total_refund=50,
            return_date=today,
            created_by=self.user,
        )
        OrderReturnItem.objects.create(
            order_return=order_return,
            product=product_a2,
            quantity=1,
            unit_price=50,
            total_price=50,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        rows = response.json()['supplier_breakdown']
        supplier_summary = response.json()['supplier_summary']
        self.assertEqual([row['supplier'] for row in rows], [supplier_a.name, supplier_b.name])
        self.assertEqual(supplier_summary['supplier_count'], 2)
        self.assertEqual(supplier_summary['product_count'], 3)
        self.assertEqual(supplier_summary['order_count'], 2)
        supplier_a_row = rows[0]
        self.assertEqual(supplier_a_row['product_count'], 2)
        self.assertEqual(supplier_a_row['order_count'], 2)
        self.assertEqual(supplier_a_row['sold_quantity'], 6.0)
        self.assertEqual(supplier_a_row['returned_quantity'], 1.0)
        self.assertEqual(supplier_a_row['net_quantity'], 5.0)
        self.assertEqual(supplier_a_row['net_revenue'], 400.0)
        self.assertEqual(supplier_a_row['cost'], 220.0)
        self.assertEqual(supplier_a_row['profit'], 180.0)
        self.assertEqual(supplier_a_row['contribution'], 66.7)
        self.assertEqual(supplier_a_row['top_products'][0]['name'], product_a1.name)
        self.assertEqual(supplier_a_row['top_products'][0]['net_quantity'], 3.0)

        export_response = self.client.get(reverse('export_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        self.assertIn('Bán hàng theo NCC', workbook.sheetnames)
        supplier_sheet = workbook['Bán hàng theo NCC']
        self.assertEqual(supplier_sheet['B2'].value, supplier_a.name)
        self.assertEqual(supplier_sheet['G2'].value, 5)
        self.assertEqual(supplier_sheet['H2'].value, 400)

    def test_api_report_sales_filter_options_include_store_users_without_orders(self):
        today = date.today()
        seller = User.objects.create_user(
            username='seller_report',
            password='pass123',
            first_name='Lan',
            last_name='Nguyen',
        )
        UserProfile.objects.create(user=seller, store=self.store)

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertIn('Lan Nguyen', payload['filter_options']['salespersons'])

    def test_api_report_sales_attributes_revenue_to_creator_unless_salesperson_is_assigned(self):
        today = date.today()
        ngoc = User.objects.create_user(
            username='ngoc_sales_report',
            password='pass123',
            first_name='Ngọc',
        )
        UserProfile.objects.create(user=ngoc, store=self.store)
        Order.objects.create(
            code='DH-RP-CREATOR-NGOC',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            creator_name='Tên lưu cũ không đúng',
            created_by=ngoc,
        )
        Order.objects.create(
            code='DH-RP-ASSIGNED-STAFF',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=250,
            final_amount=250,
            paid_amount=250,
            order_date=today,
            salesperson='Lan được gán',
            created_by=self.user,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        staff_rows = {
            row['salesperson']: row
            for row in payload['staff_breakdown']
        }
        self.assertEqual(staff_rows['Ngọc']['order_count'], 1)
        self.assertEqual(staff_rows['Ngọc']['revenue'], 100.0)
        self.assertEqual(staff_rows['Lan được gán']['order_count'], 1)
        self.assertEqual(staff_rows['Lan được gán']['revenue'], 250.0)
        self.assertNotIn('(Chưa gán NV)', staff_rows)
        self.assertEqual(
            {
                row['code']: row['salesperson']
                for row in payload['order_details']
            },
            {
                'DH-RP-ASSIGNED-STAFF': 'Lan được gán',
                'DH-RP-CREATOR-NGOC': 'Ngọc',
            },
        )

        filtered_response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'salesperson': 'Ngọc',
        })
        self.assertEqual(filtered_response.status_code, 200)
        filtered_payload = filtered_response.json()
        self.assertEqual(filtered_payload['summary']['total_orders'], 1)
        self.assertEqual(filtered_payload['summary']['total_revenue'], 100.0)
        self.assertEqual(
            [row['salesperson'] for row in filtered_payload['staff_breakdown']],
            ['Ngọc'],
        )

        staff_report_response = self.client.get(reverse('api_report_staff_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        self.assertEqual(staff_report_response.status_code, 200)
        staff_report_rows = {
            row['salesperson']: row
            for row in staff_report_response.json()['staff_data']
        }
        self.assertEqual(staff_report_rows['Ngọc']['order_count'], 1)
        self.assertEqual(staff_report_rows['Ngọc']['revenue'], 100.0)
        self.assertEqual(staff_report_rows['Lan được gán']['order_count'], 1)
        self.assertEqual(staff_report_rows['Lan được gán']['revenue'], 250.0)
        self.assertNotIn('(Chưa gán NV)', staff_report_rows)

        export_response = self.client.get(reverse('export_staff_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        sheet = workbook['BC Doanh thu NV']
        exported_revenue = {
            sheet.cell(row=row, column=2).value: sheet.cell(row=row, column=4).value
            for row in range(5, sheet.max_row)
        }
        self.assertEqual(exported_revenue['Ngọc'], 100)
        self.assertEqual(exported_revenue['Lan được gán'], 250)

    def test_api_report_sales_filters_customer_kind_wholesale(self):
        today = date.today()
        wholesale_group = CustomerGroup.objects.create(name='Khách sỉ')
        retail_group = CustomerGroup.objects.create(name='Khách lẻ')
        wholesale_customer = Customer.objects.create(
            store=self.store,
            code='KH-RP-SI',
            name='Khách mua sỉ',
            group=wholesale_group,
            created_by=self.user,
        )
        retail_customer = Customer.objects.create(
            store=self.store,
            code='KH-RP-LE',
            name='Khách mua lẻ',
            group=retail_group,
            created_by=self.user,
        )

        wholesale_order = Order.objects.create(
            code='DH-RP-SI',
            store=self.store,
            customer=wholesale_customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=wholesale_order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=60,
            total_price=100,
        )
        retail_order = Order.objects.create(
            code='DH-RP-LE',
            store=self.store,
            customer=retail_customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=80,
            final_amount=80,
            paid_amount=80,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=retail_order,
            product=self.product,
            quantity=1,
            unit_price=80,
            cost_price=50,
            total_price=80,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'customer_kind': 'wholesale',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 1)
        self.assertEqual(payload['order_details'][0]['code'], wholesale_order.code)
        self.assertEqual(payload['order_details'][0]['customer_kind'], 'wholesale')
        self.assertEqual(payload['customer_kind_breakdown'][0]['name'], 'Khách buôn / sỉ')

    def test_api_report_sales_prefers_explicit_customer_kind_field(self):
        today = date.today()
        neutral_group = CustomerGroup.objects.create(name='VIP thân thiết')
        wholesale_customer = Customer.objects.create(
            store=self.store,
            code='KH-RP-EX-SI',
            name='Khách field sỉ',
            group=neutral_group,
            customer_kind=Customer.CUSTOMER_KIND_WHOLESALE,
            created_by=self.user,
        )
        retail_customer = Customer.objects.create(
            store=self.store,
            code='KH-RP-EX-LE',
            name='Khách field lẻ',
            group=neutral_group,
            customer_kind=Customer.CUSTOMER_KIND_RETAIL,
            created_by=self.user,
        )

        wholesale_order = Order.objects.create(
            code='DH-RP-EX-SI',
            store=self.store,
            customer=wholesale_customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=120,
            final_amount=120,
            paid_amount=120,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=wholesale_order,
            product=self.product,
            quantity=1,
            unit_price=120,
            cost_price=70,
            total_price=120,
        )
        retail_order = Order.objects.create(
            code='DH-RP-EX-LE',
            store=self.store,
            customer=retail_customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=80,
            final_amount=80,
            paid_amount=80,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=retail_order,
            product=self.product,
            quantity=1,
            unit_price=80,
            cost_price=40,
            total_price=80,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'customer_kind': 'wholesale',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 1)
        self.assertEqual(payload['order_details'][0]['code'], wholesale_order.code)
        self.assertEqual(payload['order_details'][0]['customer_kind'], 'wholesale')
        self.assertEqual(payload['filter_options']['customers'][0]['name'], wholesale_customer.name)

    def test_api_report_sales_root_category_filter_includes_child_type(self):
        today = date.today()
        root_category = ProductCategory.objects.create(name='Máy móc')
        product_type = ProductCategory.objects.create(name='Máy xay', parent=root_category)
        product = Product.objects.create(
            store=self.store,
            code='SP-RP-MAY-XAY',
            name='Máy xay sinh tố',
            category=product_type,
            created_by=self.user,
        )
        order = Order.objects.create(
            code='DH-RP-ROOT-CAT',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=200,
            final_amount=200,
            paid_amount=200,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=product,
            quantity=1,
            unit_price=200,
            cost_price=120,
            total_price=200,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'category_id': root_category.id,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 1)
        self.assertEqual(payload['product_breakdown'][0]['category'], root_category.name)
        self.assertEqual(payload['product_breakdown'][0]['product_type'], product_type.name)
        self.assertEqual(payload['category_breakdown'][0]['name'], root_category.name)

    def test_api_report_sales_product_type_filter_limits_child_category(self):
        today = date.today()
        root_category = ProductCategory.objects.create(name='Nhóm máy')
        selected_type = ProductCategory.objects.create(name='Máy ép', parent=root_category)
        other_type = ProductCategory.objects.create(name='Máy xay', parent=root_category)
        selected_product = Product.objects.create(
            store=self.store,
            code='SP-RP-MAY-EP',
            name='Máy ép',
            category=selected_type,
            created_by=self.user,
        )
        other_product = Product.objects.create(
            store=self.store,
            code='SP-RP-MAY-XAY-2',
            name='Máy xay khác',
            category=other_type,
            created_by=self.user,
        )
        selected_order = Order.objects.create(
            code='DH-RP-TYPE-1',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=300,
            final_amount=300,
            paid_amount=300,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=selected_order,
            product=selected_product,
            quantity=1,
            unit_price=300,
            cost_price=200,
            total_price=300,
        )
        other_order = Order.objects.create(
            code='DH-RP-TYPE-2',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=150,
            final_amount=150,
            paid_amount=150,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=other_order,
            product=other_product,
            quantity=1,
            unit_price=150,
            cost_price=90,
            total_price=150,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'product_type_id': selected_type.id,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 1)
        self.assertEqual(payload['order_details'][0]['code'], selected_order.code)
        self.assertEqual(payload['product_breakdown'][0]['product_type'], selected_type.name)

    def test_api_report_sales_filters_line_profit_and_shows_loss_order(self):
        today = date.today()
        loss_order = Order.objects.create(
            code='DH-RP-LOSS-LINE',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=loss_order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=130,
            total_price=100,
        )
        profit_order = Order.objects.create(
            code='DH-RP-PROFIT-LINE',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=120,
            final_amount=120,
            paid_amount=120,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=profit_order,
            product=self.product,
            quantity=1,
            unit_price=120,
            cost_price=80,
            total_price=120,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'line_profit_max': -1,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['total_orders'], 1)
        self.assertEqual(payload['order_details'][0]['code'], loss_order.code)
        self.assertTrue(payload['order_details'][0]['is_loss'])
        self.assertEqual(payload['order_details'][0]['loss_product_names'], self.product.name)
        self.assertEqual(len(payload['order_details'][0]['loss_products']), 1)
        loss_product = payload['order_details'][0]['loss_products'][0]
        self.assertEqual(loss_product['product_name'], self.product.name)
        self.assertEqual(loss_product['unit_price'], 100.0)
        self.assertEqual(loss_product['gross_line_amount'], 100.0)
        self.assertEqual(loss_product['line_discount_amount'], 0.0)
        self.assertEqual(loss_product['order_discount_allocated'], 0.0)
        self.assertEqual(loss_product['shipping_fee_allocated'], 0.0)
        self.assertEqual(loss_product['other_fee_allocated'], 0.0)
        self.assertEqual(loss_product['net_revenue'], 100.0)
        self.assertEqual(loss_product['unit_revenue'], 100.0)
        self.assertEqual(loss_product['unit_cost'], 130.0)
        self.assertEqual(loss_product['total_cost'], 130.0)
        self.assertEqual(loss_product['line_profit'], -30.0)
        self.assertEqual(loss_product['loss_amount'], 30.0)
        self.assertFalse(loss_product['is_exchange_order'])
        self.assertEqual(payload['summary']['loss_count'], 1)

    def test_api_report_sales_explains_loss_warning_for_exchange_order(self):
        today = date.today()
        source_order = Order.objects.create(
            code='DH-RP-EXCHANGE-SOURCE',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            created_by=self.user,
        )
        exchange_order = Order.objects.create(
            code='DH-RP-EXCHANGE-LOSS',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            discount_amount=100,
            final_amount=0,
            paid_amount=0,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=exchange_order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=130,
            total_price=100,
        )
        order_return = OrderReturn.objects.create(
            code='TH-RP-EXCHANGE-LOSS',
            order=source_order,
            exchange_order=exchange_order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            return_amount=100,
            exchange_amount=100,
            total_refund=0,
            amount_due=0,
            reason='Giao sai địa chỉ, gửi lại hàng',
            exchange_note='Đóng lại và gửi khách',
            return_date=today,
            created_by=self.user,
        )

        response = self.client.get(reverse('api_report_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'search': exchange_order.code,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = payload['order_details'][0]
        self.assertEqual(row['code'], exchange_order.code)
        self.assertTrue(row['is_loss'])
        warning = row['loss_products'][0]
        self.assertTrue(warning['is_exchange_order'])
        self.assertEqual(warning['return_code'], order_return.code)
        self.assertEqual(warning['source_order_code'], source_order.code)
        self.assertEqual(warning['return_amount'], 100.0)
        self.assertEqual(warning['exchange_amount'], 100.0)
        self.assertEqual(warning['exchange_offset_amount'], 100.0)
        self.assertEqual(warning['amount_due'], 0.0)
        self.assertEqual(warning['return_reason'], order_return.reason)
        self.assertEqual(warning['exchange_note'], order_return.exchange_note)
        self.assertEqual(warning['order_discount_allocated'], 100.0)
        self.assertEqual(warning['net_revenue'], 0.0)
        self.assertEqual(warning['unit_cost'], 130.0)
        self.assertEqual(warning['loss_amount'], 130.0)

    def test_export_sales_excel_respects_filters_and_uses_readable_labels(self):
        today = date.today()
        wholesale_group = CustomerGroup.objects.create(name='Khách sỉ')
        retail_group = CustomerGroup.objects.create(name='Khách lẻ')
        wholesale_customer = Customer.objects.create(
            store=self.store,
            code='KH-EX-SI',
            name='Khách mua sỉ Excel',
            group=wholesale_group,
            created_by=self.user,
        )
        retail_customer = Customer.objects.create(
            store=self.store,
            code='KH-EX-LE',
            name='Khách mua lẻ Excel',
            group=retail_group,
            created_by=self.user,
        )
        beverage_root = ProductCategory.objects.create(name='Đồ uống')
        coffee_type = ProductCategory.objects.create(name='Cà phê', parent=beverage_root)
        other_root = ProductCategory.objects.create(name='Thiết bị')
        exported_product = Product.objects.create(
            store=self.store,
            code='SP-EX-COFFEE',
            name='Cà phê hạt',
            category=coffee_type,
            created_by=self.user,
        )
        excluded_product = Product.objects.create(
            store=self.store,
            code='SP-EX-DEVICE',
            name='Máy xay',
            category=other_root,
            created_by=self.user,
        )

        loss_order = Order.objects.create(
            code='DH-EX-LOSS',
            store=self.store,
            customer=wholesale_customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=100,
            final_amount=100,
            paid_amount=100,
            order_date=today,
            salesperson='Nhân viên Excel',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=loss_order,
            product=exported_product,
            quantity=1,
            unit_price=100,
            cost_price=130,
            total_price=100,
        )

        profit_order = Order.objects.create(
            code='DH-EX-PROFIT',
            store=self.store,
            customer=retail_customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=200,
            final_amount=200,
            paid_amount=200,
            order_date=today,
            salesperson='Nhân viên khác',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=profit_order,
            product=excluded_product,
            quantity=1,
            unit_price=200,
            cost_price=120,
            total_price=200,
        )

        response = self.client.get(reverse('export_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'customer_kind': 'wholesale',
            'category_id': beverage_root.id,
            'profit_filter': 'loss',
        })

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertIn('Chi tiết đơn hàng', workbook.sheetnames)
        self.assertEqual(
            workbook.active['A3'].value,
            'Bộ lọc: Xem theo: Ngày | Phạm vi đơn: Đã xuất kho + Hoàn thành | Mốc ghi nhận doanh thu: Ngày xuất kho | Kiểu khách: Khách buôn / sỉ | Nhóm mặt hàng: Đồ uống | Lợi nhuận: Báo lỗ',
        )

        order_sheet = workbook['Chi tiết đơn hàng']
        exported_order_codes = [
            row[1]
            for row in order_sheet.iter_rows(min_row=2, max_col=2, values_only=True)
            if row[1] and row[1] != 'TỔNG'
        ]
        self.assertEqual(exported_order_codes, [loss_order.code])
        order_headers = [cell.value for cell in order_sheet[1]]
        loss_product_col = order_headers.index('Sản phẩm lỗ') + 1
        self.assertEqual(order_sheet.cell(row=2, column=loss_product_col).value, exported_product.name)

        product_sheet = workbook['Mặt hàng']
        exported_product_names = [
            row[1]
            for row in product_sheet.iter_rows(min_row=2, max_col=2, values_only=True)
            if row[1]
        ]
        self.assertEqual(exported_product_names, [exported_product.name])

    def test_export_sales_excel_respects_order_scope(self):
        today = date.today()
        pending_order = Order.objects.create(
            code='DH-EX-SCOPE-PENDING',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=1,
            total_amount=100,
            final_amount=100,
            order_date=today,
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=pending_order,
            product=self.product,
            quantity=1,
            unit_price=100,
            cost_price=60,
            total_price=100,
        )

        default_response = self.client.get(reverse('export_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        default_workbook = load_workbook(BytesIO(default_response.content))
        default_codes = [
            row[1]
            for row in default_workbook['Chi tiết đơn hàng'].iter_rows(min_row=2, max_col=2, values_only=True)
            if row[1] and row[1] != 'TỔNG'
        ]
        self.assertNotIn(pending_order.code, default_codes)

        all_active_response = self.client.get(reverse('export_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
            'order_scope': 'all_active',
        })
        all_active_workbook = load_workbook(BytesIO(all_active_response.content))
        all_active_codes = [
            row[1]
            for row in all_active_workbook['Chi tiết đơn hàng'].iter_rows(min_row=2, max_col=2, values_only=True)
            if row[1] and row[1] != 'TỔNG'
        ]
        self.assertIn(pending_order.code, all_active_codes)

    def test_api_report_staff_sales_filter_options_include_store_users_without_orders(self):
        today = date.today()
        seller = User.objects.create_user(
            username='staff_sales_report',
            password='pass123',
            first_name='Minh',
            last_name='Tran',
        )
        UserProfile.objects.create(user=seller, store=self.store)

        response = self.client.get(reverse('api_report_staff_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertIn('Minh Tran', payload['salespersons'])

    def test_staff_sales_report_includes_gross_margin_in_api_and_excel(self):
        today = date.today()
        order = Order.objects.create(
            code='DH-RP-STAFF-MARGIN',
            store=self.store,
            customer=self.customer,
            warehouse=self.warehouse,
            status=5,
            payment_status=2,
            total_amount=200,
            final_amount=200,
            paid_amount=200,
            order_date=today,
            salesperson='Nhân viên biên gộp',
            created_by=self.user,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=1,
            unit_price=200,
            cost_price=120,
            total_price=200,
        )

        api_response = self.client.get(reverse('api_report_staff_sales'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })

        self.assertEqual(api_response.status_code, 200)
        row = next(
            item for item in api_response.json()['staff_data']
            if item['salesperson'] == 'Nhân viên biên gộp'
        )
        self.assertEqual(row['profit'], 80.0)
        self.assertEqual(row['gross_margin'], 40.0)

        export_response = self.client.get(reverse('export_staff_sales_excel'), {
            'from_date': today.isoformat(),
            'to_date': today.isoformat(),
        })
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        sheet = workbook['BC Doanh thu NV']
        self.assertEqual(sheet.cell(row=4, column=7).value, 'Tỷ suất lợi nhuận gộp')
        exported_row = next(
            row_number for row_number in range(5, sheet.max_row)
            if sheet.cell(row=row_number, column=2).value == 'Nhân viên biên gộp'
        )
        self.assertEqual(sheet.cell(row=exported_row, column=7).value, 0.4)
        self.assertEqual(sheet.cell(row=exported_row, column=7).number_format, '0.0%')


class QuotationProfitReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(
            username='quotation_profit_owner',
            password='pass123',
        )
        cls.brand = Brand.objects.create(
            name='Brand LN báo giá',
            owner=cls.owner,
        )
        cls.store = Store.objects.create(
            brand=cls.brand,
            name='Cửa hàng LN báo giá',
            code='QLN',
        )
        cls.other_store = Store.objects.create(
            brand=cls.brand,
            name='Cửa hàng LN khác',
            code='QLN2',
        )
        cls.manager = User.objects.create_user(
            username='quotation_profit_manager',
            password='pass123',
        )
        UserProfile.objects.create(
            user=cls.manager,
            store=cls.store,
            position='Quản lý cửa hàng',
        )
        cls.accountant = User.objects.create_user(
            username='quotation_profit_accountant',
            password='pass123',
        )
        UserProfile.objects.create(
            user=cls.accountant,
            store=cls.store,
            position='Kế toán',
        )
        cls.salesperson = User.objects.create_user(
            username='quotation_profit_salesperson',
            password='pass123',
        )
        UserProfile.objects.create(
            user=cls.salesperson,
            store=cls.store,
            position='Nhân viên bán hàng',
        )
        cls.customer = Customer.objects.create(
            store=cls.store,
            code='KH-QLN',
            name='Khách LN báo giá',
            created_by=cls.manager,
        )
        cls.other_customer = Customer.objects.create(
            store=cls.other_store,
            code='KH-QLN2',
            name='Khách LN cửa hàng khác',
            created_by=cls.owner,
        )
        cls.product = Product.objects.create(
            store=cls.store,
            code='SP-QLN',
            name='Sản phẩm LN báo giá',
            cost_price=900,
            created_by=cls.manager,
        )
        cls.other_product = Product.objects.create(
            store=cls.other_store,
            code='SP-QLN2',
            name='Sản phẩm LN cửa hàng khác',
            cost_price=400,
            created_by=cls.owner,
        )

    def setUp(self):
        self.client.force_login(self.manager)

    def _create_quotation(
        self,
        code,
        *,
        store=None,
        customer=None,
        status=1,
        total_amount=2000,
        discount_amount=200,
        shipping_fee=100,
        other_fee=50,
        created_by=None,
    ):
        return Quotation.objects.create(
            code=code,
            store=store or self.store,
            customer=customer or self.customer,
            status=status,
            total_amount=total_amount,
            discount_amount=discount_amount,
            shipping_fee=shipping_fee,
            other_fee=other_fee,
            final_amount=total_amount - discount_amount + shipping_fee + other_fee,
            quotation_date=date.today(),
            salesperson='Nguyễn Quản lý',
            created_by=created_by or self.manager,
        )

    def test_manager_can_view_report_menu_and_open_quotation_link_in_new_tab(self):
        response = self.client.get(reverse('report_quotation_profit'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'BC LN dự kiến')
        self.assertContains(response, 'Báo cáo lợi nhuận dự kiến từ báo giá')
        self.assertContains(response, 'class="quotation-code-link"')
        self.assertContains(response, 'target="_blank"')
        self.assertContains(response, 'Thử CK CTV')
        self.assertContains(response, 'không sửa hoặc lưu vào báo giá')

    def test_report_page_has_pagination_and_persistent_column_visibility(self):
        response = self.client.get(reverse('report_quotation_profit'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="qp_column_config"')
        self.assertContains(response, 'id="qp_page_size"')
        self.assertContains(response, 'id="qp_pagination_summary"')
        self.assertContains(response, 'id="qp_pagination"')
        self.assertContains(response, '/static/js/column_config.js')
        self.assertContains(response, 'ifshop_report_quotation_profit_columns_v1')
        self.assertContains(response, 'ifshop_report_quotation_profit_page_size')
        self.assertContains(response, 'function renderQuotationProfitPage()')
        self.assertContains(response, 'quotationProfitRows.slice(start, end)')
        self.assertContains(response, 'function buildPaginationButtons(currentPage, totalPages)')
        self.assertContains(response, 'data-col="code"')
        self.assertContains(response, 'data-col="remaining_profit"')
        self.assertContains(response, 'renderFooterTotals(quotationProfitRows)')

    def test_regular_salesperson_cannot_view_quotation_profit_report(self):
        self.client.force_login(self.salesperson)

        api_response = self.client.get(reverse('api_report_quotation_profit'))
        page_response = self.client.get(reverse('report_quotation_profit'))

        self.assertEqual(api_response.status_code, 403)
        self.assertEqual(api_response.json()['status'], 'error')
        self.assertEqual(page_response.status_code, 302)

    def test_accountant_and_brand_owner_can_view_quotation_profit_report(self):
        for user in (self.accountant, self.owner):
            with self.subTest(user=user.username):
                self.client.force_login(user)
                response = self.client.get(reverse('api_report_quotation_profit'))
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.json()['status'], 'ok')

    def test_report_uses_snapshot_cost_and_calculates_expected_profit(self):
        quotation = self._create_quotation('BG-QLN-SNAPSHOT')
        QuotationItem.objects.create(
            quotation=quotation,
            product=self.product,
            quantity=2,
            unit_price=1000,
            total_price=2000,
            cost_price=600,
        )

        response = self.client.get(reverse('api_report_quotation_profit'), {
            'from_date': date.today().isoformat(),
            'to_date': date.today().isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['quotation_count'], 1)
        self.assertEqual(payload['summary']['total_revenue'], 1950)
        self.assertEqual(payload['summary']['total_cost'], 1200)
        self.assertEqual(payload['summary']['total_profit'], 750)
        self.assertAlmostEqual(payload['summary']['profit_margin'], 38.46, places=2)
        row = payload['data'][0]
        self.assertEqual(row['cost_status'], 'snapshot')
        self.assertEqual(row['ctv_discount_capacity'], 750)
        self.assertEqual(
            row['quotation_url'],
            f'/order-tbl/?edit_quotation={quotation.id}',
        )

    def test_report_marks_legacy_and_missing_cost_lines(self):
        legacy = self._create_quotation(
            'BG-QLN-LEGACY',
            total_amount=1000,
            discount_amount=0,
            shipping_fee=0,
            other_fee=0,
        )
        QuotationItem.objects.create(
            quotation=legacy,
            product=self.product,
            quantity=1,
            unit_price=1000,
            total_price=1000,
            cost_price=None,
        )
        missing = self._create_quotation(
            'BG-QLN-MISSING',
            total_amount=500,
            discount_amount=0,
            shipping_fee=0,
            other_fee=0,
        )
        QuotationItem.objects.create(
            quotation=missing,
            item_name='Dịch vụ chưa khai báo giá vốn',
            is_service_line=True,
            quantity=1,
            unit_price=500,
            total_price=500,
            cost_price=None,
        )

        payload = self.client.get(
            reverse('api_report_quotation_profit'),
            {'from_date': date.today().isoformat(), 'to_date': date.today().isoformat()},
        ).json()
        rows = {row['code']: row for row in payload['data']}

        self.assertEqual(rows[legacy.code]['expected_cost'], 900)
        self.assertEqual(rows[legacy.code]['estimated_cost_lines'], 1)
        self.assertEqual(rows[legacy.code]['cost_status'], 'estimated')
        self.assertEqual(rows[missing.code]['missing_cost_lines'], 1)
        self.assertEqual(rows[missing.code]['cost_status'], 'missing')
        self.assertIsNone(rows[missing.code]['ctv_discount_capacity'])
        self.assertEqual(payload['summary']['estimated_cost_count'], 1)
        self.assertEqual(payload['summary']['missing_cost_count'], 1)

        missing_only = self.client.get(
            reverse('api_report_quotation_profit'),
            {
                'from_date': date.today().isoformat(),
                'to_date': date.today().isoformat(),
                'profit_filter': 'missing_cost',
            },
        ).json()
        self.assertEqual([row['code'] for row in missing_only['data']], [missing.code])

    def test_manager_scope_and_default_status_do_not_leak_other_store_or_converted_quotes(self):
        own_active = self._create_quotation('BG-QLN-OWN')
        QuotationItem.objects.create(
            quotation=own_active,
            product=self.product,
            quantity=1,
            unit_price=1000,
            total_price=1000,
            cost_price=500,
        )
        own_converted = self._create_quotation('BG-QLN-CONVERTED', status=3)
        QuotationItem.objects.create(
            quotation=own_converted,
            product=self.product,
            quantity=1,
            unit_price=1000,
            total_price=1000,
            cost_price=500,
        )
        other_active = self._create_quotation(
            'BG-QLN-OTHER',
            store=self.other_store,
            customer=self.other_customer,
            created_by=self.owner,
        )
        QuotationItem.objects.create(
            quotation=other_active,
            product=self.other_product,
            quantity=1,
            unit_price=1000,
            total_price=1000,
            cost_price=400,
        )

        default_payload = self.client.get(
            reverse('api_report_quotation_profit'),
            {'from_date': date.today().isoformat(), 'to_date': date.today().isoformat()},
        ).json()
        self.assertEqual([row['code'] for row in default_payload['data']], [own_active.code])

        converted_payload = self.client.get(
            reverse('api_report_quotation_profit'),
            {
                'from_date': date.today().isoformat(),
                'to_date': date.today().isoformat(),
                'status': '3',
            },
        ).json()
        self.assertEqual(
            [row['code'] for row in converted_payload['data']],
            [own_converted.code],
        )

        self.client.force_login(self.owner)
        owner_payload = self.client.get(
            reverse('api_report_quotation_profit'),
            {'from_date': date.today().isoformat(), 'to_date': date.today().isoformat()},
        ).json()
        self.assertEqual(
            {row['code'] for row in owner_payload['data']},
            {own_active.code, other_active.code},
        )

    def test_export_quotation_profit_excel_contains_calculated_values(self):
        quotation = self._create_quotation('BG-QLN-EXPORT')
        QuotationItem.objects.create(
            quotation=quotation,
            product=self.product,
            quantity=2,
            unit_price=1000,
            total_price=2000,
            cost_price=600,
        )

        response = self.client.get(reverse('export_quotation_profit_excel'), {
            'from_date': date.today().isoformat(),
            'to_date': date.today().isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook['LN dự kiến báo giá']
        self.assertEqual(sheet['B6'].value, quotation.code)
        self.assertEqual(sheet['O6'].value, 1950)
        self.assertEqual(sheet['P6'].value, 1200)
        self.assertEqual(sheet['Q6'].value, 750)


class StockAlertEmailSettingTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(
            username='stock_alert_owner',
            email='owner@example.com',
            password='pass123',
        )
        cls.brand = Brand.objects.create(name='Brand Stock Alert', owner=cls.owner)
        cls.store = Store.objects.create(brand=cls.brand, name='Kho hàng cảnh báo', code='SAS')
        UserProfile.objects.create(user=cls.owner, store=cls.store)
        cls.staff = User.objects.create_user(
            username='stock_alert_staff',
            email='warehouse@example.com',
            password='pass123',
        )
        UserProfile.objects.create(user=cls.staff, store=cls.store, position='Quản lý cửa hàng')
        cls.warehouse = Warehouse.objects.create(store=cls.store, code='KHO-SAS', name='Kho chính')
        cls.parent_category = ProductCategory.objects.create(name='Đồ uống')
        cls.child_category = ProductCategory.objects.create(name='Cà phê', parent=cls.parent_category)
        cls.other_category = ProductCategory.objects.create(name='Văn phòng phẩm')
        cls.low_product = Product.objects.create(
            store=cls.store,
            code='SP-LOW-001',
            name='Cà phê sắp hết',
            category=cls.child_category,
            min_stock=5,
            created_by=cls.owner,
        )
        cls.other_product = Product.objects.create(
            store=cls.store,
            code='SP-OTHER-001',
            name='Sản phẩm ngoài danh mục',
            category=cls.other_category,
            min_stock=5,
            created_by=cls.owner,
        )
        ProductStock.objects.create(product=cls.low_product, warehouse=cls.warehouse, quantity=2)
        ProductStock.objects.create(product=cls.other_product, warehouse=cls.warehouse, quantity=1)

    def setUp(self):
        self.client.force_login(self.owner)

    def test_setting_page_is_available_to_brand_owner(self):
        response = self.client.get(reverse('stock_alert_email_setting'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cảnh báo tồn kho qua email')
        self.assertContains(response, 'id="stock_alert_send_time"')
        self.assertContains(response, 'id="stock_alert_send_time_picker"')
        self.assertContains(response, "format: 'hh:mm A'")
        self.assertContains(response, "locale: 'en'")
        self.assertContains(response, 'dùng AM hoặc PM')
        self.assertContains(response, 'id="stock_alert_category_list"')
        self.assertContains(response, 'id="stock_alert_recipient_list"')
        self.assertContains(response, 'id="active_recipient_name"')
        self.assertContains(response, '21:00')

    def test_regular_staff_cannot_manage_setting(self):
        self.client.force_login(self.staff)
        response = self.client.get(reverse('stock_alert_email_setting'))
        self.assertEqual(response.status_code, 302)
        response = self.client.post(
            reverse('api_save_stock_alert_email_setting'),
            data='{}',
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 403)

    def test_save_and_test_email_include_selected_category_children_only(self):
        response = self.client.post(
            reverse('api_save_stock_alert_email_setting'),
            data={
                'is_active': True,
                'send_time': '21:00',
                'include_child_categories': True,
                'recipient_assignments': [
                    {
                        'user_id': self.staff.id,
                        'category_ids': [self.parent_category.id],
                    },
                    {
                        'email': 'external@example.com',
                        'category_ids': [self.parent_category.id],
                    },
                ],
            },
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200, response.content)
        config = StockAlert.objects.get(brand=self.brand)
        self.assertEqual(list(config.categories.values_list('id', flat=True)), [self.parent_category.id])
        self.assertEqual(list(config.recipient_users.values_list('id', flat=True)), [self.staff.id])
        self.assertEqual(config.email_recipient_scopes.count(), 2)

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            response = self.client.post(reverse('api_test_stock_alert_email'))

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(len(mail.outbox), 2)
        messages_by_email = {message.to[0]: message for message in mail.outbox}
        self.assertEqual(set(messages_by_email), {'warehouse@example.com', 'external@example.com'})
        for message in messages_by_email.values():
            self.assertIn('Cà phê sắp hết', message.body)
            self.assertNotIn('Sản phẩm ngoài danh mục', message.body)

    def test_each_recipient_only_receives_their_assigned_categories(self):
        response = self.client.post(
            reverse('api_save_stock_alert_email_setting'),
            data={
                'is_active': True,
                'send_time': '21:00',
                'include_child_categories': True,
                'recipient_assignments': [
                    {
                        'user_id': self.staff.id,
                        'category_ids': [self.parent_category.id],
                    },
                    {
                        'user_id': self.owner.id,
                        'category_ids': [self.other_category.id],
                    },
                ],
            },
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200, response.content)

        config = StockAlert.objects.get(brand=self.brand)
        staff_scope = StockAlertEmailRecipient.objects.get(
            stock_alert=config,
            user=self.staff,
        )
        owner_scope = StockAlertEmailRecipient.objects.get(
            stock_alert=config,
            user=self.owner,
        )
        self.assertEqual(
            set(staff_scope.categories.values_list('id', flat=True)),
            {self.parent_category.id},
        )
        self.assertEqual(
            set(owner_scope.categories.values_list('id', flat=True)),
            {self.other_category.id},
        )

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            response = self.client.post(reverse('api_test_stock_alert_email'))

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(len(mail.outbox), 2)
        messages_by_email = {message.to[0]: message for message in mail.outbox}
        self.assertIn('Cà phê sắp hết', messages_by_email['warehouse@example.com'].body)
        self.assertNotIn('Sản phẩm ngoài danh mục', messages_by_email['warehouse@example.com'].body)
        self.assertIn('Sản phẩm ngoài danh mục', messages_by_email['owner@example.com'].body)
        self.assertNotIn('Cà phê sắp hết', messages_by_email['owner@example.com'].body)

    def test_disabled_stock_recipient_is_kept_but_does_not_receive_email(self):
        response = self.client.post(
            reverse('api_save_stock_alert_email_setting'),
            data={
                'is_active': True,
                'send_time': '21:00',
                'include_child_categories': True,
                'recipient_assignments': [
                    {
                        'user_id': self.staff.id,
                        'is_active': True,
                        'category_ids': [self.parent_category.id],
                    },
                    {
                        'email': 'paused-stock@example.com',
                        'is_active': False,
                        'category_ids': [],
                    },
                ],
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        config = StockAlert.objects.get(brand=self.brand)
        paused = StockAlertEmailRecipient.objects.get(
            stock_alert=config,
            email='paused-stock@example.com',
        )
        self.assertFalse(paused.is_active)
        self.assertEqual(config.email_recipient_scopes.count(), 2)

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            response = self.client.post(reverse('api_test_stock_alert_email'))

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual([message.to for message in mail.outbox], [['warehouse@example.com']])

    def test_scheduled_processing_sends_only_once_per_day(self):
        config = StockAlert.objects.create(
            brand=self.brand,
            is_active=True,
            alert_on_min=True,
            send_time=time(21, 0),
        )
        config.recipient_users.add(self.staff)
        config.categories.add(self.parent_category)
        run_at = datetime.combine(date.today(), time(21, 0))

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            first = process_stock_alert(config.id, now=run_at)
            second = process_stock_alert(config.id, now=run_at)

        self.assertEqual(first['status'], 'sent')
        self.assertEqual(second['status'], 'skipped')
        self.assertEqual(len(mail.outbox), 1)
        config.refresh_from_db()
        self.assertEqual(config.last_status, 'sent')


class DailyEmailReportSettingTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(
            username='daily_report_owner',
            email='owner-daily@example.com',
            password='pass123',
        )
        cls.brand = Brand.objects.create(
            name='Brand Daily Email',
            owner=cls.owner,
        )
        cls.store = Store.objects.create(
            brand=cls.brand,
            name='Cửa hàng báo cáo ngày',
            code='DLY',
        )
        UserProfile.objects.create(user=cls.owner, store=cls.store)
        cls.staff = User.objects.create_user(
            username='daily_report_staff',
            email='daily-staff@example.com',
            password='pass123',
        )
        UserProfile.objects.create(
            user=cls.staff,
            store=cls.store,
            position='Kế toán',
        )
        cls.warehouse = Warehouse.objects.create(
            store=cls.store,
            code='KHO-DLY',
            name='Kho báo cáo ngày',
        )
        cls.product = Product.objects.create(
            store=cls.store,
            code='SP-DLY-001',
            name='Sản phẩm báo cáo ngày',
            cost_price=40,
            created_by=cls.owner,
        )
        cls.customer = Customer.objects.create(
            store=cls.store,
            code='KH-DLY-001',
            name='Khách báo cáo ngày',
            created_by=cls.owner,
        )
        cls.bank_account = CashBook.objects.create(name='TK ngân hàng A')
        cls.cash_account = CashBook.objects.create(name='Quỹ tiền mặt')
        cls.bank_method = PaymentMethodOption.objects.create(
            code='BANK-DAILY-REPORT',
            name='Chuyển khoản công ty La Maison Bui',
            legacy_type=2,
            default_cash_book=cls.bank_account,
        )
        cls.cash_method = PaymentMethodOption.objects.create(
            code='CASH-DAILY-REPORT',
            name='Tiền mặt tại quỹ',
            legacy_type=1,
            default_cash_book=cls.cash_account,
        )

    def setUp(self):
        self.client.force_login(self.owner)

    def _create_daily_transactions(self):
        report_date = date.today()
        order = Order.objects.create(
            code='DH-DLY-001',
            store=self.store,
            warehouse=self.warehouse,
            customer=self.customer,
            status=5,
            payment_status=1,
            total_amount=200,
            final_amount=150,
            paid_amount=100,
            order_date=report_date,
            exported_at=datetime.combine(report_date, time(10, 0)),
            created_by=self.owner,
        )
        OrderItem.objects.create(
            order=order,
            product=self.product,
            quantity=2,
            unit_price=100,
            cost_price=40,
            total_price=200,
        )
        order_return = OrderReturn.objects.create(
            code='TH-DLY-001',
            order=order,
            customer=self.customer,
            warehouse=self.warehouse,
            status=2,
            total_refund=30,
            return_date=report_date,
            created_by=self.owner,
        )
        OrderReturnItem.objects.create(
            order_return=order_return,
            product=self.product,
            quantity=1,
            unit_price=30,
            total_price=30,
        )
        Receipt.objects.create(
            code='PT-DLY-001',
            store=self.store,
            cash_book=self.bank_account,
            payment_method_option=self.bank_method,
            customer=self.customer,
            order=order,
            amount=100,
            receipt_date=report_date,
            status=1,
            created_by=self.owner,
        )
        Receipt.objects.create(
            code='PT-DLY-002',
            store=self.store,
            cash_book=self.cash_account,
            payment_method_option=self.cash_method,
            customer=self.customer,
            amount=50,
            receipt_date=report_date,
            status=1,
            created_by=self.owner,
        )

    def test_setting_page_and_menu_are_available_to_brand_owner(self):
        response = self.client.get(reverse('daily_email_report_setting'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'BC email hàng ngày')
        self.assertContains(response, 'Doanh thu')
        self.assertContains(response, 'Lợi nhuận gộp')
        self.assertContains(response, 'Tổng tiền về')
        self.assertContains(response, 'id="daily_email_report_send_time"')
        self.assertContains(response, 'id="daily_email_report_send_time_picker"')
        self.assertContains(response, "format: 'hh:mm A'")
        self.assertContains(response, "locale: 'en'")

    def test_regular_staff_cannot_manage_daily_email_report(self):
        self.client.force_login(self.staff)

        response = self.client.get(reverse('daily_email_report_setting'))
        self.assertEqual(response.status_code, 302)
        response = self.client.post(
            reverse('api_save_daily_email_report_setting'),
            data='{}',
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 403)

    def test_metrics_include_returns_cost_and_completed_receipts(self):
        self._create_daily_transactions()
        config = DailyEmailReport.objects.create(brand=self.brand)

        metrics = collect_daily_email_report_metrics(
            config,
            report_date=date.today(),
        )

        self.assertEqual(metrics['revenue'], 150)
        self.assertEqual(metrics['returns_total'], 30)
        self.assertEqual(metrics['net_revenue'], 120)
        self.assertEqual(metrics['net_cost'], 40)
        self.assertEqual(metrics['gross_profit'], 80)
        self.assertEqual(metrics['total_money_received'], 150)
        self.assertEqual(
            metrics['money_received_by_cash_book'],
            [
                {
                    'cash_book_id': self.bank_account.id,
                    'payment_method_option_id': self.bank_method.id,
                    'name': 'Chuyển khoản công ty La Maison Bui',
                    'cash_book_name': 'TK ngân hàng A',
                    'payment_method_name': 'Chuyển khoản công ty La Maison Bui',
                    'amount': 100,
                    'amount_text': '100đ',
                },
                {
                    'cash_book_id': self.cash_account.id,
                    'payment_method_option_id': self.cash_method.id,
                    'name': 'Tiền mặt tại quỹ',
                    'cash_book_name': 'Quỹ tiền mặt',
                    'payment_method_name': 'Tiền mặt tại quỹ',
                    'amount': 50,
                    'amount_text': '50đ',
                },
            ],
        )

    def test_daily_email_revenue_uses_exported_at_instead_of_order_date(self):
        report_date = date.today()
        previous_date = report_date - timedelta(days=1)
        included = Order.objects.create(
            code='DH-DLY-EXPORT-TODAY',
            store=self.store,
            warehouse=self.warehouse,
            customer=self.customer,
            status=4,
            total_amount=120,
            final_amount=120,
            order_date=previous_date,
            exported_at=datetime.combine(report_date, time(9, 0)),
            created_by=self.owner,
        )
        Order.objects.create(
            code='DH-DLY-ORDER-TODAY',
            store=self.store,
            warehouse=self.warehouse,
            customer=self.customer,
            status=5,
            total_amount=300,
            final_amount=300,
            order_date=report_date,
            exported_at=datetime.combine(previous_date, time(16, 0)),
            created_by=self.owner,
        )
        OrderItem.objects.create(
            order=included,
            product=self.product,
            quantity=1,
            unit_price=120,
            cost_price=40,
            total_price=120,
        )
        config = DailyEmailReport.objects.create(brand=self.brand)

        metrics = collect_daily_email_report_metrics(config, report_date=report_date)

        self.assertEqual(metrics['revenue'], 120)
        self.assertEqual(metrics['sales_cost'], 40)

    def test_metrics_show_completed_receipts_without_a_cash_book_separately(self):
        Receipt.objects.create(
            code='PT-DLY-UNASSIGNED',
            store=self.store,
            customer=self.customer,
            amount=25,
            receipt_date=date.today(),
            status=1,
            created_by=self.owner,
        )
        config = DailyEmailReport.objects.create(brand=self.brand)

        metrics = collect_daily_email_report_metrics(
            config,
            report_date=date.today(),
        )

        self.assertEqual(metrics['total_money_received'], 25)
        self.assertEqual(
            metrics['money_received_by_cash_book'],
            [
                {
                    'cash_book_id': None,
                    'payment_method_option_id': None,
                    'name': 'Chưa gán tài khoản',
                    'cash_book_name': 'Chưa gán tài khoản',
                    'payment_method_name': 'Chuyển khoản',
                    'amount': 25,
                    'amount_text': '25đ',
                },
            ],
        )

    def test_save_and_send_test_daily_email_report(self):
        self._create_daily_transactions()
        response = self.client.post(
            reverse('api_save_daily_email_report_setting'),
            data={
                'is_active': True,
                'send_time': '21:00',
                'recipient_user_ids': [self.staff.id],
                'email_recipients': 'external-daily@example.com',
            },
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200, response.content)
        config = DailyEmailReport.objects.get(brand=self.brand)
        self.assertTrue(config.is_active)
        self.assertEqual(
            list(config.recipient_users.values_list('id', flat=True)),
            [self.staff.id],
        )
        self.assertEqual(config.recipient_settings.count(), 2)
        self.assertFalse(
            config.recipient_settings.filter(is_active=False).exists()
        )

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            response = self.client.post(reverse('api_test_daily_email_report'))

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(len(mail.outbox), 2)
        self.assertEqual(
            {message.to[0] for message in mail.outbox},
            {'daily-staff@example.com', 'external-daily@example.com'},
        )
        self.assertIn('Doanh thu: 150đ', mail.outbox[0].body)
        self.assertIn('Lợi nhuận gộp: 80đ', mail.outbox[0].body)
        self.assertIn('Tổng tiền về: 150đ', mail.outbox[0].body)
        self.assertIn('Chi tiết tiền về theo tài khoản nhận:', mail.outbox[0].body)
        self.assertIn(
            '- Chuyển khoản công ty La Maison Bui (sổ quỹ: TK ngân hàng A): 100đ',
            mail.outbox[0].body,
        )
        self.assertIn(
            '- Tiền mặt tại quỹ (sổ quỹ: Quỹ tiền mặt): 50đ',
            mail.outbox[0].body,
        )
        html_body = mail.outbox[0].alternatives[0].content
        summary_position = html_body.index('Tổng hợp bán hàng trong ngày')
        account_detail_position = html_body.index('Chi tiết theo tài khoản nhận')
        self.assertLess(summary_position, account_detail_position)
        self.assertIn('Chi tiết theo tài khoản nhận', html_body)
        self.assertIn('Chuyển khoản công ty La Maison Bui', html_body)
        self.assertIn('Tiền mặt tại quỹ', html_body)
        self.assertNotIn('table-layout:fixed', html_body)
        self.assertNotIn('font-size:20px', html_body)
        self.assertIn('Doanh thu</td>', html_body)
        self.assertIn('Lợi nhuận gộp</td>', html_body)
        self.assertIn('Tỷ suất lợi nhuận gộp</td>', html_body)
        self.assertIn('Tổng tiền về</td>', html_body)
        self.assertIn('Tài khoản nhận', html_body)
        self.assertIn('Sổ quỹ', html_body)
        self.assertIn('TK ngân hàng A', html_body)
        self.assertIn('Quỹ tiền mặt', html_body)
        self.assertIn('Tổng hợp bán hàng trong ngày', html_body)

    def test_disabled_daily_report_recipient_is_kept_but_not_sent(self):
        response = self.client.post(
            reverse('api_save_daily_email_report_setting'),
            data={
                'is_active': True,
                'send_time': '21:00',
                'recipient_assignments': [
                    {'user_id': self.staff.id, 'is_active': True},
                    {'user_id': self.owner.id, 'is_active': False},
                    {'email': 'paused-daily@example.com', 'is_active': False},
                ],
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        config = DailyEmailReport.objects.get(brand=self.brand)
        self.assertEqual(config.recipient_settings.count(), 3)
        self.assertEqual(
            set(
                config.recipient_settings.filter(is_active=False)
                .values_list('email', flat=True)
            ),
            {'owner-daily@example.com', 'paused-daily@example.com'},
        )

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            response = self.client.post(reverse('api_test_daily_email_report'))

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual([message.to for message in mail.outbox], [['daily-staff@example.com']])

        response = self.client.get(reverse('daily_email_report_setting'))
        staff_by_id = {
            person['id']: person for person in response.context['staff_options']
        }
        self.assertTrue(staff_by_id[self.staff.id]['selected'])
        self.assertFalse(staff_by_id[self.owner.id]['selected'])
        self.assertEqual(
            response.context['extra_recipients'],
            [{'email': 'paused-daily@example.com', 'is_active': False}],
        )

    def test_scheduled_daily_report_sends_only_once_per_day(self):
        self._create_daily_transactions()
        config = DailyEmailReport.objects.create(
            brand=self.brand,
            is_active=True,
            send_time=time(21, 0),
        )
        config.recipient_users.add(self.staff)
        run_at = datetime.combine(date.today(), time(21, 0))

        with override_settings(
            EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
            DEFAULT_FROM_EMAIL='ifshop@example.com',
        ):
            first = process_daily_email_report(config.id, now=run_at)
            second = process_daily_email_report(config.id, now=run_at)

        self.assertEqual(first['status'], 'sent')
        self.assertEqual(second['status'], 'skipped')
        self.assertEqual(len(mail.outbox), 1)
        config.refresh_from_db()
        self.assertEqual(config.last_status, 'sent')


class ScheduledEmailApiTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(
            username='scheduler_api_owner',
            email='scheduler-owner@example.com',
            password='pass123',
        )
        cls.brand = Brand.objects.create(
            name='Brand Scheduler API',
            owner=cls.owner,
        )
        cls.store = Store.objects.create(
            brand=cls.brand,
            name='Cửa hàng Scheduler API',
            code='SCA',
        )
        UserProfile.objects.create(user=cls.owner, store=cls.store)
        cls.warehouse = Warehouse.objects.create(
            store=cls.store,
            code='KHO-SCA',
            name='Kho Scheduler API',
        )
        cls.category = ProductCategory.objects.create(
            name='Danh mục Scheduler API',
        )
        cls.product = Product.objects.create(
            store=cls.store,
            code='SP-SCA-001',
            name='Sản phẩm Scheduler API',
            category=cls.category,
            min_stock=5,
            created_by=cls.owner,
        )
        ProductStock.objects.create(
            product=cls.product,
            warehouse=cls.warehouse,
            quantity=1,
        )
        cls.stock_config = StockAlert.objects.create(
            brand=cls.brand,
            is_active=True,
            alert_on_min=True,
            send_time=time(0, 0),
        )
        cls.stock_config.recipient_users.add(cls.owner)
        cls.stock_config.categories.add(cls.category)
        cls.daily_config = DailyEmailReport.objects.create(
            brand=cls.brand,
            is_active=True,
            send_time=time(0, 0),
        )
        cls.daily_config.recipient_users.add(cls.owner)

    def test_scheduler_api_requires_post(self):
        response = self.client.get(reverse('api_run_scheduled_emails'))

        self.assertEqual(response.status_code, 405)

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        DEFAULT_FROM_EMAIL='ifshop@example.com',
    )
    def test_scheduler_api_runs_both_email_jobs_once_per_day(self):
        response = self.client.post(reverse('api_run_scheduled_emails'))

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload['stock_alerts']['totals']['sent'], 1)
        self.assertEqual(payload['daily_email_reports']['totals']['sent'], 1)
        self.assertEqual(len(mail.outbox), 2)

        second_response = self.client.post(reverse('api_run_scheduled_emails'))

        self.assertEqual(second_response.status_code, 200, second_response.content)
        second_payload = second_response.json()
        self.assertEqual(second_payload['stock_alerts']['totals']['skipped'], 1)
        self.assertEqual(
            second_payload['daily_email_reports']['totals']['skipped'],
            1,
        )
        self.assertEqual(len(mail.outbox), 2)
