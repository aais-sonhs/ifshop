import json
import logging
import unicodedata
from collections import defaultdict
from datetime import datetime, time, timedelta
from decimal import Decimal
from functools import wraps
from django.conf import settings
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, Sum, Count, Max, Q, F
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_time
from orders.models import (
    Order,
    OrderItem,
    OrderReturn,
    OrderReturnItem,
    Quotation,
)
from core.store_utils import (
    can_view_quotation_profit_report,
    filter_by_store,
    brand_owner_required,
    report_permission_required,
    can_view_sales_report,
    get_company_brand_for_user,
    get_managed_store_ids,
    is_brand_owner,
)
from products.models import ProductCategory
from .models import (
    DailyEmailReport,
    DailyEmailReportRecipient,
    StockAlert,
    StockAlertEmailRecipient,
)
from .daily_email_reports import (
    DailyEmailReportConfigurationError,
    get_daily_email_report_recipients,
    send_daily_email_report,
)
from .stock_alerts import (
    StockAlertConfigurationError,
    get_stock_alert_recipients,
    parse_recipient_emails,
    send_stock_alert_email,
)

logger = logging.getLogger(__name__)


CUSTOMER_KIND_OPTIONS = [
    {'value': 'retail', 'label': 'Khách lẻ'},
    {'value': 'wholesale', 'label': 'Khách buôn / sỉ'},
    {'value': 'other', 'label': 'Khác / chưa phân loại'},
]

RETAIL_GROUP_KEYWORDS = ('lẻ', 'le', 'bán lẻ', 'ban le', 'retail')
WHOLESALE_GROUP_KEYWORDS = (
    'sỉ', 'si', 'buôn', 'buon', 'bán buôn', 'ban buon',
    'đại lý', 'dai ly', 'wholesale',
)


def _stock_alert_brand_for_user(user):
    return get_company_brand_for_user(user)


def _stock_alert_users_for_brand(brand):
    if not brand:
        return User.objects.none()
    return (
        User.objects.filter(
            Q(id=brand.owner_id) | Q(profile__store__brand_id=brand.id),
            is_active=True,
        )
        .distinct()
        .order_by('first_name', 'last_name', 'username')
    )


def _stock_alert_category_options(selected_ids=None):
    selected_ids = set(selected_ids or [])
    rows = list(
        ProductCategory.objects.filter(is_active=True)
        .values('id', 'name', 'parent_id')
        .order_by('name', 'id')
    )
    children = defaultdict(list)
    known_ids = {row['id'] for row in rows}
    for row in rows:
        parent_id = row['parent_id'] if row['parent_id'] in known_ids else None
        children[parent_id].append(row)

    options = []
    visited = set()

    def append_branch(parent_id, depth):
        for row in children.get(parent_id, []):
            if row['id'] in visited:
                continue
            visited.add(row['id'])
            options.append({
                'id': row['id'],
                'name': row['name'],
                'depth': depth,
                'selected': row['id'] in selected_ids,
            })
            append_branch(row['id'], depth + 1)

    append_branch(None, 0)
    # Dữ liệu cũ có thể có vòng hoặc cha lỗi; vẫn cho phép chọn các dòng còn lại.
    for row in rows:
        if row['id'] not in visited:
            options.append({
                'id': row['id'],
                'name': row['name'],
                'depth': 0,
                'selected': row['id'] in selected_ids,
            })
    return options


def _stock_alert_for_brand(brand):
    config, _ = StockAlert.objects.get_or_create(
        brand=brand,
        defaults={
            'email_recipients': '',
            'alert_on_min': True,
            'alert_on_max': False,
            'is_active': False,
        },
    )
    return config


def _stock_alert_forbidden_json():
    return JsonResponse({
        'status': 'error',
        'message': 'Chỉ chủ thương hiệu mới được cấu hình cảnh báo tồn kho qua email.',
    }, status=403)


def _daily_email_report_for_brand(brand):
    config, _ = DailyEmailReport.objects.get_or_create(
        brand=brand,
        defaults={
            'email_recipients': '',
            'is_active': False,
        },
    )
    return config


def _daily_email_report_forbidden_json():
    return JsonResponse({
        'status': 'error',
        'message': 'Chỉ chủ thương hiệu mới được cấu hình báo cáo email hằng ngày.',
    }, status=403)


@login_required(login_url='/login/')
def daily_email_report_setting(request):
    if not is_brand_owner(request.user):
        messages.error(request, 'Chỉ chủ thương hiệu mới được cấu hình báo cáo email hằng ngày.')
        return redirect('/dashboard/')

    brand = _stock_alert_brand_for_user(request.user)
    if not brand:
        messages.error(request, 'Không tìm thấy thương hiệu để lưu cấu hình.')
        return redirect('/dashboard/')

    config = _daily_email_report_for_brand(brand)
    recipient_settings = list(
        config.recipient_settings.select_related('user').all()
    )
    has_recipient_settings = bool(recipient_settings)
    settings_by_user_id = {
        recipient.user_id: recipient
        for recipient in recipient_settings
        if recipient.user_id
    }
    legacy_user_ids = set(config.recipient_users.values_list('id', flat=True))
    staff_options = []
    for user in _stock_alert_users_for_brand(brand):
        recipient = settings_by_user_id.get(user.id)
        selected = (
            recipient.is_active
            if recipient else (
                user.id in legacy_user_ids if not has_recipient_settings else False
            )
        )
        staff_options.append({
            'id': user.id,
            'name': user.get_full_name().strip() or user.username,
            'username': user.username,
            'email': user.email or '',
            'selected': selected and bool(user.email),
        })
    if has_recipient_settings:
        extra_recipients = [
            {
                'email': recipient.email,
                'is_active': recipient.is_active,
            }
            for recipient in recipient_settings
            if not recipient.user_id
        ]
    else:
        legacy_extra_emails, _ = parse_recipient_emails(config.email_recipients)
        extra_recipients = [
            {'email': email, 'is_active': True}
            for email in legacy_extra_emails
        ]

    context = {
        'active_tab': 'daily_email_report_setting',
        'daily_email_report': config,
        'staff_options': staff_options,
        'extra_recipients': extra_recipients,
        'smtp_ready': not (
            str(getattr(settings, 'EMAIL_BACKEND', '') or '').endswith('smtp.EmailBackend')
            and not getattr(settings, 'EMAIL_HOST_USER', '')
        ),
    }
    return render(request, 'system/daily_email_report_setting.html', context)


@login_required(login_url='/login/')
def api_save_daily_email_report_setting(request):
    if not is_brand_owner(request.user):
        return _daily_email_report_forbidden_json()
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'}, status=405)

    brand = _stock_alert_brand_for_user(request.user)
    if not brand:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy thương hiệu.'}, status=400)
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Dữ liệu gửi lên không hợp lệ.'}, status=400)

    send_time = parse_time(str(data.get('send_time') or '21:00'))
    if not send_time:
        return JsonResponse({'status': 'error', 'message': 'Giờ gửi không hợp lệ.'}, status=400)

    raw_assignments = data.get('recipient_assignments')
    if not isinstance(raw_assignments, list):
        raw_assignments = [
            {'user_id': user_id, 'is_active': True}
            for user_id in _parse_id_set(data.get('recipient_user_ids'))
        ]
        legacy_emails, invalid_emails = parse_recipient_emails(
            data.get('email_recipients')
        )
        if invalid_emails:
            return JsonResponse({
                'status': 'error',
                'message': 'Email không hợp lệ: ' + ', '.join(invalid_emails),
            }, status=400)
        raw_assignments.extend(
            {'email': email, 'is_active': True}
            for email in legacy_emails
        )

    allowed_users = {
        user.id: user
        for user in _stock_alert_users_for_brand(brand).exclude(email='')
    }
    assignments = []
    seen_recipient_keys = set()
    seen_emails = set()
    for raw_assignment in raw_assignments:
        if not isinstance(raw_assignment, dict):
            return JsonResponse({
                'status': 'error',
                'message': 'Danh sách người nhận không hợp lệ.',
            }, status=400)

        user_id = raw_assignment.get('user_id')
        user = None
        if user_id not in (None, ''):
            try:
                user_id = int(user_id)
            except (TypeError, ValueError):
                user_id = 0
            user = allowed_users.get(user_id)
            if not user:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Có người nhận không thuộc thương hiệu hoặc chưa có email.',
                }, status=400)
            email = user.email.strip().lower()
            recipient_key = f'user:{user.id}'
        else:
            parsed_emails, invalid_emails = parse_recipient_emails(
                str(raw_assignment.get('email') or '').strip()
            )
            if invalid_emails or len(parsed_emails) != 1:
                invalid_value = (
                    invalid_emails[0]
                    if invalid_emails else str(raw_assignment.get('email') or '')
                )
                return JsonResponse({
                    'status': 'error',
                    'message': 'Email không hợp lệ: ' + invalid_value,
                }, status=400)
            email = parsed_emails[0]
            recipient_key = f'email:{email}'

        if recipient_key in seen_recipient_keys or email in seen_emails:
            return JsonResponse({
                'status': 'error',
                'message': f'Email {email} đang được chọn trùng.',
            }, status=400)
        seen_recipient_keys.add(recipient_key)
        seen_emails.add(email)
        assignments.append({
            'user': user,
            'email': email,
            'is_active': bool(raw_assignment.get('is_active', True)),
        })

    is_active = bool(data.get('is_active', False))
    active_assignments = [
        assignment for assignment in assignments if assignment['is_active']
    ]
    if is_active and not active_assignments:
        return JsonResponse({
            'status': 'error',
            'message': 'Vui lòng bật ít nhất một người nhận.',
        }, status=400)

    with transaction.atomic():
        config = _daily_email_report_for_brand(brand)
        config.send_time = send_time
        config.is_active = is_active
        config.email_recipients = ', '.join(
            assignment['email']
            for assignment in active_assignments
            if not assignment['user']
        )
        config.save()
        config.recipient_users.set(
            assignment['user']
            for assignment in active_assignments
            if assignment['user']
        )
        config.recipient_settings.all().delete()
        DailyEmailReportRecipient.objects.bulk_create([
            DailyEmailReportRecipient(
                daily_email_report=config,
                user=assignment['user'],
                email=assignment['email'],
                is_active=assignment['is_active'],
            )
            for assignment in assignments
        ])

    return JsonResponse({
        'status': 'ok',
        'message': 'Đã lưu cấu hình báo cáo email hằng ngày.',
        'recipient_count': len(get_daily_email_report_recipients(config)),
    })


@login_required(login_url='/login/')
def api_test_daily_email_report(request):
    if not is_brand_owner(request.user):
        return _daily_email_report_forbidden_json()
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'}, status=405)

    brand = _stock_alert_brand_for_user(request.user)
    if not brand:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy thương hiệu.'}, status=400)
    config = _daily_email_report_for_brand(brand)
    try:
        result = send_daily_email_report(config, is_test=True)
        config.last_test_sent = timezone.now()
        config.last_error = ''
        config.save(update_fields=['last_test_sent', 'last_error', 'updated_at'])
        return JsonResponse({
            'status': 'ok',
            'message': (
                f"Đã gửi báo cáo thử tới {result['sent_recipient_count']} người nhận."
            ),
        })
    except DailyEmailReportConfigurationError as exc:
        logger.warning(
            'Cấu hình báo cáo email hằng ngày chưa hợp lệ cho brand_id=%s: %s',
            brand.id,
            exc,
        )
        config.last_error = str(exc)
        config.save(update_fields=['last_error', 'updated_at'])
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)
    except Exception as exc:
        logger.exception(
            'Không thể gửi thử báo cáo email hằng ngày cho brand_id=%s',
            brand.id,
        )
        config.last_error = str(exc)
        config.save(update_fields=['last_error', 'updated_at'])
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)


@login_required(login_url='/login/')
def stock_alert_email_setting(request):
    if not is_brand_owner(request.user):
        messages.error(request, 'Chỉ chủ thương hiệu mới được cấu hình cảnh báo tồn kho qua email.')
        return redirect('/dashboard/')

    brand = _stock_alert_brand_for_user(request.user)
    if not brand:
        messages.error(request, 'Không tìm thấy thương hiệu để lưu cấu hình.')
        return redirect('/dashboard/')

    config = _stock_alert_for_brand(brand)
    scoped_recipients = list(
        config.email_recipient_scopes.select_related('user').prefetch_related('categories')
    )
    has_scoped_recipients = bool(scoped_recipients)
    scopes_by_user_id = {
        recipient.user_id: recipient
        for recipient in scoped_recipients
        if recipient.user_id
    }
    legacy_user_ids = set(config.recipient_users.values_list('id', flat=True))
    legacy_category_ids = list(config.categories.values_list('id', flat=True))
    staff_options = []
    recipient_assignments = []
    for user in _stock_alert_users_for_brand(brand):
        display_name = user.get_full_name().strip() or user.username
        recipient = scopes_by_user_id.get(user.id)
        selected = (
            recipient.is_active
            if recipient else (
                user.id in legacy_user_ids if not has_scoped_recipients else False
            )
        ) and bool(user.email)
        category_ids = (
            [category.id for category in recipient.categories.all()]
            if recipient else (legacy_category_ids if selected and not has_scoped_recipients else [])
        )
        option = {
            'id': user.id,
            'key': f'user:{user.id}',
            'name': display_name,
            'username': user.username,
            'email': user.email or '',
            'selected': selected,
            'category_ids': category_ids,
        }
        staff_options.append(option)
        recipient_assignments.append({
            'key': option['key'],
            'kind': 'user',
            'user_id': user.id,
            'email': user.email or '',
            'label': display_name,
            'selected': selected,
            'is_active': selected,
            'category_ids': category_ids,
        })

    if has_scoped_recipients:
        extra_recipients = [
            recipient for recipient in scoped_recipients if not recipient.user_id
        ]
    else:
        legacy_emails, _ = parse_recipient_emails(config.email_recipients)
        extra_recipients = [
            {'email': email, 'category_ids': legacy_category_ids}
            for email in legacy_emails
        ]

    for recipient in extra_recipients:
        if isinstance(recipient, dict):
            email = recipient['email']
            category_ids = recipient['category_ids']
        else:
            email = recipient.email
            category_ids = [category.id for category in recipient.categories.all()]
        key = f'email:{email.lower()}'
        recipient_assignments.append({
            'key': key,
            'kind': 'email',
            'user_id': None,
            'email': email,
            'label': email,
            'selected': (
                bool(recipient.get('is_active', True))
                if isinstance(recipient, dict) else recipient.is_active
            ),
            'is_active': (
                bool(recipient.get('is_active', True))
                if isinstance(recipient, dict) else recipient.is_active
            ),
            'category_ids': category_ids,
        })

    context = {
        'active_tab': 'stock_alert_email_setting',
        'stock_alert': config,
        'staff_options': staff_options,
        'recipient_assignments': recipient_assignments,
        'category_options': _stock_alert_category_options(),
        'smtp_ready': not (
            str(getattr(settings, 'EMAIL_BACKEND', '') or '').endswith('smtp.EmailBackend')
            and not getattr(settings, 'EMAIL_HOST_USER', '')
        ),
    }
    return render(request, 'system/stock_alert_email_setting.html', context)


def _parse_id_set(values):
    result = set()
    for value in values or []:
        try:
            result.add(int(value))
        except (TypeError, ValueError):
            continue
    return result


@login_required(login_url='/login/')
def api_save_stock_alert_email_setting(request):
    if not is_brand_owner(request.user):
        return _stock_alert_forbidden_json()
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'}, status=405)

    brand = _stock_alert_brand_for_user(request.user)
    if not brand:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy thương hiệu.'}, status=400)

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Dữ liệu gửi lên không hợp lệ.'}, status=400)

    send_time = parse_time(str(data.get('send_time') or '21:00'))
    if not send_time:
        return JsonResponse({'status': 'error', 'message': 'Giờ gửi không hợp lệ.'}, status=400)

    raw_assignments = data.get('recipient_assignments')
    if not isinstance(raw_assignments, list):
        # Tương thích với payload của giao diện cũ: mọi người nhận dùng chung danh mục.
        legacy_category_ids = list(_parse_id_set(data.get('category_ids')))
        raw_assignments = [
            {'user_id': user_id, 'category_ids': legacy_category_ids}
            for user_id in _parse_id_set(data.get('recipient_user_ids'))
        ]
        legacy_emails, invalid_emails = parse_recipient_emails(
            str(data.get('email_recipients') or '').strip()
        )
        if invalid_emails:
            return JsonResponse({
                'status': 'error',
                'message': 'Email không hợp lệ: ' + ', '.join(invalid_emails),
            }, status=400)
        raw_assignments.extend({
            'email': email,
            'category_ids': legacy_category_ids,
        } for email in legacy_emails)

    allowed_users = {
        user.id: user
        for user in _stock_alert_users_for_brand(brand).exclude(email='')
    }
    assignments = []
    all_category_ids = set()
    seen_recipient_keys = set()
    seen_emails = set()
    for raw_assignment in raw_assignments:
        if not isinstance(raw_assignment, dict):
            return JsonResponse({
                'status': 'error',
                'message': 'Danh sách người nhận không hợp lệ.',
            }, status=400)

        user_id = raw_assignment.get('user_id')
        user = None
        if user_id not in (None, ''):
            try:
                user_id = int(user_id)
            except (TypeError, ValueError):
                user_id = 0
            user = allowed_users.get(user_id)
            if not user:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Có người nhận không thuộc thương hiệu hoặc chưa có email.',
                }, status=400)
            email = user.email.strip().lower()
            recipient_key = f'user:{user.id}'
        else:
            parsed_emails, invalid_emails = parse_recipient_emails(
                str(raw_assignment.get('email') or '').strip()
            )
            if invalid_emails or len(parsed_emails) != 1:
                invalid_value = invalid_emails[0] if invalid_emails else str(raw_assignment.get('email') or '')
                return JsonResponse({
                    'status': 'error',
                    'message': 'Email không hợp lệ: ' + invalid_value,
                }, status=400)
            email = parsed_emails[0]
            recipient_key = f'email:{email}'

        if recipient_key in seen_recipient_keys or email in seen_emails:
            return JsonResponse({
                'status': 'error',
                'message': f'Email {email} đang được chọn trùng.',
            }, status=400)

        recipient_is_active = bool(raw_assignment.get('is_active', True))
        category_ids = _parse_id_set(raw_assignment.get('category_ids'))
        if recipient_is_active and not category_ids:
            return JsonResponse({
                'status': 'error',
                'message': f'Vui lòng chọn ít nhất một danh mục cho email {email}.',
            }, status=400)

        seen_recipient_keys.add(recipient_key)
        seen_emails.add(email)
        all_category_ids.update(category_ids)
        assignments.append({
            'user': user,
            'email': email,
            'is_active': recipient_is_active,
            'category_ids': category_ids,
        })

    selected_categories = ProductCategory.objects.filter(
        id__in=all_category_ids,
        is_active=True,
    )
    if selected_categories.count() != len(all_category_ids):
        return JsonResponse({
            'status': 'error',
            'message': 'Có danh mục không hợp lệ hoặc đã ngừng hoạt động.',
        }, status=400)

    is_active = bool(data.get('is_active', False))
    active_assignments = [
        assignment for assignment in assignments if assignment['is_active']
    ]
    if is_active and not active_assignments:
        return JsonResponse({'status': 'error', 'message': 'Vui lòng bật ít nhất một người nhận.'}, status=400)

    with transaction.atomic():
        config = _stock_alert_for_brand(brand)
        config.email_recipients = ', '.join(
            assignment['email']
            for assignment in active_assignments
            if not assignment['user']
        )
        config.include_child_categories = bool(data.get('include_child_categories', True))
        config.send_time = send_time
        config.alert_on_min = True
        config.alert_on_max = False
        config.is_active = is_active
        config.save()
        config.recipient_users.set(
            assignment['user']
            for assignment in active_assignments
            if assignment['user']
        )
        config.categories.set(selected_categories)
        config.email_recipient_scopes.all().delete()
        for assignment in assignments:
            recipient = StockAlertEmailRecipient.objects.create(
                stock_alert=config,
                user=assignment['user'],
                email=assignment['email'],
                is_active=assignment['is_active'],
            )
            recipient.categories.set(assignment['category_ids'])

    return JsonResponse({
        'status': 'ok',
        'message': 'Đã lưu cấu hình cảnh báo tồn kho qua email.',
        'recipient_count': len(get_stock_alert_recipients(config)),
        'category_count': len(all_category_ids),
    })


@login_required(login_url='/login/')
def api_test_stock_alert_email(request):
    if not is_brand_owner(request.user):
        return _stock_alert_forbidden_json()
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'}, status=405)

    brand = _stock_alert_brand_for_user(request.user)
    if not brand:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy thương hiệu.'}, status=400)
    config = _stock_alert_for_brand(brand)
    try:
        result = send_stock_alert_email(config, is_test=True)
        config.last_test_sent = timezone.now()
        config.last_error = ''
        config.save(update_fields=['last_test_sent', 'last_error', 'updated_at'])
        return JsonResponse({
            'status': 'ok',
            'message': (
                f"Đã gửi email thử tới {result['sent_recipient_count']} người nhận, "
                f"gồm {result['row_count']} sản phẩm tồn thấp."
            ),
        })
    except StockAlertConfigurationError as exc:
        logger.warning('Cấu hình email cảnh báo tồn kho chưa hợp lệ cho brand_id=%s: %s', brand.id, exc)
        config.last_error = str(exc)
        config.save(update_fields=['last_error', 'updated_at'])
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)
    except Exception as exc:
        logger.exception('Không thể gửi thử email cảnh báo tồn kho cho brand_id=%s', brand.id)
        config.last_error = str(exc)
        config.save(update_fields=['last_error', 'updated_at'])
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)


def _parse_sales_report_number(value):
    """Chuyển tham số số từ query string sang float; trả None nếu rỗng hoặc sai định dạng."""
    if value in (None, ''):
        return None
    try:
        return float(str(value).replace(',', '').strip())
    except (TypeError, ValueError):
        return None


def _normalize_report_text(value):
    text = str(value or '').casefold()
    text = ''.join(
        char for char in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(char)
    )
    return text.replace('đ', 'd')


def _parse_filter_int(value):
    if value in (None, ''):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _inventory_valuation_unit_cost(product):
    """Giá dùng để định giá tồn: ưu tiên giá vốn, fallback về giá nhập."""
    cost_price = Decimal(str(product.cost_price or 0))
    if cost_price > 0:
        return cost_price, 'cost_price'

    import_price = Decimal(str(product.import_price or 0))
    if import_price > 0:
        return import_price, 'import_price'

    return Decimal('0'), 'none'


def _build_slow_moving_inventory_payload(request, filters):
    """Cảnh báo hàng còn tồn theo lần bán thực tế gần nhất trên toàn lịch sử."""
    from products.models import Product, ProductStock

    managed_store_ids = get_managed_store_ids(request.user)
    selected_store_id = _parse_filter_int(filters.get('store_id'))
    if selected_store_id is not None:
        if selected_store_id not in managed_store_ids:
            return [], {
                'total_products': 0,
                'never_sold_count': 0,
                'over_30_count': 0,
                'over_60_count': 0,
                'over_90_count': 0,
                'over_90_stock_value': 0,
            }
        scoped_store_ids = [selected_store_id]
    else:
        scoped_store_ids = managed_store_ids

    products_qs = Product.objects.filter(
        store_id__in=scoped_store_ids,
        is_active=True,
        is_service=False,
    ).select_related('category', 'category__parent', 'supplier')
    if filters.get('category_id'):
        products_qs = products_qs.filter(
            _get_product_category_scope_q(filters['category_id'], '')
        )
    if filters.get('product_type_id'):
        products_qs = products_qs.filter(
            _get_product_category_direct_q(filters['product_type_id'], '')
        )
    if filters.get('product_id'):
        products_qs = products_qs.filter(id=filters['product_id'])

    products = list(products_qs.order_by('name', 'id'))
    product_ids = [product.id for product in products]
    if not product_ids:
        return [], {
            'total_products': 0,
            'never_sold_count': 0,
            'over_30_count': 0,
            'over_60_count': 0,
            'over_90_count': 0,
            'over_90_stock_value': 0,
        }

    stock_by_product = {
        row['product_id']: Decimal(str(row['total_stock'] or 0))
        for row in ProductStock.objects.filter(
            product_id__in=product_ids,
            warehouse__store_id__in=scoped_store_ids,
            warehouse__is_deleted=False,
        ).values('product_id').annotate(total_stock=Sum('quantity'))
    }
    stocked_product_ids = [
        product_id
        for product_id, quantity in stock_by_product.items()
        if quantity > 0
    ]
    if not stocked_product_ids:
        return [], {
            'total_products': 0,
            'never_sold_count': 0,
            'over_30_count': 0,
            'over_60_count': 0,
            'over_90_count': 0,
            'over_90_stock_value': 0,
        }

    today = datetime.now().date()
    last_sale_by_product = {}
    last_sale_rows = OrderItem.objects.filter(
            product_id__in=stocked_product_ids,
            quantity__gt=0,
            order__store_id__in=scoped_store_ids,
            order__status__in=[4, 5],
            order__is_deleted=False,
        ).filter(
            Q(order__exported_at__date__lte=today)
            | Q(order__exported_at__isnull=True, order__order_date__lte=today)
        ).values('product_id').annotate(
            last_sale_at=Max('order__exported_at'),
            legacy_last_sale_date=Max(
                'order__order_date',
                filter=Q(order__exported_at__isnull=True),
            ),
        )
    for row in last_sale_rows:
        last_sale_at = row['last_sale_at']
        exported_date = None
        if last_sale_at:
            if settings.USE_TZ and timezone.is_aware(last_sale_at):
                last_sale_at = timezone.localtime(last_sale_at)
            exported_date = last_sale_at.date()
        last_sale_by_product[row['product_id']] = max(
            (
                value
                for value in (exported_date, row['legacy_last_sale_date'])
                if value
            ),
            default=None,
        )

    rows = []
    stocked_product_id_set = set(stocked_product_ids)
    for product in products:
        if product.id not in stocked_product_id_set:
            continue

        last_sale_date = last_sale_by_product.get(product.id)
        never_sold = last_sale_date is None
        days_without_sale = (
            max((today - last_sale_date).days, 0)
            if last_sale_date else None
        )

        if never_sold:
            warning_level = 'critical'
            warning_label = 'Chưa từng bán'
            suggested_action = 'Kiểm tra giá/hiển thị và ưu tiên kích cầu'
        elif days_without_sale >= 90:
            warning_level = 'critical'
            warning_label = 'Rất chậm'
            suggested_action = 'Ưu tiên xả hàng, combo hoặc giảm giá'
        elif days_without_sale >= 60:
            warning_level = 'slow'
            warning_label = 'Chậm'
            suggested_action = 'Lập chương trình khuyến mãi'
        elif days_without_sale >= 30:
            warning_level = 'watch'
            warning_label = 'Theo dõi'
            suggested_action = 'Tăng trưng bày và kích cầu'
        else:
            warning_level = 'normal'
            warning_label = 'Bình thường'
            suggested_action = 'Duy trì bán hàng'

        category_name = ''
        if product.category:
            category_name = product.category.name
            if product.category.parent:
                category_name = f'{product.category.parent.name} / {category_name}'
        total_stock = stock_by_product[product.id]
        valuation_price, valuation_source = _inventory_valuation_unit_cost(product)
        stock_value = total_stock * valuation_price
        rows.append({
            'product_id': product.id,
            'product_code': product.code,
            'product_name': product.name,
            'category': category_name or 'Chưa phân loại',
            'supplier': product.supplier.name if product.supplier else 'Chưa có NCC',
            'last_sale_date': last_sale_date.isoformat() if last_sale_date else '',
            'last_sale_date_display': last_sale_date.strftime('%d/%m/%Y') if last_sale_date else '',
            'days_without_sale': days_without_sale,
            'never_sold': never_sold,
            'stock': float(total_stock),
            'valuation_price': float(valuation_price),
            'valuation_source': valuation_source,
            'stock_value': float(stock_value),
            'warning_level': warning_level,
            'warning_label': warning_label,
            'suggested_action': suggested_action,
        })

    # Chưa từng bán là một cảnh báo riêng; các mặt hàng còn lại ưu tiên theo số ngày.
    rows.sort(key=lambda row: (
        0 if row['never_sold'] else 1,
        -(row['days_without_sale'] or 0),
        -row['stock_value'],
        row['product_name'].casefold(),
    ))

    def is_over(row, days):
        return not row['never_sold'] and row['days_without_sale'] >= days

    over_90_rows = [row for row in rows if is_over(row, 90)]
    summary = {
        'total_products': len(rows),
        'never_sold_count': sum(1 for row in rows if row['never_sold']),
        'over_30_count': sum(1 for row in rows if is_over(row, 30)),
        'over_60_count': sum(1 for row in rows if is_over(row, 60)),
        'over_90_count': len(over_90_rows),
        'over_90_stock_value': sum(row['stock_value'] for row in over_90_rows),
    }
    return rows, summary


def _report_lookup(prefix, suffix):
    return f'{prefix}__{suffix}' if prefix else suffix


def _group_name_keyword_q(prefix, keywords):
    query = Q()
    for keyword in keywords:
        query |= Q(**{_report_lookup(prefix, 'group__name__icontains'): keyword})
    return query


def _get_sales_report_customer_kind_q(kind, prefix='customer'):
    """Ưu tiên field customer_kind; fallback sang suy luận từ nhóm KH cho dữ liệu cũ."""
    if kind not in {'retail', 'wholesale', 'other'}:
        return Q()

    kind_lookup = _report_lookup(prefix, 'customer_kind')
    explicit_kind_q = Q(**{kind_lookup: kind})
    blank_kind_q = Q(**{kind_lookup: ''})
    retail_legacy_q = blank_kind_q & _group_name_keyword_q(prefix, RETAIL_GROUP_KEYWORDS)
    wholesale_legacy_q = blank_kind_q & _group_name_keyword_q(prefix, WHOLESALE_GROUP_KEYWORDS)

    if kind == 'retail':
        if prefix:
            return explicit_kind_q | retail_legacy_q | Q(**{_report_lookup(prefix, 'isnull'): True})
        return explicit_kind_q | retail_legacy_q
    if kind == 'wholesale':
        return explicit_kind_q | wholesale_legacy_q
    known_legacy_q = retail_legacy_q | wholesale_legacy_q
    return explicit_kind_q | (blank_kind_q & ~known_legacy_q)


def _get_sales_report_return_customer_kind_q(kind):
    if kind not in {'retail', 'wholesale', 'other'}:
        return Q()
    return (
        (Q(order__isnull=False) & _get_sales_report_customer_kind_q(kind, 'order__customer')) |
        (Q(order__isnull=True) & _get_sales_report_customer_kind_q(kind, 'customer'))
    )


def _get_fully_returned_sales_order_ids(order_ids):
    """Return orders whose sellable product quantities were fully returned.

    A completed return makes the original sale no longer relevant for the
    loss-warning card. Prefer quantities because the refund price may differ
    from the original sale price, with a value-only fallback for legacy
    returns that do not have return item rows.
    """
    order_ids = list(order_ids or [])
    if not order_ids:
        return set()

    sold_by_order = defaultdict(dict)
    sold_amount_by_order = defaultdict(Decimal)
    sold_rows = (
        OrderItem.objects
        .filter(
            order_id__in=order_ids,
            product_id__isnull=False,
            product__is_service=False,
        )
        .values('order_id', 'product_id')
        .annotate(quantity=Sum('quantity'), amount=Sum('total_price'))
    )
    for row in sold_rows:
        sold_by_order[row['order_id']][row['product_id']] = Decimal(str(row['quantity'] or 0))
        sold_amount_by_order[row['order_id']] += Decimal(str(row['amount'] or 0))

    returned_by_order = defaultdict(lambda: defaultdict(Decimal))
    returned_rows = (
        OrderReturnItem.objects
        .filter(
            order_return__order_id__in=order_ids,
            order_return__status=2,
            order_return__is_deleted=False,
        )
        .values('order_return__order_id', 'product_id')
        .annotate(quantity=Sum('quantity'))
    )
    for row in returned_rows:
        returned_by_order[row['order_return__order_id']][row['product_id']] = Decimal(
            str(row['quantity'] or 0)
        )

    # Phiếu hoàn cũ có thể chỉ lưu tổng "giá trị hàng trả" mà không có dòng
    # sản phẩm. Khi tổng giá trị đó hoàn đủ tiền hàng, vẫn phải coi đơn là đã
    # hoàn toàn để không phát cảnh báo bán lỗ cho dữ liệu lịch sử.
    completed_return_amounts = {
        row['order_id']: Decimal(str(row['amount'] or 0))
        for row in OrderReturn.objects.filter(
            order_id__in=order_ids,
            status=2,
            is_deleted=False,
        ).values('order_id').annotate(amount=Sum('return_amount'))
    }
    completed_return_item_counts = {
        row['order_id']: row['item_count']
        for row in OrderReturn.objects.filter(
            order_id__in=order_ids,
            status=2,
            is_deleted=False,
        ).values('order_id').annotate(item_count=Count('items'))
    }

    fully_returned = set()
    for order_id, sold_products in sold_by_order.items():
        returned_products = returned_by_order.get(order_id, {})
        quantity_complete = sold_products and all(
            returned_products.get(product_id, Decimal('0')) >= quantity
            for product_id, quantity in sold_products.items()
        )
        amount_only_complete = (
            completed_return_item_counts.get(order_id, 0) == 0
            and sold_amount_by_order.get(order_id, Decimal('0')) > 0
            and completed_return_amounts.get(order_id, Decimal('0')) >= sold_amount_by_order[order_id]
        )
        if quantity_complete or amount_only_complete:
            fully_returned.add(order_id)
    return fully_returned


def _classify_sales_report_customer_kind(customer):
    if not customer:
        return 'retail', 'Khách lẻ'
    explicit_kind = str(getattr(customer, 'customer_kind', '') or '').strip()
    explicit_label = next(
        (option['label'] for option in CUSTOMER_KIND_OPTIONS if option['value'] == explicit_kind),
        '',
    )
    if explicit_label:
        return explicit_kind, explicit_label
    group_name = customer.group.name if getattr(customer, 'group', None) else ''
    normalized_group = _normalize_report_text(group_name)
    if any(keyword in normalized_group for keyword in ('si', 'buon', 'ban buon', 'dai ly', 'wholesale')):
        return 'wholesale', 'Khách buôn / sỉ'
    if any(keyword in normalized_group for keyword in ('le', 'ban le', 'retail')):
        return 'retail', 'Khách lẻ'
    return 'other', 'Khác / chưa phân loại'


def _get_product_category_scope_q(category_id, prefix):
    category_id = _parse_filter_int(category_id)
    if category_id is None:
        return Q()
    return (
        Q(**{f'{prefix}category_id': category_id}) |
        Q(**{f'{prefix}category__parent_id': category_id})
    )


def _get_product_category_direct_q(category_id, prefix):
    category_id = _parse_filter_int(category_id)
    if category_id is None:
        return Q()
    return Q(**{f'{prefix}category_id': category_id})


def sales_report_privileged_required(view_func):
    """Báo cáo bán hàng chỉ cho Chủ thương hiệu / Giám đốc / Kế toán."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if can_view_sales_report(request.user):
            return view_func(request, *args, **kwargs)
        message = 'Chỉ tài khoản Chủ thương hiệu, Giám đốc hoặc Kế toán mới được xem báo cáo bán hàng.'
        if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': message}, status=403)
        messages.error(request, message)
        from django.shortcuts import redirect
        return redirect('/dashboard/')
    return wrapper


def quotation_profit_report_privileged_required(view_func):
    """BC LN dự kiến: Chủ thương hiệu / Giám đốc / Kế toán / Quản lý cửa hàng."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if can_view_quotation_profit_report(request.user):
            return view_func(request, *args, **kwargs)
        message = (
            'Chỉ tài khoản Chủ thương hiệu, Giám đốc, Kế toán hoặc '
            'Quản lý cửa hàng mới được xem báo cáo lợi nhuận dự kiến.'
        )
        if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': message}, status=403)
        messages.error(request, message)
        return redirect('/dashboard/')
    return wrapper


def _get_sales_report_filters(request):
    today = datetime.now().date()
    from_date = request.GET.get('from_date') or today.replace(day=1).strftime('%Y-%m-%d')
    to_date = request.GET.get('to_date') or today.strftime('%Y-%m-%d')
    time_group = (request.GET.get('time_group') or 'day').strip().lower()
    if time_group not in ('day', 'month', 'year'):
        time_group = 'day'
    order_scope = (request.GET.get('order_scope') or 'realized').strip().lower()
    if order_scope not in ('realized', 'all_active'):
        order_scope = 'realized'
    return {
        'from_date': from_date,
        'to_date': to_date,
        'time_group': time_group,
        'order_scope': order_scope,
        'store_id': request.GET.get('store_id') or '',
        'customer_kind': request.GET.get('customer_kind', '').strip(),
        'customer_group_id': request.GET.get('customer_group_id') or '',
        'category_id': request.GET.get('category_id') or '',
        'product_type_id': request.GET.get('product_type_id') or '',
        'profit_filter': request.GET.get('profit_filter', '').strip(),
        'customer_id': request.GET.get('customer_id') or '',
        'product_id': request.GET.get('product_id') or '',
        'salesperson': request.GET.get('salesperson', '').strip(),
        'search': request.GET.get('search', '').strip(),
        'revenue_min': _parse_sales_report_number(request.GET.get('revenue_min')),
        'revenue_max': _parse_sales_report_number(request.GET.get('revenue_max')),
        'cost_min': _parse_sales_report_number(request.GET.get('cost_min')),
        'cost_max': _parse_sales_report_number(request.GET.get('cost_max')),
        'line_profit_min': _parse_sales_report_number(request.GET.get('line_profit_min')),
        'line_profit_max': _parse_sales_report_number(request.GET.get('line_profit_max')),
        'profit_min': _parse_sales_report_number(request.GET.get('profit_min')),
        'profit_max': _parse_sales_report_number(request.GET.get('profit_max')),
    }


def _report_date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    return parse_date(str(value or ''))


def _exported_at_period_q(from_date, to_date, field_prefix=''):
    """Lọc doanh thu theo thời điểm xuất kho, gồm trọn ngày kết thúc."""
    start_date = _report_date_value(from_date)
    end_date = _report_date_value(to_date)
    if not start_date or not end_date:
        return Q(pk__in=[])

    start_at = datetime.combine(start_date, time.min)
    end_at = datetime.combine(end_date + timedelta(days=1), time.min)
    if settings.USE_TZ:
        current_timezone = timezone.get_current_timezone()
        start_at = timezone.make_aware(start_at, current_timezone)
        end_at = timezone.make_aware(end_at, current_timezone)
    exported_q = Q(**{
        f'{field_prefix}exported_at__gte': start_at,
        f'{field_prefix}exported_at__lt': end_at,
    })
    # Migration 0029 đã backfill đơn cũ. Nhánh này chỉ bảo vệ dữ liệu
    # legacy/import thủ công chưa có exported_at khỏi biến mất báo cáo.
    legacy_q = Q(**{
        f'{field_prefix}exported_at__isnull': True,
        f'{field_prefix}order_date__gte': start_date,
        f'{field_prefix}order_date__lte': end_date,
    })
    return exported_q | legacy_q


def _order_revenue_date(order):
    exported_at = order.exported_at
    if not exported_at:
        return order.order_date
    if timezone.is_aware(exported_at):
        exported_at = timezone.localtime(exported_at)
    return exported_at.date()


def _get_sales_report_time_group_meta(time_group):
    """Trả metadata gom nhóm thời gian cho báo cáo bán hàng."""
    if time_group == 'month':
        return {'label': 'Tháng', 'key_format': '%Y-%m', 'display_format': '%m/%Y'}
    if time_group == 'year':
        return {'label': 'Năm', 'key_format': '%Y', 'display_format': '%Y'}
    return {'label': 'Ngày', 'key_format': '%Y-%m-%d', 'display_format': '%d/%m/%Y'}


def _get_sales_report_filter_labels(filters):
    from customers.models import Customer, CustomerGroup
    from products.models import Product, ProductCategory
    from system_management.models import Store

    def _lookup_name(model, raw_id):
        lookup_id = _parse_filter_int(raw_id)
        if lookup_id is None:
            return str(raw_id or '').strip()
        return model.objects.filter(id=lookup_id).values_list('name', flat=True).first() or str(raw_id)

    filter_labels = []
    if filters.get('time_group'):
        filter_labels.append(f"Xem theo: {_get_sales_report_time_group_meta(filters['time_group'])['label']}")
    order_scope_label = {
        'realized': 'Đã xuất kho + Hoàn thành',
        'all_active': 'Tất cả đơn chưa hủy',
    }.get(filters.get('order_scope'))
    if order_scope_label:
        filter_labels.append(f"Phạm vi đơn: {order_scope_label}")
    filter_labels.append(
        'Mốc ngày: Ngày đặt hàng'
        if filters.get('order_scope') == 'all_active'
        else 'Mốc ghi nhận doanh thu: Ngày xuất kho'
    )
    if filters.get('store_id'):
        filter_labels.append(f"Cửa hàng: {_lookup_name(Store, filters['store_id'])}")
    if filters.get('customer_kind'):
        kind_label = next(
            (option['label'] for option in CUSTOMER_KIND_OPTIONS if option['value'] == filters['customer_kind']),
            filters['customer_kind'],
        )
        filter_labels.append(f"Kiểu khách: {kind_label}")
    if filters.get('customer_group_id'):
        filter_labels.append(f"Nhóm KH: {_lookup_name(CustomerGroup, filters['customer_group_id'])}")
    if filters.get('category_id'):
        filter_labels.append(f"Nhóm mặt hàng: {_lookup_name(ProductCategory, filters['category_id'])}")
    if filters.get('product_type_id'):
        filter_labels.append(f"Loại SP: {_lookup_name(ProductCategory, filters['product_type_id'])}")
    if filters.get('customer_id'):
        customer_name = _lookup_name(Customer, filters['customer_id'])
        filter_labels.append(f"Khách hàng: {customer_name}")
    if filters.get('product_id'):
        product_name = _lookup_name(Product, filters['product_id'])
        filter_labels.append(f"Mặt hàng: {product_name}")
    if filters.get('salesperson'):
        filter_labels.append(f"Nhân viên: {filters['salesperson']}")
    if filters.get('search'):
        filter_labels.append(f"Từ khóa: {filters['search']}")
    if filters.get('profit_filter'):
        profit_label = {
            'profit': 'Có lãi',
            'loss': 'Báo lỗ',
        }.get(filters['profit_filter'], filters['profit_filter'])
        filter_labels.append(f"Lợi nhuận: {profit_label}")
    for key, label in (
        ('revenue_min', 'DT từ'),
        ('revenue_max', 'DT đến'),
        ('cost_min', 'GV từ'),
        ('cost_max', 'GV đến'),
        ('line_profit_min', 'LN dòng từ'),
        ('line_profit_max', 'LN dòng đến'),
        ('profit_min', 'LN gộp từ'),
        ('profit_max', 'LN gộp đến'),
    ):
        if filters.get(key) is not None:
            value = filters[key]
            filter_labels.append(f"{label}: {int(value) if float(value).is_integer() else value}")
    return filter_labels


def _filter_sales_returns_by_scope(queryset, request):
    """Lọc phiếu trả theo store hợp lệ.

    Ưu tiên liên kết order -> store. Với dữ liệu dev/legacy thiếu order, fallback sang
    warehouse.store rồi customer.store để không làm rơi dữ liệu vẫn còn suy luận được.
    """
    if request.user.is_superuser:
        return queryset.none()
    managed_ids = get_managed_store_ids(request.user)
    if not managed_ids:
        return queryset.none()
    return queryset.filter(
        Q(order__store_id__in=managed_ids) |
        Q(order__isnull=True, warehouse__store_id__in=managed_ids) |
        Q(order__isnull=True, warehouse__isnull=True, customer__store_id__in=managed_ids)
    ).distinct()


def _get_order_revenue_staff_name(order):
    """Người được ghi nhận doanh thu của đơn.

    NV bán hàng là giá trị chủ tài khoản có thể gán thủ công và luôn được ưu
    tiên. Nếu chưa gán, doanh thu thuộc về tài khoản thực sự đã tạo đơn.
    `creator_name` chỉ dùng làm dự phòng cho dữ liệu cũ không còn `created_by`.
    """
    assigned_name = (getattr(order, 'salesperson', None) or '').strip()
    if assigned_name:
        return assigned_name

    creator = getattr(order, 'created_by', None)
    if creator:
        creator_name = (creator.get_full_name() or creator.username or '').strip()
        if creator_name:
            return creator_name

    return (getattr(order, 'creator_name', None) or '').strip()


def _group_order_ids_by_revenue_staff(orders):
    grouped_order_ids = defaultdict(list)
    for order in orders.select_related('created_by'):
        staff_name = _get_order_revenue_staff_name(order) or '(Chưa gán NV)'
        grouped_order_ids[staff_name].append(order.id)
    return grouped_order_ids


def _get_salesperson_filter_options(request, store_id=''):
    """Danh sách nhân viên cho filter báo cáo bán hàng trong phạm vi store hợp lệ."""
    from django.contrib.auth.models import User

    managed_ids = get_managed_store_ids(request.user)
    if not managed_ids:
        return []

    scoped_store_ids = list(managed_ids)
    if store_id:
        try:
            selected_store_id = int(store_id)
        except (TypeError, ValueError):
            return []
        if selected_store_id not in scoped_store_ids:
            return []
        scoped_store_ids = [selected_store_id]

    names = set()

    users = User.objects.filter(
        is_active=True,
        profile__store_id__in=scoped_store_ids,
    ).distinct().order_by('last_name', 'first_name', 'username')
    for user in users:
        name = (user.get_full_name() or user.username or '').strip()
        if name:
            names.add(name)

    legacy_salespersons = Order.objects.filter(
        store_id__in=scoped_store_ids,
    ).exclude(status=6).exclude(
        salesperson__isnull=True
    ).exclude(salesperson='').values_list('salesperson', flat=True)
    for name in legacy_salespersons:
        normalized = (name or '').strip()
        if normalized:
            names.add(normalized)

    unassigned_orders = Order.objects.filter(
        store_id__in=scoped_store_ids,
    ).exclude(status=6).filter(
        Q(salesperson__isnull=True) | Q(salesperson='')
    )
    creator_ids = unassigned_orders.exclude(
        created_by_id__isnull=True
    ).values_list('created_by_id', flat=True).distinct()
    for creator in User.objects.filter(id__in=creator_ids):
        name = (creator.get_full_name() or creator.username or '').strip()
        if name:
            names.add(name)

    legacy_creator_names = unassigned_orders.filter(
        created_by_id__isnull=True,
    ).exclude(
        creator_name__isnull=True,
    ).exclude(
        creator_name='',
    ).values_list('creator_name', flat=True).distinct()
    for name in legacy_creator_names:
        normalized = (name or '').strip()
        if normalized:
            names.add(normalized)

    return sorted(names, key=lambda value: value.casefold())


def _build_sales_report_payload(request, include_filter_options=True):
    from customers.models import CustomerGroup, Customer
    from products.models import ProductCategory, Product, GoodsReceipt
    from system_management.models import Store

    filters = _get_sales_report_filters(request)
    line_profit_scope = filters['line_profit_min'] is not None or filters['line_profit_max'] is not None
    item_scope = bool(
        filters['category_id'] or filters['product_type_id'] or filters['product_id'] or line_profit_scope
    )

    def _matches_metric_filters(row):
        revenue = float(row.get('revenue', row.get('amount', 0)) or 0)
        cost = float(row.get('cost', 0) or 0)
        profit = float(row.get('profit', revenue - cost) or 0)
        if filters['revenue_min'] is not None and revenue < filters['revenue_min']:
            return False
        if filters['revenue_max'] is not None and revenue > filters['revenue_max']:
            return False
        if filters['cost_min'] is not None and cost < filters['cost_min']:
            return False
        if filters['cost_max'] is not None and cost > filters['cost_max']:
            return False
        if filters['profit_min'] is not None and profit < filters['profit_min']:
            return False
        if filters['profit_max'] is not None and profit > filters['profit_max']:
            return False
        return True

    def _matches_line_profit_filters(row):
        line_profit = float(row.get('line_profit', row.get('profit', 0)) or 0)
        if filters['line_profit_min'] is not None and line_profit < filters['line_profit_min']:
            return False
        if filters['line_profit_max'] is not None and line_profit > filters['line_profit_max']:
            return False
        return True

    def _effective_product_unit_cost(product):
        if not product:
            return 0.0
        for candidate in (product.cost_price, product.import_price):
            value = float(candidate or 0)
            if value > 0:
                return value
        if product.is_combo:
            combo_cost = 0.0
            for combo_item in product.combo_items.select_related('product').all():
                component_cost = float(
                    combo_item.product.cost_price
                    or combo_item.product.import_price
                    or 0
                )
                combo_cost += component_cost * float(combo_item.quantity or 0)
            if combo_cost > 0:
                return combo_cost
        return 0.0

    def _effective_item_unit_cost(item):
        candidates = [item.cost_price]
        if item.variant_id:
            candidates.extend([item.variant.cost_price, item.variant.import_price])
        for candidate in candidates:
            value = float(candidate or 0)
            if value > 0:
                return value
        return _effective_product_unit_cost(item.product if item.product_id else None)

    if filters['order_scope'] == 'all_active':
        orders_qs = Order.objects.filter(
            order_date__gte=filters['from_date'],
            order_date__lte=filters['to_date'],
        ).exclude(status=6)
    else:
        orders_qs = Order.objects.filter(
            _exported_at_period_q(filters['from_date'], filters['to_date']),
            status__in=[4, 5],
        )
    orders_qs = orders_qs.select_related(
        'customer',
        'customer__group',
        'warehouse',
        'store',
        'created_by',
        'source_return_exchange',
        'source_return_exchange__order',
    )
    orders_qs = filter_by_store(orders_qs, request)

    if filters['store_id']:
        orders_qs = orders_qs.filter(store_id=filters['store_id'])
    if filters['customer_kind']:
        orders_qs = orders_qs.filter(_get_sales_report_customer_kind_q(filters['customer_kind']))
    if filters['customer_group_id']:
        orders_qs = orders_qs.filter(customer__group_id=filters['customer_group_id'])
    if filters['customer_id']:
        orders_qs = orders_qs.filter(customer_id=filters['customer_id'])
    if filters['category_id']:
        orders_qs = orders_qs.filter(_get_product_category_scope_q(filters['category_id'], 'items__product__'))
    if filters['product_type_id']:
        orders_qs = orders_qs.filter(_get_product_category_direct_q(filters['product_type_id'], 'items__product__'))
    if filters['product_id']:
        orders_qs = orders_qs.filter(items__product_id=filters['product_id'])
    if filters['search']:
        search = filters['search']
        orders_qs = orders_qs.filter(
            Q(code__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(customer__phone__icontains=search) |
            Q(tags__icontains=search) |
            Q(note__icontains=search) |
            Q(salesperson__icontains=search) |
            Q(items__product__name__icontains=search) |
            Q(items__product__code__icontains=search)
        )
    orders_qs = orders_qs.distinct()

    order_by_date = '-order_date' if filters['order_scope'] == 'all_active' else '-exported_at'
    orders_list = list(orders_qs.order_by(order_by_date, '-id'))
    if filters['salesperson']:
        selected_salesperson = filters['salesperson'].strip().casefold()
        orders_list = [
            order for order in orders_list
            if _get_order_revenue_staff_name(order).casefold() == selected_salesperson
        ]

    order_items_qs = OrderItem.objects.filter(order__in=orders_qs).select_related(
        'product', 'product__supplier', 'product__category', 'product__category__parent', 'variant',
        'order', 'order__customer', 'order__customer__group', 'order__created_by',
        'order__source_return_exchange', 'order__source_return_exchange__order',
    )
    if filters['category_id']:
        order_items_qs = order_items_qs.filter(_get_product_category_scope_q(filters['category_id'], 'product__'))
    if filters['product_type_id']:
        order_items_qs = order_items_qs.filter(_get_product_category_direct_q(filters['product_type_id'], 'product__'))
    if filters['product_id']:
        order_items_qs = order_items_qs.filter(product_id=filters['product_id'])

    order_items = list(order_items_qs)
    fully_returned_order_ids = _get_fully_returned_sales_order_ids(
        [order.id for order in orders_list]
    )
    adjusted_item_rows = []
    for item in order_items:
        base_total = float(item.order.total_amount or 0)
        final_total = max(float(item.order.final_amount or 0), 0)
        quantity = float(item.quantity or 0)
        unit_price = float(item.unit_price or 0)
        gross_line_amount = unit_price * quantity
        line_revenue = float(item.total_price or 0)
        line_discount_amount = max(gross_line_amount - line_revenue, 0)
        unit_cost = _effective_item_unit_cost(item)
        line_cost = unit_cost * quantity
        line_share = line_revenue / base_total if base_total > 0 else 0
        order_discount_allocated = float(item.order.discount_amount or 0) * line_share
        shipping_fee_allocated = float(item.order.shipping_fee or 0) * line_share
        other_fee_allocated = float(item.order.other_fee or 0) * line_share
        if base_total > 0:
            adjusted_revenue = line_revenue * final_total / base_total
        else:
            adjusted_revenue = line_revenue

        product = item.product
        order = item.order
        source_return = getattr(order, 'source_return_exchange', None)
        category = product.category if product else None
        root_category = category.parent if category and category.parent_id else category
        product_type = category if category and category.parent_id else None
        line_profit = adjusted_revenue - line_cost
        salesperson = _get_order_revenue_staff_name(order)
        revenue_date = (
            order.order_date
            if filters['order_scope'] == 'all_active'
            else _order_revenue_date(order)
        )
        adjusted_item_rows.append({
            'id': item.id,
            'order_id': item.order_id,
            'order_code': order.code,
            'date': revenue_date.strftime('%d/%m/%Y') if revenue_date else '',
            'date_raw': revenue_date.strftime('%Y-%m-%d') if revenue_date else '',
            'customer_name': order.customer.name if order.customer else '',
            'salesperson': salesperson,
            'product_id': item.product_id,
            'product_name': product.name if product else (item.item_name or 'Dịch vụ'),
            'sku': item.variant.sku if item.variant else (product.code if product else 'DV'),
            'supplier_id': product.supplier_id if product else None,
            'supplier_name': product.supplier.name if product and product.supplier else '',
            'category_name': root_category.name if root_category else '',
            'product_type_name': product_type.name if product_type else '',
            'quantity': quantity,
            'unit_price': unit_price,
            'listed_unit_price': float(product.selling_price or 0) if product else 0,
            'gross_line_amount': gross_line_amount,
            'line_discount_amount': line_discount_amount,
            'order_discount_allocated': order_discount_allocated,
            'shipping_fee_allocated': shipping_fee_allocated,
            'other_fee_allocated': other_fee_allocated,
            'goods_amount': line_revenue,
            'revenue': adjusted_revenue,
            'unit_cost': unit_cost,
            'cost': line_cost,
            'line_profit': line_profit,
            'is_exchange_order': bool(source_return),
            'return_code': source_return.code if source_return else '',
            'source_order_code': (
                source_return.order.code
                if source_return and source_return.order
                else ''
            ),
            'return_amount': float(source_return.return_amount or 0) if source_return else 0,
            'exchange_amount': float(source_return.exchange_amount or 0) if source_return else 0,
            'exchange_offset_amount': float(order.discount_amount or 0) if source_return else 0,
            'amount_due': float(source_return.amount_due or 0) if source_return else 0,
            'return_reason': source_return.reason or '' if source_return else '',
            'exchange_note': source_return.exchange_note or '' if source_return else '',
        })

    loss_items_by_order = defaultdict(list)
    for item_row in adjusted_item_rows:
        if item_row['line_profit'] >= 0:
            continue
        quantity = float(item_row['quantity'] or 0)
        unit_revenue = item_row['revenue'] / quantity if quantity > 0 else item_row['revenue']
        unit_cost = item_row['cost'] / quantity if quantity > 0 else item_row['cost']
        loss_items_by_order[item_row['order_id']].append({
            'product_name': item_row['product_name'],
            'sku': item_row['sku'],
            'quantity': quantity,
            'unit_price': item_row['unit_price'],
            'listed_unit_price': item_row['listed_unit_price'],
            'gross_line_amount': item_row['gross_line_amount'],
            'line_discount_amount': item_row['line_discount_amount'],
            'order_discount_allocated': item_row['order_discount_allocated'],
            'shipping_fee_allocated': item_row['shipping_fee_allocated'],
            'other_fee_allocated': item_row['other_fee_allocated'],
            'goods_amount': item_row['goods_amount'],
            'net_revenue': item_row['revenue'],
            'unit_revenue': unit_revenue,
            'unit_cost': unit_cost,
            'total_cost': item_row['cost'],
            'line_profit': item_row['line_profit'],
            'loss_amount': abs(item_row['line_profit']),
            'is_exchange_order': item_row['is_exchange_order'],
            'return_code': item_row['return_code'],
            'source_order_code': item_row['source_order_code'],
            'return_amount': item_row['return_amount'],
            'exchange_amount': item_row['exchange_amount'],
            'exchange_offset_amount': item_row['exchange_offset_amount'],
            'amount_due': item_row['amount_due'],
            'return_reason': item_row['return_reason'],
            'exchange_note': item_row['exchange_note'],
        })

    if line_profit_scope:
        adjusted_item_rows = [row for row in adjusted_item_rows if _matches_line_profit_filters(row)]

    order_item_map = defaultdict(lambda: {'goods_amount': 0.0, 'revenue': 0.0, 'cost': 0.0})
    for item_row in adjusted_item_rows:
        order_item_map[item_row['order_id']]['goods_amount'] += item_row['goods_amount']
        order_item_map[item_row['order_id']]['revenue'] += item_row['revenue']
        order_item_map[item_row['order_id']]['cost'] += item_row['cost']

    order_rows = []
    for order in orders_list:
        if item_scope and order.id not in order_item_map:
            continue
        item_totals = order_item_map.get(order.id, {'goods_amount': 0, 'revenue': 0, 'cost': 0})
        goods_amount = item_totals['goods_amount'] if item_scope else float(max(order.total_amount or 0, 0))
        revenue = item_totals['revenue'] if item_scope else float(max(order.final_amount or 0, 0))
        cost = item_totals['cost']
        order_goods_total = float(max(order.total_amount or 0, 0))
        scope_ratio = (goods_amount / order_goods_total) if item_scope and order_goods_total > 0 else 1.0
        discount_amount = float(order.discount_amount or 0) * scope_ratio
        shipping_fee = float(order.shipping_fee or 0) * scope_ratio
        other_fee = float(order.other_fee or 0) * scope_ratio

        if item_scope:
            base_amount = float(order.final_amount or 0)
            if base_amount > 0:
                paid = min(float(order.paid_amount or 0) * revenue / base_amount, revenue)
            else:
                paid = 0
        else:
            paid = float(order.paid_amount or 0)

        profit = revenue - cost
        loss_products = loss_items_by_order.get(order.id, [])
        customer_kind, customer_kind_label = _classify_sales_report_customer_kind(order.customer)
        loss_products_for_order = [] if order.id in fully_returned_order_ids else loss_products
        revenue_date = (
            order.order_date
            if filters['order_scope'] == 'all_active'
            else _order_revenue_date(order)
        )
        order_rows.append({
            'id': order.id,
            'code': order.code,
            'date': revenue_date.strftime('%d/%m/%Y') if revenue_date else '',
            'date_raw': revenue_date.strftime('%Y-%m-%d') if revenue_date else '',
            'customer': order.customer.name if order.customer else '',
            'customer_id': order.customer_id,
            'customer_kind': customer_kind,
            'customer_kind_label': customer_kind_label,
            'customer_group': order.customer.group.name if order.customer and order.customer.group else '',
            'customer_group_id': order.customer.group_id if order.customer else None,
            'store_id': order.store_id,
            'store_name': order.store.name if order.store else '',
            'salesperson': _get_order_revenue_staff_name(order),
            'goods_amount': goods_amount,
            'discount_amount': discount_amount,
            'shipping_fee': shipping_fee,
            'other_fee': other_fee,
            'revenue': revenue,
            'paid': paid,
            'debt': revenue - paid,
            'cost': cost,
            'profit': profit,
            # A fully returned sale is no longer a realized loss.  Keep the
            # original order values for audit/report totals, but suppress the
            # loss warning and its loss-product details.
            'is_loss': profit < 0 and order.id not in fully_returned_order_ids,
            'loss_products': loss_products_for_order,
            'loss_product_names': ', '.join(
                row['product_name'] for row in loss_products_for_order
            ),
            'status': order.status,
            'status_display': order.get_status_display(),
            'payment_status': order.payment_status,
            'payment_status_display': order.get_payment_status_display(),
        })

    if filters['profit_filter'] == 'loss':
        order_rows = [row for row in order_rows if row['is_loss']]
    elif filters['profit_filter'] == 'profit':
        order_rows = [row for row in order_rows if row['profit'] >= 0]
    order_rows = [row for row in order_rows if _matches_metric_filters(row)]

    order_row_map = {row['id']: row for row in order_rows}

    allowed_order_ids = [row['id'] for row in order_rows]
    if allowed_order_ids:
        order_items_qs = order_items_qs.filter(order_id__in=allowed_order_ids)
    else:
        order_items_qs = order_items_qs.none()

    total_orders = len(order_rows)
    total_goods_amount = sum(row['goods_amount'] for row in order_rows)
    total_revenue = sum(row['revenue'] for row in order_rows)
    total_debt = sum(row['debt'] for row in order_rows)
    total_cost = sum(row['cost'] for row in order_rows)
    total_profit = sum(row['profit'] for row in order_rows)
    loss_count = len([row for row in order_rows if row['is_loss']])

    allowed_order_id_set = set(allowed_order_ids)
    product_map = {}
    category_map = {}
    supplier_map = {}
    sku_details = []
    for item_row in adjusted_item_rows:
        if item_row['order_id'] not in allowed_order_id_set:
            continue

        product_key = (item_row['product_name'], item_row['category_name'], item_row['product_type_name'])
        if product_key not in product_map:
            product_map[product_key] = {
                'name': item_row['product_name'],
                'category': item_row['category_name'],
                'product_type': item_row['product_type_name'],
                'qty': 0,
                'amount': 0,
                'cost': 0,
                'returned_qty': 0,
                'returns_amount': 0,
                'return_cost': 0,
            }
        product_map[product_key]['qty'] += item_row['quantity']
        product_map[product_key]['amount'] += item_row['revenue']
        product_map[product_key]['cost'] += item_row['cost']

        category_key = item_row['category_name'] or 'Không DM'
        if category_key not in category_map:
            category_map[category_key] = {
                'name': category_key,
                'qty': 0,
                'revenue': 0,
                'cost': 0,
                'returns_amount': 0,
                'return_cost': 0,
            }
        category_map[category_key]['qty'] += item_row['quantity']
        category_map[category_key]['revenue'] += item_row['revenue']
        category_map[category_key]['cost'] += item_row['cost']

        # Dòng dịch vụ không có sản phẩm không thuộc hàng hóa của nhà cung cấp.
        # Sản phẩm chưa gắn NCC vẫn được gom riêng để người dùng nhận biết dữ liệu cần bổ sung.
        if item_row['product_id']:
            supplier_key = item_row['supplier_id'] or 0
            if supplier_key not in supplier_map:
                supplier_map[supplier_key] = {
                    'supplier_id': item_row['supplier_id'],
                    'supplier': item_row['supplier_name'] or 'Chưa gán nhà cung cấp',
                    'order_ids': set(),
                    'product_ids': set(),
                    'sold_quantity': 0,
                    'returned_quantity': 0,
                    'gross_revenue': 0,
                    'returns_amount': 0,
                    'gross_cost': 0,
                    'return_cost': 0,
                    'products': {},
                }
            supplier_row = supplier_map[supplier_key]
            supplier_row['order_ids'].add(item_row['order_id'])
            supplier_row['product_ids'].add(item_row['product_id'])
            supplier_row['sold_quantity'] += item_row['quantity']
            supplier_row['gross_revenue'] += item_row['revenue']
            supplier_row['gross_cost'] += item_row['cost']
            supplier_product = supplier_row['products'].setdefault(item_row['product_id'], {
                'name': item_row['product_name'],
                'sold_quantity': 0,
                'returned_quantity': 0,
                'gross_revenue': 0,
                'returns_amount': 0,
            })
            supplier_product['sold_quantity'] += item_row['quantity']
            supplier_product['gross_revenue'] += item_row['revenue']

        if _matches_metric_filters(item_row):
            sku_details.append({
                'id': item_row['id'],
                'date': item_row['date'],
                'date_raw': item_row['date_raw'],
                'customer': item_row['customer_name'],
                'product_name': item_row['product_name'],
                'sku': item_row['sku'],
                'order_id': item_row['order_id'],
                'order_code': item_row['order_code'],
                'salesperson': item_row['salesperson'],
                'quantity': item_row['quantity'],
                'revenue': item_row['revenue'],
                'cost': item_row['cost'],
                'profit': item_row['line_profit'],
                'line_profit': item_row['line_profit'],
            })

    sku_details = sorted(
        sku_details,
        key=lambda row: (row['date_raw'], row['order_id'], row['id']),
        reverse=True,
    )

    returns_qs = OrderReturn.objects.filter(
        return_date__gte=filters['from_date'],
        return_date__lte=filters['to_date'],
    ).exclude(status=3).select_related(
        'order', 'order__store', 'order__created_by', 'customer',
        'customer__group', 'customer__store', 'warehouse', 'warehouse__store'
    )
    returns_qs = _filter_sales_returns_by_scope(returns_qs, request)
    if filters['store_id']:
        returns_qs = returns_qs.filter(
            Q(order__store_id=filters['store_id']) |
            Q(order__isnull=True, warehouse__store_id=filters['store_id']) |
            Q(order__isnull=True, warehouse__isnull=True, customer__store_id=filters['store_id'])
        )
    if filters['customer_kind']:
        returns_qs = returns_qs.filter(_get_sales_report_return_customer_kind_q(filters['customer_kind']))
    if filters['customer_group_id']:
        returns_qs = returns_qs.filter(
            Q(order__customer__group_id=filters['customer_group_id']) |
            Q(order__isnull=True, customer__group_id=filters['customer_group_id'])
        )
    if filters['customer_id']:
        returns_qs = returns_qs.filter(
            Q(order__customer_id=filters['customer_id']) |
            Q(order__isnull=True, customer_id=filters['customer_id'])
        )
    if filters['search']:
        search = filters['search']
        returns_qs = returns_qs.filter(
            Q(code__icontains=search) |
            Q(order__code__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(reason__icontains=search)
        )
    # Phiếu trả có đơn gốc phải đi theo đúng phạm vi đơn đang được báo cáo.
    # Phiếu legacy thiếu đơn gốc vẫn được giữ lại nếu suy luận được đúng cửa hàng.
    if filters['salesperson']:
        returns_qs = returns_qs.filter(order_id__in=allowed_order_ids)
    else:
        returns_qs = returns_qs.filter(
            Q(order_id__in=allowed_order_ids) | Q(order__isnull=True)
        )

    returns_total = 0
    returns_count = 0
    returns_by_date = {}
    return_cost_total = 0
    return_cost_by_date = {}

    return_items_qs = OrderReturnItem.objects.filter(order_return__in=returns_qs)
    if filters['category_id']:
        return_items_qs = return_items_qs.filter(_get_product_category_scope_q(filters['category_id'], 'product__'))
    if filters['product_type_id']:
        return_items_qs = return_items_qs.filter(_get_product_category_direct_q(filters['product_type_id'], 'product__'))
    if filters['product_id']:
        return_items_qs = return_items_qs.filter(product_id=filters['product_id'])
    return_items_for_breakdown = list(return_items_qs.select_related(
        'product',
        'product__supplier',
        'product__category',
        'product__category__parent',
        'order_return',
        'order_return__order',
        'order_return__customer',
        'order_return__order__store',
    ))

    if item_scope:
        returns_total = sum(float(item.total_price or 0) for item in return_items_for_breakdown)
        returns_count = len({item.order_return_id for item in return_items_for_breakdown})
        for item in return_items_for_breakdown:
            if not item.order_return.return_date:
                continue
            date_key = item.order_return.return_date.strftime('%Y-%m-%d')
            returns_by_date[date_key] = returns_by_date.get(date_key, 0) + float(item.total_price or 0)
    else:
        returns_total = float(returns_qs.aggregate(s=Sum('total_refund'))['s'] or 0)
        returns_count = returns_qs.count()
        for row in returns_qs.values('return_date').annotate(total=Sum('total_refund')):
            if not row['return_date']:
                continue
            returns_by_date[row['return_date'].strftime('%Y-%m-%d')] = float(row['total'] or 0)

    # Giá vốn hàng trả lấy theo snapshot giá vốn trên đơn gốc. Nếu dữ liệu cũ
    # thiếu snapshot thì fallback về giá vốn/import hiện có của sản phẩm.
    return_order_ids = {
        item.order_return.order_id
        for item in return_items_for_breakdown
        if item.order_return.order_id
    }
    return_product_ids = {item.product_id for item in return_items_for_breakdown if item.product_id}
    original_cost_map = defaultdict(lambda: {'quantity': 0.0, 'cost': 0.0})
    if return_order_ids and return_product_ids:
        original_items = OrderItem.objects.filter(
            order_id__in=return_order_ids,
            product_id__in=return_product_ids,
        ).select_related('product', 'variant')
        for original_item in original_items:
            quantity = float(original_item.quantity or 0)
            if quantity <= 0:
                continue
            key = (original_item.order_id, original_item.product_id)
            original_cost_map[key]['quantity'] += quantity
            original_cost_map[key]['cost'] += _effective_item_unit_cost(original_item) * quantity

    return_cost_by_return_id = {}
    return_cost_by_item_id = {}
    for item in return_items_for_breakdown:
        quantity = float(item.quantity or 0)
        order_id = item.order_return.order_id
        original_cost = original_cost_map.get((order_id, item.product_id))
        if original_cost and original_cost['quantity'] > 0:
            unit_cost = original_cost['cost'] / original_cost['quantity']
        else:
            unit_cost = _effective_product_unit_cost(item.product)
        item_return_cost = unit_cost * quantity
        return_cost_by_item_id[item.id] = item_return_cost
        return_cost_total += item_return_cost
        return_cost_by_return_id[item.order_return_id] = (
            return_cost_by_return_id.get(item.order_return_id, 0) + item_return_cost
        )
        if item.order_return.return_date:
            date_key = item.order_return.return_date.strftime('%Y-%m-%d')
            return_cost_by_date[date_key] = return_cost_by_date.get(date_key, 0) + item_return_cost

        product = item.product
        category = product.category if product else None
        root_category = category.parent if category and category.parent_id else category
        product_type = category if category and category.parent_id else None
        product_key = (
            product.name if product else 'Sản phẩm',
            root_category.name if root_category else '',
            product_type.name if product_type else '',
        )
        if product_key not in product_map:
            product_map[product_key] = {
                'name': product_key[0],
                'category': product_key[1],
                'product_type': product_key[2],
                'qty': 0,
                'amount': 0,
                'cost': 0,
                'returned_qty': 0,
                'returns_amount': 0,
                'return_cost': 0,
            }
        product_map[product_key]['returned_qty'] += quantity
        product_map[product_key]['returns_amount'] += float(item.total_price or 0)
        product_map[product_key]['return_cost'] += item_return_cost

        category_key = root_category.name if root_category else 'Không DM'
        if category_key not in category_map:
            category_map[category_key] = {
                'name': category_key,
                'qty': 0,
                'revenue': 0,
                'cost': 0,
                'returns_amount': 0,
                'return_cost': 0,
            }
        category_map[category_key]['returns_amount'] += float(item.total_price or 0)
        category_map[category_key]['return_cost'] += item_return_cost

        supplier_key = item.product.supplier_id or 0
        if supplier_key not in supplier_map:
            supplier_map[supplier_key] = {
                'supplier_id': item.product.supplier_id,
                'supplier': item.product.supplier.name if item.product.supplier else 'Chưa gán nhà cung cấp',
                'order_ids': set(),
                'product_ids': set(),
                'sold_quantity': 0,
                'returned_quantity': 0,
                'gross_revenue': 0,
                'returns_amount': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'products': {},
            }
        supplier_row = supplier_map[supplier_key]
        supplier_row['product_ids'].add(item.product_id)
        supplier_row['returned_quantity'] += quantity
        supplier_row['returns_amount'] += float(item.total_price or 0)
        supplier_row['return_cost'] += item_return_cost
        supplier_product = supplier_row['products'].setdefault(item.product_id, {
            'name': item.product.name,
            'sold_quantity': 0,
            'returned_quantity': 0,
            'gross_revenue': 0,
            'returns_amount': 0,
        })
        supplier_product['returned_quantity'] += quantity
        supplier_product['returns_amount'] += float(item.total_price or 0)

    purchases = GoodsReceipt.objects.filter(
        receipt_date__gte=filters['from_date'],
        receipt_date__lte=filters['to_date'],
    ).exclude(status=2)
    purchases = filter_by_store(purchases, request, field_name='warehouse__store')
    if filters['store_id']:
        purchases = purchases.filter(warehouse__store_id=filters['store_id'])
    total_purchases = float(purchases.aggregate(s=Sum('total_amount'))['s'] or 0)

    time_group_meta = _get_sales_report_time_group_meta(filters['time_group'])
    daily_map = {}
    for row in sorted(order_rows, key=lambda item: item['date_raw'] or ''):
        if not row['date_raw']:
            continue
        date_obj = datetime.strptime(row['date_raw'], '%Y-%m-%d')
        key = date_obj.strftime(time_group_meta['key_format'])
        if key not in daily_map:
            daily_map[key] = {
                'date': date_obj.strftime(time_group_meta['display_format']),
                'count': 0,
                'goods_amount': 0,
                'revenue': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'cost': 0,
                'profit': 0,
                'returns': 0,
            }
        daily_map[key]['count'] += 1
        daily_map[key]['goods_amount'] += row['goods_amount']
        daily_map[key]['revenue'] += row['revenue']
        daily_map[key]['gross_cost'] += row['cost']
    for date_key, amount in returns_by_date.items():
        date_obj = datetime.strptime(date_key, '%Y-%m-%d')
        bucket_key = date_obj.strftime(time_group_meta['key_format'])
        bucket_label = date_obj.strftime(time_group_meta['display_format'])
        if bucket_key in daily_map:
            daily_map[bucket_key]['returns'] += amount
        else:
            daily_map[bucket_key] = {
                'date': bucket_label,
                'count': 0,
                'goods_amount': 0,
                'revenue': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'cost': 0,
                'profit': 0,
                'returns': amount,
            }
    for date_key, amount in return_cost_by_date.items():
        date_obj = datetime.strptime(date_key, '%Y-%m-%d')
        bucket_key = date_obj.strftime(time_group_meta['key_format'])
        bucket_label = date_obj.strftime(time_group_meta['display_format'])
        if bucket_key not in daily_map:
            daily_map[bucket_key] = {
                'date': bucket_label,
                'count': 0,
                'goods_amount': 0,
                'revenue': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'cost': 0,
                'profit': 0,
                'returns': 0,
            }
        daily_map[bucket_key]['return_cost'] += amount
    for row in daily_map.values():
        row['cost'] = row['gross_cost'] - row['return_cost']
        row['profit'] = row['revenue'] - row['returns'] - row['cost']
        net_revenue = row['revenue'] - row['returns']
        row['profit_margin'] = round(row['profit'] / net_revenue * 100, 1) if net_revenue > 0 else 0
    daily_data = []
    for key in sorted(daily_map.keys()):
        row = daily_map[key]
        row['period_key'] = key
        daily_data.append(row)

    daily_finance_map = {}
    for row in sorted(order_rows, key=lambda item: item['date_raw'] or ''):
        date_key = row['date_raw']
        if not date_key:
            continue
        date_obj = datetime.strptime(date_key, '%Y-%m-%d')
        if date_key not in daily_finance_map:
            daily_finance_map[date_key] = {
                'date': date_obj.strftime('%d/%m/%Y'),
                'date_raw': date_key,
                'goods_amount': 0,
                'revenue': 0,
                'returns': 0,
                'net_revenue': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'cost': 0,
                'gross_profit': 0,
                'gross_margin': 0,
                'net_profit': 0,
            }
        daily_finance_map[date_key]['goods_amount'] += row['goods_amount']
        daily_finance_map[date_key]['revenue'] += row['revenue']
        daily_finance_map[date_key]['gross_cost'] += row['cost']
    for date_key, amount in returns_by_date.items():
        date_obj = datetime.strptime(date_key, '%Y-%m-%d')
        if date_key not in daily_finance_map:
            daily_finance_map[date_key] = {
                'date': date_obj.strftime('%d/%m/%Y'),
                'date_raw': date_key,
                'goods_amount': 0,
                'revenue': 0,
                'returns': 0,
                'net_revenue': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'cost': 0,
                'gross_profit': 0,
                'gross_margin': 0,
                'net_profit': 0,
            }
        daily_finance_map[date_key]['returns'] += amount
    for date_key, amount in return_cost_by_date.items():
        date_obj = datetime.strptime(date_key, '%Y-%m-%d')
        if date_key not in daily_finance_map:
            daily_finance_map[date_key] = {
                'date': date_obj.strftime('%d/%m/%Y'),
                'date_raw': date_key,
                'goods_amount': 0,
                'revenue': 0,
                'returns': 0,
                'net_revenue': 0,
                'gross_cost': 0,
                'return_cost': 0,
                'cost': 0,
                'gross_profit': 0,
                'gross_margin': 0,
                'net_profit': 0,
            }
        daily_finance_map[date_key]['return_cost'] += amount

    daily_finance = []
    for row in [daily_finance_map[key] for key in sorted(daily_finance_map.keys())]:
        revenue = float(row.get('revenue') or 0)
        returns = float(row.get('returns') or 0)
        gross_cost = float(row.get('gross_cost') or 0)
        return_cost = float(row.get('return_cost') or 0)
        cost = gross_cost - return_cost
        net_revenue = revenue - returns
        gross_profit = net_revenue - cost
        row['net_revenue'] = net_revenue
        row['cost'] = cost
        row['gross_profit'] = gross_profit
        row['gross_margin'] = round(gross_profit / net_revenue * 100, 1) if net_revenue > 0 else 0
        row['net_profit'] = gross_profit
        daily_finance.append(row)

    product_breakdown = []
    for row in sorted(product_map.values(), key=lambda row: (-row['amount'], -row['qty'], row['name'])):
        line_profit = row['amount'] - row['cost']
        net_revenue = row['amount'] - row['returns_amount']
        net_cost = row['cost'] - row['return_cost']
        gross_profit = net_revenue - net_cost
        gross_margin = round(gross_profit / net_revenue * 100, 1) if net_revenue > 0 else 0
        product_breakdown.append({
            'name': row['name'],
            'category': row['category'] or '',
            'product_type': row.get('product_type') or '',
            'qty': row['qty'],
            'amount': row['amount'],
            'cost': row['cost'],
            'profit': line_profit,
            'line_profit': line_profit,
            'returned_qty': row['returned_qty'],
            'returns_amount': row['returns_amount'],
            'return_cost': row['return_cost'],
            'net_revenue': net_revenue,
            'net_cost': net_cost,
            'gross_profit': gross_profit,
            'gross_margin': gross_margin,
        })
    product_breakdown = [row for row in product_breakdown if _matches_metric_filters(row)]

    supplier_breakdown = []
    supplier_order_ids = set()
    supplier_product_ids = set()
    for supplier_row in supplier_map.values():
        net_revenue = supplier_row['gross_revenue'] - supplier_row['returns_amount']
        net_cost = supplier_row['gross_cost'] - supplier_row['return_cost']
        profit = net_revenue - net_cost
        top_products = []
        for product_row in supplier_row['products'].values():
            product_net_quantity = product_row['sold_quantity'] - product_row['returned_quantity']
            product_net_revenue = product_row['gross_revenue'] - product_row['returns_amount']
            top_products.append({
                'name': product_row['name'],
                'net_quantity': product_net_quantity,
                'net_revenue': product_net_revenue,
            })
        top_products.sort(key=lambda row: (-row['net_quantity'], -row['net_revenue'], row['name']))
        row = {
            'supplier_id': supplier_row['supplier_id'],
            'supplier': supplier_row['supplier'],
            'product_count': len(supplier_row['product_ids']),
            'order_count': len(supplier_row['order_ids']),
            'sold_quantity': supplier_row['sold_quantity'],
            'returned_quantity': supplier_row['returned_quantity'],
            'net_quantity': supplier_row['sold_quantity'] - supplier_row['returned_quantity'],
            'gross_revenue': supplier_row['gross_revenue'],
            'returns_amount': supplier_row['returns_amount'],
            'net_revenue': net_revenue,
            'gross_cost': supplier_row['gross_cost'],
            'return_cost': supplier_row['return_cost'],
            'cost': net_cost,
            'profit': profit,
            'top_products': top_products[:3],
        }
        if _matches_metric_filters({'revenue': net_revenue, 'cost': net_cost, 'profit': profit}):
            supplier_breakdown.append(row)
            supplier_order_ids.update(supplier_row['order_ids'])
            supplier_product_ids.update(supplier_row['product_ids'])

    supplier_breakdown.sort(
        key=lambda row: (-row['net_quantity'], -row['net_revenue'], row['supplier'])
    )
    supplier_net_revenue = sum(row['net_revenue'] for row in supplier_breakdown)
    for row in supplier_breakdown:
        row['contribution'] = round(
            row['net_revenue'] / supplier_net_revenue * 100, 1
        ) if supplier_net_revenue > 0 else 0
    supplier_summary = {
        'supplier_count': len(supplier_breakdown),
        'product_count': len(supplier_product_ids),
        'order_count': len(supplier_order_ids),
        'sold_quantity': sum(row['sold_quantity'] for row in supplier_breakdown),
        'returned_quantity': sum(row['returned_quantity'] for row in supplier_breakdown),
        'net_quantity': sum(row['net_quantity'] for row in supplier_breakdown),
        'net_revenue': supplier_net_revenue,
        'cost': sum(row['cost'] for row in supplier_breakdown),
        'profit': sum(row['profit'] for row in supplier_breakdown),
    }

    category_breakdown = []
    for row in sorted(category_map.values(), key=lambda row: (-row['revenue'], row['name'])):
        profit = row['revenue'] - row['cost']
        net_revenue = row['revenue'] - row['returns_amount']
        net_cost = row['cost'] - row['return_cost']
        gross_profit = net_revenue - net_cost
        category_breakdown.append({
            'name': row['name'],
            'qty': row['qty'],
            'revenue': row['revenue'],
            'cost': row['cost'],
            'profit': profit,
            'returns_amount': row['returns_amount'],
            'return_cost': row['return_cost'],
            'net_revenue': net_revenue,
            'net_cost': net_cost,
            'gross_profit': gross_profit,
            'gross_margin': (
                round(gross_profit / net_revenue * 100, 1)
                if net_revenue > 0 else 0
            ),
        })
    category_breakdown = [row for row in category_breakdown if _matches_metric_filters(row)]

    customer_map = {}
    customer_kind_map = {}
    group_map = {}
    staff_map = {}
    store_map = {}
    for row in order_rows:
        customer_key = row['customer_id'] or ('guest:' + (row['customer'] or 'Khách lẻ'))
        if customer_key not in customer_map:
            customer_map[customer_key] = {
                'name': row['customer'] or 'Khách lẻ',
                'customer_kind': row['customer_kind'],
                'customer_kind_label': row['customer_kind_label'],
                'group': row['customer_group'] or '',
                'orders': 0,
                'amount': 0,
                'cost': 0,
                'profit': 0,
                'paid': 0,
                'debt': 0,
            }
        customer_map[customer_key]['orders'] += 1
        customer_map[customer_key]['amount'] += row['revenue']
        customer_map[customer_key]['cost'] += row['cost']
        customer_map[customer_key]['profit'] += row['profit']
        customer_map[customer_key]['paid'] += row['paid']
        customer_map[customer_key]['debt'] += row['debt']

        customer_kind_key = row['customer_kind']
        if customer_kind_key not in customer_kind_map:
            customer_kind_map[customer_kind_key] = {
                'key': customer_kind_key,
                'name': row['customer_kind_label'],
                'orders': 0,
                'amount': 0,
                'cost': 0,
                'profit': 0,
                'paid': 0,
                'debt': 0,
            }
        customer_kind_map[customer_kind_key]['orders'] += 1
        customer_kind_map[customer_kind_key]['amount'] += row['revenue']
        customer_kind_map[customer_kind_key]['cost'] += row['cost']
        customer_kind_map[customer_kind_key]['profit'] += row['profit']
        customer_kind_map[customer_kind_key]['paid'] += row['paid']
        customer_kind_map[customer_kind_key]['debt'] += row['debt']

        group_name = row['customer_group'] or 'Không nhóm'
        if group_name not in group_map:
            group_map[group_name] = {
                'name': group_name,
                'orders': 0,
                'amount': 0,
                'cost': 0,
                'profit': 0,
                'paid': 0,
                'debt': 0,
            }
        group_map[group_name]['orders'] += 1
        group_map[group_name]['amount'] += row['revenue']
        group_map[group_name]['cost'] += row['cost']
        group_map[group_name]['profit'] += row['profit']
        group_map[group_name]['paid'] += row['paid']
        group_map[group_name]['debt'] += row['debt']

        staff_name = row['salesperson'] or '(Chưa gán NV)'
        if staff_name not in staff_map:
            staff_map[staff_name] = {
                'salesperson': staff_name,
                'order_count': 0,
                'revenue': 0,
                'cost': 0,
                'profit': 0,
                'returns_amount': 0,
            }
        staff_map[staff_name]['order_count'] += 1
        staff_map[staff_name]['revenue'] += row['revenue']
        staff_map[staff_name]['cost'] += row['cost']
        staff_map[staff_name]['profit'] += row['profit']

        store_key = row['store_id'] or 0
        if store_key not in store_map:
            store_map[store_key] = {
                'store_id': row['store_id'],
                'store_name': row['store_name'] or 'Chưa gán cửa hàng',
                'orders': 0,
                'revenue': 0,
                'cost': 0,
                'profit': 0,
                'debt': 0,
                'paid': 0,
            }
        store_map[store_key]['orders'] += 1
        store_map[store_key]['revenue'] += row['revenue']
        store_map[store_key]['cost'] += row['cost']
        store_map[store_key]['profit'] += row['profit']
        store_map[store_key]['debt'] += row['debt']
        store_map[store_key]['paid'] += row['paid']

    if returns_qs.exists():
        for ret in returns_qs.select_related('order', 'order__created_by'):
            staff_name = (
                _get_order_revenue_staff_name(ret.order)
                if ret.order else ''
            ) or '(Chưa gán NV)'
            if staff_name not in staff_map:
                staff_map[staff_name] = {
                    'salesperson': staff_name,
                    'order_count': 0,
                    'revenue': 0,
                    'cost': 0,
                    'profit': 0,
                    'returns_amount': 0,
                }
            if item_scope:
                scoped_return_items = OrderReturnItem.objects.filter(order_return=ret)
                if filters['category_id']:
                    scoped_return_items = scoped_return_items.filter(
                        _get_product_category_scope_q(filters['category_id'], 'product__')
                    )
                if filters['product_type_id']:
                    scoped_return_items = scoped_return_items.filter(
                        _get_product_category_direct_q(filters['product_type_id'], 'product__')
                    )
                if filters['product_id']:
                    scoped_return_items = scoped_return_items.filter(product_id=filters['product_id'])
                ret_amount = float(scoped_return_items.aggregate(s=Sum('total_price'))['s'] or 0)
            else:
                ret_amount = float(ret.total_refund or 0)
            staff_map[staff_name]['returns_amount'] += ret_amount

    customer_breakdown = sorted(customer_map.values(), key=lambda row: (-row['amount'], -row['orders'], row['name']))
    customer_breakdown = [row for row in customer_breakdown if _matches_metric_filters(row)][:200]
    top_customers = customer_breakdown[:10]
    group_breakdown = sorted(group_map.values(), key=lambda row: (-row['amount'], row['name']))
    group_breakdown = [row for row in group_breakdown if _matches_metric_filters(row)]
    customer_kind_breakdown = sorted(customer_kind_map.values(), key=lambda row: (-row['amount'], row['name']))
    customer_kind_breakdown = [row for row in customer_kind_breakdown if _matches_metric_filters(row)]
    for row in customer_breakdown:
        row['contribution'] = round(row['amount'] / total_revenue * 100, 1) if total_revenue > 0 else 0
    for row in group_breakdown:
        row['contribution'] = round(row['amount'] / total_revenue * 100, 1) if total_revenue > 0 else 0
    for row in customer_kind_breakdown:
        row['contribution'] = round(row['amount'] / total_revenue * 100, 1) if total_revenue > 0 else 0

    staff_breakdown = sorted(staff_map.values(), key=lambda row: (-row['revenue'], row['salesperson']))
    staff_breakdown = [row for row in staff_breakdown if _matches_metric_filters(row)]
    for row in staff_breakdown:
        row['contribution'] = round(row['revenue'] / total_revenue * 100, 1) if total_revenue > 0 else 0

    order_status_map = {}
    payment_status_map = {}
    for row in order_rows:
        status_key = row['status_display'] or 'Khác'
        if status_key not in order_status_map:
            order_status_map[status_key] = {'name': status_key, 'count': 0, 'revenue': 0}
        order_status_map[status_key]['count'] += 1
        order_status_map[status_key]['revenue'] += row['revenue']

        payment_key = row['payment_status_display'] or 'Khác'
        if payment_key not in payment_status_map:
            payment_status_map[payment_key] = {'name': payment_key, 'count': 0, 'revenue': 0, 'debt': 0}
        payment_status_map[payment_key]['count'] += 1
        payment_status_map[payment_key]['revenue'] += row['revenue']
        payment_status_map[payment_key]['debt'] += row['debt']

    order_status_breakdown = sorted(order_status_map.values(), key=lambda row: (-row['count'], row['name']))
    payment_status_breakdown = sorted(payment_status_map.values(), key=lambda row: (-row['count'], row['name']))

    return_amount_by_return_id = {}
    return_qty_by_return_id = {}
    return_product_map = {}
    for item in return_items_for_breakdown:
        refund = float(item.total_price or 0)
        qty = float(item.quantity or 0)
        return_amount_by_return_id[item.order_return_id] = return_amount_by_return_id.get(item.order_return_id, 0) + refund
        return_qty_by_return_id[item.order_return_id] = return_qty_by_return_id.get(item.order_return_id, 0) + qty

        product_key = item.product_id or f"product:{item.product.name if item.product else 'N/A'}"
        if product_key not in return_product_map:
            return_product_map[product_key] = {
                'product_id': item.product_id,
                'name': item.product.name if item.product else 'N/A',
                'qty': 0,
                'amount': 0,
                'cost': 0,
                'return_ids': set(),
            }
        return_product_map[product_key]['qty'] += qty
        return_product_map[product_key]['amount'] += refund
        return_product_map[product_key]['cost'] += return_cost_by_item_id.get(item.id, 0)
        return_product_map[product_key]['return_ids'].add(item.order_return_id)

    return_order_rows = []
    for ret in returns_qs.select_related(
        'order', 'customer', 'warehouse', 'order__store', 'warehouse__store', 'customer__store'
    ).order_by('-return_date', '-id'):
        refund = return_amount_by_return_id.get(ret.id, float(ret.total_refund or 0))
        qty = return_qty_by_return_id.get(ret.id, 0)
        if item_scope and refund <= 0 and qty <= 0:
            continue
        order_row = order_row_map.get(ret.order_id, {})
        if ret.order and ret.order.store:
            store_name = ret.order.store.name
        elif ret.warehouse and ret.warehouse.store:
            store_name = ret.warehouse.store.name
        elif ret.customer and ret.customer.store:
            store_name = ret.customer.store.name
        else:
            store_name = 'Chưa gán cửa hàng'
        return_order_rows.append({
            'id': ret.id,
            'code': ret.code,
            'date': ret.return_date.strftime('%d/%m/%Y') if ret.return_date else '',
            'order_code': ret.order.code if ret.order else '(Thiếu đơn gốc)',
            'customer': ret.customer.name if ret.customer else 'Khách chưa gán',
            'salesperson': (
                _get_order_revenue_staff_name(ret.order)
                if ret.order else ''
            ) or '(Chưa gán NV)',
            'store_name': store_name,
            'qty': qty,
            'amount': refund,
            'cost': return_cost_by_return_id.get(ret.id, 0),
            'order_revenue': float(order_row.get('revenue') or (ret.order.final_amount if ret.order else 0) or 0),
            'status': ret.status,
            'status_display': ret.get_status_display(),
            'reason': ret.reason or '',
        })

    return_product_breakdown = sorted(
        [{
            'product_id': row['product_id'],
            'name': row['name'],
            'qty': row['qty'],
            'amount': row['amount'],
            'cost': row['cost'],
            'return_count': len(row['return_ids']),
        } for row in return_product_map.values()],
        key=lambda row: (-row['amount'], -row['qty'], row['name'])
    )

    top_products = product_breakdown[:10]

    managed_ids = get_managed_store_ids(request.user)
    managed_stores = Store.objects.filter(id__in=managed_ids).select_related('brand')
    has_multiple = managed_stores.count() > 1
    stores_list = [{'id': store.id, 'name': store.name, 'brand': store.brand.name if store.brand else ''} for store in managed_stores]

    store_breakdown = []
    if has_multiple and not filters['store_id']:
        store_breakdown = sorted(store_map.values(), key=lambda row: (-row['revenue'], row['store_name']))
        store_breakdown = [row for row in store_breakdown if _matches_metric_filters(row)]

    total_net_revenue = total_revenue - returns_total
    total_net_cost = total_cost - return_cost_total
    total_gross_profit = total_net_revenue - total_net_cost
    gross_margin = round(total_gross_profit / total_net_revenue * 100, 1) if total_net_revenue > 0 else 0
    slow_moving_products, slow_moving_summary = _build_slow_moving_inventory_payload(
        request,
        filters,
    )

    payload = {
        'has_multiple_stores': has_multiple,
        'stores': stores_list,
        'summary': {
            'total_orders': total_orders,
            'total_goods_amount': total_goods_amount,
            'total_revenue': total_revenue,
            'total_net_revenue': total_net_revenue,
            'total_cost': total_net_cost,
            'total_sales_cost': total_cost,
            'total_return_cost': return_cost_total,
            'total_net_cost': total_net_cost,
            'total_order_profit': total_profit,
            'total_profit': total_gross_profit,
            'total_gross_profit': total_gross_profit,
            'total_net_profit': total_gross_profit,
            'profit_margin': gross_margin,
            'gross_margin': gross_margin,
            'total_returns': returns_total,
            'returns_count': returns_count,
            'total_debt': total_debt,
            'total_purchases': total_purchases,
            'loss_count': loss_count,
        },
        'timeline': daily_data,
        'daily': daily_data,
        'daily_finance': daily_finance,
        'time_group': filters['time_group'],
        'time_group_label': time_group_meta['label'],
        'order_details': order_rows,
        'store_breakdown': store_breakdown,
        'customer_kind_breakdown': customer_kind_breakdown,
        'group_breakdown': group_breakdown,
        'category_breakdown': category_breakdown,
        'order_status_breakdown': order_status_breakdown,
        'payment_status_breakdown': payment_status_breakdown,
        'top_products': top_products,
        'top_customers': top_customers,
        'product_breakdown': product_breakdown,
        'supplier_breakdown': supplier_breakdown,
        'supplier_summary': supplier_summary,
        'sku_details': sku_details,
        'customer_breakdown': customer_breakdown,
        'staff_breakdown': staff_breakdown,
        'return_orders': return_order_rows,
        'return_products': return_product_breakdown,
        'return_summary': {
            'total_returns': returns_total,
            'return_count': returns_count,
            'return_products': len(return_product_breakdown),
            'returned_qty': sum(row['qty'] for row in return_product_breakdown),
            'return_cost': return_cost_total,
            'return_rate': round(returns_total / total_revenue * 100, 1) if total_revenue > 0 else 0,
        },
        'slow_moving_products': slow_moving_products,
        'slow_moving_summary': slow_moving_summary,
        'filters_applied': filters,
    }

    if include_filter_options:
        groups = list(CustomerGroup.objects.filter(is_active=True).values('id', 'name').order_by('name'))
        categories = list(
            ProductCategory.objects.filter(is_active=True)
            .values('id', 'name', 'parent_id')
            .order_by('parent__name', 'name')
        )
        root_categories = [category for category in categories if not category['parent_id']]
        product_types = [category for category in categories if category['parent_id']]
        if filters['category_id']:
            selected_category_id = _parse_filter_int(filters['category_id'])
            if selected_category_id is not None:
                product_types = [category for category in product_types if category['parent_id'] == selected_category_id]

        category_name_map = {category['id']: category['name'] for category in categories}
        for category in product_types:
            category['parent_name'] = category_name_map.get(category['parent_id'], '')

        customers_qs = filter_by_store(Customer.objects.filter(is_active=True), request)
        products_qs = filter_by_store(Product.objects.filter(is_active=True), request)
        if filters['store_id']:
            customers_qs = customers_qs.filter(store_id=filters['store_id'])
            products_qs = products_qs.filter(store_id=filters['store_id'])
        if filters['customer_kind']:
            customers_qs = customers_qs.filter(_get_sales_report_customer_kind_q(filters['customer_kind'], ''))
        if filters['customer_group_id']:
            customers_qs = customers_qs.filter(group_id=filters['customer_group_id'])
        if filters['category_id']:
            products_qs = products_qs.filter(_get_product_category_scope_q(filters['category_id'], ''))
        if filters['product_type_id']:
            products_qs = products_qs.filter(_get_product_category_direct_q(filters['product_type_id'], ''))

        customers = list(customers_qs.values('id', 'code', 'name').order_by('name')[:300])
        products = list(products_qs.values('id', 'code', 'name').order_by('name')[:300])
        salespersons = _get_salesperson_filter_options(request, filters['store_id'])

        payload['filter_options'] = {
            'customer_kinds': CUSTOMER_KIND_OPTIONS,
            'customer_groups': groups,
            'categories': root_categories,
            'product_types': product_types,
            'customers': customers,
            'products': products,
            'salespersons': salespersons,
        }

    return payload


def _get_quotation_profit_filters(request):
    today = datetime.now().date()
    default_from = today.replace(day=1)
    from_date = parse_date(request.GET.get('from_date') or '') or default_from
    to_date = parse_date(request.GET.get('to_date') or '') or today
    if from_date > to_date:
        from_date, to_date = to_date, from_date

    status = (request.GET.get('status') or 'active').strip().lower()
    allowed_statuses = {'active', 'non_cancelled', 'all', '0', '1', '2', '3', '4'}
    if status not in allowed_statuses:
        status = 'active'

    validity = (request.GET.get('validity') or 'all').strip().lower()
    if validity not in {'all', 'current', 'expired', 'no_expiry'}:
        validity = 'all'

    profit_filter = (request.GET.get('profit_filter') or '').strip().lower()
    if profit_filter not in {'', 'profit', 'loss', 'missing_cost', 'estimated_cost'}:
        profit_filter = ''

    sort = (request.GET.get('sort') or 'date_desc').strip().lower()
    if sort not in {'date_desc', 'revenue_desc', 'profit_asc', 'profit_desc', 'margin_asc'}:
        sort = 'date_desc'

    source = (request.GET.get('source') or 'all').strip().lower()
    if source not in {'all', 'quotation', 'order'}:
        source = 'all'

    return {
        'from_date': from_date,
        'to_date': to_date,
        'status': status,
        'validity': validity,
        'profit_filter': profit_filter,
        'sort': sort,
        'source': source,
        'store_id': _parse_filter_int(request.GET.get('store_id')),
        'customer_id': _parse_filter_int(request.GET.get('customer_id')),
        'product_id': _parse_filter_int(request.GET.get('product_id')),
        'salesperson': (request.GET.get('salesperson') or '').strip(),
        'search': (request.GET.get('search') or '').strip(),
    }


def _quotation_product_unit_cost(product, visited_product_ids=None):
    if not product:
        return Decimal('0')
    for candidate in (product.cost_price, product.import_price):
        value = Decimal(str(candidate or 0))
        if value > 0:
            return value
    if not product.is_combo:
        return Decimal('0')

    visited_product_ids = set(visited_product_ids or ())
    if product.id in visited_product_ids:
        return Decimal('0')
    visited_product_ids.add(product.id)

    combo_cost = Decimal('0')
    for combo_item in product.combo_items.all():
        combo_cost += (
            _quotation_product_unit_cost(combo_item.product, visited_product_ids)
            * Decimal(str(combo_item.quantity or 0))
        )
    return combo_cost


def _quotation_item_current_unit_cost(item):
    if item.variant_id:
        for candidate in (item.variant.cost_price, item.variant.import_price):
            value = Decimal(str(candidate or 0))
            if value > 0:
                return value
    return _quotation_product_unit_cost(item.product if item.product_id else None)


def _build_expected_profit_row(document, source_type, today):
    """Chuẩn hóa báo giá và đơn chưa xuất kho về cùng một dòng báo cáo."""
    items = list(document.items.all())
    expected_cost = Decimal('0')
    missing_cost_lines = 0
    estimated_cost_lines = 0
    product_names = []

    for item in items:
        if item.product_id:
            product_label = f'{item.product.code} - {item.product.name}'
        else:
            product_label = item.item_name or 'Dịch vụ'
        if product_label not in product_names:
            product_names.append(product_label)

        if item.cost_price is not None and Decimal(str(item.cost_price or 0)) > 0:
            unit_cost = Decimal(str(item.cost_price))
        else:
            unit_cost = _quotation_item_current_unit_cost(item)
            if unit_cost > 0:
                estimated_cost_lines += 1
            else:
                missing_cost_lines += 1
        expected_cost += unit_cost * Decimal(str(item.quantity or 0))

    expected_revenue = Decimal(str(document.final_amount or 0))
    expected_profit = expected_revenue - expected_cost
    expected_margin = (
        expected_profit / expected_revenue * Decimal('100')
        if expected_revenue > 0
        else Decimal('0')
    )

    if source_type == 'quotation':
        document_date = document.quotation_date
        valid_until = document.valid_until
        source_display = 'Báo giá'
        document_url = f'/order-tbl/?edit_quotation={document.id}'
        validity_status = 'no_expiry'
        validity_display = 'Không đặt hạn'
        if valid_until:
            if valid_until < today:
                validity_status = 'expired'
                validity_display = 'Hết hạn'
            else:
                validity_status = 'current'
                validity_display = 'Còn hạn'
    else:
        document_date = document.order_date
        valid_until = None
        source_display = 'Đơn báo giá' if document.status == 0 else 'Đơn chưa xuất kho'
        document_url = f'/order-tbl/?open_order={document.id}'
        validity_status = 'pending_order'
        validity_display = 'Chưa xuất kho'

    if missing_cost_lines:
        cost_status = 'missing'
        cost_status_display = 'Thiếu giá vốn'
    elif estimated_cost_lines:
        cost_status = 'estimated'
        cost_status_display = 'Giá vốn hiện tại'
    else:
        cost_status = 'snapshot'
        cost_status_display = 'Giá vốn đã chụp'

    creator_display = ''
    if document.created_by:
        creator_display = (
            document.created_by.get_full_name()
            or document.created_by.username
        )

    return {
        'id': document.id,
        'source_type': source_type,
        'source_display': source_display,
        'code': document.code,
        'document_url': document_url,
        # Giữ khóa cũ để các tích hợp đang dùng API không bị gãy.
        'quotation_url': document_url,
        'date': document_date.strftime('%d/%m/%Y') if document_date else '',
        'date_raw': document_date.isoformat() if document_date else '',
        'valid_until': valid_until.strftime('%d/%m/%Y') if valid_until else '',
        'valid_until_raw': valid_until.isoformat() if valid_until else '',
        'validity_status': validity_status,
        'validity_display': validity_display,
        'status': document.status,
        'status_display': document.get_status_display(),
        'store_id': document.store_id,
        'store_name': document.store.name if document.store else '',
        'customer_id': document.customer_id,
        'customer_name': document.customer.name if document.customer else '',
        'salesperson': document.salesperson or creator_display,
        'item_count': len(items),
        'product_names': product_names,
        'product_summary': ', '.join(product_names),
        'goods_amount': float(document.total_amount or 0),
        'discount_amount': float(document.discount_amount or 0),
        'shipping_fee': float(document.shipping_fee or 0),
        'other_fee': float(document.other_fee or 0),
        'expected_revenue': float(expected_revenue),
        'expected_cost': float(expected_cost),
        'expected_profit': float(expected_profit),
        'expected_margin': round(float(expected_margin), 2),
        'ctv_discount_capacity': (
            None
            if missing_cost_lines
            else float(max(expected_profit, Decimal('0')))
        ),
        'is_loss': expected_profit < 0,
        'missing_cost_lines': missing_cost_lines,
        'estimated_cost_lines': estimated_cost_lines,
        'cost_status': cost_status,
        'cost_status_display': cost_status_display,
    }


def _build_quotation_profit_payload(request, include_filter_options=True):
    from customers.models import Customer
    from products.models import Product
    from system_management.models import Store

    filters = _get_quotation_profit_filters(request)
    today = datetime.now().date()

    scoped_quotations = filter_by_store(
        Quotation.objects.filter(is_deleted=False),
        request,
    )
    scoped_pending_orders = filter_by_store(
        Order.objects.filter(is_deleted=False, status__in=(0, 1, 2, 3)),
        request,
    )
    quotations = scoped_quotations.filter(
        quotation_date__gte=filters['from_date'],
        quotation_date__lte=filters['to_date'],
    )
    pending_orders = scoped_pending_orders.filter(
        order_date__gte=filters['from_date'],
        order_date__lte=filters['to_date'],
    )

    if filters['status'] == 'active':
        quotations = quotations.filter(status__in=(0, 1, 2))
    elif filters['status'] == 'non_cancelled':
        quotations = quotations.exclude(status=4)
    elif filters['status'] != 'all':
        quotations = quotations.filter(status=int(filters['status']))
        # Trạng thái số trên bộ lọc cũ là trạng thái riêng của báo giá.
        pending_orders = pending_orders.none()

    if filters['source'] == 'quotation':
        pending_orders = pending_orders.none()
    elif filters['source'] == 'order':
        quotations = quotations.none()

    if filters['validity'] == 'current':
        quotations = quotations.filter(Q(valid_until__isnull=True) | Q(valid_until__gte=today))
    elif filters['validity'] == 'expired':
        quotations = quotations.filter(valid_until__lt=today)
    elif filters['validity'] == 'no_expiry':
        quotations = quotations.filter(valid_until__isnull=True)

    if filters['store_id']:
        quotations = quotations.filter(store_id=filters['store_id'])
        pending_orders = pending_orders.filter(store_id=filters['store_id'])
    if filters['customer_id']:
        quotations = quotations.filter(customer_id=filters['customer_id'])
        pending_orders = pending_orders.filter(customer_id=filters['customer_id'])
    if filters['product_id']:
        quotations = quotations.filter(items__product_id=filters['product_id'])
        pending_orders = pending_orders.filter(items__product_id=filters['product_id'])
    if filters['salesperson']:
        quotations = quotations.filter(salesperson=filters['salesperson'])
        pending_orders = pending_orders.filter(salesperson=filters['salesperson'])
    if filters['search']:
        search = filters['search']
        quotations = quotations.filter(
            Q(code__icontains=search)
            | Q(customer__name__icontains=search)
            | Q(customer__phone__icontains=search)
            | Q(salesperson__icontains=search)
            | Q(tags__icontains=search)
            | Q(note__icontains=search)
            | Q(items__product__code__icontains=search)
            | Q(items__product__name__icontains=search)
            | Q(items__item_name__icontains=search)
        )
        pending_orders = pending_orders.filter(
            Q(code__icontains=search)
            | Q(customer__name__icontains=search)
            | Q(customer__phone__icontains=search)
            | Q(salesperson__icontains=search)
            | Q(creator_name__icontains=search)
            | Q(tags__icontains=search)
            | Q(note__icontains=search)
            | Q(items__product__code__icontains=search)
            | Q(items__product__name__icontains=search)
            | Q(items__item_name__icontains=search)
        )

    # Khi báo giá đã tạo ra một đơn chưa xuất kho đang nằm trong cùng kết quả,
    # chỉ giữ dòng đơn hàng để doanh thu/lợi nhuận không bị cộng hai lần.
    linked_pending_quotation_ids = pending_orders.exclude(
        quotation_id__isnull=True,
    ).values_list('quotation_id', flat=True)
    quotations = quotations.exclude(id__in=linked_pending_quotation_ids)

    quotations = (
        quotations.select_related('customer', 'store', 'created_by')
        .prefetch_related(
            'items__variant',
            'items__product',
            'items__product__combo_items__product',
        )
        .distinct()
    )
    pending_orders = (
        pending_orders.select_related('customer', 'store', 'created_by')
        .prefetch_related(
            'items__variant',
            'items__product',
            'items__product__combo_items__product',
        )
        .distinct()
    )

    rows = [
        *[
            _build_expected_profit_row(quotation, 'quotation', today)
            for quotation in quotations
        ],
        *[
            _build_expected_profit_row(order, 'order', today)
            for order in pending_orders
        ],
    ]

    if filters['profit_filter'] == 'profit':
        rows = [row for row in rows if row['expected_profit'] >= 0 and not row['missing_cost_lines']]
    elif filters['profit_filter'] == 'loss':
        rows = [row for row in rows if row['expected_profit'] < 0]
    elif filters['profit_filter'] == 'missing_cost':
        rows = [row for row in rows if row['missing_cost_lines']]
    elif filters['profit_filter'] == 'estimated_cost':
        rows = [row for row in rows if row['estimated_cost_lines']]

    sorters = {
        'date_desc': lambda row: (row['date_raw'], row['id']),
        'revenue_desc': lambda row: (row['expected_revenue'], row['id']),
        'profit_asc': lambda row: (row['expected_profit'], row['id']),
        'profit_desc': lambda row: (row['expected_profit'], row['id']),
        'margin_asc': lambda row: (row['expected_margin'], row['id']),
    }
    reverse = filters['sort'] in {'date_desc', 'revenue_desc', 'profit_desc'}
    rows.sort(key=sorters[filters['sort']], reverse=reverse)

    total_revenue = sum(row['expected_revenue'] for row in rows)
    total_cost = sum(row['expected_cost'] for row in rows)
    total_profit = total_revenue - total_cost
    summary = {
        'record_count': len(rows),
        'quotation_count': sum(1 for row in rows if row['source_type'] == 'quotation'),
        'pending_order_count': sum(1 for row in rows if row['source_type'] == 'order'),
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'profit_margin': round(total_profit / total_revenue * 100, 2) if total_revenue > 0 else 0,
        'loss_count': sum(1 for row in rows if row['is_loss']),
        'missing_cost_count': sum(1 for row in rows if row['missing_cost_lines']),
        'estimated_cost_count': sum(1 for row in rows if row['estimated_cost_lines']),
        'warning_count': sum(
            1 for row in rows if row['is_loss'] or row['missing_cost_lines']
        ),
        'refreshed_at': datetime.now().strftime('%d/%m/%Y %H:%M'),
    }

    payload = {
        'filters': {
            **filters,
            'from_date': filters['from_date'].isoformat(),
            'to_date': filters['to_date'].isoformat(),
        },
        'summary': summary,
        'data': rows,
    }

    if include_filter_options:
        managed_store_ids = get_managed_store_ids(request.user)
        stores = list(
            Store.objects.filter(id__in=managed_store_ids)
            .order_by('name')
            .values('id', 'name')
        )
        customers = list(
            Customer.objects.filter(
                Q(quotations__in=scoped_quotations)
                | Q(orders__in=scoped_pending_orders),
            )
            .distinct()
            .order_by('name')
            .values('id', 'name')
        )
        products = list(
            Product.objects.filter(
                Q(quotation_items__quotation__in=scoped_quotations)
                | Q(order_items__order__in=scoped_pending_orders),
            )
            .distinct()
            .order_by('name')
            .values('id', 'code', 'name')
        )
        quotation_salespersons = set(
            scoped_quotations.exclude(salesperson__isnull=True)
            .exclude(salesperson='')
            .values_list('salesperson', flat=True)
            .distinct()
        )
        order_salespersons = set(
            scoped_pending_orders.exclude(salesperson__isnull=True)
            .exclude(salesperson='')
            .values_list('salesperson', flat=True)
            .distinct()
        )
        salespersons = sorted(quotation_salespersons | order_salespersons)
        payload['options'] = {
            'stores': stores,
            'has_multiple_stores': len(stores) > 1,
            'customers': customers,
            'products': products,
            'salespersons': salespersons,
            'statuses': [
                {'value': 'active', 'label': 'BG đang chào + đơn chưa xuất kho'},
                {'value': 'non_cancelled', 'label': 'Tất cả BG trừ hủy + đơn chưa xuất kho'},
                {'value': 'all', 'label': 'Tất cả BG + đơn chưa xuất kho'},
                *[
                    {'value': str(value), 'label': f'Chỉ BG: {label}'}
                    for value, label in Quotation.STATUS_CHOICES
                ],
            ],
        }
    return payload


@login_required(login_url="/login/")
@report_permission_required
@quotation_profit_report_privileged_required
def report_quotation_profit(request):
    return render(
        request,
        'reports/report_quotation_profit.html',
        {'active_tab': 'report_quotation_profit'},
    )


@login_required(login_url="/login/")
@report_permission_required
@quotation_profit_report_privileged_required
def api_report_quotation_profit(request):
    return JsonResponse({
        'status': 'ok',
        **_build_quotation_profit_payload(request, include_filter_options=True),
    })


@login_required(login_url="/login/")
@report_permission_required
@quotation_profit_report_privileged_required
def export_quotation_profit_excel(request):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    payload = _build_quotation_profit_payload(request, include_filter_options=False)
    filters = payload['filters']
    summary = payload['summary']
    rows = payload['data']

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'LN dự kiến báo giá'

    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    warning_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    money_format = '#,##0'

    sheet.merge_cells('A1:T1')
    sheet['A1'] = 'BÁO CÁO LỢI NHUẬN DỰ KIẾN: BÁO GIÁ VÀ ĐƠN CHƯA XUẤT KHO'
    sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    sheet['A1'].fill = title_fill
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A2:T2')
    sheet['A2'] = f"Từ {filters['from_date']} đến {filters['to_date']}"
    sheet['A2'].font = Font(italic=True)
    sheet['A2'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A3:T3')
    sheet['A3'] = (
        f"Số chứng từ: {summary['record_count']} "
        f"(Báo giá: {summary['quotation_count']}, đơn chưa xuất kho: {summary['pending_order_count']}) | "
        f"Doanh thu dự kiến: {summary['total_revenue']:,.0f}đ | "
        f"Giá vốn dự kiến: {summary['total_cost']:,.0f}đ | "
        f"LN dự kiến: {summary['total_profit']:,.0f}đ | "
        f"Biên LN: {summary['profit_margin']:.2f}%"
    )
    sheet['A3'].font = Font(bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center')

    headers = [
        'STT', 'Mã chứng từ', 'Ngày chứng từ', 'Loại / Hiệu lực', 'Trạng thái', 'Cửa hàng',
        'Khách hàng', 'Nhân viên', 'Sản phẩm', 'Số dòng', 'Tiền hàng',
        'Chiết khấu', 'Phí vận chuyển', 'Chi phí khác', 'Doanh thu dự kiến',
        'Giá vốn dự kiến', 'LN dự kiến', 'Biên LN (%)', 'Nguồn giá vốn', 'Cảnh báo',
    ]
    for column, heading in enumerate(headers, 1):
        cell = sheet.cell(row=5, column=column, value=heading)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for index, item in enumerate(rows, 1):
        warnings = []
        if item['is_loss']:
            warnings.append('Báo lỗ')
        if item['missing_cost_lines']:
            warnings.append(f"Thiếu GV {item['missing_cost_lines']} dòng")
        if item['estimated_cost_lines']:
            warnings.append(f"GV tạm tính {item['estimated_cost_lines']} dòng")
        values = [
            index,
            item['code'],
            item['date'],
            (
                item['source_display']
                if item['source_type'] == 'order'
                else f"Báo giá · {item['valid_until'] or 'Không đặt hạn'}"
            ),
            item['status_display'],
            item['store_name'],
            item['customer_name'],
            item['salesperson'],
            item['product_summary'],
            item['item_count'],
            item['goods_amount'],
            item['discount_amount'],
            item['shipping_fee'],
            item['other_fee'],
            item['expected_revenue'],
            item['expected_cost'],
            item['expected_profit'],
            item['expected_margin'],
            item['cost_status_display'],
            '; '.join(warnings),
        ]
        row_number = index + 5
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical='top', wrap_text=column in (7, 8, 9, 19, 20))
            if 11 <= column <= 17:
                cell.number_format = money_format
            if column == 18:
                cell.number_format = '0.00'
            if item['is_loss'] or item['missing_cost_lines']:
                cell.fill = warning_fill

    total_row = len(rows) + 6
    sheet.cell(row=total_row, column=1, value='TỔNG')
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
    totals = [
        summary['total_revenue'],
        summary['total_cost'],
        summary['total_profit'],
        summary['profit_margin'],
    ]
    for column in range(1, 21):
        cell = sheet.cell(row=total_row, column=column)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = thin
    for column, value in zip((15, 16, 17, 18), totals):
        sheet.cell(row=total_row, column=column, value=value)
        sheet.cell(row=total_row, column=column).number_format = (
            '0.00' if column == 18 else money_format
        )

    sheet.freeze_panes = 'A6'
    widths = [6, 17, 13, 15, 18, 22, 28, 24, 45, 10, 16, 16, 16, 16, 20, 18, 18, 13, 20, 24]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f"attachment; filename=\"BC_LN_du_kien_BG_DH_chua_XK_{filters['from_date']}_{filters['to_date']}.xlsx\""
    )
    workbook.save(response)
    return response


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
@sales_report_privileged_required
def report_sales(request):
    """Báo cáo bán hàng"""
    context = {'active_tab': 'report_sales'}
    return render(request, "reports/report_sales.html", context)


@login_required(login_url="/login/")
@report_permission_required
@sales_report_privileged_required
def api_report_sales(request):
    """API báo cáo bán hàng — chuẩn hóa theo bộ lọc chung cho mọi tab."""
    payload = _build_sales_report_payload(request, include_filter_options=True)
    return JsonResponse({'status': 'ok', **payload})


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
def report_purchases(request):
    from products.models import Supplier

    store_ids = get_managed_store_ids(request.user)
    suppliers = Supplier.objects.filter(
        goods_receipts__warehouse__store_id__in=store_ids,
        goods_receipts__is_deleted=False,
    ).distinct().order_by('name').values('id', 'name')
    context = {
        'active_tab': 'report_purchases',
        'suppliers': list(suppliers),
    }
    return render(request, "reports/report_purchases.html", context)


def _purchase_report_receipts(request, from_date, to_date, supplier_id=None):
    from products.models import GoodsReceipt

    receipts = GoodsReceipt.objects.filter(
        receipt_date__gte=from_date,
        receipt_date__lte=to_date,
    ).select_related('supplier', 'warehouse')
    receipts = filter_by_store(receipts, request, field_name='warehouse__store')
    if supplier_id:
        receipts = receipts.filter(supplier_id=supplier_id)
    return receipts


def _purchase_supplier_summary(receipts):
    rows = (
        receipts.filter(status=1)
        .values('supplier_id', 'supplier__name')
        .annotate(receipt_count=Count('id'), total_amount=Sum('total_amount'))
        .order_by('-total_amount', 'supplier__name')
    )
    return [{
        'supplier_id': row['supplier_id'],
        'supplier': row['supplier__name'] or 'Chưa chọn NCC',
        'receipt_count': row['receipt_count'],
        'total_amount': float(row['total_amount'] or 0),
    } for row in rows]


@login_required(login_url="/login/")
@report_permission_required
def api_report_purchases(request):
    """API báo cáo nhập hàng"""
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    today = datetime.now().date()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    supplier_id = _parse_filter_int(request.GET.get('supplier_id'))
    receipts = _purchase_report_receipts(
        request,
        from_date,
        to_date,
        supplier_id=supplier_id,
    ).order_by('-receipt_date', '-id')
    supplier_summary = _purchase_supplier_summary(receipts)

    data = [{
        'id': r.id, 'code': r.code,
        'date': r.receipt_date.strftime('%d/%m/%Y') if r.receipt_date else '',
        'supplier': r.supplier.name if r.supplier else '',
        'warehouse': r.warehouse.name if r.warehouse else '',
        'total_amount': float(r.total_amount),
        'status': r.status, 'status_display': r.get_status_display(),
    } for r in receipts]

    total = sum(d['total_amount'] for d in data if d['status'] == 1)
    count = len([d for d in data if d['status'] == 1])

    return JsonResponse({
        'status': 'ok', 'data': data,
        'summary': {
            'total_amount': total,
            'total_count': count,
            'total_suppliers': sum(1 for row in supplier_summary if row['supplier_id'] is not None),
            'refreshed_at': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        },
        'supplier_summary': supplier_summary,
    })


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
def report_inventory(request):
    today = datetime.now().date()
    context = {
        'active_tab': 'report_inventory',
        'inventory_movement_from_date': today.replace(day=1).isoformat(),
        'inventory_movement_to_date': today.isoformat(),
    }
    return render(request, "reports/report_inventory.html", context)


@login_required(login_url="/login/")
@report_permission_required
def api_report_inventory(request):
    """API báo cáo tồn kho"""
    from products.models import ProductStock, Warehouse, ProductCategory
    from core.store_utils import filter_by_store
    warehouse_id = request.GET.get('warehouse_id')
    search = request.GET.get('search', '').strip()
    category_id = request.GET.get('category_id')
    product_type_id = request.GET.get('product_type_id')
    stock_status = request.GET.get('stock_status', '')  # positive, zero, negative

    stocks = ProductStock.objects.select_related(
        'product', 'product__category', 'product__category__parent',
        'product__supplier', 'warehouse',
    ).filter(
        product__is_deleted=False,
    )
    stocks = filter_by_store(stocks, request, field_name='warehouse__store')
    if warehouse_id:
        stocks = stocks.filter(warehouse_id=warehouse_id)
    if search:
        stocks = stocks.filter(
            Q(product__name__icontains=search) |
            Q(product__code__icontains=search) |
            Q(product__barcode__icontains=search)
        )
    if category_id:
        stocks = stocks.filter(
            Q(product__category_id=category_id) |
            Q(product__category__parent_id=category_id)
        )
    if product_type_id:
        stocks = stocks.filter(product__category_id=product_type_id)

    data = []
    for s in stocks:
        quantity = Decimal(str(s.quantity or 0))
        cost_price = Decimal(str(s.product.cost_price or 0))
        import_price = Decimal(str(s.product.import_price or 0))
        valuation_price, valuation_source = _inventory_valuation_unit_cost(s.product)
        # Tồn âm là chênh lệch cần xử lý, không phải tài sản âm để khấu trừ
        # khỏi giá trị của các hàng hóa thực tế đang còn trong kho.
        stock_value = max(quantity, Decimal('0')) * valuation_price
        qty = float(quantity)
        min_stock = float(s.product.min_stock or 0)
        max_stock = float(s.product.max_stock or 0)
        alert = ''
        alert_type = ''
        if qty < min_stock:
            alert = 'Dưới tối thiểu'
            alert_type = 'danger'
        elif max_stock > 0 and qty > max_stock:
            alert = 'Trên tối đa'
            alert_type = 'warning'

        product_category = s.product.category
        root_category = (
            product_category.parent
            if product_category and product_category.parent_id
            else product_category
        )
        data.append({
            'product_id': s.product_id,
            'product_code': s.product.code,
            'product_name': s.product.name,
            'supplier': s.product.supplier.name if s.product.supplier else '',
            'category': s.product.category.name if s.product.category else '',
            'category_id': root_category.id if root_category else None,
            'category_name': root_category.name if root_category else '',
            'warehouse': s.warehouse.name,
            'warehouse_id': s.warehouse_id,
            'quantity': qty,
            'min_stock': min_stock,
            'max_stock': max_stock,
            'restock_needed': max(min_stock - qty, 0) if alert_type == 'danger' else 0,
            'unit': s.product.unit or '',
            'cost_price': float(cost_price),
            'import_price': float(import_price),
            'valuation_price': float(valuation_price),
            'valuation_source': valuation_source,
            'stock_value': float(stock_value),
            'alert': alert,
            'alert_type': alert_type,
        })

    # Client-side stock status filter
    if stock_status == 'positive':
        data = [d for d in data if d['quantity'] > 0]
    elif stock_status == 'zero':
        data = [d for d in data if d['quantity'] == 0]
    elif stock_status == 'negative':
        data = [d for d in data if d['quantity'] < 0]

    warehouses_qs = Warehouse.objects.filter(is_active=True)
    warehouses_qs = filter_by_store(warehouses_qs, request)
    warehouses = [{'id': w.id, 'name': w.name} for w in warehouses_qs]

    categories_qs = ProductCategory.objects.filter(is_active=True, parent__isnull=True).order_by('name')
    categories = [{'id': c.id, 'name': c.name} for c in categories_qs]
    product_types_qs = ProductCategory.objects.filter(
        is_active=True,
        parent__isnull=False,
        parent__is_deleted=False,
    ).select_related('parent').order_by('parent__name', 'name')
    product_types = [{
        'id': item.id,
        'name': item.name,
        'parent_id': item.parent_id,
        'parent_name': item.parent.name,
    } for item in product_types_qs]

    total_value = sum(d['stock_value'] for d in data)
    total_items = sum(d['quantity'] for d in data)
    alert_count = len([d for d in data if d['alert']])
    low_stock_count = len([d for d in data if d['alert_type'] == 'danger'])
    high_stock_count = len([d for d in data if d['alert_type'] == 'warning'])

    return JsonResponse({
        'status': 'ok', 'data': data, 'warehouses': warehouses,
        'categories': categories, 'product_types': product_types,
        'summary': {
            'total_value': total_value,
            'total_items': total_items,
            'alert_count': alert_count,
            'low_stock_count': low_stock_count,
            'high_stock_count': high_stock_count,
        }
    })


def _inventory_movement_date_range(request):
    """Chuẩn hóa kỳ báo cáo nhập xuất tồn, mặc định từ đầu tháng đến hôm nay."""
    today = datetime.now().date()
    raw_from_date = (request.GET.get('from_date') or '').strip()
    raw_to_date = (request.GET.get('to_date') or '').strip()
    from_date = parse_date(raw_from_date) if raw_from_date else today.replace(day=1)
    to_date = parse_date(raw_to_date) if raw_to_date else today
    if raw_from_date and not from_date:
        raise ValueError('Từ ngày không hợp lệ.')
    if raw_to_date and not to_date:
        raise ValueError('Đến ngày không hợp lệ.')
    if from_date > to_date:
        raise ValueError('Từ ngày phải nhỏ hơn hoặc bằng đến ngày.')
    return from_date, to_date


def _inventory_movement_filtered_stocks(request):
    """Các dòng sản phẩm/kho thuộc phạm vi và bộ lọc chung của báo cáo kho."""
    from products.models import ProductStock

    stocks = ProductStock.objects.select_related(
        'product', 'product__category', 'product__category__parent', 'warehouse',
    ).filter(
        product__is_deleted=False,
        warehouse__is_deleted=False,
    )
    stocks = filter_by_store(stocks, request, field_name='warehouse__store')

    warehouse_id = request.GET.get('warehouse_id')
    search = (request.GET.get('search') or '').strip()
    category_id = request.GET.get('category_id')
    product_type_id = request.GET.get('product_type_id')
    if warehouse_id:
        stocks = stocks.filter(warehouse_id=warehouse_id)
    if search:
        stocks = stocks.filter(
            Q(product__name__icontains=search)
            | Q(product__code__icontains=search)
            | Q(product__barcode__icontains=search)
        )
    if category_id:
        stocks = stocks.filter(
            Q(product__category_id=category_id)
            | Q(product__category__parent_id=category_id)
        )
    if product_type_id:
        stocks = stocks.filter(product__category_id=product_type_id)
    return stocks.order_by('product__name', 'product_id', 'warehouse__name', 'warehouse_id')


def _build_inventory_movement_payload(request):
    """Tổng hợp nhập, xuất và tồn theo sản phẩm/kho từ các chứng từ đã áp dụng."""
    from orders.models import OrderItem, OrderReturnItem
    from products.models import (
        ComboItem,
        GoodsReceiptItem,
        PurchaseReturnItem,
        StockCheckItem,
        StockTransferItem,
    )

    from_date, to_date = _inventory_movement_date_range(request)
    stocks = list(_inventory_movement_filtered_stocks(request))
    rows = {}
    product_costs = {}
    for stock in stocks:
        product = stock.product
        category = product.category
        root_category = category.parent if category and category.parent_id else category
        valuation_price, valuation_source = _inventory_valuation_unit_cost(product)
        product_costs[product.id] = valuation_price
        current_quantity = Decimal(str(stock.quantity or 0))
        rows[(stock.product_id, stock.warehouse_id)] = {
            'product_id': stock.product_id,
            'product_code': product.code,
            'product_name': product.name,
            'category': category.name if category else '',
            'category_id': root_category.id if root_category else None,
            'category_name': root_category.name if root_category else '',
            'unit': product.unit or '',
            'warehouse_id': stock.warehouse_id,
            'warehouse': stock.warehouse.name,
            'valuation_price': valuation_price,
            'valuation_source': valuation_source,
            'opening_quantity': current_quantity,
            'import_quantity': Decimal('0'),
            'import_value': Decimal('0'),
            'export_quantity': Decimal('0'),
            'export_value': Decimal('0'),
        }

    if not rows:
        return {
            'from_date': from_date,
            'to_date': to_date,
            'data': [],
            'summary': {
                'opening_quantity': 0.0,
                'opening_value': 0.0,
                'import_quantity': 0.0,
                'import_value': 0.0,
                'export_quantity': 0.0,
                'export_value': 0.0,
                'closing_quantity': 0.0,
                'closing_value': 0.0,
            },
        }

    product_ids = {key[0] for key in rows}
    warehouse_ids = {key[1] for key in rows}

    def unit_cost(product_id, preferred_cost=None):
        preferred_cost = Decimal(str(preferred_cost or 0))
        if preferred_cost > 0:
            return preferred_cost
        return product_costs.get(product_id, Decimal('0'))

    def apply_movement(product_id, warehouse_id, movement_date, signed_quantity, movement_value):
        row = rows.get((product_id, warehouse_id))
        if not row or not movement_date or movement_date < from_date:
            return
        signed_quantity = Decimal(str(signed_quantity or 0))
        if signed_quantity == 0:
            return
        movement_value = abs(Decimal(str(movement_value or 0)))

        # Lùi từ tồn hiện tại qua mọi chứng từ kể từ đầu kỳ để suy ra tồn đầu.
        # Các chứng từ sau ngày cuối kỳ cũng được lùi, nhờ đó có thể xem kỳ cũ.
        row['opening_quantity'] -= signed_quantity
        if movement_date > to_date:
            return
        if signed_quantity > 0:
            row['import_quantity'] += signed_quantity
            row['import_value'] += movement_value
        else:
            row['export_quantity'] += abs(signed_quantity)
            row['export_value'] += movement_value

    receipt_items = GoodsReceiptItem.objects.filter(
        product_id__in=product_ids,
        goods_receipt__warehouse_id__in=warehouse_ids,
        goods_receipt__status=1,
        goods_receipt__is_deleted=False,
        goods_receipt__receipt_date__gte=from_date,
    ).select_related('goods_receipt')
    for item in receipt_items:
        quantity = Decimal(str(item.quantity or 0))
        apply_movement(
            item.product_id,
            item.goods_receipt.warehouse_id,
            item.goods_receipt.receipt_date,
            quantity,
            quantity * Decimal(str(item.unit_price or 0)),
        )

    check_items = StockCheckItem.objects.filter(
        product_id__in=product_ids,
        stock_check__warehouse_id__in=warehouse_ids,
        stock_check__status=1,
        stock_check__stock_applied=True,
        stock_check__is_deleted=False,
        stock_check__check_date__gte=from_date,
    ).select_related('stock_check')
    for item in check_items:
        difference = Decimal(str(item.difference or 0))
        apply_movement(
            item.product_id,
            item.stock_check.warehouse_id,
            item.stock_check.check_date,
            difference,
            abs(difference) * unit_cost(item.product_id),
        )

    transfer_items = StockTransferItem.objects.filter(
        Q(transfer__from_warehouse_id__in=warehouse_ids)
        | Q(transfer__to_warehouse_id__in=warehouse_ids),
        product_id__in=product_ids,
        transfer__status=2,
        transfer__is_deleted=False,
        transfer__transfer_date__gte=from_date,
    ).select_related('transfer')
    for item in transfer_items:
        quantity = Decimal(str(item.quantity or 0))
        value = quantity * unit_cost(item.product_id)
        apply_movement(
            item.product_id,
            item.transfer.from_warehouse_id,
            item.transfer.transfer_date,
            -quantity,
            value,
        )
        apply_movement(
            item.product_id,
            item.transfer.to_warehouse_id,
            item.transfer.transfer_date,
            quantity,
            value,
        )

    purchase_return_items = PurchaseReturnItem.objects.filter(
        product_id__in=product_ids,
        purchase_return__warehouse_id__in=warehouse_ids,
        purchase_return__status=1,
        purchase_return__stock_applied=True,
        purchase_return__is_deleted=False,
        purchase_return__return_date__gte=from_date,
    ).select_related('purchase_return')
    for item in purchase_return_items:
        quantity = Decimal(str(item.quantity or 0))
        apply_movement(
            item.product_id,
            item.purchase_return.warehouse_id,
            item.purchase_return.return_date,
            -quantity,
            quantity * unit_cost(item.product_id, item.unit_price),
        )

    order_items = list(OrderItem.objects.filter(
        Q(product_id__in=product_ids)
        | Q(product__is_combo=True, product__combo_items__product_id__in=product_ids),
        order__warehouse_id__in=warehouse_ids,
        order__status__in=(4, 5),
        order__is_deleted=False,
    ).filter(
        Q(order__exported_at__date__gte=from_date)
        | Q(order__exported_at__isnull=True, order__order_date__gte=from_date)
    ).select_related('order', 'product').distinct())

    return_items = list(OrderReturnItem.objects.filter(
        Q(product_id__in=product_ids)
        | Q(product__is_combo=True, product__combo_items__product_id__in=product_ids),
        order_return__warehouse_id__in=warehouse_ids,
        order_return__status=2,
        order_return__is_deleted=False,
        order_return__return_date__gte=from_date,
    ).select_related('order_return', 'order_return__order', 'product').distinct())

    combo_ids = {
        item.product_id
        for item in [*order_items, *return_items]
        if item.product and item.product.is_combo
    }
    combo_components = defaultdict(list)
    if combo_ids:
        for component in ComboItem.objects.filter(
            combo_id__in=combo_ids,
            product_id__in=product_ids,
        ).select_related('product'):
            if not component.product.is_service:
                combo_components[component.combo_id].append(component)

    def expanded_stock_movements(product, quantity, preferred_cost=None):
        quantity = Decimal(str(quantity or 0))
        if not product or product.is_service:
            return []
        if product.is_combo:
            return [
                (
                    component.product_id,
                    quantity * Decimal(str(component.quantity or 0)),
                    unit_cost(component.product_id),
                )
                for component in combo_components.get(product.id, [])
            ]
        return [(product.id, quantity, unit_cost(product.id, preferred_cost))]

    for item in order_items:
        order = item.order
        if order.exported_at:
            exported_at = order.exported_at
            if timezone.is_aware(exported_at):
                exported_at = timezone.localtime(exported_at)
            movement_date = exported_at.date()
        else:
            movement_date = order.order_date
        for product_id, quantity, cost in expanded_stock_movements(
            item.product, item.quantity, item.cost_price,
        ):
            apply_movement(
                product_id,
                order.warehouse_id,
                movement_date,
                -quantity,
                quantity * cost,
            )

    original_order_costs = {}
    original_order_ids = {
        item.order_return.order_id
        for item in return_items
        if item.order_return.order_id
    }
    if original_order_ids:
        cost_buckets = defaultdict(lambda: [Decimal('0'), Decimal('0')])
        for quantity, cost_price, order_id, product_id in OrderItem.objects.filter(
            order_id__in=original_order_ids,
        ).values_list('quantity', 'cost_price', 'order_id', 'product_id'):
            quantity = Decimal(str(quantity or 0))
            cost = Decimal(str(cost_price or 0))
            if product_id and quantity > 0 and cost > 0:
                bucket = cost_buckets[(order_id, product_id)]
                bucket[0] += quantity
                bucket[1] += quantity * cost
        for key, (quantity, value) in cost_buckets.items():
            original_order_costs[key] = value / quantity if quantity > 0 else Decimal('0')

    for item in return_items:
        order_return = item.order_return
        preferred_cost = original_order_costs.get(
            (order_return.order_id, item.product_id),
            Decimal('0'),
        )
        for product_id, quantity, cost in expanded_stock_movements(
            item.product, item.quantity, preferred_cost,
        ):
            apply_movement(
                product_id,
                order_return.warehouse_id,
                order_return.return_date,
                quantity,
                quantity * cost,
            )

    data = []
    for row in rows.values():
        row['closing_quantity'] = (
            row['opening_quantity'] + row['import_quantity'] - row['export_quantity']
        )
        row['opening_value'] = max(row['opening_quantity'], Decimal('0')) * row['valuation_price']
        row['closing_value'] = max(row['closing_quantity'], Decimal('0')) * row['valuation_price']
        data.append({
            key: float(value) if isinstance(value, Decimal) else value
            for key, value in row.items()
        })

    summary_fields = (
        'opening_quantity', 'opening_value', 'import_quantity', 'import_value',
        'export_quantity', 'export_value', 'closing_quantity', 'closing_value',
    )
    summary = {
        field: float(sum(Decimal(str(row[field] or 0)) for row in rows.values()))
        for field in summary_fields
    }
    return {
        'from_date': from_date,
        'to_date': to_date,
        'data': data,
        'summary': summary,
    }


def _build_inventory_movement_category_rows(product_rows):
    """Gộp các dòng nhập xuất tồn sản phẩm/kho theo danh mục gốc."""
    numeric_fields = (
        'opening_quantity', 'opening_value', 'import_quantity', 'import_value',
        'export_quantity', 'export_value', 'closing_quantity', 'closing_value',
    )
    category_map = {}
    for item in product_rows:
        category_id = item.get('category_id')
        key = str(category_id) if category_id not in (None, '') else '__uncategorized__'
        row = category_map.setdefault(key, {
            'category_id': category_id,
            'category': item.get('category_name') or 'Chưa phân loại',
            '_product_ids': set(),
            '_warehouse_ids': set(),
            **{field: Decimal('0') for field in numeric_fields},
        })
        row['_product_ids'].add(item.get('product_id'))
        row['_warehouse_ids'].add(item.get('warehouse_id'))
        for field in numeric_fields:
            row[field] += Decimal(str(item.get(field) or 0))

    data = []
    for row in category_map.values():
        data.append({
            'category_id': row['category_id'],
            'category': row['category'],
            'product_count': len(row.pop('_product_ids')),
            'warehouse_count': len(row.pop('_warehouse_ids')),
            **{field: float(row[field]) for field in numeric_fields},
        })
    return sorted(
        data,
        key=lambda item: (
            item['category_id'] is None,
            str(item['category']).casefold(),
        ),
    )


@login_required(login_url="/login/")
@report_permission_required
def api_report_inventory_movement(request):
    """API báo cáo nhập xuất tồn theo kỳ."""
    try:
        payload = _build_inventory_movement_payload(request)
    except ValueError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)
    return JsonResponse({
        'status': 'ok',
        'from_date': payload['from_date'].isoformat(),
        'to_date': payload['to_date'].isoformat(),
        'data': payload['data'],
        'category_data': _build_inventory_movement_category_rows(payload['data']),
        'summary': payload['summary'],
    })


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
def report_finance(request):
    context = {'active_tab': 'report_finance'}
    return render(request, "reports/report_finance.html", context)


def _get_finance_order_queryset(request, from_date, to_date, store_id=None):
    """Các đơn hàng tạo nên số liệu đã thu và công nợ trên báo cáo tài chính."""
    orders = Order.objects.select_related('customer', 'store').filter(
        order_date__gte=from_date,
        order_date__lte=to_date,
    ).exclude(status=6)
    orders = filter_by_store(orders, request)
    if store_id:
        orders = orders.filter(store_id=store_id)
    return orders


def _with_positive_order_debt(orders):
    """Chỉ giữ đơn còn nợ và tính số nợ theo từng đơn, không để đơn dư tiền bù trừ."""
    return orders.filter(final_amount__gt=F('paid_amount')).annotate(
        debt_amount=ExpressionWrapper(
            F('final_amount') - F('paid_amount'),
            output_field=DecimalField(max_digits=18, decimal_places=0),
        ),
    )


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
def report_finance_order_debt(request):
    """Bảng chi tiết các đơn hàng còn công nợ trong kỳ báo cáo."""
    today = datetime.now().date()
    from_date = parse_date(request.GET.get('from_date') or '') or today.replace(day=1)
    to_date = parse_date(request.GET.get('to_date') or '') or today
    if from_date > to_date:
        from_date, to_date = to_date, from_date

    store_id = _parse_filter_int(request.GET.get('store_id'))
    keyword = (request.GET.get('q') or '').strip()[:100]

    orders = _with_positive_order_debt(
        _get_finance_order_queryset(request, from_date, to_date, store_id),
    )
    if keyword:
        orders = orders.filter(
            Q(code__icontains=keyword)
            | Q(customer__code__icontains=keyword)
            | Q(customer__name__icontains=keyword)
            | Q(customer__phone__icontains=keyword)
            | Q(shipping_phone__icontains=keyword)
        )

    totals = orders.aggregate(
        order_count=Count('id'),
        final_amount=Sum('final_amount'),
        paid_amount=Sum('paid_amount'),
        total_debt_amount=Sum('debt_amount'),
    )
    totals['debt_amount'] = totals.pop('total_debt_amount') or Decimal('0')
    for key in ('final_amount', 'paid_amount'):
        totals[key] = totals[key] or Decimal('0')

    debt_descending = (F('final_amount') - F('paid_amount')).desc()
    paginator = Paginator(orders.order_by(debt_descending, '-order_date', '-id'), 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    from system_management.models import Store
    stores = Store.objects.filter(
        id__in=get_managed_store_ids(request.user),
    ).select_related('brand').order_by('name', 'id')

    context = {
        'active_tab': 'report_finance',
        'page_obj': page_obj,
        'stores': stores,
        'totals': totals,
        'filters': {
            'from_date': from_date.isoformat(),
            'to_date': to_date.isoformat(),
            'store_id': store_id,
            'q': keyword,
            'sort': 'debt_desc',
        },
    }
    return render(request, 'reports/report_finance_order_debt.html', context)


def _get_finance_report_querysets(request, from_date, to_date, store_id=None):
    """Lấy đúng các chứng từ tạo nên báo cáo tài chính.

    Dùng chung hàm này cho API và Excel để phạm vi ngày/cửa hàng/trạng thái
    luôn giống nhau. Việc tách dòng tiền, chi phí khác và hàng nhập sau KM được
    thực hiện ở `_get_finance_expense_metrics`.
    """
    from finance.models import Payment, Receipt
    from products.models import GoodsReceipt

    receipts = Receipt.objects.filter(
        receipt_date__gte=from_date,
        receipt_date__lte=to_date,
        status=1,
    )
    receipts = filter_by_store(receipts, request)

    payments = Payment.objects.filter(
        payment_date__gte=from_date,
        payment_date__lte=to_date,
        status=1,
    )
    payments = filter_by_store(payments, request)

    goods_receipts = GoodsReceipt.objects.filter(
        receipt_date__gte=from_date,
        receipt_date__lte=to_date,
        status=1,
    )
    goods_receipts = filter_by_store(
        goods_receipts,
        request,
        field_name='warehouse__store',
    )

    if store_id:
        receipts = receipts.filter(store_id=store_id)
        payments = payments.filter(store_id=store_id)
        goods_receipts = goods_receipts.filter(warehouse__store_id=store_id)

    return receipts, payments, goods_receipts


def _get_finance_expense_metrics(payments, goods_receipts):
    """Tách chi phí khỏi dòng tiền để phiếu nhập không bị cộng hai lần.

    - Phiếu chi gắn phiếu nhập là dòng tiền thanh toán, không phải một chi phí
      mới vì giá trị hàng đã được ghi nhận từ phiếu nhập.
    - KM nhà cung cấp làm giảm giá trị hàng nhập nhưng không làm tăng số dư quỹ.
    - Phiếu chi không gắn phiếu nhập vẫn là chi phí khác trong kỳ theo ngày chi.
    """
    from finance.models import Payment

    money_zero = Decimal('0')
    cash_payment_expense = payments.aggregate(total=Sum('amount'))['total'] or money_zero
    other_payments = payments.filter(goods_receipt_id__isnull=True)
    other_payment_expense = other_payments.aggregate(total=Sum('amount'))['total'] or money_zero
    goods_receipt_amounts = {
        receipt_id: Decimal(str(total_amount or 0))
        for receipt_id, total_amount in goods_receipts.values_list('id', 'total_amount').iterator()
    }
    promotion_by_receipt = {
        item['goods_receipt_id']: Decimal(str(item['total'] or 0))
        for item in Payment.objects.filter(
            status=1,
            goods_receipt_id__in=goods_receipt_amounts,
        ).values('goods_receipt_id').annotate(total=Sum('promotion_amount')).iterator()
    }
    goods_receipt_gross_expense = sum(goods_receipt_amounts.values(), money_zero)
    # Chặn riêng từng phiếu nhập để dữ liệu cũ/chi nhiều lần không thể tạo ra
    # chi phí âm và tổng API luôn khớp tổng các dòng trong Excel.
    applied_promotion_by_receipt = {
        receipt_id: min(
            max(promotion_by_receipt.get(receipt_id, money_zero), money_zero),
            gross_amount,
        )
        for receipt_id, gross_amount in goods_receipt_amounts.items()
    }
    supplier_promotion = sum(applied_promotion_by_receipt.values(), money_zero)
    goods_receipt_expense = goods_receipt_gross_expense - supplier_promotion
    total_expense = other_payment_expense + goods_receipt_expense
    return {
        'cash_payment_expense': cash_payment_expense,
        'payment_expense': other_payment_expense,
        'goods_receipt_gross_expense': goods_receipt_gross_expense,
        'supplier_promotion': supplier_promotion,
        'goods_receipt_expense': goods_receipt_expense,
        'total_expense': total_expense,
        'other_payments': other_payments,
        'promotion_by_receipt': applied_promotion_by_receipt,
    }


@login_required(login_url="/login/")
@report_permission_required
def api_report_finance(request):
    """API báo cáo tài chính — hỗ trợ filter theo store + breakdown"""
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    store_id = request.GET.get('store_id')
    today = datetime.now().date()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    from finance.models import Payment, Receipt
    from products.models import GoodsReceipt

    receipts, payments, goods_receipts = _get_finance_report_querysets(
        request,
        from_date,
        to_date,
        store_id,
    )

    # Phiếu thu (hoàn thành)
    total_income = float(receipts.aggregate(s=Sum('amount'))['s'] or 0)

    # Thu theo danh mục
    income_by_cat = receipts.values('category__name').annotate(
        amount=Sum('amount')
    ).order_by('-amount')

    expense_metrics = _get_finance_expense_metrics(payments, goods_receipts)
    cash_payment_expense = float(expense_metrics['cash_payment_expense'])
    payment_expense = float(expense_metrics['payment_expense'])
    goods_receipt_gross_expense = float(expense_metrics['goods_receipt_gross_expense'])
    supplier_promotion = float(expense_metrics['supplier_promotion'])
    goods_receipt_expense = float(expense_metrics['goods_receipt_expense'])
    total_expense = float(expense_metrics['total_expense'])

    # Chỉ phiếu chi không gắn phiếu nhập mới là một khoản chi phí riêng.
    expense_by_cat = expense_metrics['other_payments'].values('category__name').annotate(
        amount=Sum('amount')
    ).order_by('-amount')

    # Doanh thu từ đơn hàng
    orders_revenue = _get_finance_order_queryset(
        request,
        from_date,
        to_date,
        store_id,
    )
    order_revenue = float(orders_revenue.aggregate(s=Sum('paid_amount'))['s'] or 0)
    order_debt = float(
        _with_positive_order_debt(orders_revenue).aggregate(s=Sum('debt_amount'))['s'] or 0
    )

    rows = []
    for c in income_by_cat:
        rows.append({'name': c['category__name'] or 'Khác', 'income': float(c['amount'] or 0), 'expense': 0})
    for c in expense_by_cat:
        existing = next((r for r in rows if r['name'] == (c['category__name'] or 'Khác')), None)
        if existing:
            existing['expense'] = float(c['amount'] or 0)
        else:
            rows.append({'name': c['category__name'] or 'Khác', 'income': 0, 'expense': float(c['amount'] or 0)})

    # Đưa hàng nhập sau KM vào bảng danh mục để khớp Tổng chi.
    if goods_receipt_expense:
        goods_receipt_category = 'Hàng nhập sau KM nhà cung cấp'
        existing = next((r for r in rows if r['name'] == goods_receipt_category), None)
        if existing:
            existing['expense'] += goods_receipt_expense
        else:
            rows.append({
                'name': goods_receipt_category,
                'income': 0,
                'expense': goods_receipt_expense,
            })

    # === STORE BREAKDOWN ===
    from core.store_utils import get_managed_store_ids
    from system_management.models import Store
    managed_ids = get_managed_store_ids(request.user)
    managed_stores = Store.objects.filter(id__in=managed_ids).select_related('brand')
    has_multiple = managed_stores.count() > 1

    stores_list = [{'id': s.id, 'name': s.name, 'brand': s.brand.name if s.brand else ''} for s in managed_stores]

    store_breakdown = []
    if has_multiple and not store_id:
        for st in managed_stores:
            st_receipts = Receipt.objects.filter(
                receipt_date__gte=from_date, receipt_date__lte=to_date, status=1, store=st
            )
            st_payments = Payment.objects.filter(
                payment_date__gte=from_date, payment_date__lte=to_date, status=1, store=st
            )
            st_goods_receipts = GoodsReceipt.objects.filter(
                receipt_date__gte=from_date,
                receipt_date__lte=to_date,
                status=1,
                warehouse__store=st,
            )
            st_income = float(st_receipts.aggregate(s=Sum('amount'))['s'] or 0)
            st_metrics = _get_finance_expense_metrics(st_payments, st_goods_receipts)
            st_cash_payment_expense = float(st_metrics['cash_payment_expense'])
            st_payment_expense = float(st_metrics['payment_expense'])
            st_goods_receipt_gross_expense = float(st_metrics['goods_receipt_gross_expense'])
            st_supplier_promotion = float(st_metrics['supplier_promotion'])
            st_goods_receipt_expense = float(st_metrics['goods_receipt_expense'])
            st_expense = float(st_metrics['total_expense'])
            store_breakdown.append({
                'store_id': st.id,
                'store_name': st.name,
                'brand_name': st.brand.name if st.brand else '',
                'income': st_income,
                'cash_payment_expense': st_cash_payment_expense,
                'payment_expense': st_payment_expense,
                'goods_receipt_gross_expense': st_goods_receipt_gross_expense,
                'supplier_promotion': st_supplier_promotion,
                'goods_receipt_expense': st_goods_receipt_expense,
                'expense': st_expense,
                'net': st_income - st_expense,
            })

    return JsonResponse({
        'status': 'ok',
        'has_multiple_stores': has_multiple,
        'stores': stores_list,
        'summary': {
            'total_income': total_income,
            'cash_payment_expense': cash_payment_expense,
            'payment_expense': payment_expense,
            'goods_receipt_gross_expense': goods_receipt_gross_expense,
            'supplier_promotion': supplier_promotion,
            'goods_receipt_expense': goods_receipt_expense,
            'total_expense': total_expense,
            'net_profit': total_income - total_expense,
            'order_revenue': order_revenue,
            'order_debt': order_debt,
            'income_cash': float(receipts.filter(payment_method=1).aggregate(s=Sum('amount'))['s'] or 0),
            'income_transfer': float(receipts.filter(payment_method=2).aggregate(s=Sum('amount'))['s'] or 0),
            'expense_cash': float(payments.filter(payment_method=1).aggregate(s=Sum('amount'))['s'] or 0),
            'expense_transfer': float(payments.filter(payment_method=2).aggregate(s=Sum('amount'))['s'] or 0),
        },
        'categories': rows,
        'store_breakdown': store_breakdown,
    })


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
def report_customers(request):
    context = {'active_tab': 'report_customers'}
    return render(request, "reports/report_customers.html", context)


@login_required(login_url="/login/")
@report_permission_required
def api_report_customers(request):
    """API báo cáo khách hàng — tổng hợp theo lô, không truy vấn từng khách."""
    from customers.models import Customer
    from system_management.models import Store

    store_id = request.GET.get('store_id')
    managed_ids = get_managed_store_ids(request.user)
    scoped_store_ids = list(managed_ids)
    if store_id:
        selected_store_id = _parse_filter_int(store_id)
        if selected_store_id is None or selected_store_id not in managed_ids:
            return JsonResponse({
                'status': 'error',
                'message': 'Cửa hàng không thuộc phạm vi được phép xem.',
            }, status=403)
        scoped_store_ids = [selected_store_id]

    customers = list(
        Customer.objects.filter(
            is_active=True,
            store_id__in=scoped_store_ids,
        ).select_related('group', 'store').order_by('id')
    )
    customer_ids = [customer.id for customer in customers]

    # Doanh thu/tổng mua chỉ lấy đơn đã xuất kho hoặc hoàn thành. Công nợ và
    # số đơn vẫn lấy mọi đơn chưa hủy để không bỏ sót khoản khách còn phải trả.
    realized_metrics = {
        row['customer_id']: row
        for row in Order.objects.filter(
            customer_id__in=customer_ids,
            status__in=[4, 5],
        ).values('customer_id').annotate(
            total_purchased=Sum('final_amount'),
            last_exported_at=Max('exported_at'),
            legacy_last_order_date=Max('order_date'),
        )
    }
    active_metrics = {
        row['customer_id']: row
        for row in Order.objects.filter(
            customer_id__in=customer_ids,
        ).exclude(status=6).values('customer_id').annotate(
            total_amount=Sum('final_amount'),
            total_paid=Sum('paid_amount'),
            order_count=Count('id'),
        )
    }

    data = []
    for customer in customers:
        realized = realized_metrics.get(customer.id)
        active = active_metrics.get(customer.id)
        live_total = float(realized['total_purchased'] or 0) if realized else 0
        live_debt = max(
            float(active['total_amount'] or 0) - float(active['total_paid'] or 0),
            0,
        ) if active else 0
        live_order_count = int(active['order_count'] or 0) if active else 0

        cached_total = float(customer.total_purchased or 0)
        cached_debt = float(customer.total_debt or 0)
        cached_order_count = int(customer.order_count or 0)
        if customer.imported_legacy_metrics:
            total = cached_total + live_total
            debt = cached_debt + live_debt
            order_count = cached_order_count + live_order_count
        else:
            total = live_total if realized else cached_total
            debt = live_debt if active else cached_debt
            order_count = live_order_count if active else cached_order_count

        last_date_value = None
        if realized:
            last_exported_at = realized['last_exported_at']
            if last_exported_at:
                if timezone.is_aware(last_exported_at):
                    last_exported_at = timezone.localtime(last_exported_at)
                last_date_value = last_exported_at.date()
            else:
                last_date_value = realized['legacy_last_order_date']
        if last_date_value is None and customer.last_purchase_at:
            last_purchase_at = customer.last_purchase_at
            if timezone.is_aware(last_purchase_at):
                last_purchase_at = timezone.localtime(last_purchase_at)
            last_date_value = last_purchase_at.date()

        data.append({
            'code': customer.code,
            'name': customer.name,
            'group': customer.group.name if customer.group else '',
            'phone': customer.phone or '',
            'email': customer.email or '',
            'store_id': customer.store_id,
            'store_name': customer.store.name if customer.store else '',
            'order_count': order_count,
            'total_purchased': total,
            'total_debt': debt,
            'last_order_date': last_date_value.strftime('%d/%m/%Y') if last_date_value else '',
        })

    data.sort(key=lambda x: -x['total_purchased'])
    total_revenue = sum(d['total_purchased'] for d in data)
    total_debt = sum(d['total_debt'] for d in data)

    managed_stores = Store.objects.filter(id__in=managed_ids).select_related('brand')
    has_multiple = managed_stores.count() > 1
    stores_list = [{'id': s.id, 'name': s.name, 'brand': s.brand.name if s.brand else ''} for s in managed_stores]

    store_breakdown = []
    if has_multiple and not store_id:
        for st in managed_stores:
            st_customers = [d for d in data if d['store_id'] == st.id]
            st_count = len(st_customers)
            st_revenue = sum(d['total_purchased'] for d in st_customers)
            st_debt = sum(d['total_debt'] for d in st_customers)
            store_breakdown.append({
                'store_id': st.id,
                'store_name': st.name,
                'brand_name': st.brand.name if st.brand else '',
                'customer_count': st_count,
                'revenue': st_revenue,
                'debt': st_debt,
            })
        # Khách chưa gán CH
        no_store = [d for d in data if not d['store_id']]
        if no_store:
            store_breakdown.append({
                'store_id': None,
                'store_name': 'Chưa gán cửa hàng',
                'brand_name': '',
                'customer_count': len(no_store),
                'revenue': sum(d['total_purchased'] for d in no_store),
                'debt': sum(d['total_debt'] for d in no_store),
            })

    return JsonResponse({
        'status': 'ok', 'data': data,
        'has_multiple_stores': has_multiple,
        'stores': stores_list,
        'store_breakdown': store_breakdown,
        'summary': {'total_customers': len(data), 'total_revenue': total_revenue, 'total_debt': total_debt}
    })


@login_required(login_url="/login/")
@brand_owner_required
@report_permission_required
def report_staff_sales(request):
    """Báo cáo doanh thu nhân viên bán hàng"""
    context = {'active_tab': 'report_staff_sales'}
    return render(request, "reports/report_staff_sales.html", context)


@login_required(login_url="/login/")
@report_permission_required
def api_report_staff_sales(request):
    """API báo cáo doanh thu theo nhân viên bán hàng — phục vụ tính KPI & lương"""
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    store_id = request.GET.get('store_id')
    salesperson_filter = request.GET.get('salesperson', '')
    customer_kind = request.GET.get('customer_kind', '').strip()
    if customer_kind not in {'retail', 'wholesale'}:
        customer_kind = ''

    today = datetime.now().date()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    # Doanh thu chỉ được ghi nhận khi đơn đã xuất kho, theo exported_at.
    orders = Order.objects.filter(
        _exported_at_period_q(from_date, to_date),
        status__in=[4, 5],
    )
    orders = filter_by_store(orders, request)
    if store_id:
        orders = orders.filter(store_id=store_id)
    if customer_kind:
        orders = orders.filter(_get_sales_report_customer_kind_q(customer_kind))

    staff_order_ids = _group_order_ids_by_revenue_staff(orders)

    # Tổng doanh thu toàn bộ (dùng để tính tỷ lệ đóng góp)
    grand_total_revenue = float(orders.aggregate(s=Sum('final_amount'))['s'] or 0)

    staff_data = []

    def calc_staff(name, staff_orders):
        order_count = staff_orders.count()
        if order_count == 0:
            return None

        revenue = float(staff_orders.aggregate(s=Sum('final_amount'))['s'] or 0)
        paid = float(staff_orders.aggregate(s=Sum('paid_amount'))['s'] or 0)
        bonus = float(staff_orders.aggregate(s=Sum('bonus_amount'))['s'] or 0)
        discount = float(staff_orders.aggregate(s=Sum('discount_amount'))['s'] or 0)

        # Giá vốn
        cost_data = OrderItem.objects.filter(order__in=staff_orders).aggregate(
            total_cost=Sum(F('cost_price') * F('quantity'))
        )
        cost = float(cost_data['total_cost'] or 0)
        profit = revenue - cost
        gross_margin = round(profit / revenue * 100, 1) if revenue > 0 else 0

        # Trả hàng liên quan (theo customer từ order)
        staff_order_ids = list(staff_orders.values_list('id', flat=True))
        returns_data = OrderReturn.objects.filter(
            order_id__in=staff_order_ids,
            return_date__gte=from_date, return_date__lte=to_date
        ).exclude(status=3).aggregate(
            total_refund=Sum('total_refund'),
            count=Count('id')
        )
        returns_amount = float(returns_data['total_refund'] or 0)
        returns_count = returns_data['count'] or 0

        net_revenue = revenue - returns_amount
        debt = revenue - paid
        contribution = (revenue / grand_total_revenue * 100) if grand_total_revenue > 0 else 0
        avg_per_order = revenue / order_count if order_count > 0 else 0

        # Top 3 sản phẩm bán chạy của NV này
        top_products = OrderItem.objects.filter(
            order__in=staff_orders
        ).values('product__name').annotate(
            total_qty=Sum('quantity'),
            total_amount=Sum('total_price')
        ).order_by('-total_qty')[:3]

        # Đơn hàng chi tiết (cho phần mở rộng)
        order_details = [{
            'code': o.code,
            'date': (
                _order_revenue_date(o).strftime('%d/%m/%Y')
                if _order_revenue_date(o) else ''
            ),
            'customer': o.customer.name if o.customer else 'N/A',
            'final_amount': float(o.final_amount),
            'paid_amount': float(o.paid_amount),
            'bonus_amount': float(o.bonus_amount),
            'status': o.status,
            'status_display': o.get_status_display(),
        } for o in staff_orders.select_related('customer').order_by('-exported_at', '-id')[:50]]

        return {
            'salesperson': name,
            'order_count': order_count,
            'revenue': revenue,
            'cost': cost,
            'profit': profit,
            'gross_margin': gross_margin,
            'discount': discount,
            'returns_amount': returns_amount,
            'returns_count': returns_count,
            'net_revenue': net_revenue,
            'bonus': bonus,
            'debt': debt,
            'paid': paid,
            'contribution': round(contribution, 1),
            'avg_per_order': round(avg_per_order),
            'top_products': [
                {'name': p['product__name'], 'qty': float(p['total_qty'] or 0), 'amount': float(p['total_amount'] or 0)}
                for p in top_products
            ],
            'orders': order_details,
        }

    # Tính cho từng NV. NV được gán thủ công được ưu tiên; nếu trống thì lấy
    # tài khoản tạo đơn.
    for sp_name in sorted(staff_order_ids):
        if salesperson_filter and salesperson_filter.casefold() != sp_name.casefold():
            continue
        sp_orders = orders.filter(id__in=staff_order_ids[sp_name])
        result = calc_staff(sp_name, sp_orders)
        if result:
            staff_data.append(result)

    # Sắp xếp theo doanh thu giảm dần
    staff_data.sort(key=lambda x: -x['revenue'])

    # Gán rank
    for i, d in enumerate(staff_data):
        d['rank'] = i + 1

    # Tổng cộng
    summary = {
        'total_staff': len([d for d in staff_data if d['salesperson'] != '(Chưa gán NV)']),
        'grand_revenue': grand_total_revenue,
        'grand_cost': sum(d['cost'] for d in staff_data),
        'grand_profit': sum(d['profit'] for d in staff_data),
        'grand_orders': sum(d['order_count'] for d in staff_data),
        'grand_returns': sum(d['returns_amount'] for d in staff_data),
        'grand_bonus': sum(d['bonus'] for d in staff_data),
        'grand_debt': sum(d['debt'] for d in staff_data),
        'grand_paid': sum(d['paid'] for d in staff_data),
    }

    # Danh sách NV cho dropdown filter lấy từ user/store, không phụ thuộc kỳ báo cáo.
    all_salespersons = _get_salesperson_filter_options(request, store_id)

    # Store list
    from core.store_utils import get_managed_store_ids
    from system_management.models import Store
    managed_ids = get_managed_store_ids(request.user)
    managed_stores = Store.objects.filter(id__in=managed_ids).select_related('brand')
    has_multiple = managed_stores.count() > 1
    stores_list = [{'id': s.id, 'name': s.name, 'brand': s.brand.name if s.brand else ''} for s in managed_stores]

    return JsonResponse({
        'status': 'ok',
        'has_multiple_stores': has_multiple,
        'stores': stores_list,
        'salespersons': all_salespersons,
        'customer_kinds': [
            option for option in CUSTOMER_KIND_OPTIONS
            if option['value'] in {'retail', 'wholesale'}
        ],
        'selected_customer_kind': customer_kind,
        'staff_data': staff_data,
        'summary': summary,
    })


@login_required(login_url="/login/")
@report_permission_required
def export_staff_sales_excel(request):
    """Xuất báo cáo doanh thu nhân viên ra Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    store_id = request.GET.get('store_id')
    salesperson_filter = request.GET.get('salesperson', '').strip()
    customer_kind = request.GET.get('customer_kind', '').strip()
    if customer_kind not in {'retail', 'wholesale'}:
        customer_kind = ''

    today = datetime.now().date()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    # Lấy dữ liệu (tái sử dụng logic)
    orders = Order.objects.filter(
        _exported_at_period_q(from_date, to_date),
        status__in=[4, 5],
    )
    orders = filter_by_store(orders, request)
    if store_id:
        orders = orders.filter(store_id=store_id)
    if customer_kind:
        orders = orders.filter(_get_sales_report_customer_kind_q(customer_kind))

    staff_order_ids = _group_order_ids_by_revenue_staff(orders)
    grand_total_revenue = float(orders.aggregate(s=Sum('final_amount'))['s'] or 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BC Doanh thu NV'

    # Styles
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_header_font = Font(bold=True, size=10, color='FFFFFF')
    sub_header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    money_format = '#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_font = Font(bold=True, size=10)

    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = 'BÁO CÁO DOANH THU NHÂN VIÊN BÁN HÀNG'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = f'Từ ngày xuất kho {from_date} đến ngày xuất kho {to_date}'
    if customer_kind:
        customer_kind_label = next(
            option['label'] for option in CUSTOMER_KIND_OPTIONS
            if option['value'] == customer_kind
        )
        ws['A2'].value += f' · Nhóm khách: {customer_kind_label}'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        'STT', 'Nhân viên', 'Số đơn', 'Doanh thu', 'Giá vốn', 'Lợi nhuận gộp',
        'Tỷ suất lợi nhuận gộp', 'Trả hàng', 'DT ròng', 'Bonus', 'Công nợ',
        'Đã thu', 'Tỷ lệ (%)',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    row = 5
    sorted_names = [
        name for name in sorted(staff_order_ids)
        if not salesperson_filter or salesperson_filter.casefold() == name.casefold()
    ]

    grand = {'orders': 0, 'revenue': 0, 'cost': 0, 'profit': 0,
             'returns': 0, 'net': 0, 'bonus': 0, 'debt': 0, 'paid': 0}

    for idx, sp_name in enumerate(sorted_names, 1):
        sp_orders = orders.filter(id__in=staff_order_ids[sp_name])

        count = sp_orders.count()
        if count == 0:
            continue
        revenue = float(sp_orders.aggregate(s=Sum('final_amount'))['s'] or 0)
        paid = float(sp_orders.aggregate(s=Sum('paid_amount'))['s'] or 0)
        bonus = float(sp_orders.aggregate(s=Sum('bonus_amount'))['s'] or 0)
        cost_data = OrderItem.objects.filter(order__in=sp_orders).aggregate(
            c=Sum(F('cost_price') * F('quantity'))
        )
        cost = float(cost_data['c'] or 0)
        profit = revenue - cost
        gross_margin = profit / revenue if revenue > 0 else 0

        sp_ids = list(sp_orders.values_list('id', flat=True))
        ret = OrderReturn.objects.filter(
            order_id__in=sp_ids, return_date__gte=from_date, return_date__lte=to_date
        ).exclude(status=3).aggregate(s=Sum('total_refund'))
        returns_amt = float(ret['s'] or 0)

        net_revenue = revenue - returns_amt
        debt = revenue - paid
        contribution = (revenue / grand_total_revenue * 100) if grand_total_revenue > 0 else 0

        data_row = [
            idx, sp_name, count, revenue, cost, profit, gross_margin,
            returns_amt, net_revenue, bonus, debt, paid, round(contribution, 1),
        ]
        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in (4, 5, 6, 8, 9, 10, 11, 12):
                cell.number_format = money_format
            if col == 7:
                cell.number_format = '0.0%'
            if col == 13:
                cell.number_format = '0.0'
            if col in (1, 3):
                cell.alignment = Alignment(horizontal='center')

        grand['orders'] += count
        grand['revenue'] += revenue
        grand['cost'] += cost
        grand['profit'] += profit
        grand['returns'] += returns_amt
        grand['net'] += net_revenue
        grand['bonus'] += bonus
        grand['debt'] += debt
        grand['paid'] += paid
        row += 1

    # Total row
    grand_gross_margin = (
        grand['profit'] / grand['revenue'] if grand['revenue'] > 0 else 0
    )
    total_row = [
        '', 'TỔNG CỘNG', grand['orders'], grand['revenue'], grand['cost'],
        grand['profit'], grand_gross_margin, grand['returns'], grand['net'],
        grand['bonus'], grand['debt'], grand['paid'], 100,
    ]
    for col, val in enumerate(total_row, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        if col in (4, 5, 6, 8, 9, 10, 11, 12):
            cell.number_format = money_format
        if col == 7:
            cell.number_format = '0.0%'
        if col == 13:
            cell.number_format = '0.0'

    # Column widths
    col_widths = [6, 25, 10, 18, 18, 18, 24, 15, 18, 15, 15, 15, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'BC_Doanh_thu_NV_{from_date}_{to_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
@sales_report_privileged_required
def export_sales_excel(request):
    """Xuất báo cáo bán hàng ra Excel theo đúng bộ lọc hiện tại."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    filters = _get_sales_report_filters(request)
    payload = _build_sales_report_payload(request, include_filter_options=False)
    summary = payload.get('summary', {})
    daily = payload.get('daily', [])
    daily_finance = payload.get('daily_finance') or daily
    supplier_breakdown = payload.get('supplier_breakdown', [])
    supplier_summary = payload.get('supplier_summary', {})
    product_breakdown = payload.get('product_breakdown', [])
    sku_details = payload.get('sku_details', [])
    category_breakdown = payload.get('category_breakdown', [])
    customer_kind_breakdown = payload.get('customer_kind_breakdown', [])
    group_breakdown = payload.get('group_breakdown', [])
    order_details = payload.get('order_details', [])
    slow_moving_products = payload.get('slow_moving_products', [])
    slow_moving_summary = payload.get('slow_moving_summary', {})
    time_group_label = payload.get('time_group_label', 'Ngày')

    # Styles
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Doanh thu theo {time_group_label.lower()}"[:31]
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_font = Font(bold=True, size=10, color='FFFFFF')
    sub_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    loss_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    money_fmt = '#,##0'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'BÁO CÁO BÁN HÀNG'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Từ {filters['from_date']} đến {filters['to_date']}"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    filter_labels = _get_sales_report_filter_labels(filters)
    if filter_labels:
        ws.merge_cells('A3:G3')
        ws['A3'] = 'Bộ lọc: ' + ' | '.join(filter_labels)
        ws['A3'].font = Font(italic=True, size=9)
        ws['A3'].alignment = Alignment(horizontal='center')

    headers = ['STT', time_group_label, 'Số ĐH', 'Doanh thu', 'Giá vốn thuần', 'Lợi nhuận', 'Trả hàng']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    row = 5
    for idx, d in enumerate(daily, 1):
        vals = [idx, d['date'], d['count'], d['revenue'], d['cost'], d['profit'], d['returns']]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col >= 4:
                cell.number_format = money_fmt
            # Highlight negative profit
            if col == 6 and val < 0:
                cell.font = Font(bold=True, color='FF0000')
        row += 1

    # Total row
    totals = [
        '', 'TỔNG', summary.get('total_orders', 0), summary.get('total_revenue', 0),
        summary.get('total_cost', 0), summary.get('total_profit', 0), summary.get('total_returns', 0)
    ]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col >= 4:
            cell.number_format = money_fmt

    for i, w in enumerate([6, 20, 12, 18, 18, 18, 15], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 2: Tổng hợp ngày =====
    ws_daily = wb.create_sheet('Tổng hợp ngày')
    daily_headers = [
        'Ngày', 'Tiền hàng', 'Doanh thu', 'Hàng bị trả lại', 'Doanh thu thuần', 'Giá vốn thuần',
        'Lợi nhuận gộp', 'Tỷ suất lợi nhuận gộp', 'Lợi nhuận ròng',
    ]
    for col, h in enumerate(daily_headers, 1):
        cell = ws_daily.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    daily_totals = {
        'goods_amount': 0,
        'revenue': 0,
        'returns': 0,
        'net_revenue': 0,
        'cost': 0,
        'gross_profit': 0,
        'net_profit': 0,
    }
    for idx, row in enumerate(daily_finance, 2):
        goods_amount = float(row.get('goods_amount') or 0)
        revenue = float(row.get('revenue') or 0)
        returns = float(row.get('returns') or 0)
        net_revenue = float(row.get('net_revenue') or 0)
        cost = float(row.get('cost') or 0)
        gross_profit = float(row.get('gross_profit') or 0)
        gross_margin = float(row.get('gross_margin') or 0)
        net_profit = float(row.get('net_profit') or 0)
        values = [
            row.get('date') or '',
            goods_amount,
            revenue,
            returns,
            net_revenue,
            cost,
            gross_profit,
            gross_margin / 100,
            net_profit,
        ]
        for col, val in enumerate(values, 1):
            cell = ws_daily.cell(row=idx, column=col, value=val)
            cell.border = thin
            if col in (2, 3, 4, 5, 6, 7, 9):
                cell.number_format = money_fmt
            if col == 8:
                cell.number_format = '0.0%'
            if col in (7, 9) and float(val or 0) < 0:
                cell.font = Font(bold=True, color='FF0000')
        daily_totals['goods_amount'] += goods_amount
        daily_totals['revenue'] += revenue
        daily_totals['returns'] += returns
        daily_totals['net_revenue'] += net_revenue
        daily_totals['cost'] += cost
        daily_totals['gross_profit'] += gross_profit
        daily_totals['net_profit'] += net_profit

    total_margin = (
        daily_totals['gross_profit'] / daily_totals['net_revenue']
        if daily_totals['net_revenue'] > 0 else 0
    )
    daily_total_row = len(daily_finance) + 2
    for col, val in enumerate([
        'TỔNG',
        daily_totals['goods_amount'],
        daily_totals['revenue'],
        daily_totals['returns'],
        daily_totals['net_revenue'],
        daily_totals['cost'],
        daily_totals['gross_profit'],
        total_margin,
        daily_totals['net_profit'],
    ], 1):
        cell = ws_daily.cell(row=daily_total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col in (2, 3, 4, 5, 6, 7, 9):
            cell.number_format = money_fmt
        if col == 8:
            cell.number_format = '0.0%'

    for i, w in enumerate([16, 18, 18, 18, 18, 18, 18, 20, 18], 1):
        ws_daily.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 3: Bán hàng theo nhà cung cấp =====
    ws_supplier = wb.create_sheet('Bán hàng theo NCC')
    supplier_headers = [
        'STT', 'Nhà cung cấp', 'Số mặt hàng', 'Số đơn hàng', 'SL bán', 'SL trả',
        'Tiêu thụ ròng', 'Doanh thu thuần', 'Giá vốn thuần', 'Lợi nhuận',
        'Tỷ trọng doanh thu', 'Mặt hàng bán chạy',
    ]
    for col, h in enumerate(supplier_headers, 1):
        cell = ws_supplier.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    for idx, supplier in enumerate(supplier_breakdown, 2):
        top_products = ', '.join(
            f"{product.get('name') or ''} ({float(product.get('net_quantity') or 0):g})"
            for product in supplier.get('top_products', [])
        )
        values = [
            idx - 1,
            supplier.get('supplier') or 'Chưa gán nhà cung cấp',
            supplier.get('product_count', 0),
            supplier.get('order_count', 0),
            supplier.get('sold_quantity', 0),
            supplier.get('returned_quantity', 0),
            supplier.get('net_quantity', 0),
            supplier.get('net_revenue', 0),
            supplier.get('cost', 0),
            supplier.get('profit', 0),
            float(supplier.get('contribution') or 0) / 100,
            top_products,
        ]
        for col, val in enumerate(values, 1):
            cell = ws_supplier.cell(row=idx, column=col, value=val)
            cell.border = thin
            if col in (8, 9, 10):
                cell.number_format = money_fmt
            if col == 11:
                cell.number_format = '0.0%'
            if col == 10 and float(val or 0) < 0:
                cell.font = Font(bold=True, color='FF0000')
    supplier_total_row = len(supplier_breakdown) + 2
    supplier_total_values = [
        '', 'TỔNG', supplier_summary.get('product_count', 0), supplier_summary.get('order_count', 0),
        supplier_summary.get('sold_quantity', 0), supplier_summary.get('returned_quantity', 0),
        supplier_summary.get('net_quantity', 0), supplier_summary.get('net_revenue', 0),
        supplier_summary.get('cost', 0), supplier_summary.get('profit', 0),
        1 if supplier_breakdown else 0, '',
    ]
    for col, val in enumerate(supplier_total_values, 1):
        cell = ws_supplier.cell(row=supplier_total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col in (8, 9, 10):
            cell.number_format = money_fmt
        if col == 11:
            cell.number_format = '0.0%'
    for i, width in enumerate([6, 30, 14, 14, 12, 12, 16, 20, 20, 18, 20, 55], 1):
        ws_supplier.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ===== Sheet 4: Mặt hàng =====
    ws2 = wb.create_sheet('Mặt hàng')
    sp_headers = [
        'STT', 'Sản phẩm', 'Nhóm mặt hàng', 'Loại SP', 'SL bán', 'Doanh thu',
        'Giá vốn', 'LN dòng', 'Biên LN', 'Lợi nhuận gộp', 'Tỉ suất LN gộp',
    ]
    for col, h in enumerate(sp_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    for idx, p in enumerate(product_breakdown, 1):
        amt = float(p.get('amount') or 0)
        cst = float(p.get('cost') or 0)
        profit = float(p.get('profit') or 0)
        margin = round((profit / amt) * 100, 1) if amt > 0 else 0
        gross_profit = float(p.get('gross_profit', profit) or 0)
        gross_margin = float(p.get('gross_margin') or 0)
        ws2.cell(row=idx + 1, column=1, value=idx).border = thin
        ws2.cell(row=idx + 1, column=2, value=p.get('name') or '').border = thin
        ws2.cell(row=idx + 1, column=3, value=p.get('category') or '').border = thin
        ws2.cell(row=idx + 1, column=4, value=p.get('product_type') or '').border = thin
        ws2.cell(row=idx + 1, column=5, value=float(p.get('qty') or 0)).border = thin
        c = ws2.cell(row=idx + 1, column=6, value=amt)
        c.number_format = money_fmt
        c.border = thin
        c = ws2.cell(row=idx + 1, column=7, value=cst)
        c.number_format = money_fmt
        c.border = thin
        c = ws2.cell(row=idx + 1, column=8, value=profit)
        c.number_format = money_fmt
        c.border = thin
        ws2.cell(row=idx + 1, column=9, value=margin / 100).number_format = '0.0%'
        ws2.cell(row=idx + 1, column=9).border = thin
        c_gross_profit = ws2.cell(row=idx + 1, column=10, value=gross_profit)
        c_gross_profit.number_format = money_fmt
        c_gross_profit.border = thin
        ws2.cell(row=idx + 1, column=11, value=gross_margin / 100).number_format = '0.0%'
        ws2.cell(row=idx + 1, column=11).border = thin
        if profit < 0:
            c.font = Font(bold=True, color='FF0000')
        if gross_profit < 0:
            c_gross_profit.font = Font(bold=True, color='FF0000')

    for i, w in enumerate([6, 35, 20, 18, 12, 18, 18, 18, 12, 18, 16], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 5: Chi tiết SKU =====
    ws_sku = wb.create_sheet('Chi tiết SKU')
    sku_headers = [
        'Ngày', 'Tên khách hàng', 'Tên sản phẩm', 'Mã SKU', 'Mã đơn hàng',
        'Tên nhân viên', 'Doanh thu thuần', 'Tiền vốn', 'Lợi nhuận gộp',
    ]
    for col, h in enumerate(sku_headers, 1):
        cell = ws_sku.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    sku_totals = {'revenue': 0, 'cost': 0, 'profit': 0}
    for idx, row in enumerate(sku_details, 2):
        revenue = float(row.get('revenue') or 0)
        cost = float(row.get('cost') or 0)
        profit = float(row.get('profit') or 0)
        values = [
            row.get('date') or '',
            row.get('customer') or '',
            row.get('product_name') or '',
            row.get('sku') or '',
            row.get('order_code') or '',
            row.get('salesperson') or '',
            revenue,
            cost,
            profit,
        ]
        for col, val in enumerate(values, 1):
            cell = ws_sku.cell(row=idx, column=col, value=val)
            cell.border = thin
            if col in (7, 8, 9):
                cell.number_format = money_fmt
            if profit < 0:
                cell.fill = loss_fill
            if col == 9 and profit < 0:
                cell.font = Font(bold=True, color='FF0000')
        sku_totals['revenue'] += revenue
        sku_totals['cost'] += cost
        sku_totals['profit'] += profit

    sku_total_row = len(sku_details) + 2
    for col, val in enumerate([
        '', '', '', '', '', 'TỔNG',
        sku_totals['revenue'], sku_totals['cost'], sku_totals['profit'],
    ], 1):
        cell = ws_sku.cell(row=sku_total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col in (7, 8, 9):
            cell.number_format = money_fmt

    for i, w in enumerate([14, 28, 36, 18, 16, 22, 18, 18, 18], 1):
        ws_sku.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 5: Nhóm mặt hàng =====
    ws3 = wb.create_sheet('Nhóm mặt hàng')
    cat_headers = [
        'STT', 'Nhóm mặt hàng', 'SL bán', 'Doanh thu', 'Giá vốn',
        'Lợi nhuận', 'Biên LN', 'Lợi nhuận gộp', 'Tỉ suất LN gộp',
    ]
    for col, h in enumerate(cat_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    for idx, row in enumerate(category_breakdown, 1):
        revenue = float(row.get('revenue') or 0)
        cost = float(row.get('cost') or 0)
        profit = float(row.get('profit') or 0)
        margin = round((profit / revenue) * 100, 1) if revenue > 0 else 0
        gross_profit = float(row.get('gross_profit', profit) or 0)
        gross_margin = float(row.get('gross_margin') or 0)
        values = [
            idx, row.get('name') or '', float(row.get('qty') or 0),
            revenue, cost, profit, margin / 100, gross_profit,
            gross_margin / 100,
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=idx + 1, column=col, value=val)
            cell.border = thin
            if col in (4, 5, 6, 8):
                cell.number_format = money_fmt
            if col in (7, 9):
                cell.number_format = '0.0%'
            if (col == 6 and profit < 0) or (col == 8 and gross_profit < 0):
                cell.font = Font(bold=True, color='FF0000')

    for i, w in enumerate([6, 28, 12, 18, 18, 18, 12, 18, 16], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 6: Khách buôn/lẻ =====
    ws_kind = wb.create_sheet('Khách buôn lẻ')
    kind_headers = ['STT', 'Kiểu khách', 'Số ĐH', 'Doanh thu', 'Giá vốn', 'Lợi nhuận', 'Đã thu', 'Công nợ', 'Tỷ trọng']
    for col, h in enumerate(kind_headers, 1):
        cell = ws_kind.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    for idx, row in enumerate(customer_kind_breakdown, 1):
        values = [
            idx,
            row.get('name') or '',
            int(row.get('orders') or 0),
            float(row.get('amount') or 0),
            float(row.get('cost') or 0),
            float(row.get('profit') or 0),
            float(row.get('paid') or 0),
            float(row.get('debt') or 0),
            (float(row.get('contribution') or 0) / 100),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_kind.cell(row=idx + 1, column=col, value=val)
            cell.border = thin
            if col in (4, 5, 6, 7, 8):
                cell.number_format = money_fmt
            if col == 9:
                cell.number_format = '0.0%'
            if col == 6 and float(row.get('profit') or 0) < 0:
                cell.font = Font(bold=True, color='FF0000')

    for i, w in enumerate([6, 22, 12, 18, 18, 18, 18, 18, 12], 1):
        ws_kind.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 7: Nhóm khách hàng =====
    ws4 = wb.create_sheet('Nhóm khách hàng')
    grp_headers = ['STT', 'Nhóm KH', 'Số ĐH', 'Doanh thu', 'Giá vốn', 'Lợi nhuận', 'Đã thu', 'Công nợ', 'Tỷ trọng']
    for col, h in enumerate(grp_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    for idx, row in enumerate(group_breakdown, 1):
        values = [
            idx,
            row.get('name') or '',
            int(row.get('orders') or 0),
            float(row.get('amount') or 0),
            float(row.get('cost') or 0),
            float(row.get('profit') or 0),
            float(row.get('paid') or 0),
            float(row.get('debt') or 0),
            (float(row.get('contribution') or 0) / 100),
        ]
        for col, val in enumerate(values, 1):
            cell = ws4.cell(row=idx + 1, column=col, value=val)
            cell.border = thin
            if col in (4, 5, 6, 7, 8):
                cell.number_format = money_fmt
            if col == 9:
                cell.number_format = '0.0%'
            if col == 6 and float(row.get('profit') or 0) < 0:
                cell.font = Font(bold=True, color='FF0000')

    for i, w in enumerate([6, 24, 12, 18, 18, 18, 18, 18, 12], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet 8: Chi tiết đơn hàng =====
    ws5 = wb.create_sheet('Chi tiết đơn hàng')
    od_headers = [
        'STT', 'Mã ĐH', 'Ngày', 'Khách hàng', 'Kiểu khách', 'Nhóm KH',
        'Sản phẩm lỗ', 'Doanh thu', 'Đã thu', 'Công nợ', 'Giá vốn', 'Lợi nhuận', 'Báo lỗ',
        'TT đơn', 'TT thanh toán',
    ]
    for col, h in enumerate(od_headers, 1):
        cell = ws5.cell(row=1, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin

    od_row = 2
    od_grand = {'revenue': 0, 'paid': 0, 'debt': 0, 'cost': 0, 'profit': 0}
    for idx, o in enumerate(order_details, 1):
        is_loss = bool(o.get('is_loss'))
        vals = [
            idx,
            o.get('code') or '',
            o.get('date') or '',
            o.get('customer') or '',
            o.get('customer_kind_label') or '',
            o.get('customer_group') or '',
            o.get('loss_product_names') or '',
            float(o.get('revenue') or 0),
            float(o.get('paid') or 0),
            float(o.get('debt') or 0),
            float(o.get('cost') or 0),
            float(o.get('profit') or 0),
            'Lỗ' if is_loss else '',
            o.get('status_display') or '',
            o.get('payment_status_display') or '',
        ]
        for col, val in enumerate(vals, 1):
            cell = ws5.cell(row=od_row, column=col, value=val)
            cell.border = thin
            if col in (8, 9, 10, 11, 12):
                cell.number_format = money_fmt
            if is_loss:
                cell.fill = loss_fill
            if col == 12 and is_loss:
                cell.font = Font(bold=True, color='FF0000')

        od_grand['revenue'] += float(o.get('revenue') or 0)
        od_grand['paid'] += float(o.get('paid') or 0)
        od_grand['debt'] += float(o.get('debt') or 0)
        od_grand['cost'] += float(o.get('cost') or 0)
        od_grand['profit'] += float(o.get('profit') or 0)
        od_row += 1

    # Total row for order detail
    for col, val in enumerate([
        '', 'TỔNG', '', '', '', '', '',
        od_grand['revenue'], od_grand['paid'], od_grand['debt'],
        od_grand['cost'], od_grand['profit'], '', '', '',
    ], 1):
        cell = ws5.cell(row=od_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col in (8, 9, 10, 11, 12):
            cell.number_format = money_fmt

    for i, w in enumerate([6, 12, 12, 25, 18, 15, 36, 18, 18, 18, 18, 18, 10, 15, 18], 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ===== Sheet: Cảnh báo hàng chậm =====
    ws_slow = wb.create_sheet('Cảnh báo hàng chậm')
    ws_slow.merge_cells('A1:I1')
    ws_slow['A1'] = 'CẢNH BÁO HÀNG CHẬM'
    ws_slow['A1'].font = header_font
    ws_slow['A1'].fill = header_fill
    ws_slow['A1'].alignment = Alignment(horizontal='center')
    ws_slow.merge_cells('A2:I2')
    ws_slow['A2'] = (
        'Ngày chưa bán tính từ lần bán gần nhất của đơn đã xuất kho/hoàn thành; '
        'hàng chưa từng bán được hiển thị riêng và không quy đổi thành số ngày.'
    )
    ws_slow['A2'].font = Font(italic=True, size=10)
    ws_slow.merge_cells('A3:I3')
    ws_slow['A3'] = (
        f"Chưa từng bán: {slow_moving_summary.get('never_sold_count', 0)} | "
        f"Từ 30 ngày: {slow_moving_summary.get('over_30_count', 0)} | "
        f"Từ 60 ngày: {slow_moving_summary.get('over_60_count', 0)} | "
        f"Từ 90 ngày: {slow_moving_summary.get('over_90_count', 0)} | "
        f"Giá trị tồn từ 90 ngày: {slow_moving_summary.get('over_90_stock_value', 0):,.0f}đ"
    )
    ws_slow['A3'].font = Font(bold=True, size=10)

    slow_headers = [
        'STT', 'Mã sản phẩm', 'Tên sản phẩm', 'Danh mục', 'Nhà cung cấp',
        'Lần bán gần nhất', 'Ngày chưa bán', 'Tồn hiện tại', 'Giá trị tồn',
    ]
    for col, heading in enumerate(slow_headers, 1):
        cell = ws_slow.cell(row=4, column=col, value=heading)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')

    alert_rows = [
        row
        for row in slow_moving_products
        if row.get('never_sold') or float(row.get('days_without_sale') or 0) >= 30
    ]
    for idx, alert_row in enumerate(alert_rows, 1):
        last_sale = alert_row.get('last_sale_date_display') or 'Chưa từng bán'
        if alert_row.get('never_sold'):
            last_sale = 'Chưa từng bán (tính ngày từ ngày tạo)'
        values = [
            idx,
            alert_row.get('product_code') or '',
            alert_row.get('product_name') or '',
            alert_row.get('category') or '',
            alert_row.get('supplier') or '',
            last_sale,
            float(alert_row.get('days_without_sale') or 0),
            float(alert_row.get('stock') or 0),
            float(alert_row.get('stock_value') or 0),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_slow.cell(row=idx + 4, column=col, value=val)
            cell.border = thin
            if col == 9:
                cell.number_format = money_fmt
            if alert_row.get('warning_level') == 'critical':
                cell.fill = loss_fill
            elif alert_row.get('warning_level') in ('slow', 'watch'):
                cell.fill = total_fill

    ws_slow.freeze_panes = 'A5'
    for i, width in enumerate([6, 18, 36, 28, 26, 28, 16, 16, 20], 1):
        ws_slow.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"BC_Ban_hang_{filters['from_date']}_{filters['to_date']}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
def export_inventory_excel(request):
    """Xuất báo cáo tồn kho ra Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from products.models import ProductStock

    warehouse_id = request.GET.get('warehouse_id')
    search = request.GET.get('search', '').strip()
    category_id = request.GET.get('category_id')
    product_type_id = request.GET.get('product_type_id')

    stocks = ProductStock.objects.select_related('product', 'product__category', 'warehouse').filter(
        product__is_deleted=False,
    )
    stocks = filter_by_store(stocks, request, field_name='warehouse__store')
    if warehouse_id:
        stocks = stocks.filter(warehouse_id=warehouse_id)
    if search:
        stocks = stocks.filter(
            Q(product__name__icontains=search) |
            Q(product__code__icontains=search) |
            Q(product__barcode__icontains=search)
        )
    if category_id:
        stocks = stocks.filter(
            Q(product__category_id=category_id) |
            Q(product__category__parent_id=category_id)
        )
    if product_type_id:
        stocks = stocks.filter(product__category_id=product_type_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tồn kho'
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_font = Font(bold=True, size=10, color='FFFFFF')
    sub_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    danger_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    warning_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    money_fmt = '#,##0'

    ws.merge_cells('A1:L1')
    ws['A1'] = 'BÁO CÁO TỒN KHO'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Ngày xuất: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['STT', 'Mã SP', 'Tên sản phẩm', 'Danh mục', 'ĐVT', 'Kho', 'Tồn kho',
               'Tối thiểu', 'Tối đa', 'Giá tính tồn', 'Giá trị tồn', 'Cảnh báo']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    row = 5
    total_value = 0
    total_qty = 0
    for idx, s in enumerate(stocks, 1):
        qty = float(s.quantity)
        valuation_price, _ = _inventory_valuation_unit_cost(s.product)
        valuation_price = float(valuation_price)
        value = valuation_price * max(qty, 0)
        total_value += value
        total_qty += qty

        alert = ''
        fill = None
        if s.product.min_stock and qty < s.product.min_stock:
            alert = 'Dưới tối thiểu'
            fill = danger_fill
        elif s.product.max_stock and qty > s.product.max_stock:
            alert = 'Trên tối đa'
            fill = warning_fill

        cat_name = s.product.category.name if s.product.category else ''
        vals = [idx, s.product.code, s.product.name, cat_name, s.product.unit or '',
                s.warehouse.name, qty, s.product.min_stock or 0,
                s.product.max_stock or 0, valuation_price, value, alert]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col in (10, 11):
                cell.number_format = money_fmt
            if fill:
                cell.fill = fill
        row += 1

    # Total
    ws.cell(row=row, column=1, value='').border = thin
    c = ws.cell(row=row, column=2, value='TỔNG CỘNG')
    c.font = Font(bold=True)
    c.border = thin
    for col in range(3, 7):
        ws.cell(row=row, column=col, value='').border = thin
    c = ws.cell(row=row, column=7, value=total_qty)
    c.font = Font(bold=True)
    c.border = thin
    for col in range(8, 11):
        ws.cell(row=row, column=col, value='').border = thin
    c = ws.cell(row=row, column=11, value=total_value)
    c.font = Font(bold=True)
    c.number_format = money_fmt
    c.border = thin
    ws.cell(row=row, column=12, value='').border = thin

    col_widths = [6, 12, 30, 18, 8, 15, 12, 12, 12, 15, 18, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'BC_Ton_kho_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
def export_inventory_movement_excel(request):
    """Xuất báo cáo nhập xuất tồn theo đúng kỳ và bộ lọc đang xem."""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    try:
        payload = _build_inventory_movement_payload(request)
    except ValueError as exc:
        return HttpResponse(str(exc), status=400, content_type='text/plain; charset=utf-8')

    group_by_category = (request.GET.get('group_by') or '').strip() == 'category'
    report_rows = (
        _build_inventory_movement_category_rows(payload['data'])
        if group_by_category else payload['data']
    )
    if group_by_category:
        identity_headers = ['STT', 'Danh mục', 'Số SP', 'Số kho']
        identity_widths = (6, 30, 12, 12)
        report_title = 'BÁO CÁO NHẬP XUẤT TỒN THEO DANH MỤC'
    else:
        identity_headers = ['STT', 'Mã SP', 'Tên sản phẩm', 'Danh mục', 'ĐVT', 'Kho']
        identity_widths = (6, 15, 32, 20, 10, 20)
        report_title = 'BÁO CÁO NHẬP XUẤT TỒN THEO SẢN PHẨM'
    identity_column_count = len(identity_headers)
    movement_start_column = identity_column_count + 1
    last_column = identity_column_count + 8
    last_column_letter = openpyxl.utils.get_column_letter(last_column)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NXT theo danh mục' if group_by_category else 'NXT theo sản phẩm'
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    group_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    sub_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    white_bold_font = Font(bold=True, color='FFFFFF')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    money_fmt = '#,##0'
    quantity_fmt = '#,##0.##'

    ws.merge_cells(f'A1:{last_column_letter}1')
    ws['A1'] = report_title
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A2:{last_column_letter}2')
    ws['A2'] = (
        f"Kỳ báo cáo: {payload['from_date'].strftime('%d/%m/%Y')} - "
        f"{payload['to_date'].strftime('%d/%m/%Y')}"
    )
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    for column, label in enumerate(identity_headers, 1):
        ws.merge_cells(start_row=4, start_column=column, end_row=5, end_column=column)
        cell = ws.cell(row=4, column=column, value=label)
        cell.fill = group_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        ws.cell(row=5, column=column).border = thin

    for pair_index, label in enumerate((
        'Tồn đầu kỳ', 'Nhập trong kỳ', 'Xuất trong kỳ', 'Tồn cuối kỳ',
    )):
        start_column = movement_start_column + pair_index * 2
        ws.merge_cells(
            start_row=4,
            start_column=start_column,
            end_row=4,
            end_column=start_column + 1,
        )
        cell = ws.cell(row=4, column=start_column, value=label)
        cell.fill = group_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin
        ws.cell(row=4, column=start_column + 1).border = thin
        for offset, sub_label in enumerate(('Số lượng', 'Giá trị')):
            sub_cell = ws.cell(row=5, column=start_column + offset, value=sub_label)
            sub_cell.fill = sub_fill
            sub_cell.font = white_bold_font
            sub_cell.alignment = Alignment(horizontal='center')
            sub_cell.border = thin

    field_names = (
        'opening_quantity', 'opening_value', 'import_quantity', 'import_value',
        'export_quantity', 'export_value', 'closing_quantity', 'closing_value',
    )
    row_number = 6
    for index, item in enumerate(report_rows, 1):
        if group_by_category:
            identity_values = [
                index, item['category'], item['product_count'], item['warehouse_count'],
            ]
        else:
            identity_values = [
                index, item['product_code'], item['product_name'], item['category'],
                item['unit'], item['warehouse'],
            ]
        values = [*identity_values, *[item[field] for field in field_names]]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.border = thin
            if column >= movement_start_column and (column - movement_start_column) % 2 == 0:
                cell.number_format = quantity_fmt
            elif column >= movement_start_column:
                cell.number_format = money_fmt
        row_number += 1

    summary = payload['summary']
    ws.merge_cells(
        start_row=row_number,
        start_column=1,
        end_row=row_number,
        end_column=identity_column_count,
    )
    total_cell = ws.cell(row=row_number, column=1, value='TỔNG CỘNG')
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='right')
    total_cell.border = thin
    for column in range(2, identity_column_count + 1):
        ws.cell(row=row_number, column=column).border = thin
    for offset, field in enumerate(field_names, movement_start_column):
        cell = ws.cell(row=row_number, column=offset, value=summary[field])
        cell.font = Font(bold=True)
        cell.border = thin
        cell.number_format = (
            quantity_fmt if (offset - movement_start_column) % 2 == 0 else money_fmt
        )

    column_widths = (*identity_widths, 14, 18, 14, 18, 14, 18, 14, 18)
    for index, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    ws.freeze_panes = 'A6'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename_prefix = 'BC_Nhap_xuat_ton_theo_danh_muc' if group_by_category else 'BC_Nhap_xuat_ton_theo_SP'
    filename = (
        f"{filename_prefix}_{payload['from_date'].strftime('%Y%m%d')}_"
        f"{payload['to_date'].strftime('%Y%m%d')}.xlsx"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
def export_orders_excel(request):
    """Xuất danh sách đơn hàng ra Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    status = request.GET.get('status')
    payment_status = request.GET.get('payment_status')

    orders = Order.objects.select_related('customer', 'warehouse').all()
    orders = filter_by_store(orders, request)
    if from_date:
        orders = orders.filter(order_date__gte=from_date)
    if to_date:
        orders = orders.filter(order_date__lte=to_date)
    if status:
        orders = orders.filter(status=int(status))
    if payment_status:
        orders = orders.filter(payment_status=int(payment_status))
    orders = orders.order_by('-order_date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Danh sách đơn hàng'
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_font = Font(bold=True, size=10, color='FFFFFF')
    sub_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    money_fmt = '#,##0'

    ws.merge_cells('A1:J1')
    ws['A1'] = 'DANH SÁCH ĐƠN HÀNG'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    date_range = ''
    if from_date and to_date:
        date_range = f'Từ {from_date} đến {to_date}'
    elif from_date:
        date_range = f'Từ {from_date}'
    elif to_date:
        date_range = f'Đến {to_date}'
    else:
        date_range = 'Tất cả'
    ws.merge_cells('A2:J2')
    ws['A2'] = date_range
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['STT', 'Mã ĐH', 'Khách hàng', 'Kho', 'Ngày đặt',
               'Tổng tiền', 'Đã thanh toán', 'Còn nợ', 'Trạng thái', 'TT thanh toán']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    row = 5
    grand = {'total': 0, 'paid': 0, 'debt': 0}
    for idx, o in enumerate(orders, 1):
        total = float(o.final_amount)
        paid = float(o.paid_amount)
        debt = total - paid
        grand['total'] += total
        grand['paid'] += paid
        grand['debt'] += debt

        vals = [
            idx, o.code,
            o.customer.name if o.customer else '',
            o.warehouse.name if o.warehouse else '',
            o.order_date.strftime('%d/%m/%Y') if o.order_date else '',
            total, paid, debt,
            o.get_status_display(),
            o.get_payment_status_display(),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col in (6, 7, 8):
                cell.number_format = money_fmt
        row += 1

    # Total row
    totals = ['', 'TỔNG', '', '', '', grand['total'], grand['paid'], grand['debt'], '', '']
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col in (6, 7, 8):
            cell.number_format = money_fmt

    col_widths = [6, 15, 25, 15, 12, 18, 18, 18, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'DS_Don_hang_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
def export_customers_excel(request):
    """Xuất báo cáo khách hàng ra Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from customers.models import Customer

    store_id = request.GET.get('store_id')

    customers = Customer.objects.filter(is_active=True).select_related('group', 'store')
    customers = filter_by_store(customers, request)
    if store_id:
        customers = customers.filter(store_id=store_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Khách hàng'
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_font = Font(bold=True, size=10, color='FFFFFF')
    sub_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    debt_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    money_fmt = '#,##0'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'BÁO CÁO KHÁCH HÀNG'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Ngày xuất: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['STT', 'Mã KH', 'Tên KH', 'SĐT', 'Email', 'Nhóm',
               'Số ĐH', 'Tổng mua', 'Công nợ']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    row = 5
    grand = {'orders': 0, 'revenue': 0, 'debt': 0}
    for idx, c in enumerate(customers, 1):
        orders = Order.objects.filter(customer=c).exclude(status=6)
        order_count = orders.count()
        total = float(orders.aggregate(s=Sum('final_amount'))['s'] or 0)
        paid = float(orders.aggregate(s=Sum('paid_amount'))['s'] or 0)
        debt = total - paid
        grand['orders'] += order_count
        grand['revenue'] += total
        grand['debt'] += debt

        vals = [idx, c.code, c.name, c.phone or '', c.email or '',
                c.group.name if c.group else '', order_count, total, debt]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col in (8, 9):
                cell.number_format = money_fmt
            if debt > 0:
                cell.fill = debt_fill
        row += 1

    # Total
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    totals = ['', 'TỔNG', '', '', '', '', grand['orders'], grand['revenue'], grand['debt']]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin
        if col in (8, 9):
            cell.number_format = money_fmt

    col_widths = [6, 12, 25, 15, 25, 15, 10, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'BC_Khach_hang_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
def export_purchases_excel(request):
    """Xuất báo cáo nhập hàng ra Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    today = datetime.now().date()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    supplier_id = _parse_filter_int(request.GET.get('supplier_id'))
    receipts = _purchase_report_receipts(
        request,
        from_date,
        to_date,
        supplier_id=supplier_id,
    ).order_by('-receipt_date', '-id')
    supplier_summary = _purchase_supplier_summary(receipts)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Nhập hàng'
    hf = Font(bold=True, size=14, color='FFFFFF')
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sf = Font(bold=True, size=10, color='FFFFFF')
    sfill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    cancel_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    tfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    mfmt = '#,##0'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'BÁO CÁO NHẬP HÀNG'
    ws['A1'].font = hf
    ws['A1'].fill = hfill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Từ {from_date} đến {to_date}'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['STT', 'Mã phiếu', 'Ngày', 'Nhà cung cấp', 'Kho', 'Tổng tiền', 'Trạng thái']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = sf
        c.fill = sfill
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    row = 5
    total = 0
    for idx, r in enumerate(receipts, 1):
        amt = float(r.total_amount)
        is_cancel = (r.status == 2)
        if r.status == 1:
            total += amt
        vals = [idx, r.code,
                r.receipt_date.strftime('%d/%m/%Y') if r.receipt_date else '',
                r.supplier.name if r.supplier else '',
                r.warehouse.name if r.warehouse else '',
                amt, r.get_status_display()]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin
            if col == 6:
                c.number_format = mfmt
            if is_cancel:
                c.fill = cancel_fill
        row += 1

    totals = ['', 'TỔNG', '', '', '', total, '']
    for col, val in enumerate(totals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = tfill
        c.border = thin
        if col == 6:
            c.number_format = mfmt

    for i, w in enumerate([6, 15, 12, 25, 15, 18, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    summary_ws = wb.create_sheet('Tổng hợp NCC')
    summary_ws.merge_cells('A1:D1')
    summary_ws['A1'] = 'TỔNG HỢP NHẬP HÀNG THEO NHÀ CUNG CẤP'
    summary_ws['A1'].font = hf
    summary_ws['A1'].fill = hfill
    summary_ws['A1'].alignment = Alignment(horizontal='center')
    summary_ws.merge_cells('A2:D2')
    summary_ws['A2'] = f'Từ {from_date} đến {to_date} · Chỉ tính phiếu hoàn thành'
    summary_ws['A2'].font = Font(italic=True, size=10)
    summary_ws['A2'].alignment = Alignment(horizontal='center')

    summary_headers = ['STT', 'Nhà cung cấp', 'Số phiếu hoàn thành', 'Tổng tiền hàng']
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=4, column=col, value=header)
        cell.font = sf
        cell.fill = sfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    summary_row = 5
    for index, item in enumerate(supplier_summary, 1):
        values = [index, item['supplier'], item['receipt_count'], item['total_amount']]
        for col, value in enumerate(values, 1):
            cell = summary_ws.cell(row=summary_row, column=col, value=value)
            cell.border = thin
            if col == 4:
                cell.number_format = mfmt
        summary_row += 1

    summary_totals = ['', 'TỔNG', sum(item['receipt_count'] for item in supplier_summary), total]
    for col, value in enumerate(summary_totals, 1):
        cell = summary_ws.cell(row=summary_row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.fill = tfill
        cell.border = thin
        if col == 4:
            cell.number_format = mfmt
    for index, width in enumerate([6, 32, 22, 20], 1):
        summary_ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="BC_Nhap_hang_{from_date}_{to_date}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url="/login/")
@report_permission_required
def export_finance_excel(request):
    """Xuất báo cáo tài chính ra Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    store_id = request.GET.get('store_id')
    today = datetime.now().date()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    receipts, payments, goods_receipts = _get_finance_report_querysets(
        request,
        from_date,
        to_date,
        store_id,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Thu chi'
    hf = Font(bold=True, size=14, color='FFFFFF')
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sf = Font(bold=True, size=10, color='FFFFFF')
    sfill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    goods_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    tfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    mfmt = '#,##0'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'BÁO CÁO TÀI CHÍNH'
    ws['A1'].font = hf
    ws['A1'].fill = hfill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Từ {from_date} đến {to_date}'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Summary: chi phí hàng nhập chỉ ghi một lần; phiếu chi gắn phiếu nhập là
    # dòng tiền và không được cộng thêm vào chi phí.
    total_income = float(receipts.aggregate(s=Sum('amount'))['s'] or 0)
    expense_metrics = _get_finance_expense_metrics(payments, goods_receipts)
    cash_payment_expense = float(expense_metrics['cash_payment_expense'])
    payment_expense = float(expense_metrics['payment_expense'])
    goods_receipt_gross_expense = float(expense_metrics['goods_receipt_gross_expense'])
    supplier_promotion = float(expense_metrics['supplier_promotion'])
    goods_receipt_expense = float(expense_metrics['goods_receipt_expense'])
    total_expense = float(expense_metrics['total_expense'])
    net = total_income - total_expense
    ws['A3'] = (
        f'Tổng thu: {total_income:,.0f}đ  |  '
        f'Thực chi qua phiếu: {cash_payment_expense:,.0f}đ  |  '
        f'Chi phí khác: {payment_expense:,.0f}đ'
    )
    ws['A3'].font = Font(bold=True, size=10)
    ws.merge_cells('A3:G3')
    ws['A4'] = (
        f'Hàng nhập: {goods_receipt_gross_expense:,.0f}đ - KM NCC: {supplier_promotion:,.0f}đ'
        f' = {goods_receipt_expense:,.0f}đ  |  Tổng chi phí: {total_expense:,.0f}đ  |  '
        f'Lãi/Lỗ: {net:,.0f}đ'
    )
    ws['A4'].font = Font(bold=True, size=10)
    ws.merge_cells('A4:G4')

    headers = ['STT', 'Loại', 'Mã phiếu', 'Ngày', 'Danh mục', 'Diễn giải', 'Số tiền']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = sf
        c.fill = sfill
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    row = 6
    idx = 1
    # Ghi phiếu thu
    for r in receipts.select_related('category').order_by('-receipt_date'):
        vals = [idx, 'THU', r.code,
                r.receipt_date.strftime('%d/%m/%Y') if r.receipt_date else '',
                r.category.name if r.category else '',
                r.description or '', float(r.amount)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin
            c.fill = green_fill
            if col == 7:
                c.number_format = mfmt
        idx += 1
        row += 1

    # Chỉ ghi phiếu chi không gắn phiếu nhập vào bảng chi phí. Phiếu chi gắn
    # phiếu nhập vẫn đã nằm trong chỉ tiêu dòng tiền ở phần tổng hợp phía trên.
    for p in expense_metrics['other_payments'].select_related('category').order_by('-payment_date'):
        vals = [idx, 'CHI', p.code,
                p.payment_date.strftime('%d/%m/%Y') if p.payment_date else '',
                p.category.name if p.category else '',
                p.description or '', float(p.amount)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin
            c.fill = red_fill
            if col == 7:
                c.number_format = mfmt
        idx += 1
        row += 1

    promotion_by_receipt = expense_metrics['promotion_by_receipt']

    # Ghi phiếu nhập hàng theo giá trị sau KM, nhưng vẫn nêu rõ giá trị gốc.
    for goods_receipt in goods_receipts.select_related(
        'supplier',
        'warehouse',
    ).order_by('-receipt_date'):
        description_parts = []
        if goods_receipt.supplier:
            description_parts.append(f'NCC: {goods_receipt.supplier.name}')
        if goods_receipt.warehouse:
            description_parts.append(f'Kho: {goods_receipt.warehouse.name}')
        if goods_receipt.note:
            description_parts.append(goods_receipt.note)
        gross_amount = Decimal(str(goods_receipt.total_amount or 0))
        promotion_amount = promotion_by_receipt.get(goods_receipt.id, Decimal('0'))
        net_goods_amount = max(gross_amount - promotion_amount, Decimal('0'))
        if promotion_amount:
            description_parts.append(
                f'Giá trị gốc: {gross_amount:,.0f}đ · KM NCC: {promotion_amount:,.0f}đ'
            )
        vals = [
            idx,
            'NHẬP HÀNG',
            goods_receipt.code,
            goods_receipt.receipt_date.strftime('%d/%m/%Y') if goods_receipt.receipt_date else '',
            'Hàng nhập sau KM nhà cung cấp',
            ' · '.join(description_parts),
            float(net_goods_amount),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin
            c.fill = goods_fill
            if col == 7:
                c.number_format = mfmt
        idx += 1
        row += 1

    # Total rows
    for label, amt, fill in [
        ('TỔNG THU', total_income, green_fill),
        ('THỰC CHI QUA PHIẾU', cash_payment_expense, red_fill),
        ('TỔNG CHI PHÍ KHÁC', payment_expense, red_fill),
        ('HÀNG NHẬP TRƯỚC KM', goods_receipt_gross_expense, goods_fill),
        ('KM NHÀ CUNG CẤP', supplier_promotion, green_fill),
        ('HÀNG NHẬP SAU KM', goods_receipt_expense, goods_fill),
        ('TỔNG CHI PHÍ', total_expense, red_fill),
    ]:
        for col, val in enumerate(['', label, '', '', '', '', amt], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(bold=True)
            c.fill = fill
            c.border = thin
            if col == 7:
                c.number_format = mfmt
        row += 1
    for col, val in enumerate(['', 'LÃI/LỖ', '', '', '', '', net], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color='006600' if net >= 0 else 'CC0000')
        c.fill = tfill
        c.border = thin
        if col == 7:
            c.number_format = mfmt

    for i, w in enumerate([6, 13, 18, 12, 24, 38, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="BC_Thu_chi_{from_date}_{to_date}.xlsx"'
    wb.save(response)
    return response
