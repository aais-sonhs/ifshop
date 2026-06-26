#!/usr/bin/env python3
import argparse
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_FLOOR
from pathlib import Path

import openpyxl


HEADER_ALIASES = {
    'Tên khách hàng *': 'name',
    'Mã khách hàng': 'code',
    'Mã nhóm khách hàng': 'group_code',
    'Áp dụng ưu đãi': 'promotion_policy',
    'Email': 'email',
    'Điện thoại': 'phone',
    'Ngày sinh': 'date_of_birth',
    'Giới tính': 'gender',
    'Người liên hệ': 'contact_person',
    'Người liên hệ - SĐT': 'contact_phone',
    'Người liên hệ - Email': 'contact_email',
    'Địa chỉ': 'address',
    'Tỉnh thành': 'province',
    'Quận huyện': 'district',
    'Phường xã': 'ward',
    'Website': 'website',
    'Fax': 'fax',
    'Mã số thuế': 'tax_code',
    'Mô tả': 'note',
    'Chính sách giá mặc định': 'default_price_policy',
    'Chiết khấu mặc định (%)': 'default_discount_percent',
    'Phương thức thanh toán mặc định': 'default_payment_method',
    'Nợ hiện tại': 'total_debt',
    'Tổng chi tiêu': 'total_purchased',
    'SL đơn hàng': 'order_count',
    'Tổng SL sản phẩm đã mua': 'total_product_quantity',
    'Tổng SL sản phẩm hoàn trả': 'total_returned_product_quantity',
    'Ngày mua cuối cùng': 'last_purchase_at',
    'Điểm hiện tại': 'points',
    'Hạng thẻ hiện tại': 'membership_level',
    'Ngày hết hạn thẻ': 'membership_expiry_date',
    'Giá trị còn lại để lên hạng': 'amount_to_next_membership',
}

GROUP_NAME_MAP = {
    'BANLE': 'Bán lẻ',
    'BANBUON': 'Bán buôn',
}

GENDER_MAP = {
    'nam': 1,
    'nu': 2,
    'nữ': 2,
    'khac': 0,
    'khác': 0,
}

MEMBERSHIP_MAP = {
    'thuong': 0,
    'thường': 0,
    'bac': 1,
    'bạc': 1,
    'vang': 2,
    'vàng': 2,
    'bach kim': 3,
    'bạch kim': 3,
    'kim cuong': 4,
    'kim cương': 4,
}

COMPANY_KEYWORDS = (
    'cty', 'công ty', 'cong ty', 'tnhh', 'ctcp', 'cổ phần', 'co phan',
    'jsc', 'corp', 'company', 'doanh nghiệp', 'doanh nghiep',
)
HOUSEHOLD_KEYWORDS = ('hộ kinh doanh', 'ho kinh doanh', 'hkd')


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Import customer Excel file into the Django system with preview/apply modes.'
    )
    parser.add_argument('--excel', required=True, help='Path to the customer Excel file')
    parser.add_argument('--store-id', type=int, help='Store id for imported customers')
    parser.add_argument('--created-by-id', type=int, help='User id recorded as creator')
    parser.add_argument('--apply', action='store_true', help='Actually write data to DB')
    parser.add_argument(
        '--report-file',
        default='logs/import_customers_excel_report.json',
        help='Path to write import report JSON',
    )
    return parser.parse_args()


def setup_django():
    project_root = Path(__file__).resolve().parents[1]
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
    import django
    django.setup()


def clean_text(value) -> str:
    return str(value if value is not None else '').strip()


def normalize_text(value: str) -> str:
    return re.sub(r'\s+', ' ', clean_text(value).lower())


def parse_decimal(value, default='0', integer=False) -> Decimal:
    if value in (None, ''):
        return Decimal(default)
    if isinstance(value, Decimal):
        parsed = value
    elif isinstance(value, (int, float)):
        parsed = Decimal(str(value))
    else:
        raw = clean_text(value)
        raw = raw.replace('\xa0', '').replace(' ', '')
        raw = raw.replace('%', '').replace('₫', '').replace('đ', '').replace('Đ', '')
        raw = re.sub(r'[^0-9,.\-]', '', raw)
        if raw in ('', '-'):
            return Decimal(default)
        if raw.count(',') > 1 or raw.count('.') > 1:
            raw = raw.replace(',', '').replace('.', '')
        elif ',' in raw and '.' in raw:
            raw = raw.replace(',', '').replace('.', '')
        elif ',' in raw:
            left, right = raw.rsplit(',', 1)
            raw = left + right if len(right) == 3 else left + '.' + right
        elif '.' in raw:
            left, right = raw.rsplit('.', 1)
            raw = left + right if len(right) == 3 else raw
        try:
            parsed = Decimal(raw)
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)
    if integer:
        return parsed.quantize(Decimal('1'), rounding=ROUND_FLOOR)
    return parsed


def parse_int(value, default=0) -> int:
    try:
        return int(parse_decimal(value, default=str(default), integer=True))
    except (TypeError, ValueError):
        return default


def parse_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    raw = clean_text(value)
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime_value(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value
    raw = clean_text(value)
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    date_value = parse_date(value)
    if date_value:
        return datetime.combine(date_value, datetime.min.time())
    return None


def parse_gender(value) -> int:
    return GENDER_MAP.get(normalize_text(clean_text(value)), 0)


def parse_membership(value) -> int:
    return MEMBERSHIP_MAP.get(normalize_text(clean_text(value)), 0)


def infer_customer_type(name: str, tax_code: str) -> int:
    normalized_name = normalize_text(name)
    if any(token in normalized_name for token in HOUSEHOLD_KEYWORDS):
        return 3
    if tax_code or any(token in normalized_name for token in COMPANY_KEYWORDS):
        return 2
    return 1


def find_header(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {}
    for index, value in enumerate(header_row, 1):
        key = HEADER_ALIASES.get(clean_text(value))
        if key:
            headers[key] = index
    required = {'name', 'code'}
    if not required.issubset(headers):
        raise ValueError('Không tìm thấy cột bắt buộc "Tên khách hàng *" và "Mã khách hàng".')
    return headers


def row_value(row, headers, key):
    col = headers.get(key)
    if not col or col > len(row):
        return None
    return row[col - 1]


def load_rows(excel_path: Path):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = find_header(sheet)

    rows = []
    skipped = []
    duplicate_codes = Counter()
    seen_codes = set()

    for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, '') for value in row):
            continue

        name = clean_text(row_value(row, headers, 'name'))
        code = clean_text(row_value(row, headers, 'code'))
        if not name and not code:
            skipped.append({'excel_row': excel_row, 'reason': 'missing_name_and_code'})
            continue
        if not name or not code:
            skipped.append({'excel_row': excel_row, 'reason': 'missing_required_field', 'name': name, 'code': code})
            continue
        if code in seen_codes:
            duplicate_codes[code] += 1
        else:
            seen_codes.add(code)

        tax_code = clean_text(row_value(row, headers, 'tax_code'))
        customer_type = infer_customer_type(name, tax_code)
        address = clean_text(row_value(row, headers, 'address'))
        rows.append({
            'excel_row': excel_row,
            'code': code[:50],
            'name': name[:255],
            'customer_type': customer_type,
            'phone': clean_text(row_value(row, headers, 'phone'))[:20] or None,
            'email': clean_text(row_value(row, headers, 'email'))[:254] or None,
            'address': address or None,
            'company': name[:255] if customer_type in (2, 3) else None,
            'id_number': None,
            'tax_code': tax_code[:20] or None,
            'company_address': address if customer_type == 2 else None,
            'owner_tax_code': None,
            'promotion_policy': clean_text(row_value(row, headers, 'promotion_policy'))[:100] or None,
            'contact_person': clean_text(row_value(row, headers, 'contact_person'))[:255] or None,
            'contact_phone': clean_text(row_value(row, headers, 'contact_phone'))[:20] or None,
            'contact_email': clean_text(row_value(row, headers, 'contact_email'))[:254] or None,
            'province': clean_text(row_value(row, headers, 'province'))[:100] or None,
            'district': clean_text(row_value(row, headers, 'district'))[:100] or None,
            'ward': clean_text(row_value(row, headers, 'ward'))[:100] or None,
            'website': clean_text(row_value(row, headers, 'website'))[:255] or None,
            'fax': clean_text(row_value(row, headers, 'fax'))[:50] or None,
            'default_price_policy': clean_text(row_value(row, headers, 'default_price_policy'))[:255] or None,
            'default_discount_percent': parse_decimal(row_value(row, headers, 'default_discount_percent')),
            'default_payment_method': clean_text(row_value(row, headers, 'default_payment_method'))[:255] or None,
            'total_debt': parse_decimal(row_value(row, headers, 'total_debt'), integer=True),
            'total_purchased': parse_decimal(row_value(row, headers, 'total_purchased'), integer=True),
            'order_count': parse_int(row_value(row, headers, 'order_count')),
            'total_product_quantity': parse_decimal(row_value(row, headers, 'total_product_quantity')),
            'total_returned_product_quantity': parse_decimal(row_value(row, headers, 'total_returned_product_quantity')),
            'last_purchase_at': parse_datetime_value(row_value(row, headers, 'last_purchase_at')),
            'points': parse_int(row_value(row, headers, 'points')),
            'membership_level': parse_membership(row_value(row, headers, 'membership_level')),
            'membership_expiry_date': parse_date(row_value(row, headers, 'membership_expiry_date')),
            'amount_to_next_membership': parse_decimal(row_value(row, headers, 'amount_to_next_membership'), integer=True),
            'date_of_birth': parse_date(row_value(row, headers, 'date_of_birth')),
            'gender': parse_gender(row_value(row, headers, 'gender')),
            'group_code': clean_text(row_value(row, headers, 'group_code'))[:50] or None,
            'note': clean_text(row_value(row, headers, 'note')) or None,
        })

    return {
        'sheet': sheet.title,
        'rows': rows,
        'skipped': skipped,
        'duplicate_codes': [{'code': code, 'extra_occurrences': count} for code, count in duplicate_codes.items()],
    }


def resolve_defaults(args):
    from django.contrib.auth.models import User
    from system_management.models import Store

    store = Store.objects.filter(id=args.store_id).first() if args.store_id else Store.objects.order_by('id').first()
    if not store:
        raise SystemExit('Không tìm thấy store để import khách hàng.')

    creator = User.objects.filter(id=args.created_by_id).first() if args.created_by_id else User.objects.order_by('id').first()
    if not creator:
        raise SystemExit('Không tìm thấy user để gán created_by cho khách hàng import.')
    return store, creator


def resolve_group(group_code, groups_created):
    from customers.models import CustomerGroup

    if not group_code:
        return None
    group = CustomerGroup.objects.filter(code=group_code).first()
    if group:
        if not group.is_active:
            group.is_active = True
            group.save(update_fields=['is_active'])
        return group
    group = CustomerGroup.objects.create(
        code=group_code,
        name=GROUP_NAME_MAP.get(group_code, group_code),
        is_active=True,
    )
    groups_created.append(group.code)
    return group


def upsert_customers(rows, store, creator):
    from django.db import transaction
    from customers.models import Customer

    summary = {
        'created': 0,
        'updated': 0,
        'restored': 0,
        'groups_created': [],
        'created_codes': [],
        'updated_codes': [],
    }

    with transaction.atomic():
        for row in rows:
            customer = Customer.all_objects.filter(code=row['code']).first()
            created = customer is None
            restored = False
            if created:
                customer = Customer(code=row['code'], created_by=creator, store=store)
            elif customer.is_deleted:
                customer.is_deleted = False
                customer.deleted_at = None
                restored = True

            group = resolve_group(row['group_code'], summary['groups_created'])

            for field in (
                'name', 'customer_type', 'phone', 'email', 'address', 'company',
                'id_number', 'tax_code', 'company_address', 'owner_tax_code',
                'promotion_policy', 'contact_person', 'contact_phone', 'contact_email',
                'province', 'district', 'ward', 'website', 'fax',
                'default_price_policy', 'default_discount_percent',
                'default_payment_method', 'total_debt', 'total_purchased',
                'order_count', 'total_product_quantity',
                'total_returned_product_quantity', 'last_purchase_at', 'points',
                'membership_level', 'membership_expiry_date',
                'amount_to_next_membership', 'date_of_birth', 'gender', 'note',
            ):
                setattr(customer, field, row[field])
            customer.group = group
            customer.store = customer.store or store
            customer.created_by = customer.created_by or creator
            customer.imported_legacy_metrics = True
            customer.is_active = True
            customer.save()

            if created:
                summary['created'] += 1
                if len(summary['created_codes']) < 50:
                    summary['created_codes'].append(customer.code)
            else:
                summary['updated'] += 1
                if len(summary['updated_codes']) < 50:
                    summary['updated_codes'].append(customer.code)
            if restored:
                summary['restored'] += 1

    summary['groups_created'] = sorted(set(summary['groups_created']))
    return summary


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel).resolve()
    report_path = Path(args.report_file).resolve()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        raise SystemExit(f'Excel file not found: {excel_path}')

    setup_django()
    loaded = load_rows(excel_path)
    rows = loaded['rows']
    report = {
        'excel_file': str(excel_path),
        'sheet': loaded['sheet'],
        'mode': 'apply' if args.apply else 'preview',
        'total_valid_rows': len(rows),
        'skipped_rows': len(loaded['skipped']),
        'skipped_samples': loaded['skipped'][:50],
        'duplicate_codes': loaded['duplicate_codes'],
    }

    if loaded['duplicate_codes']:
        report['status'] = 'error'
        report['message'] = 'File có mã khách hàng trùng, dừng import để tránh ghi sai dữ liệu.'
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
        print(report['message'])
        print(f'Chi tiết report: {report_path}')
        return 1

    from customers.models import Customer
    existing_codes = set(Customer.all_objects.filter(code__in=[row['code'] for row in rows]).values_list('code', flat=True))
    group_codes = sorted({row['group_code'] for row in rows if row['group_code']})
    report['existing_code_count'] = len(existing_codes)
    report['group_codes_in_file'] = group_codes

    if not args.apply:
        report['status'] = 'ok'
        report['message'] = (
            f'Preview: {len(rows)} dòng hợp lệ, '
            f'{len(existing_codes)} mã sẽ update, '
            f'{len(rows) - len(existing_codes)} mã sẽ tạo mới.'
        )
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
        print(report['message'])
        print(f'Chi tiết report: {report_path}')
        return 0

    store, creator = resolve_defaults(args)
    summary = upsert_customers(rows, store, creator)
    report.update({
        'status': 'ok',
        'store_id': store.id,
        'created_by_id': creator.id,
        'created': summary['created'],
        'updated': summary['updated'],
        'restored': summary['restored'],
        'groups_created': summary['groups_created'],
        'created_codes': summary['created_codes'],
        'updated_codes': summary['updated_codes'],
    })
    report['message'] = (
        f'Đã import {summary["created"] + summary["updated"]} khách hàng '
        f'({summary["created"]} tạo mới, {summary["updated"]} cập nhật).'
    )
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    print(report['message'])
    print(f'Chi tiết report: {report_path}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
