#!/usr/bin/env python3
"""
Gắn nhà cung cấp cho sản phẩm từ file Excel cũ.

File Excel cần có:
- Cột mã sản phẩm: Mã SP / Mã sản phẩm / Mã hàng / SKU / Code.
- Cột nhà cung cấp: Nhãn hiệu / NCC / Nhà cung cấp / Supplier.

Mặc định script chỉ xem trước, không ghi database. Thêm --apply để cập nhật thật.
Sản phẩm được đối chiếu bằng mã sản phẩm, không dùng ID nên có thể chạy trên
database khác.
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl


HEADER_ALIASES = {
    'ma sp': 'product_code',
    'ma san pham': 'product_code',
    'ma hang': 'product_code',
    'sku': 'product_code',
    'code': 'product_code',
    'nhan hieu': 'supplier_name',
    'ncc': 'supplier_name',
    'nha cung cap': 'supplier_name',
    'supplier': 'supplier_name',
}


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            'Gắn NCC cho sản phẩm từ cột Nhãn hiệu trong Excel. '
            'Mặc định chỉ preview; thêm --apply để ghi database.'
        ),
    )
    parser.add_argument('--excel', required=True, help='Đường dẫn file .xlsx hoặc .xlsm')
    parser.add_argument('--sheet', help='Tên sheet; mặc định dùng sheet đầu tiên')
    parser.add_argument(
        '--store-id',
        type=int,
        help='Chỉ cập nhật sản phẩm thuộc cửa hàng này (khuyến nghị nếu hệ thống có nhiều cửa hàng)',
    )
    parser.add_argument(
        '--created-by-id',
        type=int,
        help='User ghi nhận là người tạo NCC mới; nếu bỏ trống sẽ dùng người tạo sản phẩm',
    )
    parser.add_argument('--apply', action='store_true', help='Cập nhật thật vào database')
    parser.add_argument('--report-file', help='Tùy chọn: đường dẫn lưu báo cáo JSON')
    return parser.parse_args()


def setup_django():
    project_root = Path(__file__).resolve().parents[1]
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
    import django

    django.setup()


def clean_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(value):
    text = clean_text(value).lower().replace('đ', 'd')
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return re.sub(r'[^a-z0-9]+', ' ', text).strip()


def find_header(sheet, max_rows=20):
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        headers = {}
        for column_number, value in enumerate(row, start=1):
            key = HEADER_ALIASES.get(normalize_text(value))
            if key and key not in headers:
                headers[key] = column_number
        if {'product_code', 'supplier_name'}.issubset(headers):
            return row_number, headers
    raise ValueError(
        'Không tìm thấy dòng tiêu đề có cả cột Mã SP và Nhãn hiệu/NCC trong 20 dòng đầu.'
    )


def row_value(row, headers, key):
    column_number = headers.get(key)
    if not column_number or column_number > len(row):
        return None
    return row[column_number - 1]


def load_rows(excel_path, sheet_name=None):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f'Không có sheet "{sheet_name}". Các sheet hiện có: {", ".join(workbook.sheetnames)}'
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    header_row, headers = find_header(sheet)
    rows = []
    skipped = []
    code_counts = Counter()

    for excel_row, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        code = clean_text(row_value(row, headers, 'product_code'))
        supplier_name = clean_text(row_value(row, headers, 'supplier_name'))
        if not code and not supplier_name:
            continue
        if not code:
            skipped.append({
                'excel_row': excel_row,
                'reason': 'missing_product_code',
                'supplier_name': supplier_name,
            })
            continue
        if not supplier_name:
            skipped.append({
                'excel_row': excel_row,
                'reason': 'missing_supplier_name',
                'product_code': code,
            })
            continue

        code_key = code.casefold()
        code_counts[code_key] += 1
        rows.append({
            'excel_row': excel_row,
            'product_code': code,
            'product_code_key': code_key,
            'supplier_name': supplier_name[:255],
        })

    duplicate_keys = {key for key, count in code_counts.items() if count > 1}
    duplicates = [
        {
            'excel_row': row['excel_row'],
            'product_code': row['product_code'],
            'supplier_name': row['supplier_name'],
        }
        for row in rows
        if row['product_code_key'] in duplicate_keys
    ]
    return {
        'sheet': sheet.title,
        'header_row': header_row,
        'rows': rows,
        'skipped': skipped,
        'duplicates': duplicates,
    }


def validate_scope(store_id=None, created_by_id=None):
    from django.contrib.auth.models import User
    from system_management.models import Store

    store = None
    if store_id is not None:
        store = Store.objects.filter(id=store_id).first()
        if not store:
            raise ValueError(f'Không tìm thấy cửa hàng có ID {store_id}.')

    creator = None
    if created_by_id is not None:
        creator = User.objects.filter(id=created_by_id).first()
        if not creator:
            raise ValueError(f'Không tìm thấy user có ID {created_by_id}.')
    return store, creator


def product_queryset(store=None, include_deleted=False):
    from products.models import Product

    queryset = Product.all_objects.all() if include_deleted else Product.objects.all()
    if store is not None:
        queryset = queryset.filter(store=store)
    return queryset


def find_product(code, store=None, for_update=False):
    queryset = product_queryset(store=store)
    if for_update:
        queryset = queryset.select_for_update()
    return queryset.filter(code__iexact=code).order_by('id').first()


def find_supplier(name):
    from products.models import Supplier

    return Supplier.all_objects.filter(name__iexact=name).order_by('is_deleted', 'id').first()


def next_supplier_code():
    from products.models import Supplier

    max_number = 0
    for code in Supplier.all_objects.filter(code__istartswith='NCC').values_list('code', flat=True):
        match = re.match(r'^NCC-?(\d+)$', code or '', re.IGNORECASE)
        if match:
            max_number = max(max_number, int(match.group(1)))

    next_number = max_number + 1
    while True:
        candidate = f'NCC{next_number:03d}'
        if not Supplier.all_objects.filter(code__iexact=candidate).exists():
            return candidate
        next_number += 1


def classify_rows(rows, store=None):
    planned = []
    missing = []
    unchanged = []
    supplier_names_to_create = set()

    for row in rows:
        product = find_product(row['product_code'], store=store)
        if not product:
            any_product = product_queryset(include_deleted=True).filter(
                code__iexact=row['product_code'],
            ).order_by('id').first()
            if any_product and any_product.is_deleted:
                reason = 'product_deleted'
            elif any_product and store is not None and any_product.store_id != store.id:
                reason = 'product_outside_store'
            else:
                reason = 'product_not_found'
            missing.append({
                'excel_row': row['excel_row'],
                'product_code': row['product_code'],
                'supplier_name': row['supplier_name'],
                'reason': reason,
            })
            continue

        supplier = find_supplier(row['supplier_name'])
        current_supplier_name = product.supplier.name if product.supplier_id else ''
        same_supplier = supplier is not None and product.supplier_id == supplier.id
        supplier_needs_restore = bool(supplier and supplier.is_deleted)
        supplier_needs_activate = bool(supplier and not supplier.is_active)

        item = {
            'excel_row': row['excel_row'],
            'product_id': product.id,
            'product_code': product.code,
            'product_name': product.name,
            'old_supplier': current_supplier_name,
            'new_supplier': row['supplier_name'],
            'supplier_will_be_created': supplier is None,
            'supplier_will_be_restored': supplier_needs_restore,
            'supplier_will_be_activated': supplier_needs_activate,
        }
        if supplier is None:
            supplier_names_to_create.add(normalize_text(row['supplier_name']))

        if same_supplier and not supplier_needs_restore and not supplier_needs_activate:
            unchanged.append(item)
        else:
            planned.append(item)

    return {
        'planned': planned,
        'missing': missing,
        'unchanged': unchanged,
        'supplier_create_count': len(supplier_names_to_create),
    }


def get_or_create_supplier(name, creator=None):
    from products.models import Supplier

    supplier = find_supplier(name)
    created = False
    restored = False
    activated = False
    if supplier is None:
        supplier = Supplier.objects.create(
            code=next_supplier_code(),
            name=name,
            created_by=creator,
            is_active=True,
        )
        created = True
    else:
        update_fields = []
        if supplier.is_deleted:
            supplier.is_deleted = False
            supplier.deleted_at = None
            update_fields.extend(['is_deleted', 'deleted_at'])
            restored = True
        if not supplier.is_active:
            supplier.is_active = True
            update_fields.append('is_active')
            activated = True
        if update_fields:
            supplier.save(update_fields=update_fields)
    return supplier, created, restored, activated


def apply_rows(rows, store=None, creator=None):
    from django.db import transaction

    summary = {
        'updated': 0,
        'unchanged': 0,
        'missing': 0,
        'suppliers_created': 0,
        'suppliers_restored': 0,
        'suppliers_activated': 0,
        'updated_samples': [],
        'missing_samples': [],
    }

    with transaction.atomic():
        for row in rows:
            product = find_product(row['product_code'], store=store, for_update=True)
            if not product:
                summary['missing'] += 1
                if len(summary['missing_samples']) < 100:
                    summary['missing_samples'].append({
                        'excel_row': row['excel_row'],
                        'product_code': row['product_code'],
                        'supplier_name': row['supplier_name'],
                    })
                continue

            supplier, created, restored, activated = get_or_create_supplier(
                row['supplier_name'],
                creator=creator or product.created_by,
            )
            summary['suppliers_created'] += int(created)
            summary['suppliers_restored'] += int(restored)
            summary['suppliers_activated'] += int(activated)

            old_supplier = product.supplier.name if product.supplier_id else ''
            if product.supplier_id == supplier.id:
                summary['unchanged'] += 1
                continue

            product.supplier = supplier
            product.save(update_fields=['supplier', 'updated_at'])
            summary['updated'] += 1
            if len(summary['updated_samples']) < 100:
                summary['updated_samples'].append({
                    'product_code': product.code,
                    'product_name': product.name,
                    'old_supplier': old_supplier,
                    'new_supplier': supplier.name,
                })
    return summary


def write_report(report, report_file=None):
    if not report_file:
        return
    report_path = Path(report_file).resolve()
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    print(f'Báo cáo chi tiết: {report_path}')


def main():
    args = parse_args()
    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        raise SystemExit(f'Không tìm thấy file Excel: {excel_path}')
    if excel_path.suffix.lower() not in {'.xlsx', '.xlsm'}:
        raise SystemExit('Chỉ hỗ trợ file .xlsx hoặc .xlsm.')

    try:
        loaded = load_rows(excel_path, sheet_name=args.sheet)
    except Exception as exc:
        raise SystemExit(f'Không đọc được file Excel: {exc}') from exc

    base_report = {
        'status': 'ok',
        'mode': 'apply' if args.apply else 'preview',
        'excel_file': str(excel_path),
        'sheet': loaded['sheet'],
        'header_row': loaded['header_row'],
        'valid_rows': len(loaded['rows']),
        'skipped_rows': len(loaded['skipped']),
        'skipped_samples': loaded['skipped'][:100],
        'duplicate_rows': loaded['duplicates'][:100],
        'store_id': args.store_id,
    }

    if loaded['duplicates']:
        base_report.update({
            'status': 'error',
            'message': 'File có Mã SP trùng. Đã dừng để tránh gắn sai NCC.',
        })
        write_report(base_report, args.report_file)
        print(base_report['message'])
        return 1

    setup_django()
    try:
        store, creator = validate_scope(args.store_id, args.created_by_id)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc

    if not args.apply:
        plan = classify_rows(loaded['rows'], store=store)
        base_report.update({
            'products_to_update': len(plan['planned']),
            'products_unchanged': len(plan['unchanged']),
            'products_not_found': len(plan['missing']),
            'suppliers_to_create': plan['supplier_create_count'],
            'update_samples': plan['planned'][:100],
            'missing_samples': plan['missing'][:100],
            'message': (
                f'PREVIEW: {len(plan["planned"])} sản phẩm sẽ cập nhật, '
                f'{len(plan["unchanged"])} không đổi, '
                f'{len(plan["missing"])} không tìm thấy; '
                f'{plan["supplier_create_count"]} NCC sẽ tạo mới. '
                'Database chưa được thay đổi.'
            ),
        })
        write_report(base_report, args.report_file)
        print(base_report['message'])
        return 0

    applied = apply_rows(loaded['rows'], store=store, creator=creator)
    base_report.update(applied)
    base_report['message'] = (
        f'ĐÃ CẬP NHẬT: {applied["updated"]} sản phẩm, '
        f'{applied["unchanged"]} không đổi, '
        f'{applied["missing"]} không tìm thấy; '
        f'{applied["suppliers_created"]} NCC được tạo mới.'
    )
    write_report(base_report, args.report_file)
    print(base_report['message'])
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
